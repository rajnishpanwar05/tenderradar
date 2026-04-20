"""
Portfolio similarity scoring for tender relevance ranking.

Builds a TF-IDF corpus from:
1) "PDS in Excel" (historical completed assignments)
2) "OPL" rows with Submitted status (recent wins/submissions)

Runtime output: 0-100 similarity score (higher = closer to historical portfolio).
"""

from __future__ import annotations

import json
import logging
import pickle
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import load_workbook

logger = logging.getLogger("tenderradar.portfolio_similarity")

try:
    import joblib as _joblib
except Exception:
    _joblib = None

try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
except Exception:
    TfidfVectorizer = None
    cosine_similarity = None

_BASE = Path(__file__).resolve().parent.parent
_SCORER_PATH = _BASE / "artifacts" / "portfolio_similarity.joblib"
_META_PATH = _BASE / "artifacts" / "portfolio_similarity_meta.json"

_instance: Optional["PortfolioSimilarityScorer"] = None


def get_portfolio_scorer() -> "PortfolioSimilarityScorer":
    global _instance
    if _instance is None:
        _instance = PortfolioSimilarityScorer()
        _instance.load()
    return _instance


def _save_obj(obj: Any, path: Path) -> None:
    if _joblib is not None:
        _joblib.dump(obj, path)
        return
    with path.open("wb") as fh:
        pickle.dump(obj, fh)


def _load_obj(path: Path) -> Any:
    if _joblib is not None:
        return _joblib.load(path)
    with path.open("rb") as fh:
        return pickle.load(fh)


def _norm_key(v: Any) -> str:
    return str(v or "").strip().lower()


def _find_col(columns: List[str], *aliases: str) -> Optional[str]:
    idx = {_norm_key(c): c for c in columns}
    for a in aliases:
        hit = idx.get(_norm_key(a))
        if hit:
            return hit
    return None


def _read_sheet_rows(path: str, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return []
    headers = [str(h).strip() if h is not None else "" for h in header]
    out: List[Dict[str, Any]] = []
    for row in rows:
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            rec[h] = row[i] if i < len(row) else None
        if any(v not in (None, "") for v in rec.values()):
            out.append(rec)
    wb.close()
    return out


class PortfolioSimilarityScorer:
    def __init__(self) -> None:
        self._ready = False
        self._meta: Dict[str, Any] = {}
        self._vectorizer = None
        self._matrix = None
        self._token_sets: List[set[str]] = []
        self._corpus: List[str] = []

    def build(self, workbook_path: str) -> Dict[str, Any]:
        corpus: List[str] = []

        pds_rows = _read_sheet_rows(workbook_path, "PDS in Excel")
        if pds_rows:
            cols = list(pds_rows[0].keys())
            col_assignment = _find_col(cols, "Assignment Name")
            col_country = _find_col(cols, "Country")
            col_narrative = _find_col(cols, "Narrative Description of Project:")
            col_services = _find_col(cols, "Description of Actual Services Provided By Your Staff:")
            col_employer = _find_col(cols, "Name of Employer")
            for r in pds_rows:
                parts = [
                    str(r.get(col_assignment) or ""),
                    str(r.get(col_country) or ""),
                    str(r.get(col_employer) or ""),
                    str(r.get(col_narrative) or ""),
                    str(r.get(col_services) or ""),
                ]
                txt = " ".join(p for p in parts if p).strip()
                if len(txt) > 20:
                    corpus.append(txt)

        opl_rows = _read_sheet_rows(workbook_path, "OPL")
        if opl_rows:
            cols = list(opl_rows[0].keys())
            col_status = _find_col(cols, "Status")
            col_title = _find_col(cols, "Project Name", "Project Name ")
            col_client = _find_col(cols, "Client")
            col_loc = _find_col(cols, "Location")
            for r in opl_rows:
                status = _norm_key(r.get(col_status))
                if "submit" not in status:
                    continue
                txt = " ".join(
                    [
                        str(r.get(col_title) or ""),
                        str(r.get(col_client) or ""),
                        str(r.get(col_loc) or ""),
                    ]
                ).strip()
                if len(txt) > 10:
                    corpus.append(txt)

        # Deduplicate while preserving order
        seen = set()
        unique_corpus = []
        for txt in corpus:
            key = " ".join(txt.lower().split())
            if key in seen:
                continue
            seen.add(key)
            unique_corpus.append(txt)

        if len(unique_corpus) < 20:
            raise ValueError(
                f"Insufficient portfolio corpus size: {len(unique_corpus)} entries"
            )

        self._corpus = unique_corpus
        self._token_sets = []

        if TfidfVectorizer is not None and cosine_similarity is not None:
            self._vectorizer = TfidfVectorizer(
                lowercase=True,
                strip_accents="unicode",
                stop_words="english",
                ngram_range=(1, 2),
                min_df=1,
                max_features=8000,
            )
            self._matrix = self._vectorizer.fit_transform(self._corpus)
        else:
            for txt in self._corpus:
                self._token_sets.append(set(_norm_key(txt).split()))

        self._ready = True
        self._meta = {
            "workbook_path": workbook_path,
            "corpus_size": len(self._corpus),
            "method": "tfidf" if self._vectorizer is not None else "token_jaccard",
        }
        logger.info(
            "[portfolio_similarity] Built corpus (%d entries) via %s",
            len(self._corpus),
            self._meta["method"],
        )
        return dict(self._meta)

    def score(self, tender: Dict[str, Any]) -> float:
        if not self._ready:
            return 50.0

        text = " ".join(
            [
                str(tender.get("title") or tender.get("Project Name") or ""),
                str(tender.get("description") or ""),
                str(tender.get("organization") or tender.get("Client") or ""),
                str(tender.get("country") or tender.get("Location") or ""),
                str(tender.get("sector") or ""),
            ]
        ).strip()
        if len(text) < 6:
            return 50.0

        try:
            if self._vectorizer is not None and self._matrix is not None:
                qv = self._vectorizer.transform([text])
                sims = cosine_similarity(qv, self._matrix)[0]
                if sims.size == 0:
                    return 50.0
                topk = np.sort(sims)[-5:]
                score = float(np.mean(topk)) * 100.0
                return round(max(0.0, min(100.0, score)), 1)

            # Fallback: token Jaccard max
            q = set(_norm_key(text).split())
            if not q:
                return 50.0
            best = 0.0
            for s in self._token_sets:
                if not s:
                    continue
                inter = len(q & s)
                union = len(q | s)
                if union == 0:
                    continue
                best = max(best, inter / union)
            return round(best * 100.0, 1)
        except Exception as exc:
            logger.debug("[portfolio_similarity] score() error: %s", exc)
            return 50.0

    def save(self) -> None:
        _SCORER_PATH.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "meta": self._meta,
            "ready": self._ready,
            "vectorizer": self._vectorizer,
            "matrix": self._matrix,
            "token_sets": self._token_sets,
            "corpus": self._corpus,
        }
        _save_obj(payload, _SCORER_PATH)
        _META_PATH.write_text(json.dumps(self._meta, indent=2))
        logger.info("[portfolio_similarity] Saved to %s", _SCORER_PATH)

    def load(self) -> bool:
        if not _SCORER_PATH.exists():
            logger.info(
                "[portfolio_similarity] No saved scorer at %s",
                _SCORER_PATH,
            )
            return False
        try:
            payload = _load_obj(_SCORER_PATH)
            self._meta = dict(payload.get("meta") or {})
            self._ready = bool(payload.get("ready"))
            self._vectorizer = payload.get("vectorizer")
            self._matrix = payload.get("matrix")
            self._token_sets = list(payload.get("token_sets") or [])
            self._corpus = list(payload.get("corpus") or [])
            logger.info(
                "[portfolio_similarity] Loaded scorer (%s, corpus=%d)",
                self._meta.get("method", "unknown"),
                self._meta.get("corpus_size", len(self._corpus)),
            )
            return True
        except Exception as exc:
            logger.warning("[portfolio_similarity] Load failed: %s", exc)
            return False

    @property
    def is_ready(self) -> bool:
        return self._ready

    @property
    def meta(self) -> Dict[str, Any]:
        return dict(self._meta)
