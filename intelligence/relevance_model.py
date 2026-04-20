"""
IDCG relevance model trained from the OPL workbook.

This module intentionally avoids pandas so training can run in lean
environments where only openpyxl + sklearn are installed.
"""

from __future__ import annotations

import json
import logging
import pickle
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
from openpyxl import load_workbook

_log = logging.getLogger("tenderradar.relevance_model")
try:
    import joblib as _joblib
except Exception:
    _joblib = None

# ── Paths ─────────────────────────────────────────────────────────────────────
_BASE = Path(__file__).resolve().parent.parent
_MODEL_PATH = _BASE / "artifacts" / "relevance_model.joblib"
_META_PATH = _BASE / "artifacts" / "relevance_model_meta.json"

# ── Singleton ─────────────────────────────────────────────────────────────────
_model_instance: Optional["IDCGRelevanceModel"] = None


def get_model() -> "IDCGRelevanceModel":
    """Return the singleton model, loading from disk if available."""
    global _model_instance
    if _model_instance is None:
        _model_instance = IDCGRelevanceModel()
        _model_instance.load()
    return _model_instance


def _save_obj(obj: Any, path: Path) -> None:
    if _joblib is not None:
        _joblib.dump(obj, path)
        return
    with path.open("wb") as fh:
        pickle.dump(obj, fh)


def _load_obj(path: Path) -> Any:
    if _joblib is not None:
        return _joblib.load(path)
    with path.open("rb") as fh:
        return pickle.load(fh)


# ── Feature engineering ───────────────────────────────────────────────────────

# Clients IDCG consistently submits to (high signal)
_HIGH_SUBMIT_CLIENTS = {
    "giz", "undp", "adb", "afd", "world bank", "wb", "unicef", "gef",
    "ifad", "unops", "unfpa", "usaid", "afdb", "iucn", "ilo",
    "dept of social justice", "niti aayog", "dmeo", "nabard",
    "tata trusts", "plan india", "oxfam", "save the children",
    "meghalaya basin", "tnc", "ci", "wwf", "care india",
}

# Geographies IDCG submits to frequently
_INDIA_REGIONS = {
    "india", "new delhi", "delhi", "mumbai", "bengaluru", "bangalore",
    "hyderabad", "chennai", "kolkata", "lucknow", "bhopal", "jaipur",
    "chandigarh", "gandhinagar", "bhubaneswar", "patna", "ranchi",
    "meghalaya", "assam", "odisha", "jharkhand", "rajasthan",
    "madhya pradesh", "uttarakhand", "uttar pradesh", "gujarat",
    "maharashtra", "karnataka", "kerala", "andhra pradesh", "telangana",
    "himachal pradesh", "bihar", "chhattisgarh", "tripura", "manipur",
    "nagaland", "mizoram", "arunachal",
}

# Core IDCG sectors by keyword clusters in title
_SECTOR_KEYWORDS = {
    "evaluation":    ["evaluat", "assessment", "review", "mid-term", "midterm",
                      "endline", "baseline", "impact", "third party", "tpm",
                      "third-party", "monitoring", "m&e", "mne"],
    "health":        ["health", "nutrition", "wash", "water sanitation", "hiv",
                      "aids", "malaria", "tb", "tuberculosis", "immuniz",
                      "maternal", "child health", "reproductive", "anemia",
                      "stunting", "wasting", "iycf", "sncu", "hepatitis"],
    "education":     ["education", "school", "learning", "ecce", "eccd",
                      "child development", "skill", "training", "vocational",
                      "literacy", "numeracy", "teacher", "curriculum"],
    "climate":       ["climate", "environment", "forest", "biodiversity",
                      "carbon", "green", "renewable", "energy", "solar",
                      "clean energy", "watershed", "land degradation",
                      "redd", "ecosystem", "nature based"],
    "livelihood":    ["livelihood", "agriculture", "rural", "farm", "crop",
                      "food security", "value chain", "market", "sme",
                      "enterprise", "income generation", "microfinance"],
    "governance":    ["governance", "policy", "institutional", "capacity",
                      "public finance", "pfm", "accountability", "civic",
                      "decentralization", "municipality", "panchayat"],
    "gender":        ["gender", "women", "girl", "inclusion", "disability",
                      "marginalized", "vulnerable", "social protection"],
    "research":      ["research", "study", "survey", "data collect",
                      "knowledge", "documentation", "evidence", "analysis",
                      "feasibility", "diagnostic", "landscape"],
}

# Tender type signals
_TYPE_MAP = {"rfp": 2, "rft": 2, "eoi": 1, "eoi/rfp": 1.5, "rfq": 0.5,
             "scqs": 1, "ppt": 1.5, "unknown": 1}


def _normalize_client(client: str) -> str:
    c = str(client or "").lower().strip()
    c = re.sub(r"\s+", " ", c)
    return c


def _geography_score(location: str) -> float:
    """1.0=India (highest IDCG submit rate), 0.7=South/SE Asia, 0.5=Africa/Global, 0.3=Other"""
    loc = str(location or "").lower()
    if any(r in loc for r in _INDIA_REGIONS):
        return 1.0
    if any(r in loc for r in ["nepal", "bangladesh", "sri lanka", "myanmar",
                               "cambodia", "vietnam", "laos", "thailand",
                               "indonesia", "philippines", "pakistan"]):
        return 0.7
    if any(r in loc for r in ["ethiopia", "kenya", "ghana", "nigeria",
                               "tanzania", "zambia", "zimbabwe", "malawi",
                               "mozambique", "senegal", "mali", "niger",
                               "guinea", "gambia", "botswana", "namibia",
                               "africa", "sub-saharan"]):
        return 0.65
    if any(r in loc for r in ["global", "international", "multi-country",
                               "regional", "asia", "south asia"]):
        return 0.55
    return 0.4


def _client_tier(client: str) -> float:
    """1.0=tier-1 multilateral, 0.7=bilateral/INGO, 0.4=PSU/state, 0.2=unknown"""
    c = _normalize_client(client)
    tier1 = {"undp", "world bank", "wb", "adb", "giz", "afd", "afdb", "ifc",
              "unicef", "unfpa", "unops", "ilo", "who", "fao", "wfp",
              "ifad", "gef", "usaid", "dfid", "fcdo", "eu", "iucn"}
    tier2 = {"tata trusts", "ford foundation", "bill melinda", "bmgf",
              "oxfam", "save the children", "plan india", "care india",
              "wwf", "tnc", "ci", "action aid", "terre des hommes",
              "welthungerhilfe", "hki", "fhi360", "psi", "jhpiego",
              "ipas", "msh", "path", "engenderhealth", "brac",
              "dept of social justice", "niti aayog", "dmeo", "nabard",
              "sidbi", "mnrega", "ministry", "government of"}
    if any(t in c for t in tier1):
        return 1.0
    if any(t in c for t in tier2):
        return 0.7
    if any(t in c for t in ["state", "district", "corporation", "authority",
                             "board", "commission", "municipal"]):
        return 0.4
    return 0.3


def _sector_score(title: str) -> Dict[str, float]:
    """Return binary sector flags for each sector."""
    t = str(title or "").lower()
    return {
        f"sec_{k}": float(any(kw in t for kw in kws))
        for k, kws in _SECTOR_KEYWORDS.items()
    }


def _title_features(title: str) -> Dict[str, float]:
    """Extract hand-crafted features from title text."""
    t = str(title or "").lower()
    return {
        "has_evaluation":   float(any(w in t for w in ["evaluat", "assess", "review", "apprais"])),
        "has_baseline":     float(any(w in t for w in ["baseline", "endline", "midline", "midterm"])),
        "has_survey":       float(any(w in t for w in ["survey", "data collect", "cati", "capi"])),
        "has_research":     float(any(w in t for w in ["research", "study", "analysis", "diagnostic"])),
        "has_capacity":     float(any(w in t for w in ["capacity", "training", "facilitat", "workshop"])),
        "has_consulting":   float(any(w in t for w in ["consult", "advisor", "technical assist", "ta "])),
        "has_goods_noise":  float(any(w in t for w in ["supply", "procurement", "equipment", "civil work",
                                                         "construction", "infrastructure", "road", "bridge",
                                                         "building", "hardware", "material", "goods"])),
        "title_len":        min(len(t.split()), 30) / 30.0,
    }


def extract_features(tender: Dict[str, Any]) -> np.ndarray:
    """
    Convert a tender dict into a fixed-length feature vector.
    Works at both training time (from OPL rows) and inference time (from DB rows).
    """
    title      = str(tender.get("title") or tender.get("Project Name") or "")
    client     = str(tender.get("organization") or tender.get("Client") or "")
    location   = str(tender.get("country") or tender.get("Location") or "")
    ttype      = str(tender.get("tender_type") or tender.get("Type") or "unknown").lower()
    portal     = str(tender.get("source_portal") or tender.get("Website") or "unknown").lower()
    sector     = str(tender.get("sector") or "")

    # Numeric features
    geo        = _geography_score(location)
    cli        = _client_tier(client)
    type_score = _TYPE_MAP.get(ttype.strip(), 1.0)

    # Portal tier — development portals score higher than state infrastructure
    portal_tier = 1.0 if any(p in portal for p in [
        "undp", "world bank", "wb", "adb", "giz", "ungm", "afdb", "afd",
        "iucn", "ilo", "dtvp", "ted", "ec", "gem", "devnet", "ngobox",
        "welthungerhilfe", "whh", "sam", "icfre", "jtds", "sidbi"
    ]) else 0.3

    title_feats  = _title_features(title)
    sector_feats = _sector_score(title + " " + sector)

    # Client known flag
    client_known = float(any(c in _normalize_client(client) for c in _HIGH_SUBMIT_CLIENTS))

    vec = [
        geo,
        cli,
        type_score / 2.0,
        portal_tier,
        client_known,
    ]
    vec += list(title_feats.values())
    vec += list(sector_feats.values())

    return np.array(vec, dtype=np.float32)


def _find_col(columns: List[str], *aliases: str) -> Optional[str]:
    """
    Resolve a column by case-insensitive alias matching.
    Returns original column key if found.
    """
    norm = {str(c).strip().lower(): c for c in columns}
    for a in aliases:
        k = norm.get(a.strip().lower())
        if k is not None:
            return k
    return None


def _read_opl_rows(opl_path: str, sheet_name: str = "OPL") -> List[Dict[str, Any]]:
    """
    Read workbook rows from OPL sheet as dictionaries keyed by header row.
    """
    wb = load_workbook(opl_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    out: List[Dict[str, Any]] = []
    for row in rows_iter:
        rec = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            rec[h] = row[i] if i < len(row) else None
        if any(v not in (None, "") for v in rec.values()):
            out.append(rec)
    wb.close()
    return out


# ── Model class ───────────────────────────────────────────────────────────────

class IDCGRelevanceModel:
    """
    Gradient Boosting classifier trained on IDCG OPL decisions.
    Outputs 0–100 relevance score.
    """

    def __init__(self) -> None:
        self._clf   = None
        self._ready = False
        self._meta: Dict[str, Any] = {}

    # ── Training ──────────────────────────────────────────────────────────────

    def train(self, opl_path: str) -> Dict[str, Any]:
        """
        Train on OPL sheet. Returns training metrics.
        """
        from sklearn.ensemble import GradientBoostingClassifier
        from sklearn.model_selection import cross_val_score, StratifiedKFold
        _log.info("[relevance_model] Loading OPL from %s", opl_path)
        raw_rows = _read_opl_rows(opl_path, sheet_name="OPL")
        if not raw_rows:
            raise ValueError("No rows found in OPL sheet")

        cols = list(raw_rows[0].keys())
        col_status = _find_col(cols, "Status")
        col_title = _find_col(cols, "Project Name", "Project Name ")
        col_client = _find_col(cols, "Client")
        col_location = _find_col(cols, "Location")
        col_type = _find_col(cols, "Type")
        col_website = _find_col(cols, "Website")

        if not col_status:
            raise ValueError("Could not find 'Status' column in OPL sheet")
        if not col_title:
            raise ValueError("Could not find 'Project Name' column in OPL sheet")

        X_rows: List[np.ndarray] = []
        y_list: List[int] = []

        for r in raw_rows:
            status_s = str(r.get(col_status) or "").strip().lower()
            if "submit" in status_s:
                label = 1
            elif "drop" in status_s:
                label = 0
            else:
                continue

            tender = {
                "title": r.get(col_title) or "",
                "organization": r.get(col_client) or "",
                "country": r.get(col_location) or "",
                "tender_type": r.get(col_type) or "unknown",
                "source_portal": r.get(col_website) or "unknown",
            }
            X_rows.append(extract_features(tender))
            y_list.append(label)

        if len(y_list) < 40:
            raise ValueError(f"Insufficient labeled rows for training: {len(y_list)}")

        X = np.vstack(X_rows)
        y = np.array(y_list, dtype=np.int32)

        submitted_count = int(y.sum())
        dropped_count = int((y == 0).sum())
        _log.info(
            "[relevance_model] Training set: %d rows (%d submitted, %d dropped)",
            len(y_list),
            submitted_count,
            dropped_count,
        )

        # Gradient Boosting — handles mixed features well, robust on small datasets
        clf = GradientBoostingClassifier(
            n_estimators=200,
            learning_rate=0.08,
            max_depth=4,
            min_samples_leaf=5,
            subsample=0.8,
            random_state=42,
        )

        # Cross-validation for honest accuracy estimate
        cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
        cv_scores = cross_val_score(clf, X, y, cv=cv, scoring="roc_auc")

        # Train on full dataset
        clf.fit(X, y)
        self._clf   = clf
        self._ready = True

        self._meta = {
            "training_rows": int(len(y_list)),
            "submitted_count": submitted_count,
            "dropped_count": dropped_count,
            "cv_roc_auc_mean": round(float(cv_scores.mean()), 4),
            "cv_roc_auc_std": round(float(cv_scores.std()), 4),
            "opl_path": opl_path,
            "feature_count": int(X.shape[1]),
            "weights": {"ml": 0.50, "portfolio": 0.30, "keywords": 0.20},
        }

        _log.info(
            "[relevance_model] Trained. CV AUC=%.4f ± %.4f",
            cv_scores.mean(), cv_scores.std()
        )
        return self._meta

    # ── Inference ─────────────────────────────────────────────────────────────

    def score(self, tender: Dict[str, Any]) -> float:
        """
        Return 0–100 relevance score for a single tender.
        Falls back to 50 (neutral) if model not ready.
        """
        if not self._ready or self._clf is None:
            return 50.0
        try:
            x = extract_features(tender).reshape(1, -1)
            prob = float(self._clf.predict_proba(x)[0][1])
            return round(prob * 100, 1)
        except Exception as exc:
            _log.debug("[relevance_model] score() error: %s", exc)
            return 50.0

    def score_batch(self, tenders: list) -> list:
        """Score a list of tenders. Returns list of floats."""
        if not self._ready or self._clf is None:
            return [50.0] * len(tenders)
        try:
            X = np.vstack([extract_features(t) for t in tenders])
            probs = self._clf.predict_proba(X)[:, 1]
            return [round(float(p) * 100, 1) for p in probs]
        except Exception as exc:
            _log.debug("[relevance_model] score_batch() error: %s", exc)
            return [50.0] * len(tenders)

    # ── Persistence ───────────────────────────────────────────────────────────

    def save(self) -> None:
        _MODEL_PATH.parent.mkdir(parents=True, exist_ok=True)
        _save_obj(self._clf, _MODEL_PATH)
        _META_PATH.write_text(json.dumps(self._meta, indent=2))
        _log.info("[relevance_model] Saved to %s", _MODEL_PATH)

    def load(self) -> bool:
        if not _MODEL_PATH.exists():
            _log.info("[relevance_model] No saved model at %s — run train_relevance_model.py first", _MODEL_PATH)
            return False
        try:
            self._clf = _load_obj(_MODEL_PATH)
            self._ready = True
            if _META_PATH.exists():
                self._meta = json.loads(_META_PATH.read_text())
            _log.info(
                "[relevance_model] Loaded. CV AUC=%.4f (trained on %d rows)",
                self._meta.get("cv_roc_auc_mean", 0),
                self._meta.get("training_rows", 0),
            )
            return True
        except Exception as exc:
            _log.warning("[relevance_model] Load failed: %s", exc)
            return False

    @property
    def is_ready(self) -> bool:
        return self._ready

    @property
    def meta(self) -> Dict[str, Any]:
        return self._meta
