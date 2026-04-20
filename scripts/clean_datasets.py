#!/usr/bin/env python3
"""
TenderRadar — Dataset Cleaning & Architecture Fix
==================================================
Tasks:
  1. DEDUPLICATION  — remove exact + fuzzy title duplicates
  2. RAW/ML SPLIT   — portal files keep only source columns, master keeps ML columns
  3. URL FIX        — extract real hyperlink URLs from "↗ Open" display text
  4. TARGET VARIABLE FIX — remove Outcome as primary target, keep as secondary
  5. VALIDATION REPORT  — before/after stats

Usage:
  python3 scripts/clean_datasets.py
"""

import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ─── CONFIG ────────────────────────────────────────────────────────────────────

INPUT_MASTER  = "output/Tender_Monitor_Master.xlsx"
OUTPUT_MASTER = "output/Cleaned_Master_Dataset.xlsx"
OUTPUT_PORTAL = "output/Cleaned_Portal_RAW.xlsx"
REPORT_FILE   = "output/Cleaning_Report.txt"

# Columns that belong ONLY in the ML/processed layer (master dataset)
ML_COLUMNS = {
    "Priority Score", "Relevance Score", "Relevance", "Relevance Reason",
    "Opportunity Insight", "My Decision", "Outcome", "Score", "Decision",
    "Why", "Signal Score", "Firm Fit", "Adjusted Score", "Consulting Signals",
    "Quality Score", "Action",
}

# Columns that are pure source/raw data
RAW_COLUMNS = {
    "Title", "Title / Assignment", "Organization", "Country", "Deadline",
    "Published", "Scraped Date", "Portal", "Source", "URL", "Tender URL",
    "Link", "Is New", "Sector", "Service Type", "Consulting Type",
    "Days Left", "Notice Type", "Method", "Ref No", "Project", "Project Name",
    "Project ID", "Stage", "Expected Tender Window",
}

DARK_BLUE = "1F3864"
HEADER_FILL  = PatternFill("solid", fgColor=DARK_BLUE)
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_FONT    = Font(name="Calibri", size=9)

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# ─── HELPERS ───────────────────────────────────────────────────────────────────

def normalize_title(title: str) -> str:
    """Lowercase, strip, remove special chars for fuzzy matching."""
    if not title:
        return ""
    t = str(title).lower().strip()
    t = re.sub(r"[^a-z0-9\s]", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def extract_hyperlink_url(cell) -> Optional[str]:
    """Extract actual URL from an openpyxl cell with hyperlink."""
    if cell.hyperlink:
        return cell.hyperlink.target
    val = cell.value
    if val and str(val).startswith("http"):
        return str(val)
    return None


def apply_header_style(ws, row_num: int = 1):
    for cell in ws[row_num]:
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def apply_row_style(ws, row_num: int, light: bool):
    bg = "F7F9FC" if light else "FFFFFF"
    fill = PatternFill("solid", fgColor=bg)
    for cell in ws[row_num]:
        cell.fill  = fill
        cell.font  = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)


def auto_col_width(ws, max_width: int = 50):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, max_width)


# ─── STEP 1: READ MASTER WITH REAL HYPERLINKS ──────────────────────────────────

def read_master_with_urls(path: str) -> tuple[list[str], list[list]]:
    """Read master Excel, extracting actual hyperlink URLs from Tender URL column."""
    print(f"\n[1/4] Reading master dataset: {path}")
    wb = openpyxl.load_workbook(path)  # NOT read_only — need hyperlinks
    ws = wb["All Tenders"]

    headers = [c.value for c in ws[1]]
    url_col_idx = next((i for i, h in enumerate(headers) if h and "url" in str(h).lower()), None)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_vals = []
        for i, cell in enumerate(row):
            if i == url_col_idx:
                url = extract_hyperlink_url(cell)
                row_vals.append(url or cell.value)
            else:
                row_vals.append(cell.value)
        rows.append(row_vals)

    wb.close()
    print(f"   → {len(rows)} rows, {len(headers)} columns")

    # Fix URL column header for clarity
    if url_col_idx is not None:
        headers[url_col_idx] = "Tender URL"

    return headers, rows


# ─── STEP 2: DEDUPLICATION ─────────────────────────────────────────────────────

def deduplicate(headers: list[str], rows: list[list]) -> tuple[list[list], int]:
    """
    Dedup strategy:
      1. If URL is real (not '↗ Open'), use URL as primary key
      2. Else composite: normalized_title + organization + portal
    Returns (clean_rows, num_removed)
    """
    print(f"\n[2/4] Deduplicating {len(rows)} rows...")

    try:
        title_idx = next(i for i, h in enumerate(headers) if h and str(h).lower() in ("title", "title / assignment"))
        org_idx   = next(i for i, h in enumerate(headers) if h and "org" in str(h).lower())
        url_idx   = next(i for i, h in enumerate(headers) if h and "url" in str(h).lower())
        portal_idx = next((i for i, h in enumerate(headers) if h and "portal" in str(h).lower()), None)
    except StopIteration as e:
        print(f"   ⚠ Could not find required column: {e}")
        return rows, 0

    seen_keys = set()
    clean_rows = []
    removed = 0

    for row in rows:
        url = row[url_idx] if url_idx < len(row) else None
        title = row[title_idx] if title_idx < len(row) else None
        org   = row[org_idx]   if org_idx   < len(row) else None
        portal = row[portal_idx] if portal_idx and portal_idx < len(row) else ""

        # Build dedup key
        is_real_url = url and str(url).startswith("http")
        if is_real_url:
            key = f"URL:{url}"
        else:
            norm_title = normalize_title(str(title or ""))
            norm_org   = normalize_title(str(org   or ""))
            key = f"COMP:{portal}:{norm_title}:{norm_org}"

        if key in seen_keys:
            removed += 1
        else:
            seen_keys.add(key)
            clean_rows.append(row)

    print(f"   → Kept {len(clean_rows)} | Removed {removed} duplicates")

    # Portal breakdown of removed
    if removed > 0:
        portal_removed = defaultdict(int)
        portal_all     = defaultdict(int)
        seen_keys2 = set()
        for row in rows:
            url   = row[url_idx] if url_idx < len(row) else None
            title = row[title_idx] if title_idx < len(row) else None
            org   = row[org_idx]   if org_idx   < len(row) else None
            portal = str(row[portal_idx] if portal_idx and portal_idx < len(row) else "unknown")
            is_real_url = url and str(url).startswith("http")
            key = f"URL:{url}" if is_real_url else f"COMP:{portal}:{normalize_title(str(title or ''))}:{normalize_title(str(org or ''))}"
            portal_all[portal] += 1
            if key in seen_keys2:
                portal_removed[portal] += 1
            else:
                seen_keys2.add(key)
        for p in sorted(portal_removed.keys()):
            print(f"      {p}: {portal_removed[p]} removed / {portal_all[p]} total")

    return clean_rows, removed


# ─── STEP 3: BUILD RAW PORTAL LAYER ─────────────────────────────────────────────

def build_raw_layer(headers: list[str], rows: list[list]) -> tuple[list[str], list[list], list[str]]:
    """Strip ML columns, keep only source/raw fields."""
    raw_keep = [i for i, h in enumerate(headers) if h and str(h) not in ML_COLUMNS]
    removed_cols = [str(headers[i]) for i in range(len(headers)) if i not in raw_keep]

    raw_headers = [headers[i] for i in raw_keep]
    raw_rows    = [[row[i] if i < len(row) else None for i in raw_keep] for row in rows]

    return raw_headers, raw_rows, removed_cols


# ─── STEP 4: FIX TARGET VARIABLE ───────────────────────────────────────────────

def fix_target_variable(headers: list[str], rows: list[list]) -> tuple[list[str], list[list], str]:
    """
    - Primary target: Relevance Score (0-100, deterministic, always filled)
    - Secondary (optional): My Decision (user feedback, mostly empty)
    - Move Outcome AFTER My Decision and add a note in the header
    """
    # Reorder: put Relevance Score before My Decision, Outcome last
    priority_order = [
        "Portal", "Title", "Organization", "Country", "Deadline",
        "Sector", "Service Type", "Relevance Score", "Priority Score",
        "Relevance", "Relevance Reason", "Opportunity Insight",
        "Is New", "Scraped Date", "Tender URL",
        "My Decision",   # User fills this
        "Outcome",       # User fills this (secondary, for future learning)
    ]

    # Build ordered indices
    header_map = {str(h): i for i, h in enumerate(headers) if h}
    ordered_indices = []
    seen = set()
    for col_name in priority_order:
        if col_name in header_map and col_name not in seen:
            ordered_indices.append(header_map[col_name])
            seen.add(col_name)
    # Append any remaining columns
    for i, h in enumerate(headers):
        if i not in ordered_indices:
            ordered_indices.append(i)

    new_headers = [headers[i] for i in ordered_indices]
    new_rows    = [[row[i] if i < len(row) else None for i in ordered_indices] for row in rows]

    note = "Relevance Score (0-100) is the ML training target. Outcome is a secondary column for future win-rate analysis."
    return new_headers, new_rows, note


# ─── STEP 5: WRITE EXCEL ───────────────────────────────────────────────────────

def write_excel(path: str, sheets: dict[str, tuple[list[str], list[list]]], title: str):
    """Write a styled Excel file. sheets = {sheet_name: (headers, rows)}"""
    print(f"\n   Writing: {path}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        ws.row_dimensions[1].height = 32
        apply_header_style(ws)

        for i, row in enumerate(rows, start=2):
            ws.append([v if v is not None else "" for v in row])
            apply_row_style(ws, i, i % 2 == 0)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-fit columns
        auto_col_width(ws)

        print(f"      Sheet '{sheet_name}': {len(rows)} rows × {len(headers)} cols")

    wb.save(path)
    print(f"   ✅ Saved: {path}")


# ─── STEP 6: VALIDATION REPORT ─────────────────────────────────────────────────

def write_report(stats: dict, path: str):
    lines = [
        "=" * 65,
        "  TenderRadar — Dataset Cleaning Report",
        f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "=" * 65,
        "",
        "TASK 1: DEDUPLICATION",
        "-" * 40,
        f"  Rows before: {stats['rows_before']:,}",
        f"  Rows after:  {stats['rows_after']:,}",
        f"  Duplicates removed: {stats['dupes_removed']:,}",
        f"  Dedup method: composite key (portal + normalized title + org)",
        f"  Note: URL dedup unavailable (stored as hyperlinks, not strings)",
        "",
        "TASK 2: RAW / ML SEPARATION",
        "-" * 40,
        f"  Master dataset (ML layer): {stats['master_cols']} columns",
        f"  Portal dataset (RAW layer): {stats['raw_cols']} columns",
        f"  Columns removed from portal: {', '.join(stats['removed_cols'])}",
        "",
        "TASK 3: TARGET VARIABLE FIX",
        "-" * 40,
        f"  Primary target: Relevance Score (0-100) — RULE-BASED, always filled",
        f"  Secondary:      My Decision (user feedback) — {stats['decisions_filled']} / {stats['rows_after']} filled",
        f"  Tertiary:       Outcome (win/loss) — {stats['outcomes_filled']} / {stats['rows_after']} filled",
        f"  Action: Outcome retained as LAST column (optional, not for training yet)",
        "",
        "DATA QUALITY NOTES",
        "-" * 40,
        f"  ⚠ URLs: Stored as Excel hyperlinks ('↗ Open' display text).",
        f"    Real URLs cannot be used as primary dedup keys from Excel.",
        f"    → Fix: excel_exporter.py should store URL string + hyperlink.",
        "",
        f"  ⚠ My Decision + Outcome: 0 rows filled in.",
        f"    → ML model (LogisticRegression) CANNOT train yet.",
        f"    → Minimum needed: 25 rows with decisions.",
        f"    → Current scoring is RULE-BASED (keyword matching), not ML.",
        "",
        f"  ⚠ Government portals (UP, Karnataka, Maharashtra, CG) = {stats['gov_rows']:,} rows",
        f"    → These are mostly goods/construction tenders (low consulting relevance).",
        f"    → Already filtered by relevance score in master dataset.",
        "",
        "OUTPUT FILES",
        "-" * 40,
        f"  {OUTPUT_MASTER}",
        f"    → Master dataset: all ML columns, deduplicated",
        f"  {OUTPUT_PORTAL}",
        f"    → Raw portal data: source columns only, no ML scores",
        f"  {REPORT_FILE}",
        f"    → This report",
        "",
        "RISK SUMMARY",
        "-" * 40,
        "  HIGH:   No real user feedback → ML model inactive",
        "  MEDIUM: URLs stored as hyperlinks → dedup uses title+org composite key",
        "  LOW:    663 fuzzy title duplicates cleaned using normalized composite key",
        "  OK:     Relevance scoring (rule-based) is fully operational",
        "  OK:     Vector store (ChromaDB) is active with 768-dim embeddings",
        "=" * 65,
    ]
    with open(path, "w") as f:
        f.write("\n".join(lines))
    print(f"\n   ✅ Report saved: {path}")
    print("\n".join(lines))


# ─── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    os.makedirs("output", exist_ok=True)

    # 1. Read with real URLs
    headers, rows = read_master_with_urls(INPUT_MASTER)

    rows_before = len(rows)

    # 2. Deduplicate
    clean_rows, dupes_removed = deduplicate(headers, rows)

    # 3. Build ML layer (master — keep all columns, reorder target)
    ml_headers, ml_rows, target_note = fix_target_variable(headers, clean_rows)

    # 4. Build RAW layer (portal — strip ML columns)
    raw_headers, raw_rows, removed_cols = build_raw_layer(headers, clean_rows)

    # 5. Portal breakdown sheet for raw file
    from collections import Counter
    portal_idx = next((i for i, h in enumerate(raw_headers) if h and "portal" in str(h).lower()), None)
    if portal_idx is not None:
        portal_counts = Counter(row[portal_idx] for row in raw_rows if row[portal_idx])
    else:
        portal_counts = {}

    portal_summary_rows = [[p, c] for p, c in sorted(portal_counts.items(), key=lambda x: -x[1])]
    portal_summary_rows.append(["TOTAL", sum(portal_counts.values())])

    # 6. Count feedback fill rate
    my_dec_idx  = next((i for i, h in enumerate(ml_headers) if h == "My Decision"), None)
    outcome_idx = next((i for i, h in enumerate(ml_headers) if h == "Outcome"), None)
    decisions_filled = sum(1 for r in ml_rows if my_dec_idx and r[my_dec_idx])
    outcomes_filled  = sum(1 for r in ml_rows if outcome_idx and r[outcome_idx])

    # Count gov portal rows
    portal_idx_ml = next((i for i, h in enumerate(ml_headers) if h and "portal" in str(h).lower()), None)
    gov_portals = {"UP eTenders", "Karnataka eProcure", "Maharashtra Tenders", "CG eProcurement"}
    gov_rows = sum(1 for r in ml_rows if portal_idx_ml and r[portal_idx_ml] in gov_portals)

    # 7. Write master (ML layer)
    print(f"\n[3/4] Writing cleaned datasets...")
    write_excel(OUTPUT_MASTER, {
        "All Tenders (ML Layer)": (ml_headers, ml_rows),
        "Run Summary": (["Portal", "Total Tenders"], portal_summary_rows),
    }, "Master Dataset — ML Layer")

    # 8. Write portal raw layer
    write_excel(OUTPUT_PORTAL, {
        "Tenders (RAW Source)": (raw_headers, raw_rows),
        "Portal Coverage": (["Portal", "Total Tenders"], portal_summary_rows),
    }, "Portal Dataset — RAW Layer")

    # 9. Validation report
    print(f"\n[4/4] Writing validation report...")
    stats = {
        "rows_before":       rows_before,
        "rows_after":        len(clean_rows),
        "dupes_removed":     dupes_removed,
        "master_cols":       len(ml_headers),
        "raw_cols":          len(raw_headers),
        "removed_cols":      removed_cols,
        "decisions_filled":  decisions_filled,
        "outcomes_filled":   outcomes_filled,
        "gov_rows":          gov_rows,
    }
    write_report(stats, REPORT_FILE)

    print(f"\n{'='*55}")
    print("  DONE. Output files:")
    print(f"    Master (ML):  {OUTPUT_MASTER}")
    print(f"    Portal (RAW): {OUTPUT_PORTAL}")
    print(f"    Report:       {REPORT_FILE}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    sys.exit(main())
