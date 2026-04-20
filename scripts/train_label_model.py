"""
Train the IDCG shadow label model from the master workbook.

Usage:
    # From project root, using the project venv:
    .venv/bin/python scripts/train_label_model.py

    # Optionally specify workbook path:
    .venv/bin/python scripts/train_label_model.py --workbook path/to/workbook.xlsx

Output:
    artifacts/label_model.joblib   (trained model)
    artifacts/label_model_meta.json (metrics and configuration)

Prints:
    Label distribution, CV metrics, confusion matrix, feature importance.
"""

from __future__ import annotations

import argparse
import json
import sys
import os
from datetime import datetime
from pathlib import Path

# Ensure project root is on sys.path
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_PROJECT_ROOT))

import numpy as np
from openpyxl import load_workbook

from intelligence.label_model import (
    IDCGLabelModel,
    LABEL_MAP,
    build_features,
    FEATURE_NAMES,
)

# ── Default workbook path ──────────────────────────────────────────────────────
_DEFAULT_WORKBOOK = _PROJECT_ROOT / "output" / "Tender_Monitor_Master.xlsx"


def _is_yes(value: str) -> bool:
    return str(value or "").strip().lower() == "yes"


def load_labeled_rows(
    workbook_path: Path,
    approved_only: bool = True,
    include_ai_fallback: bool = False,
) -> list:
    """
    Load human-reviewed labeled rows from the master workbook.

    Label priority:
      1. Approved Human_Label rows (canonical workflow)
      2. Approved AI_Suggested_Label rows only when AI fallback is enabled

    Rows without any label are skipped.
    Rows labeled 'Not Relevant' are mapped to 'Irrelevant'.
    """
    print(f"[train] Reading workbook: {workbook_path}")
    wb = load_workbook(workbook_path, read_only=True)

    if "All Tenders" in wb.sheetnames:
        ws = wb["All Tenders"]
    else:
        ws = wb.active

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    required = {"Title", "Priority Score", "Relevance Score"}
    missing = required - set(col)
    if missing:
        print(f"[train] WARNING: missing columns {missing}")

    rows = []
    skipped_no_label = 0
    skipped_uncertain = 0
    skipped_unapproved = 0
    source_counts = {"human": 0, "ai": 0}

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        def _get(key, default=""):
            idx = col.get(key)
            return row_vals[idx] if idx is not None and row_vals[idx] is not None else default

        human_label = str(_get("Human_Label", "")).strip()
        ai_label    = str(_get("AI_Suggested_Label", "")).strip()
        approved    = str(_get("Training_Approved", "")).strip()
        approved_yes = _is_yes(approved)

        if approved_only and not approved_yes:
            skipped_unapproved += 1
            continue

        # Choose label source
        if human_label and human_label not in ("", "None", "nan"):
            label = human_label
            source = "human"
        elif include_ai_fallback and ai_label and ai_label not in ("", "None", "nan"):
            label = ai_label
            source = "ai"
        else:
            skipped_no_label += 1
            continue

        # Normalize label
        label = label.replace("Not Relevant", "Irrelevant").strip()
        if label not in LABEL_MAP:
            skipped_uncertain += 1
            continue

        row = {
            "label":          label,
            "label_source":   source,
            "training_approved": "Yes" if approved_yes else approved,
            "title":          str(_get("Title", "")),
            "portal":         str(_get("Portal", "")),
            "sector":         str(_get("Sector", "")),
            "service_type":   str(_get("Service Type", "")),
            "org":            str(_get("Organization", "")),
            "country":        str(_get("Country", "")),
            "priority_score": float(_get("Priority Score", 0) or 0),
            "relevance_score": float(_get("Relevance Score", 0) or 0),
            "deep_scope":     str(_get("Deep Scope", "")),
            "ai_summary":     str(_get("AI Summary", "")),
        }
        rows.append(row)
        source_counts[source] += 1

    wb.close()
    print(f"[train] Loaded {len(rows)} labeled rows "
          f"(skipped: {skipped_no_label} no-label, {skipped_uncertain} uncertain, "
          f"{skipped_unapproved} unapproved)")
    print(f"[train] Source counts: {source_counts}")
    return rows


def print_metrics(meta: dict) -> None:
    """Pretty-print training metrics."""
    print("\n" + "=" * 60)
    print("  IDCG SHADOW LABEL MODEL — TRAINING SUMMARY")
    print("=" * 60)
    print(f"  Trained at:   {meta.get('trained_at', 'N/A')}")
    print(f"  N samples:    {meta['n_samples']}")
    print(f"  N features:   {meta['n_features']}")

    dist = meta["label_distribution"]
    print(f"\n  Label distribution:")
    for lbl in ["Relevant", "Borderline", "Irrelevant"]:
        n = dist.get(lbl, 0)
        pct = 100 * n / meta["n_samples"] if meta["n_samples"] else 0
        bar = "█" * int(pct / 2)
        print(f"    {lbl:12s}: {n:3d} ({pct:4.1f}%)  {bar}")

    cv = meta["cv_summary"]
    print(f"\n  5-fold CV Results:")
    print(f"  {'Class':12s}  {'Precision':>10s}  {'Recall':>8s}  {'F1':>6s}  {'Support':>8s}")
    print("  " + "-" * 50)
    for cls in ["Relevant", "Borderline", "Irrelevant"]:
        m = cv.get(cls, {})
        print(f"  {cls:12s}  {m.get('precision',0):>10.3f}  "
              f"{m.get('recall',0):>8.3f}  {m.get('f1',0):>6.3f}  "
              f"{m.get('support',0):>8d}")
    print("  " + "-" * 50)
    print(f"  {'Accuracy':12s}  {cv.get('accuracy',0):>10.3f}")
    print(f"  {'Macro F1':12s}  {cv.get('macro_f1',0):>10.3f}")

    cm = meta["confusion_matrix"]
    print(f"\n  Confusion Matrix (5-fold CV):")
    print(f"  {'':18s}  {'Pred Irr':>10s}  {'Pred Brd':>10s}  {'Pred Rel':>10s}")
    for i, lbl in enumerate(["True Irrel.", "True Border.", "True Relev."]):
        print(f"  {lbl:18s}  {cm[i][0]:>10d}  {cm[i][1]:>10d}  {cm[i][2]:>10d}")

    fi = meta.get("feature_importance_top10", [])
    print(f"\n  Top 10 Feature Importances:")
    for name, imp in fi:
        bar = "█" * int(imp * 200)
        print(f"    {name:35s}: {imp:.4f}  {bar}")

    print(f"\n  Production blend weight: {meta.get('blend_weight', 0.35):.0%}")
    print("=" * 60 + "\n")


def main():
    parser = argparse.ArgumentParser(description="Train IDCG shadow label model")
    parser.add_argument(
        "--workbook", type=str,
        default=str(_DEFAULT_WORKBOOK),
        help="Path to master workbook with labeled rows",
    )
    parser.add_argument(
        "--min-samples", type=int, default=30,
        help="Minimum labeled rows required to train (default: 30)",
    )
    parser.add_argument(
        "--include-unapproved", action="store_true",
        help="Include rows without Training_Approved=Yes",
    )
    parser.add_argument(
        "--include-ai-fallback", action="store_true",
        help="Allow approved AI_Suggested_Label rows when Human_Label is blank",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"[train] ERROR: workbook not found: {workbook_path}")
        sys.exit(1)

    # Load labeled data
    rows = load_labeled_rows(
        workbook_path,
        approved_only=not args.include_unapproved,
        include_ai_fallback=args.include_ai_fallback,
    )

    if len(rows) < args.min_samples:
        print(f"[train] ERROR: only {len(rows)} rows found, need ≥ {args.min_samples}")
        sys.exit(1)

    # Label distribution check
    from collections import Counter
    dist = Counter(r["label"] for r in rows)
    print(f"[train] Label distribution: {dict(dist)}")
    for lbl in ["Relevant", "Borderline", "Irrelevant"]:
        if dist.get(lbl, 0) < 5:
            print(f"[train] WARNING: very few '{lbl}' rows ({dist.get(lbl, 0)}). "
                  f"CV may be unreliable for this class.")

    # Train
    print(f"[train] Training shadow label model on {len(rows)} rows...")
    model = IDCGLabelModel()
    trained_at = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    meta = model.train(rows, trained_at=trained_at)
    source_counts = Counter(str(r.get("label_source") or "unknown") for r in rows)
    meta["approved_only"] = not args.include_unapproved
    meta["ai_fallback_enabled"] = bool(args.include_ai_fallback)
    meta["approved_rows_used"] = int(
        sum(1 for r in rows if _is_yes(str(r.get("training_approved") or "")))
    )
    meta["label_source_counts"] = {
        "human": int(source_counts.get("human", 0)),
        "ai": int(source_counts.get("ai", 0)),
    }

    # Save
    model.save()
    print(f"[train] Model saved to artifacts/label_model.joblib")
    print(f"[train] Meta saved to artifacts/label_model_meta.json")
    print(
        f"[train] Approved rows used: {meta['approved_rows_used']} "
        f"(human={meta['label_source_counts']['human']}, ai={meta['label_source_counts']['ai']})"
    )

    # Print metrics
    print_metrics(meta)

    # Smoke test: score a few rows
    print("[train] Smoke test: scoring 3 sample rows...")
    for r in rows[:3]:
        score = model.predict_score(r)
        label = model.predict_label(r)
        print(f"  [{r['label']:12s}] → predicted {label:12s} (score={score:.1f}) "
              f"| {r['title'][:60]}")

    print("\n[train] Done. Run the pipeline to activate blended scoring.")


if __name__ == "__main__":
    main()
