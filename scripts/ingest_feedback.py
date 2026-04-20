#!/usr/bin/env python3
"""
ingest_feedback.py — Sync Excel feedback (My Decision, Outcome) to bid_pipeline table

Run weekly via cron:
  0 2 * * 0 /usr/bin/python3 ~/tender_system/scripts/ingest_feedback.py

This reads the master Excel columns P (My Decision) and Q (Outcome),
validates them, and writes to the bid_pipeline table for ML training.
"""

import sys
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database.db import get_connection, get_tender
from config.config import UNIFIED_EXCEL_PATH
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)

# Valid values for decision and outcome columns (match Excel dropdown labels)
VALID_DECISIONS = {"bid", "no bid", "review later", ""}
VALID_OUTCOMES = {"won", "lost", "no submission", "pending", ""}


def load_excel_feedback(excel_path: str) -> list:
    """
    Load feedback from Excel master file.
    Returns list of dicts: {tender_url, decision, outcome, notes}
    """
    if not os.path.exists(excel_path):
        logger.error(f"Excel file not found: {excel_path}")
        return []

    feedback = []
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # Find column indices
        # Assuming columns: ... O=Tender URL, P=My Decision, Q=Outcome
        # Adjust if your column layout is different
        header_row = 1
        col_map = {}

        for col_idx, cell in enumerate(ws[header_row], start=1):
            if not cell.value:
                continue
            header = str(cell.value).strip().lower()
            if "url" in header or "link" in header:
                col_map["url"] = col_idx
            elif "decision" in header:
                col_map["decision"] = col_idx
            elif "outcome" in header:
                col_map["outcome"] = col_idx

        if not col_map.get("url"):
            logger.warning("Could not find URL column in Excel. Skipping.")
            return []

        # Read data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            url_cell = row[col_map.get("url", 14) - 1]  # Default to column O
            decision_cell = row[col_map.get("decision", 15) - 1]  # Default to column P
            outcome_cell = row[col_map.get("outcome", 16) - 1]  # Default to column Q

            tender_url = str(url_cell.value).strip() if url_cell.value else None
            decision = str(decision_cell.value).strip().lower() if decision_cell.value else ""
            outcome = str(outcome_cell.value).strip().lower() if outcome_cell.value else ""

            # Skip empty rows
            if not tender_url:
                continue

            # Validate
            if decision and decision not in VALID_DECISIONS:
                logger.warning(f"Row {row_idx}: Invalid decision '{decision}', skipping")
                continue

            if outcome and outcome not in VALID_OUTCOMES:
                logger.warning(f"Row {row_idx}: Invalid outcome '{outcome}', skipping")
                continue

            # Only include rows with actual feedback
            if decision or outcome:
                feedback.append({
                    "tender_url": tender_url,
                    "decision": decision,
                    "outcome": outcome,
                    "row": row_idx
                })

        logger.info(f"Loaded {len(feedback)} feedback entries from Excel")
        return feedback

    except Exception as e:
        logger.error(f"Error loading Excel: {e}")
        return []


def ingest_to_database(feedback: list) -> None:
    """
    Write feedback to bid_pipeline table.
    """
    if not feedback:
        logger.info("No feedback to ingest.")
        return

    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)

        ingested = 0
        skipped = 0

        for item in feedback:
            tender_url = item["tender_url"]
            decision = item["decision"]
            outcome = item["outcome"]

            # Find tender by URL
            tender = get_tender(cursor, url=tender_url)
            if not tender:
                logger.warning(f"Tender not found: {tender_url}")
                skipped += 1
                continue

            tender_id = tender["id"]

            # Map decision to is_bid (Bid/Review Later -> True, No Bid/Empty -> False)
            is_bid = decision in ("bid", "review later")

            # Map outcome (won/lost/rejected/pending)
            # Insert or update bid_pipeline
            try:
                cursor.execute(
                    """
                    INSERT INTO bid_pipeline (tender_id, is_bid, outcome, feedback_date, notes)
                    VALUES (%s, %s, %s, NOW(), %s)
                    ON DUPLICATE KEY UPDATE
                        is_bid = VALUES(is_bid),
                        outcome = VALUES(outcome),
                        feedback_date = NOW()
                    """,
                    (tender_id, is_bid, outcome if outcome else None, "")
                )
                ingested += 1
            except Exception as e:
                logger.error(f"Error inserting bid_pipeline for tender {tender_id}: {e}")
                skipped += 1

        conn.commit()
        cursor.close()
        conn.close()

        logger.info(f"✓ Ingested {ingested} feedback entries, skipped {skipped}")

    except Exception as e:
        logger.error(f"Database error: {e}")


def main():
    logger.info("=" * 80)
    logger.info("Starting Excel → bid_pipeline feedback ingestion")
    logger.info("=" * 80)

    feedback = load_excel_feedback(UNIFIED_EXCEL_PATH)
    ingest_to_database(feedback)

    logger.info("=" * 80)
    logger.info("Feedback ingestion complete")
    logger.info("=" * 80)


if __name__ == "__main__":
    main()
