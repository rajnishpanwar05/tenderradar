"""
build_intelligence_hub.py — v2.0
Consulting-grade Tender Intelligence Dashboard.

Status logic:
  ACTIVE:        deadline > today + 7 days
  CLOSING_SOON:  0 ≤ days to deadline ≤ 7
  EXPIRED:       deadline < today
  UNKNOWN:       no parseable deadline

Sheets:
  1. Executive_Dashboard  — KPIs, urgency metrics, charts, top scoring
  2. Live_Tenders         — ACTIVE + CLOSING_SOON, sorted by Score ↓ Deadline ↑
  3. Closing_Soon         — deadline ≤ 7 days, urgent view
  4. Expired              — archive, sorted by Score ↓ Deadline ↓
  5. Source_Performance   — per-portal scraping health
"""

import os, re, sys
from datetime import date, datetime, timedelta
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule

# ═══ PATHS ════════════════════════════════════════════════════════════════════
BASE     = os.path.expanduser("~/tender_system")
MASTER   = os.path.join(BASE, "output", "Tender_Monitor_Master.xlsx")
OUT_PATH = os.path.join(BASE, "output", "Tender_Intelligence_Hub.xlsx")
TODAY    = date.today()

# ═══ COLOUR PALETTE ═══════════════════════════════════════════════════════════
C_NAVY   = "1F3864"
C_WHITE  = "FFFFFF"
C_DARK   = "111827"
C_LIGHT  = "F9FAFB"
C_MID    = "F3F4F6"
C_ACCENT = "1D4ED8"

# Status badge palette  (bg, text)
_STATUS_COLORS = {
    "ACTIVE":       ("D1FAE5", "065F46"),   # green
    "CLOSING_SOON": ("FEF3C7", "92400E"),   # amber/orange
    "EXPIRED":      ("FEE2E2", "991B1B"),   # red
    "UNKNOWN":      ("F3F4F6", "6B7280"),   # grey
}

# Score badge (per user spec: ≥90 green, 70-89 yellow, <70 grey)
def _score_colors(s: int):
    if s >= 90: return ("D1FAE5", "065F46")
    if s >= 70: return ("FEF3C7", "92400E")
    return ("F3F4F6", "9CA3AF")

FONT_NAME = "Arial"

# ═══ PORTAL METADATA ══════════════════════════════════════════════════════════
PORTAL_META = {
    "ICFRE Tenders":       ("India",         "ICFRE (Govt. of India)"),
    "Meghalaya MBDA":      ("India",         "Meghalaya Basin Dev. Authority"),
    "World Bank":          ("International", "World Bank Group"),
    "IUCN Procurement":    ("International", "IUCN"),
    "DTVP Germany":        ("Germany",       "Deutschen Vergabeportal"),
    "JTDS Jharkhand":      ("India",         "JTDS Jharkhand"),
    "TED EU":              ("EU",            "European Commission"),
    "AfDB Consultants":    ("International", "African Dev. Bank"),
    "UNGM":                ("International", "UN Global Marketplace"),
    "UNDP Procurement":    ("International", "UNDP"),
    "DevNet India":        ("India",         "DevNet Jobs India"),
    "NGO Box":             ("India",         "NGO Box"),
    "GIZ India":           ("India",         "GIZ"),
    "AFD France":          ("International", "Agence Française de Développement"),
    "GeM BidPlus":         ("India",         "GeM (Govt. of India)"),
    "TANEPS Tanzania":     ("Tanzania",      "TANEPS"),
    "Karnataka eProcure":  ("India",         "Govt. of Karnataka"),
    "CG eProcurement":     ("India",         "Central Govt. eProcurement"),
    "SAM.gov":             ("USA",           "US Federal Govt."),
    "USAID":               ("International", "USAID"),
    "SIDBI":               ("India",         "SIDBI"),
    "PHFI":                ("India",         "PHFI"),
    "Welthungerhilfe":     ("International", "Deutsche Welthungerhilfe"),
}

PORTAL_GROUP = {
    "World Bank":          "🌍 International",
    "UNGM":                "🌍 International",
    "UNDP Procurement":    "🌍 International",
    "AfDB Consultants":    "🌍 International",
    "AFD France":          "🌍 International",
    "TED EU":              "🌍 International",
    "DTVP Germany":        "🌍 International",
    "TANEPS Tanzania":     "🌍 International",
    "IUCN Procurement":    "🌍 International",
    "Welthungerhilfe":     "🌍 International",
    "USAID":               "🌍 International",
    "ICFRE Tenders":       "🇮🇳 India – Central",
    "GeM BidPlus":         "🇮🇳 India – Central",
    "DevNet India":        "🇮🇳 India – Central",
    "CG eProcurement":     "🇮🇳 India – Central",
    "SIDBI":               "🇮🇳 India – Central",
    "PHFI":                "🇮🇳 India – Central",
    "JTDS Jharkhand":      "🏛 India – State",
    "Karnataka eProcure":  "🏛 India – State",
    "Meghalaya MBDA":      "🏛 India – State",
    "NGO Box":             "🤝 NGO / Civil Society",
    "GIZ India":           "🤝 NGO / Civil Society",
}


# ═══ BORDER & CELL HELPERS ════════════════════════════════════════════════════

def _thin(color="D1D5DB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_only(color=C_NAVY):
    return Border(bottom=Side(style="medium", color=color))

def _cell(ws, row, col, value=None, bold=False, size=10, color=C_DARK,
          bg=None, halign="left", valign="center", wrap=False,
          border=None, fmt=None):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font      = Font(name=FONT_NAME, bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = border
    if fmt:
        c.number_format = fmt
    return c


# ═══ 1. DATA LOADING & CLEANING ═══════════════════════════════════════════════

def _parse_deadline(raw):
    if pd.isna(raw):
        return None
    s = str(raw).strip()
    if s in ("", "-", "N/A", "nan", "None", "TBD", "Ongoing", "Rolling"):
        return None
    s = re.sub(r"(\d+)(st|nd|rd|th)\b", r"\1", s, flags=re.I)
    s = re.sub(r"\s+um\s+\d+[:.]\d+\s*(Uhr)?.*$", "", s, flags=re.I)
    s = re.sub(r"\s+(by|at|upto)\s+\d+[:.]\d+.*$", "", s, flags=re.I)
    s = re.sub(r"\s*\(.*?\)", "", s)
    s = re.sub(r"\s*\d{1,2}:\d{2}(:\d{2})?(\s*(CET|IST|UTC|GMT|AM|PM|hrs|hours|New York time))?.*$",
               "", s, flags=re.I)
    s = re.sub(r"\s*(CET|IST|UTC|GMT|New York time)\s*$", "", s, flags=re.I)
    s = re.sub(r"\s+upto\s+.*$", "", s, flags=re.I)
    s = s.strip().rstrip(",.;:").strip()
    s = re.sub(r"(\d{4})\d{2}:\d{2}.*$", r"\1", s)

    fmts = [
        "%d %B %Y", "%d %b %Y",  "%d %B, %Y", "%B %d %Y", "%B %d, %Y",
        "%d-%b-%y", "%d-%b-%Y",  "%d.%m.%Y",  "%d.%m.%y", "%Y-%m-%d",
        "%d/%m/%Y", "%m/%d/%Y",  "%d-%m-%Y",  "%d %b %y", "%d, %B %Y",
        "%d-%b-%y", "%b %d, %Y",
    ]
    for fmt in fmts:
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _determine_status(d) -> str:
    if d is None:
        return "UNKNOWN"
    if d < TODAY:
        return "EXPIRED"
    if d <= TODAY + timedelta(days=7):
        return "CLOSING_SOON"
    return "ACTIVE"


def load_and_clean() -> pd.DataFrame:
    df = pd.read_excel(MASTER)
    df = df[df["Title"].notna() & (df["Title"].str.strip() != "")].copy()

    # Impute Country / Organization from portal metadata
    for idx, row in df.iterrows():
        portal = str(row.get("Portal", "")).strip()
        meta   = PORTAL_META.get(portal, (None, None))
        if pd.isna(row.get("Country")) or str(row.get("Country", "")).strip() in ("", "-"):
            df.at[idx, "Country"] = meta[0] or "Unknown"
        if pd.isna(row.get("Organization")) or str(row.get("Organization", "")).strip() in ("", "-"):
            df.at[idx, "Organization"] = meta[1] or portal

    df["_parsed_deadline"] = df["Deadline"].apply(_parse_deadline)
    df["Status"]           = df["_parsed_deadline"].apply(_determine_status)

    def _fmt_dl(row):
        if row["_parsed_deadline"]:
            return row["_parsed_deadline"].isoformat()
        raw = str(row["Deadline"]).strip()
        return "" if raw in ("-", "nan", "None", "") else raw

    df["Deadline_Display"]  = df.apply(_fmt_dl, axis=1)
    df["Days_To_Deadline"]  = df.apply(
        lambda r: (r["_parsed_deadline"] - TODAY).days if r["_parsed_deadline"] else None,
        axis=1
    )

    key_cols = ["Title", "Organization", "Country", "Deadline", "Sector", "Tender URL"]
    df["Completeness"] = df[key_cols].apply(
        lambda r: round(
            sum(1 for v in r if pd.notna(v) and str(v).strip() not in ("", "-"))
            / len(key_cols) * 100
        ), axis=1
    )

    # Recompute relevance score for freshness
    try:
        sys.path.insert(0, BASE)
        from intelligence.keywords import score_tender_numeric as _sfn

        def _score_row(row):
            return _sfn(
                str(row.get("Title", "")),
                str(row.get("Relevance", "")),
                str(row.get("Country", ""))
            )
        _sr = df.apply(_score_row, axis=1)
        df["Relevance Score"]  = _sr.apply(lambda x: x[0])
        df["Relevance Reason"] = _sr.apply(lambda x: x[1])
    except Exception as e:
        print(f"[hub] WARNING: scoring unavailable — {e}")
        if "Relevance Score" not in df.columns:
            df["Relevance Score"] = 0
        if "Relevance Reason" not in df.columns:
            df["Relevance Reason"] = ""

    return df


# ═══ 2. SHARED LAYOUT HELPERS ═════════════════════════════════════════════════

def _banner(ws, title: str, subtitle: str, ncols: int):
    ws.row_dimensions[1].height = 44
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    _cell(ws, 1, 1, title,    bold=True,  size=16, color=C_WHITE,  bg=C_NAVY,
          halign="center", valign="center")
    _cell(ws, 2, 1, subtitle, bold=False, size=9,  color="A5B4FC", bg=C_NAVY,
          halign="center", valign="center")


def _col_headers(ws, row: int, headers: list, bg=C_NAVY):
    ws.row_dimensions[row].height = 26
    for ci, h in enumerate(headers, 1):
        _cell(ws, row, ci, h, bold=True, size=9, color=C_WHITE,
              bg=bg, halign="center", valign="center", border=_thin(C_NAVY))


def _kpi_card(ws, row: int, col: int, label: str, value, bg: str, tc: str, span=2):
    ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+span-1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+span-1)
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+span-1)
    _cell(ws, row,   col, label, bold=False, size=8,  color="6B7280",
          bg=bg, halign="center", border=_thin())
    _cell(ws, row+1, col, value, bold=True,  size=24, color=tc,
          bg=bg, halign="center", valign="center")
    _cell(ws, row+2, col, "",   bg=bg, border=_thin())
    ws.row_dimensions[row].height   = 18
    ws.row_dimensions[row+1].height = 42
    ws.row_dimensions[row+2].height = 8


def _section_header(ws, row: int, title: str, start_col=2, end_col=11):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row,   end_column=end_col)
    _cell(ws, row, start_col, title, bold=True, size=11,
          color=C_WHITE, bg=C_NAVY, halign="left", valign="center")


# ═══ 3. EXECUTIVE DASHBOARD ═══════════════════════════════════════════════════

def build_dashboard(wb: Workbook, df: pd.DataFrame):
    ws = wb.active
    ws.title = "Executive_Dashboard"
    ws.sheet_view.showGridLines = False

    # Column widths: A=2 (margin), B-K=2 cols per KPI × 5 KPIs, L=2 (margin)
    widths = [2, 18, 14, 18, 14, 18, 14, 18, 14, 18, 14, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Core metrics ──────────────────────────────────────────────────────────
    active_df  = df[df["Status"].isin(["ACTIVE", "CLOSING_SOON"])].copy()
    total_act  = len(active_df)
    close7     = int((active_df["Days_To_Deadline"].notna() &
                      (active_df["Days_To_Deadline"] <= 7)).sum())
    close30    = int((active_df["Days_To_Deadline"].notna() &
                      (active_df["Days_To_Deadline"] <= 30)).sum())
    new_run    = int((df.get("Is New", pd.Series(dtype=str)) == "YES").sum()) \
                 if "Is New" in df.columns else 0
    missing_dl = int((df["Status"] == "UNKNOWN").sum())

    # ── Banner ────────────────────────────────────────────────────────────────
    _banner(ws,
        "🏆  TenderRadar Intelligence Hub — Executive Dashboard",
        f"Generated: {TODAY.isoformat()}   ·   Total Dataset: {len(df)} tenders"
        f"   ·   Active: {total_act}   ·   IDCG Consulting",
        ncols=12)

    ws.row_dimensions[4].height = 12
    for c in range(1, 13):
        _cell(ws, 4, c, "", bg=C_WHITE)

    # ── Row 1 of KPI cards (row 5–7) ─────────────────────────────────────────
    _kpi_card(ws, 5, 2,  "ACTIVE TENDERS",       total_act,  "D1FAE5", "065F46")
    _kpi_card(ws, 5, 4,  "⚡ CLOSING ≤ 7 DAYS",  close7,     "FEE2E2", "991B1B")
    _kpi_card(ws, 5, 6,  "📅 CLOSING ≤ 30 DAYS", close30,    "FEF3C7", "92400E")
    _kpi_card(ws, 5, 8,  "🆕 NEW THIS RUN",       new_run,    "DBEAFE", "1E40AF")
    _kpi_card(ws, 5, 10, "❓ MISSING DEADLINE",   missing_dl, "F3F4F6", "6B7280")

    ws.row_dimensions[8].height = 12
    for c in range(1, 13):
        _cell(ws, 8, c, "", bg=C_WHITE)

    # ── Deadline urgency strip (row 9–11) ─────────────────────────────────────
    _section_header(ws, 9, "⏱  DEADLINE URGENCY — ACTIVE TENDERS")

    d7  = close7
    d14 = int((active_df["Days_To_Deadline"].notna() &
               (active_df["Days_To_Deadline"] <= 14)).sum())
    d30 = close30
    d60 = int((active_df["Days_To_Deadline"].notna() &
               (active_df["Days_To_Deadline"] <= 60)).sum())

    urgency_bands = [
        ("Closing ≤ 7 days",  d7,  "EF4444", "FEE2E2"),
        ("Closing ≤ 14 days", d14, "F97316", "FFF7ED"),
        ("Closing ≤ 30 days", d30, "F59E0B", "FEF3C7"),
        ("Closing ≤ 60 days", d60, "10B981", "D1FAE5"),
    ]
    for i, (label, val, tc, bg) in enumerate(urgency_bands):
        col = [2, 4, 7, 9][i]
        _kpi_card(ws, 10, col, label, val, bg, tc)

    ws.row_dimensions[13].height = 12
    for c in range(1, 13):
        _cell(ws, 13, c, "", bg=C_WHITE)

    # ── Top Portals table (row 14–26) ─────────────────────────────────────────
    _section_header(ws, 14, "🏛  TOP PORTALS BY TENDER COUNT")

    portal_stats = df.groupby("Portal").agg(
        Total   = ("Title",  "count"),
        Active  = ("Status", lambda x: x.isin(["ACTIVE","CLOSING_SOON"]).sum()),
        Closing = ("Status", lambda x: (x == "CLOSING_SOON").sum()),
        Expired = ("Status", lambda x: (x == "EXPIRED").sum()),
    ).sort_values("Total", ascending=False).head(10)

    phdr = 15
    ws.row_dimensions[phdr].height = 22
    for ci, (h, c_off) in enumerate(
        zip(["Portal", "Total", "🟢 Active", "⚡ Closing", "🔴 Expired", "Live %"],
            [2, 4, 5, 7, 9, 10]), 1
    ):
        _cell(ws, phdr, c_off, h, bold=True, size=9, color=C_WHITE,
              bg="374151", halign="center", border=_thin())

    # Merge portal name across two cols, and live % across last two
    ws.merge_cells(start_row=phdr, start_column=2, end_row=phdr, end_column=3)
    ws.merge_cells(start_row=phdr, start_column=10, end_row=phdr, end_column=11)

    for ri, (portal, r) in enumerate(portal_stats.iterrows(), phdr + 1):
        alt      = "F8FAFC" if ri % 2 == 0 else C_WHITE
        live_pct = f"{r.Active / r.Total * 100:.0f}%" if r.Total else "—"
        ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=3)
        ws.merge_cells(start_row=ri, start_column=10, end_row=ri, end_column=11)
        for val, c_off, ha in [
            (portal,          2,  "left"),
            (int(r.Total),    4,  "center"),
            (int(r.Active),   5,  "center"),
            (int(r.Closing),  7,  "center"),
            (int(r.Expired),  9,  "center"),
            (live_pct,        10, "center"),
        ]:
            tc = C_DARK
            if val == live_pct:
                pv = r.Active / r.Total * 100 if r.Total else 0
                tc = "065F46" if pv >= 50 else ("92400E" if pv >= 20 else "991B1B")
            _cell(ws, ri, c_off, val, size=9, color=tc,
                  bg=alt, halign=ha, border=_thin("E5E7EB"))
        ws.row_dimensions[ri].height = 18

    # ── Top Scoring Active Tenders (after portal table) ───────────────────────
    r3 = phdr + len(portal_stats) + 3
    _section_header(ws, r3, "🏅  TOP SCORING ACTIVE OPPORTUNITIES")

    top5 = (df[df["Status"].isin(["ACTIVE", "CLOSING_SOON"])]
            .sort_values("Relevance Score", ascending=False)
            .head(5))

    top_hdr = r3 + 1
    ws.row_dimensions[top_hdr].height = 22
    for h, c_off in [("Score", 2), ("Title", 4), ("Portal", 9), ("Deadline", 11)]:
        _cell(ws, top_hdr, c_off, h, bold=True, size=9, color=C_WHITE,
              bg="374151", halign="center", border=_thin())
    ws.merge_cells(start_row=top_hdr, start_column=4, end_row=top_hdr, end_column=8)

    for ri_t, (_, trow) in enumerate(top5.iterrows(), top_hdr + 1):
        raw_s = trow.get("Relevance Score", 0)
        sv    = int(raw_s) if pd.notna(raw_s) else 0
        s_bg, s_tc = _score_colors(sv)
        _cell(ws, ri_t, 2, sv, bold=True, size=12,
              color=s_tc, bg=s_bg, halign="center", border=_thin())
        ws.merge_cells(start_row=ri_t, start_column=4, end_row=ri_t, end_column=8)
        _cell(ws, ri_t, 4,  str(trow.get("Title",   ""))[:80], size=9,
              bg=C_LIGHT, halign="left", border=_thin(), wrap=True)
        _cell(ws, ri_t, 9,  str(trow.get("Portal",  "")), size=9,
              bg=C_LIGHT, halign="center", border=_thin())
        status = str(trow.get("Status", "ACTIVE"))
        sbg, stc = _STATUS_COLORS.get(status, ("F3F4F6", "6B7280"))
        _cell(ws, ri_t, 11, str(trow.get("Deadline_Display", "")), size=9,
              bg=sbg, halign="center", border=_thin(), color=stc)
        ws.row_dimensions[ri_t].height = 24

    # ── CHARTS ────────────────────────────────────────────────────────────────
    # Write chart data to hidden columns P, R, T
    # Chart 1: Pie — Tenders by Portal (top 12)
    _chart_pie_portals(ws, df)

    # Chart 2: Bar — Tenders by Country (top 10 active)
    _chart_bar_countries(ws, active_df)

    # Chart 3: Bar — Deadline Distribution next 30 days
    _chart_bar_deadlines(ws, active_df)

    # Hide chart data columns
    for col_letter in ("P", "Q", "R", "S", "T", "U"):
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "B5"
    ws.sheet_properties.tabColor = C_NAVY


def _chart_pie_portals(ws, df: pd.DataFrame):
    """Pie chart: all tenders by portal (top 12), anchored at N2."""
    portal_counts = df["Portal"].value_counts().head(12)
    base_row = 2

    ws.cell(base_row, 16, "Portal")
    ws.cell(base_row, 17, "Count")
    for i, (portal, cnt) in enumerate(portal_counts.items(), 1):
        ws.cell(base_row + i, 16, portal[:20])
        ws.cell(base_row + i, 17, int(cnt))

    n = len(portal_counts)
    pie = PieChart()
    pie.title  = "Tenders by Portal"
    pie.style  = 10
    pie.width  = 16
    pie.height = 12
    data_ref  = Reference(ws, min_col=17, min_row=base_row,     max_row=base_row + n)
    label_ref = Reference(ws, min_col=16, min_row=base_row + 1, max_row=base_row + n)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(label_ref)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, "N2")


def _chart_bar_countries(ws, active_df: pd.DataFrame):
    """Bar chart: top 10 countries by active tender count, anchored at N18."""
    country_counts = (
        active_df["Country"]
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown"})
        .value_counts()
        .head(10)
    )
    base_row = 20

    ws.cell(base_row, 19, "Country")
    ws.cell(base_row, 20, "Active")
    for i, (country, cnt) in enumerate(country_counts.items(), 1):
        ws.cell(base_row + i, 19, str(country)[:18])
        ws.cell(base_row + i, 20, int(cnt))

    n = len(country_counts)
    bar = BarChart()
    bar.type   = "bar"
    bar.title  = "Active Tenders by Country (Top 10)"
    bar.style  = 10
    bar.width  = 16
    bar.height = 12
    bar.y_axis.title = "Tenders"
    data_ref  = Reference(ws, min_col=20, min_row=base_row,     max_row=base_row + n)
    label_ref = Reference(ws, min_col=19, min_row=base_row + 1, max_row=base_row + n)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(label_ref)
    ws.add_chart(bar, "N18")


def _chart_bar_deadlines(ws, active_df: pd.DataFrame):
    """Bar chart: deadlines in next 30 days grouped by week, anchored at N35."""
    bands = [
        ("Week 1 (1–7d)",   1,  7),
        ("Week 2 (8–14d)",  8,  14),
        ("Week 3 (15–21d)", 15, 21),
        ("Week 4 (22–30d)", 22, 30),
    ]
    counts = []
    for label, lo, hi in bands:
        n = int((
            active_df["Days_To_Deadline"].notna() &
            (active_df["Days_To_Deadline"] >= lo) &
            (active_df["Days_To_Deadline"] <= hi)
        ).sum())
        counts.append((label, n))

    base_row = 38
    ws.cell(base_row, 21, "Period")
    ws.cell(base_row, 22, "Count")
    for i, (label, cnt) in enumerate(counts, 1):
        ws.cell(base_row + i, 21, label)
        ws.cell(base_row + i, 22, cnt)

    bar2 = BarChart()
    bar2.type   = "col"
    bar2.title  = "Deadlines in Next 30 Days"
    bar2.style  = 10
    bar2.width  = 16
    bar2.height = 10
    bar2.y_axis.title = "Tenders Closing"
    data_ref  = Reference(ws, min_col=22, min_row=base_row,     max_row=base_row + 4)
    label_ref = Reference(ws, min_col=21, min_row=base_row + 1, max_row=base_row + 4)
    bar2.add_data(data_ref, titles_from_data=True)
    bar2.set_categories(label_ref)
    ws.add_chart(bar2, "N35")


# ═══ 4. DATA SHEET BUILDER ════════════════════════════════════════════════════

# Live/Closing columns per spec:
# Score | Title | Portal | Country | Deadline | Days_Left | Sector | Relevance_Reason | Detail_Link
LIVE_COLS = [
    ("Score",            10),
    ("Title",            55),
    ("Portal",           18),
    ("Country",          14),
    ("Status",           14),
    ("Deadline",         16),
    ("Days Left",        10),
    ("Sector",           28),
    ("Relevance Reason", 45),
    ("Tender URL",       40),
]

EXPIRED_COLS = [
    ("Score",            10),
    ("Title",            55),
    ("Portal",           18),
    ("Country",          14),
    ("Deadline",         16),
    ("Sector",           28),
    ("Relevance Reason", 45),
    ("Tender URL",       40),
]

SOURCE_COLS = [
    ("Portal",          22),
    ("Rows Scraped",    13),
    ("🟢 Active",       10),
    ("⚡ Closing",      11),
    ("🔴 Expired",      12),
    ("❓ Unknown",      12),
    ("New Tenders",     13),
    ("Live %",          10),
    ("Completeness",    14),
    ("Last Scraped",    18),
    ("Country",         14),
    ("Group",           20),
]


def _write_data_sheet(ws, df_sheet: pd.DataFrame, header_bg: str,
                      columns: list, include_status_col: bool = False,
                      status_row_color: bool = False):
    """Generic data sheet writer with conditional formatting."""
    ws.sheet_view.showGridLines = False
    ncols = len(columns)

    for ci, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = width

    hdr_row = 3
    _col_headers(ws, hdr_row, [c[0] for c in columns], bg=header_bg)
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = (
        f"A{hdr_row}:{get_column_letter(ncols)}{hdr_row + max(1, len(df_sheet))}"
    )

    col_map = {c[0]: i + 1 for i, c in enumerate(columns)}

    # ── Add conditional formatting rules ──────────────────────────────────────
    data_start = hdr_row + 1
    data_end   = data_start + max(1, len(df_sheet)) - 1

    # Score conditional formatting (user spec: ≥90 green, 70-89 yellow, <70 grey)
    sc_col = col_map.get("Score")
    if sc_col:
        sc_letter = get_column_letter(sc_col)
        sc_range  = f"{sc_letter}{data_start}:{sc_letter}{data_end}"
        ws.conditional_formatting.add(sc_range, CellIsRule(
            operator="greaterThanOrEqual", formula=["90"],
            fill=PatternFill("solid", fgColor="D1FAE5"),
            font=Font(name=FONT_NAME, bold=True, color="065F46")
        ))
        ws.conditional_formatting.add(sc_range, CellIsRule(
            operator="between", formula=["70", "89"],
            fill=PatternFill("solid", fgColor="FEF3C7"),
            font=Font(name=FONT_NAME, bold=True, color="92400E")
        ))
        ws.conditional_formatting.add(sc_range, CellIsRule(
            operator="lessThan", formula=["70"],
            fill=PatternFill("solid", fgColor="F3F4F6"),
            font=Font(name=FONT_NAME, color="9CA3AF")
        ))

    # Days Left conditional formatting (≤7 → red, 8-30 → amber)
    dl_col = col_map.get("Days Left")
    if dl_col:
        dl_letter = get_column_letter(dl_col)
        dl_range  = f"{dl_letter}{data_start}:{dl_letter}{data_end}"
        ws.conditional_formatting.add(dl_range, CellIsRule(
            operator="lessThanOrEqual", formula=["7"],
            fill=PatternFill("solid", fgColor="FEE2E2"),
            font=Font(name=FONT_NAME, bold=True, color="991B1B")
        ))
        ws.conditional_formatting.add(dl_range, CellIsRule(
            operator="between", formula=["8", "30"],
            fill=PatternFill("solid", fgColor="FEF3C7"),
            font=Font(name=FONT_NAME, color="92400E")
        ))

    # ── Write data rows ────────────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df_sheet.iterrows(), hdr_row + 1):
        status = str(row.get("Status", "UNKNOWN"))
        if status_row_color:
            bg, _ = _STATUS_COLORS.get(status, ("F3F4F6", "6B7280"))
            alt    = bg
        else:
            alt = "F8FAFC" if ri % 2 == 0 else C_WHITE
        ws.row_dimensions[ri].height = 40

        def _w(col_name, value, **kw):
            ci = col_map.get(col_name)
            if ci is None:
                return
            defaults = dict(size=9, color=C_DARK, bg=alt,
                            border=_thin("E5E7EB"), valign="top", wrap=True)
            defaults.update(kw)
            _cell(ws, ri, ci, value, **defaults)

        _w("Portal",       str(row.get("Portal",       "")) or "")
        _w("Title",        str(row.get("Title",        "")) or "")
        _w("Organization", str(row.get("Organization", "")) or "")
        _w("Country",      str(row.get("Country",      "")) or "")
        _w("Deadline",     row.get("Deadline_Display",  "")  or "")
        _w("Scraped Date", str(row.get("Scraped Date", "")) or "")
        _w("Sector",       str(row.get("Sector",       "")) or "")
        _w("Service Type", str(row.get("Service Type", "")) or "")
        _w("Relevance",    str(row.get("Relevance",    "")) or "")
        _w("Relevance Reason", str(row.get("Relevance Reason", "")) or "",
           size=9, color="4B5563")

        # Status badge
        ci_s = col_map.get("Status")
        if ci_s:
            s_bg, s_tc = _STATUS_COLORS.get(status, ("F3F4F6", "6B7280"))
            c = ws.cell(row=ri, column=ci_s)
            c.value     = status.replace("_", " ")
            c.font      = Font(name=FONT_NAME, size=9, bold=True, color=s_tc)
            c.fill      = PatternFill("solid", fgColor=s_bg)
            c.alignment = Alignment(horizontal="center", vertical="top")
            c.border    = _thin("E5E7EB")

        # Score badge (per-cell coloring supplements conditional formatting)
        ci_sc = col_map.get("Score")
        if ci_sc:
            raw_s = row.get("Relevance Score", 0)
            try:
                sv = int(raw_s) if pd.notna(raw_s) else 0
            except (ValueError, TypeError):
                sv = 0
            s_bg, s_tc = _score_colors(sv)
            c = ws.cell(row=ri, column=ci_sc)
            c.value     = sv
            c.font      = Font(name=FONT_NAME, size=10, bold=(sv >= 70), color=s_tc)
            c.fill      = PatternFill("solid", fgColor=s_bg)
            c.alignment = Alignment(horizontal="center", vertical="top")
            c.border    = _thin("E5E7EB")

        # Days Left badge
        ci_dl = col_map.get("Days Left")
        if ci_dl:
            days = row.get("Days_To_Deadline")
            if days is not None and not (isinstance(days, float) and pd.isna(days)):
                dv = int(days)
                dc = ("991B1B" if dv <= 7 else "92400E" if dv <= 30 else "065F46")
                _w("Days Left", dv, bold=(dv <= 30), color=dc, wrap=False,
                   halign="center")
            else:
                _w("Days Left", "—", halign="center")

        # Tender URL as hyperlink
        ci_url = col_map.get("Tender URL")
        if ci_url:
            url = str(row.get("Tender URL", "")).strip()
            c   = ws.cell(row=ri, column=ci_url)
            c.border    = _thin("E5E7EB")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.fill      = PatternFill("solid", fgColor=alt)
            if url and url.startswith("http"):
                c.hyperlink = url
                c.value     = "🔗 View"
                c.font      = Font(name=FONT_NAME, size=9, color="1D4ED8",
                                   underline="single")
            else:
                c.value = "—"
                c.font  = Font(name=FONT_NAME, size=9, color="9CA3AF")


# ═══ 5. SHEET 2 — LIVE TENDERS ════════════════════════════════════════════════

def build_live_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Live_Tenders")
    live = df[df["Status"].isin(["ACTIVE", "CLOSING_SOON"])].copy()
    live = live.sort_values(
        ["Relevance Score", "_parsed_deadline"],
        ascending=[False, True],
        na_position="last"
    )

    _banner(ws,
        f"🟢  LIVE TENDERS — {len(live)} Active Opportunities",
        f"Showing ACTIVE + CLOSING SOON  ·  Sorted: Score ↓  Deadline ↑  ·  {TODAY.isoformat()}",
        ncols=len(LIVE_COLS))

    _write_data_sheet(ws, live, "065F46", LIVE_COLS, include_status_col=True)
    ws.sheet_properties.tabColor = "10B981"


# ═══ 6. SHEET 3 — CLOSING SOON ════════════════════════════════════════════════

def build_closing_soon_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Closing_Soon")
    closing = df[df["Status"] == "CLOSING_SOON"].copy()
    closing = closing.sort_values(
        ["_parsed_deadline", "Relevance Score"],
        ascending=[True, False],
        na_position="last"
    )

    _banner(ws,
        f"⚡  CLOSING SOON — {len(closing)} Tenders Due in ≤ 7 Days",
        f"URGENT: Act immediately on these opportunities  ·  {TODAY.isoformat()}",
        ncols=len(LIVE_COLS))

    _write_data_sheet(ws, closing, "DC2626", LIVE_COLS,
                      include_status_col=True, status_row_color=True)
    ws.sheet_properties.tabColor = "EF4444"


# ═══ 7. SHEET 4 — EXPIRED ═════════════════════════════════════════════════════

def build_expired_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Expired")
    expired = df[df["Status"] == "EXPIRED"].copy()
    expired = expired.sort_values(
        ["Relevance Score", "_parsed_deadline"],
        ascending=[False, False],
        na_position="last"
    )

    _banner(ws,
        f"🔴  EXPIRED / CLOSED — {len(expired)} Archived Tenders",
        f"Sorted: Score ↓  Deadline ↓ (most recent first)  ·  {TODAY.isoformat()}",
        ncols=len(EXPIRED_COLS))

    _write_data_sheet(ws, expired, "991B1B", EXPIRED_COLS)
    ws.sheet_properties.tabColor = "EF4444"


# ═══ 8. SHEET 5 — SOURCE PERFORMANCE ════════════════════════════════════════

def build_source_performance_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Source_Performance")
    ws.sheet_view.showGridLines = False

    _banner(ws,
        "📈  SOURCE PERFORMANCE — Pipeline Health & Scraping Coverage",
        f"Per-portal statistics as of {TODAY.isoformat()}",
        ncols=len(SOURCE_COLS))

    for ci, (_, w) in enumerate(SOURCE_COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    hdr_row = 3
    _col_headers(ws, hdr_row, [c[0] for c in SOURCE_COLS], bg=C_NAVY)
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(SOURCE_COLS))}{hdr_row}"

    portal_df = df.groupby("Portal").agg(
        Total    = ("Title",    "count"),
        Active   = ("Status",   lambda x: x.isin(["ACTIVE","CLOSING_SOON"]).sum()),
        Closing  = ("Status",   lambda x: (x == "CLOSING_SOON").sum()),
        Expired  = ("Status",   lambda x: (x == "EXPIRED").sum()),
        Unknown  = ("Status",   lambda x: (x == "UNKNOWN").sum()),
        New      = ("Is New",   lambda x: (x == "YES").sum()) if "Is New" in df.columns
                   else ("Title", lambda x: 0),
        Compl    = ("Completeness", "mean"),
        LastScr  = ("Scraped Date", "max"),
    ).sort_values("Total", ascending=False)

    col_order = [c[0] for c in SOURCE_COLS]
    col_idx   = {c: i + 1 for i, c in enumerate(col_order)}

    for ri, (portal, r) in enumerate(portal_df.iterrows(), hdr_row + 1):
        alt     = "F8FAFC" if ri % 2 == 0 else C_WHITE
        lpct    = f"{r.Active / r.Total * 100:.0f}%" if r.Total else "—"
        comp    = f"{r.Compl:.0f}%"
        group   = PORTAL_GROUP.get(portal, "Other")
        country = PORTAL_META.get(portal, (None,))[0] or "—"
        last_s  = str(r.LastScr)[:16]

        row_data = {
            "Portal":       portal,
            "Rows Scraped": int(r.Total),
            "🟢 Active":    int(r.Active),
            "⚡ Closing":   int(r.Closing),
            "🔴 Expired":   int(r.Expired),
            "❓ Unknown":   int(r.Unknown),
            "New Tenders":  int(r.New) if hasattr(r, "New") else 0,
            "Live %":       lpct,
            "Completeness": comp,
            "Last Scraped": last_s,
            "Country":      country,
            "Group":        group,
        }

        for col_name, ci in col_idx.items():
            val    = row_data.get(col_name, "")
            halign = "center" if col_name not in ("Portal", "Last Scraped",
                                                   "Country", "Group") else "left"
            tc = C_DARK
            if col_name == "Live %":
                pv = r.Active / r.Total * 100 if r.Total else 0
                tc = "065F46" if pv >= 50 else ("92400E" if pv >= 20 else "991B1B")
            if col_name == "Completeness":
                cv = r.Compl if pd.notna(r.Compl) else 0
                tc = "065F46" if cv >= 80 else ("92400E" if cv >= 50 else "991B1B")
            _cell(ws, ri, ci, val, size=9, color=tc,
                  bg=alt, halign=halign, border=_thin("E5E7EB"))
        ws.row_dimensions[ri].height = 20

    ws.sheet_properties.tabColor = "3B82F6"


# ═══ 9. MAIN ══════════════════════════════════════════════════════════════════

def main():
    print(f"[hub] Loading: {MASTER}")
    if not os.path.exists(MASTER):
        print(f"[hub] ERROR: master file not found at {MASTER}")
        sys.exit(1)

    df = load_and_clean()
    print(f"[hub] Loaded {len(df)} rows")

    status_counts = df["Status"].value_counts().to_dict()
    active    = status_counts.get("ACTIVE",       0)
    closing   = status_counts.get("CLOSING_SOON", 0)
    expired   = status_counts.get("EXPIRED",      0)
    unknown   = status_counts.get("UNKNOWN",      0)
    print(f"[hub] Status: {active} ACTIVE · {closing} CLOSING_SOON"
          f" · {expired} EXPIRED · {unknown} UNKNOWN")

    parsed = df["_parsed_deadline"].notna().sum()
    print(f"[hub] Deadline parsed: {parsed}/{len(df)} ({parsed/len(df)*100:.0f}%)")

    wb = Workbook()

    print("[hub] Sheet 1: Executive Dashboard...")
    build_dashboard(wb, df)

    print("[hub] Sheet 2: Live Tenders...")
    build_live_sheet(wb, df)

    print("[hub] Sheet 3: Closing Soon...")
    build_closing_soon_sheet(wb, df)

    print("[hub] Sheet 4: Expired...")
    build_expired_sheet(wb, df)

    print("[hub] Sheet 5: Source Performance...")
    build_source_performance_sheet(wb, df)

    wb.save(OUT_PATH)
    size_kb = os.path.getsize(OUT_PATH) // 1024
    print(f"\n[hub] ✅  Saved: {OUT_PATH}  ({size_kb} KB)")
    print(f"[hub] Sheets: {wb.sheetnames}")

    # ── Score band summary ─────────────────────────────────────────────────────
    if "Relevance Score" in df.columns:
        active_df = df[df["Status"].isin(["ACTIVE", "CLOSING_SOON"])]
        h90 = (active_df["Relevance Score"] >= 90).sum()
        h70 = ((active_df["Relevance Score"] >= 70) & (active_df["Relevance Score"] < 90)).sum()
        lo  = (active_df["Relevance Score"] < 70).sum()
        avg = active_df["Relevance Score"].mean()
        print(f"[hub] Score bands (active): "
              f"{h90} HIGH(≥90) · {h70} MED(70-89) · {lo} LOW(<70) · avg {avg:.1f}")


if __name__ == "__main__":
    main()
