#!/usr/bin/env python3
"""
Summarize label persistence in the TenderRadar master workbook.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = PROJECT_ROOT / "output" / "Tender_Monitor_Master.xlsx"


def _is_yes(value: str) -> bool:
    return str(value or "").strip().lower() == "yes"


def summarize_workbook(workbook_path: Path) -> dict[str, object]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["All Tenders"] if "All Tenders" in wb.sheetnames else wb.active

    header_row = next(ws.iter_rows(max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        raise RuntimeError("Workbook has no header row.")

    headers = [str(cell or "").strip() for cell in header_row]
    col = {name: idx for idx, name in enumerate(headers) if name}

    total_rows = 0
    human_label_rows = 0
    approved_yes_rows = 0
    unlabeled_rows = 0
    new_unlabeled_rows = 0

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        human_label = str(row_vals[col["Human_Label"]] or "").strip() if "Human_Label" in col else ""
        training_approved = str(row_vals[col["Training_Approved"]] or "").strip() if "Training_Approved" in col else ""
        is_new = str(row_vals[col["Is New"]] or "").strip().upper() if "Is New" in col else ""

        if human_label:
            human_label_rows += 1
        else:
            unlabeled_rows += 1
            if is_new == "YES":
                new_unlabeled_rows += 1

        if _is_yes(training_approved):
            approved_yes_rows += 1

    wb.close()
    return {
        "workbook": str(workbook_path),
        "total_rows": total_rows,
        "human_label_rows": human_label_rows,
        "training_approved_yes_rows": approved_yes_rows,
        "unlabeled_rows": unlabeled_rows,
        "new_unlabeled_rows": new_unlabeled_rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Summarize workbook label persistence.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to Tender_Monitor_Master.xlsx")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"ERROR: workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    summary = summarize_workbook(workbook_path)
    if args.json:
        print(json.dumps(summary, ensure_ascii=True))
    else:
        print("Workbook label summary")
        print(f"  workbook                  : {summary['workbook']}")
        print(f"  total rows                : {summary['total_rows']}")
        print(f"  rows with Human_Label     : {summary['human_label_rows']}")
        print(f"  rows with Training_Approved=Yes : {summary['training_approved_yes_rows']}")
        print(f"  unlabeled rows            : {summary['unlabeled_rows']}")
        print(f"  new unlabeled rows        : {summary['new_unlabeled_rows']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
