#!/usr/bin/env python3
"""
IDCG TenderRadar Dashboard Builder
Generates a world-class executive Excel dashboard from Tender_Monitor_Master.xlsx
"""

import re
import sys
import logging
from datetime import date, datetime
from collections import defaultdict

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
_log = logging.getLogger("tenderradar.dashboard")

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
SOURCE = "/Users/rajnishpanwar/tender_system/output/Tender_Monitor_Master.xlsx"
OUTPUT = "/Users/rajnishpanwar/tender_system/output/IDCG_TenderRadar_Dashboard.xlsx"
TODAY = date(2026, 3, 21)

# Colors
NAVY       = "1F3864"
NAVY2      = "2E4D7B"
RED        = "C00000"
ORANGE     = "E2711D"
GREEN_DK   = "375623"
GOLD       = "C9973B"
PURPLE     = "7030A0"
GRAY_DK    = "595959"
GRAY_MD    = "767676"
GRAY_LT    = "D6DCE4"
WHITE      = "FFFFFF"
OFF_WHITE  = "F8F9FA"
BLUE_HDR   = "BDD7EE"
BLUE_LIGHT = "EBF3FB"
GREEN_LT   = "E8F5E9"
RED_LT     = "FFEBEE"
ORANGE_LT  = "FFF3E0"
GOLD_LT    = "FFF8E1"
PURPLE_LT  = "F3E5F5"
YELLOW_LT  = "FFFDE7"

# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(name="Arial", size=10, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def make_align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def make_border(top=None, bottom=None, left=None, right=None, color=GRAY_LT):
    def side(style):
        if style:
            return Side(style=style, color=color)
        return Side(style=None)
    return Border(top=side(top), bottom=side(bottom), left=side(left), right=side(right))

def thin_border_all(color=GRAY_LT):
    s = Side(style="thin", color=color)
    return Border(top=s, bottom=s, left=s, right=s)

def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    if number_format: cell.number_format = number_format
    return cell

def merge_set(ws, r1, c1, r2, c2, value=None, font=None, fill=None, alignment=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    if value is not None: cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border
    # Fill all merged cells (openpyxl quirk)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            if fill:
                ws.cell(row=r, column=c).fill = fill
    return cell

def col_letter(n):
    return get_column_letter(n)

def trunc(s, n):
    if not s:
        return ""
    s = str(s)
    return s[:n] + "…" if len(s) > n else s

# ─────────────────────────────────────────────
# DEADLINE PARSING
# ─────────────────────────────────────────────

GERMAN_NO_DEADLINE_PHRASES = [
    "nicht vorhanden", "keine frist", "unbefristet", "laufend",
    "keine deadline", "no deadline"
]

def parse_deadline(raw):
    """Returns a date object or None."""
    if not raw:
        return None
    s = str(raw).strip()

    # German no-deadline phrases
    sl = s.lower()
    for phrase in GERMAN_NO_DEADLINE_PHRASES:
        if phrase in sl:
            return None

    # JSON list format from UNDP/UNGM: ['2026-04-06T10:00:00+02:00']
    m = re.search(r"(\d{4}-\d{2}-\d{2})T", s)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except (ValueError, TypeError):
            pass

    # German: "31.03.2026 um 12:00 Uhr"
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except (ValueError, TypeError):
            pass

    # GeM: "28-03-2026 16:00" or "24-03-2026 17:00"
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except (ValueError, TypeError):
            pass

    # ISO: "2026-03-31"
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except (ValueError, TypeError):
            pass

    # "05-Mar-26 07:00 AM" — short year
    m = re.search(r"(\d{1,2})-([A-Za-z]{3})-(\d{2})\b", s)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                  "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        mo = months.get(m.group(2).lower())
        yr = int(m.group(3))
        yr = 2000 + yr if yr < 100 else yr
        if mo:
            try:
                return date(yr, mo, int(m.group(1)))
            except (ValueError, TypeError):
                pass

    # "19-Mar-2026 06:30..." long year
    m = re.search(r"(\d{1,2})-([A-Za-z]{3})-(\d{4})", s)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                  "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        mo = months.get(m.group(2).lower())
        if mo:
            try:
                return date(int(m.group(3)), mo, int(m.group(1)))
            except (ValueError, TypeError):
                pass

    # "31 Mar 2026" or "22 Mar 2026" or "23 Mar. 2026"
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3})\.?\s+(\d{4})", s)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                  "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        mo = months.get(m.group(2).lower())
        if mo:
            try:
                return date(int(m.group(3)), mo, int(m.group(1)))
            except (ValueError, TypeError):
                pass

    # "28th April 2022" / "25th November, 2019"
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?\s+([A-Za-z]+),?\s+(\d{4})", s, re.IGNORECASE)
    if m:
        months = {"january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
                  "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
                  "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,
                  "sep":9,"oct":10,"nov":11,"dec":12}
        mo = months.get(m.group(2).lower())
        if mo:
            try:
                return date(int(m.group(3)), mo, int(m.group(1)))
            except (ValueError, TypeError):
                pass

    # "24 December 2022 (12:00 PM)"  — already caught above if month is alpha

    # Plain dd/mm/yyyy or mm/dd/yyyy
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except (ValueError, TypeError):
            pass

    return None


def deadline_status(dl, today=TODAY):
    if dl is None:
        return "UNKNOWN", None
    days = (dl - today).days
    if days < 0:
        return "EXPIRED", days
    elif days <= 7:
        return "CRITICAL", days
    elif days <= 30:
        return "URGENT", days
    else:
        return "ACTIVE", days


def format_deadline(raw, parsed):
    if parsed:
        return parsed.strftime("%d %b %Y")
    if raw:
        s = str(raw).strip()
        # shorten long strings
        return s[:30] if len(s) > 30 else s
    return "—"


# ─────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────

_log.info("Loading source data...")
wb_src = openpyxl.load_workbook(SOURCE)
ws_all = wb_src["All Tenders"]
ws_run = wb_src["Run Summary"]

# Read all tenders
headers = [ws_all.cell(1, c).value for c in range(1, ws_all.max_column+1)]
hmap = {h: i+1 for i, h in enumerate(headers)}

def gcol(name):
    return hmap.get(name, None)

rows = []
for r in range(2, ws_all.max_row+1):
    row = {h: ws_all.cell(r, hmap[h]).value for h in headers}
    rows.append(row)

_log.info(f"  Loaded {len(rows)} tenders")

# Compute deadline info
for row in rows:
    raw = row.get("Deadline")
    parsed = parse_deadline(raw)
    status, days_left = deadline_status(parsed)
    row["_parsed_deadline"] = parsed
    row["_status"] = status
    row["_days_left"] = days_left
    row["_deadline_fmt"] = format_deadline(raw, parsed)

# KPIs
total_tenders = len(rows)
active_now = sum(1 for r in rows if r["_status"] in ("ACTIVE", "CRITICAL", "URGENT"))
closing_7  = sum(1 for r in rows if r["_status"] == "CRITICAL")
closing_30 = sum(1 for r in rows if r["_status"] in ("CRITICAL", "URGENT"))
high_priority = sum(1 for r in rows if (r.get("Priority Score") or 0) >= 75)
portals_count = len(set(r["Portal"] for r in rows if r["Portal"]))

# Sectors
all_sectors = []
for row in rows:
    s = row.get("Sector") or ""
    for sec in re.split(r"[,;]+", s):
        sec = sec.strip()
        if sec:
            all_sectors.append(sec)

from collections import Counter
sector_counts = Counter(all_sectors)
top_sectors = sector_counts.most_common(12)
sectors_count = len(sector_counts)

# Avg score (non-zero)
scores = [r.get("Priority Score") or 0 for r in rows]
nonzero_scores = [s for s in scores if s > 0]
avg_score = round(sum(nonzero_scores) / len(nonzero_scores), 1) if nonzero_scores else 0

_log.info(f"  KPIs: total={total_tenders}, active={active_now}, critical={closing_7}, urgent={closing_30-closing_7}")
_log.info(f"  high_priority={high_priority}, portals={portals_count}, avg_score={avg_score}")

# Top 25 opportunities
top_opps = sorted([r for r in rows if (r.get("Priority Score") or 0) > 0],
                  key=lambda r: r.get("Priority Score", 0), reverse=True)[:25]

# Portal stats
portal_stats = {}
for row in rows:
    p = row.get("Portal") or "Unknown"
    if p not in portal_stats:
        portal_stats[p] = {"total": 0, "active": 0, "critical": 0, "urgent": 0,
                           "expired": 0, "unknown": 0, "scores": [], "high_prio": 0,
                           "last_scraped": None}
    ps = portal_stats[p]
    ps["total"] += 1
    st = row["_status"]
    ps[st.lower()] += 1
    if st in ("ACTIVE", "CRITICAL", "URGENT"):
        ps["active"] += 1
    sc = row.get("Priority Score") or 0
    if sc > 0:
        ps["scores"].append(sc)
    if sc >= 75:
        ps["high_prio"] += 1
    sd = row.get("Scraped Date")
    if sd:
        ps["last_scraped"] = sd

for p, ps in portal_stats.items():
    ps["avg_score"] = round(sum(ps["scores"]) / len(ps["scores"]), 1) if ps["scores"] else 0

portals_sorted = sorted(portal_stats.items(), key=lambda x: x[1]["total"], reverse=True)

# Urgency breakdown
urgency_bands = {
    "≤ 7 days":  0,
    "8–14 days": 0,
    "15–30 days": 0,
    "31–60 days": 0,
    "> 60 days":  0,
    "Unknown":    0,
}
for row in rows:
    dl = row["_days_left"]
    st = row["_status"]
    if st == "UNKNOWN":
        urgency_bands["Unknown"] += 1
    elif st == "EXPIRED":
        pass  # expired, not in urgency bands
    elif dl is not None:
        if dl <= 7:
            urgency_bands["≤ 7 days"] += 1
        elif dl <= 14:
            urgency_bands["8–14 days"] += 1
        elif dl <= 30:
            urgency_bands["15–30 days"] += 1
        elif dl <= 60:
            urgency_bands["31–60 days"] += 1
        else:
            urgency_bands["> 60 days"] += 1

# Active tenders for sheet 2
active_rows = [r for r in rows if r["_status"] in ("ACTIVE", "CRITICAL", "URGENT")]
active_rows = sorted(active_rows, key=lambda r: (-(r.get("Priority Score") or 0), r["_days_left"] or 9999))

_log.info(f"  Active tenders: {len(active_rows)}")

# Portal group mapping
PORTAL_GROUPS = {
    "ICFRE Tenders": "India Central",
    "GeM BidPlus": "India Central",
    "CG eProcurement": "India Central",
    "DevNet India": "India Central",
    "Karnataka eProcure": "India State",
    "Meghalaya MBDA": "India State",
    "JTDS Jharkhand": "India State",
    "DTVP Germany": "International",
    "TED EU": "International",
    "AfDB Consultants": "International",
    "GIZ India": "International",
    "UNGM": "International",
    "UNDP Procurement": "International",
    "AFD France": "International",
    "TANEPS Tanzania": "International",
    "World Bank": "International",
    "NGO Box": "NGO/Civil Society",
    "IUCN Procurement": "International",
}

# ─────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────

_log.info("Building Excel workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ─────────────────────────────────────────────
# SHEET 1: Command Center
# ─────────────────────────────────────────────

ws1 = wb.create_sheet("🎯 Command Center")
ws1.sheet_properties.tabColor = NAVY

# Column widths (A=1, B=2 ... N=14)
ws1.column_dimensions["A"].width = 2
for c in range(2, 14):  # B through M
    ws1.column_dimensions[col_letter(c)].width = 12
ws1.column_dimensions["N"].width = 2

# ── Row 1: Main header ──
ws1.row_dimensions[1].height = 35
ws1.merge_cells("A1:N1")
c = ws1["A1"]
c.value = "  TenderRadar™   |   IDCG Consulting   |   Procurement Intelligence Command Center"
c.font = make_font("Arial", 18, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")

# ── Row 2: Subtitle ──
ws1.row_dimensions[2].height = 20
ws1.merge_cells("A2:N2")
c = ws1["A2"]
c.value = (f"  18 Portals Monitored   ·   {total_tenders} Tenders in Database"
           f"   ·   Last Updated: 21 March 2026   ·   Automated Intelligence System")
c.font = make_font("Arial", 10, italic=True, color=WHITE)
c.fill = make_fill(NAVY2)
c.alignment = make_align("left", "center")

# ── Row 3: Spacer ──
ws1.row_dimensions[3].height = 12
for c in range(1, 15):
    ws1.cell(3, c).fill = make_fill("F2F2F2")

# ── KPI Tiles (Rows 4-6) ──
# Each tile: 2 columns wide, rows 4-6
# Tile layout: row 4 = number, row 5 = label, row 6 = sub-desc
ws1.row_dimensions[4].height = 28
ws1.row_dimensions[5].height = 16
ws1.row_dimensions[6].height = 14

KPI_TILES = [
    # (col_start, label, value, sub, accent_color, bg_color)
    (2,  "TOTAL IN DB",       str(total_tenders),  "All portals tracked",  "2E75B6", BLUE_LIGHT),
    (4,  "ACTIVE NOW",        str(active_now),      "Live opportunities",   GREEN_DK, GREEN_LT),
    (6,  "CLOSING ≤ 7 DAYS",  str(closing_7),       "Critical deadlines",   RED,      RED_LT),
    (8,  "CLOSING ≤ 30 DAYS", str(closing_30),      "Urgent pipeline",      ORANGE,   ORANGE_LT),
    (10, "HIGH PRIORITY",     str(high_priority),   "Score ≥ 75",           GOLD,     GOLD_LT),
    (12, "PORTALS ACTIVE",    str(portals_count),   "Source portals",       PURPLE,   PURPLE_LT),
]

for (cs, label, val, sub, accent, bg) in KPI_TILES:
    ce = cs + 1  # end column of tile
    bg_fill = make_fill(bg)
    border_color = accent

    # Fill all cells in tile
    for row in range(4, 7):
        for col in range(cs, ce+1):
            ws1.cell(row, col).fill = bg_fill

    # Row 4: Big number
    ws1.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce)
    c = ws1.cell(4, cs)
    c.value = val
    c.font = make_font("Arial", 26, bold=True, color=accent)
    c.fill = bg_fill
    c.alignment = make_align("center", "center")

    # Row 5: Label
    ws1.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce)
    c = ws1.cell(5, cs)
    c.value = label
    c.font = make_font("Arial", 9, bold=True, color="595959")
    c.fill = bg_fill
    c.alignment = make_align("center", "center")

    # Row 6: Sub-description
    ws1.merge_cells(start_row=6, start_column=cs, end_row=6, end_column=ce)
    c = ws1.cell(6, cs)
    c.value = sub
    c.font = make_font("Arial", 8, color="767676")
    c.fill = bg_fill
    c.alignment = make_align("center", "center")

    # Add border around tile
    bc = GRAY_LT
    for row in range(4, 7):
        for col in range(cs, ce+1):
            top    = "thin" if row == 4 else None
            bottom = "thin" if row == 6 else None
            left   = "thin" if col == cs else None
            right  = "thin" if col == ce else None
            ws1.cell(row, col).border = make_border(top, bottom, left, right, bc)

# ── Rows 7-8: Spacers ──
ws1.row_dimensions[7].height = 8
ws1.row_dimensions[8].height = 8
for row in range(7, 9):
    for c in range(1, 15):
        ws1.cell(row, c).fill = make_fill(WHITE)

# ── Row 9: Section Header - Top Opportunities ──
ws1.row_dimensions[9].height = 22
ws1.merge_cells("B9:M9")
c = ws1["B9"]
c.value = "  🔥  TOP PRIORITY OPPORTUNITIES"
c.font = make_font("Arial", 13, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")

# ── Row 10: Column headers for opportunities table ──
ws1.row_dimensions[10].height = 18
opp_headers = [
    (2, 6, "TITLE"),
    (7, 8, "PORTAL"),
    (9, 10, "DEADLINE"),
    (11, 11, "SCORE"),
    (12, 13, "SECTOR"),
]
for (c1, c2, label) in opp_headers:
    if c1 == c2:
        ws1.cell(10, c1).value = label
        ws1.cell(10, c1).font = make_font("Arial", 9, bold=True, color=NAVY)
        ws1.cell(10, c1).fill = make_fill(BLUE_HDR)
        ws1.cell(10, c1).alignment = make_align("center", "center")
        ws1.cell(10, c1).border = make_border(bottom="medium", color=NAVY)
    else:
        ws1.merge_cells(start_row=10, start_column=c1, end_row=10, end_column=c2)
        c = ws1.cell(10, c1)
        c.value = label
        c.font = make_font("Arial", 9, bold=True, color=NAVY)
        c.fill = make_fill(BLUE_HDR)
        c.alignment = make_align("center", "center")
        c.border = make_border(bottom="medium", color=NAVY)
        for col in range(c1, c2+1):
            ws1.cell(10, col).fill = make_fill(BLUE_HDR)
            ws1.cell(10, col).border = make_border(bottom="medium", color=NAVY)

# ── Rows 11-35: Top 25 Opportunities ──
for i, tender in enumerate(top_opps):
    row = 11 + i
    ws1.row_dimensions[row].height = 16
    row_fill = make_fill(WHITE) if i % 2 == 0 else make_fill(OFF_WHITE)
    border = make_border(bottom="thin", color=GRAY_LT)

    # Title (B:F merged)
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    c = ws1.cell(row, 2)
    c.value = trunc(tender.get("Title") or "", 90)
    c.font = make_font("Arial", 9)
    c.fill = row_fill
    c.alignment = make_align("left", "center", wrap=True)
    c.border = border

    # Portal (G:H merged)
    ws1.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    c = ws1.cell(row, 7)
    c.value = tender.get("Portal") or ""
    c.font = make_font("Arial", 9)
    c.fill = row_fill
    c.alignment = make_align("center", "center")
    c.border = border

    # Deadline (I:J merged)
    ws1.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
    c = ws1.cell(row, 9)
    c.value = tender.get("_deadline_fmt") or "—"
    c.font = make_font("Arial", 9)
    c.fill = row_fill
    c.alignment = make_align("center", "center")
    c.border = border

    # Score (K)
    score = tender.get("Priority Score") or 0
    c = ws1.cell(row, 11)
    c.value = score
    if score >= 90:
        c.font = make_font("Arial", 11, bold=True, color=RED)
        c.fill = make_fill(RED_LT)
    elif score >= 75:
        c.font = make_font("Arial", 11, bold=True, color=GOLD)
        c.fill = make_fill(GOLD_LT)
    elif score >= 50:
        c.font = make_font("Arial", 11, bold=True, color="2E75B6")
        c.fill = make_fill(BLUE_LIGHT)
    else:
        c.font = make_font("Arial", 11, bold=True, color=GRAY_DK)
        c.fill = row_fill
    c.alignment = make_align("center", "center")
    c.border = border

    # Sector (L:M merged)
    ws1.merge_cells(start_row=row, start_column=12, end_row=row, end_column=13)
    c = ws1.cell(row, 12)
    sec_raw = tender.get("Sector") or ""
    sec_first = sec_raw.split(",")[0].strip() if sec_raw else "—"
    c.value = sec_first
    c.font = make_font("Arial", 8)
    c.fill = row_fill
    c.alignment = make_align("center", "center")
    c.border = border

    # Fill side cells
    for col in [6, 8, 10, 12, 13]:
        ws1.cell(row, col).fill = row_fill

# ── Row 36: Spacer ──
ws1.row_dimensions[36].height = 8

# ── Row 37: Portal Health header ──
ws1.row_dimensions[37].height = 22
ws1.merge_cells("B37:M37")
c = ws1["B37"]
c.value = "  📊  PORTAL HEALTH MATRIX"
c.font = make_font("Arial", 13, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")

# ── Row 38: Portal table headers ──
ws1.row_dimensions[38].height = 18
portal_headers = [
    (2, 4, "PORTAL"),
    (5, 5, "TOTAL"),
    (6, 6, "ACTIVE"),
    (7, 7, "CRITICAL"),
    (8, 8, "AVG SCORE"),
    (9, 9, "HIGH PRIO"),
    (10, 10, "STATUS"),
    (11, 13, "LAST SCRAPED"),
]
for (c1, c2, label) in portal_headers:
    if c1 == c2:
        ws1.cell(38, c1).value = label
        ws1.cell(38, c1).font = make_font("Arial", 9, bold=True, color=NAVY)
        ws1.cell(38, c1).fill = make_fill(BLUE_HDR)
        ws1.cell(38, c1).alignment = make_align("center", "center")
        ws1.cell(38, c1).border = make_border(bottom="medium", color=NAVY)
    else:
        ws1.merge_cells(start_row=38, start_column=c1, end_row=38, end_column=c2)
        cc = ws1.cell(38, c1)
        cc.value = label
        cc.font = make_font("Arial", 9, bold=True, color=NAVY)
        cc.fill = make_fill(BLUE_HDR)
        cc.alignment = make_align("center", "center")
        cc.border = make_border(bottom="medium", color=NAVY)
        for col in range(c1, c2+1):
            ws1.cell(38, col).fill = make_fill(BLUE_HDR)
            ws1.cell(38, col).border = make_border(bottom="medium", color=NAVY)

# ── Rows 39+: Portal data ──
for i, (portal, ps) in enumerate(portals_sorted[:18]):
    row = 39 + i
    ws1.row_dimensions[row].height = 15
    rf = make_fill(WHITE) if i % 2 == 0 else make_fill(OFF_WHITE)
    bd = make_border(bottom="thin", color=GRAY_LT)

    # Portal name (B:D)
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws1.cell(row, 2)
    c.value = portal
    c.font = make_font("Arial", 9, bold=True)
    c.fill = rf
    c.alignment = make_align("left", "center")
    c.border = bd
    for col in [3, 4]:
        ws1.cell(row, col).fill = rf

    # Total
    c = ws1.cell(row, 5)
    c.value = ps["total"]
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # Active
    c = ws1.cell(row, 6)
    c.value = ps["active"]
    c.font = make_font("Arial", 9, color=GREEN_DK if ps["active"] > 0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # Critical
    c = ws1.cell(row, 7)
    c.value = ps["critical"]
    c.font = make_font("Arial", 9, bold=ps["critical"]>0, color=RED if ps["critical"]>0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # Avg Score
    c = ws1.cell(row, 8)
    c.value = ps["avg_score"] if ps["avg_score"] > 0 else "—"
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # High priority
    c = ws1.cell(row, 9)
    c.value = ps["high_prio"]
    c.font = make_font("Arial", 9, bold=ps["high_prio"]>0, color=GOLD if ps["high_prio"]>0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # Status badge
    c = ws1.cell(row, 10)
    if ps["active"] > 0:
        c.value = "✅ LIVE"
        c.font = make_font("Arial", 9, bold=True, color=GREEN_DK)
    elif ps["total"] > 0:
        c.value = "⚠️ STALE"
        c.font = make_font("Arial", 9, bold=True, color=ORANGE)
    else:
        c.value = "❌ EMPTY"
        c.font = make_font("Arial", 9, bold=True, color=RED)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # Last scraped (K:M)
    ws1.merge_cells(start_row=row, start_column=11, end_row=row, end_column=13)
    c = ws1.cell(row, 11)
    ls = ps.get("last_scraped") or ""
    if ls:
        ls_str = str(ls)[:16]
    else:
        ls_str = "—"
    c.value = ls_str
    c.font = make_font("Arial", 8, color=GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd
    for col in [12, 13]:
        ws1.cell(row, col).fill = rf

# ── Row 57: Spacer ──
last_portal_row = 39 + min(len(portals_sorted), 18) - 1
spacer_row = last_portal_row + 1
ws1.row_dimensions[spacer_row].height = 8

# ── Deadline Urgency + Sectors section ──
header_row = spacer_row + 1
ws1.row_dimensions[header_row].height = 22
ws1.merge_cells(start_row=header_row, start_column=2, end_row=header_row, end_column=13)
c = ws1.cell(header_row, 2)
c.value = "  📅  DEADLINE URGENCY     |     🌐  TOP SECTORS"
c.font = make_font("Arial", 13, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")
for col in range(2, 14):
    ws1.cell(header_row, col).fill = make_fill(NAVY)

# Urgency table headers
urg_hdr = header_row + 1
ws1.row_dimensions[urg_hdr].height = 16
urgency_colors = {
    "≤ 7 days":   RED,
    "8–14 days":  ORANGE,
    "15–30 days": "C9973B",
    "31–60 days": GREEN_DK,
    "> 60 days":  "1F6B6B",
    "Unknown":    GRAY_DK,
}
urgency_bgs = {
    "≤ 7 days":   RED_LT,
    "8–14 days":  ORANGE_LT,
    "15–30 days": GOLD_LT,
    "31–60 days": GREEN_LT,
    "> 60 days":  BLUE_LIGHT,
    "Unknown":    "F5F5F5",
}

# Urgency header row
for col, (c1, c2, label) in enumerate([(2,2,"WINDOW"), (3,4,"TENDERS"), (5,6,"BAR")]):
    if c1 == c2:
        ws1.cell(urg_hdr, c1).value = label
        ws1.cell(urg_hdr, c1).font = make_font("Arial", 8, bold=True, color=NAVY)
        ws1.cell(urg_hdr, c1).fill = make_fill(BLUE_HDR)
        ws1.cell(urg_hdr, c1).alignment = make_align("center", "center")
    else:
        ws1.merge_cells(start_row=urg_hdr, start_column=c1, end_row=urg_hdr, end_column=c2)
        cc = ws1.cell(urg_hdr, c1)
        cc.value = label
        cc.font = make_font("Arial", 8, bold=True, color=NAVY)
        cc.fill = make_fill(BLUE_HDR)
        cc.alignment = make_align("center", "center")
        for col in range(c1, c2+1):
            ws1.cell(urg_hdr, col).fill = make_fill(BLUE_HDR)

# Sector header row
for (c1, c2, label) in [(8,9,"SECTOR"), (10,11,"COUNT"), (12,13,"SHARE")]:
    if c1 == c2:
        ws1.cell(urg_hdr, c1).value = label
        ws1.cell(urg_hdr, c1).font = make_font("Arial", 8, bold=True, color=NAVY)
        ws1.cell(urg_hdr, c1).fill = make_fill(BLUE_HDR)
        ws1.cell(urg_hdr, c1).alignment = make_align("center", "center")
    else:
        ws1.merge_cells(start_row=urg_hdr, start_column=c1, end_row=urg_hdr, end_column=c2)
        cc = ws1.cell(urg_hdr, c1)
        cc.value = label
        cc.font = make_font("Arial", 8, bold=True, color=NAVY)
        cc.fill = make_fill(BLUE_HDR)
        cc.alignment = make_align("center", "center")
        for col in range(c1, c2+1):
            ws1.cell(urg_hdr, col).fill = make_fill(BLUE_HDR)

# Urgency data rows
total_active_urg = sum(urgency_bands.values())
max_urg = max(urgency_bands.values()) if urgency_bands else 1

for i, (band, count) in enumerate(urgency_bands.items()):
    row = urg_hdr + 1 + i
    ws1.row_dimensions[row].height = 14
    color = urgency_colors.get(band, GRAY_DK)
    bg = urgency_bgs.get(band, "F5F5F5")

    # Window label
    c = ws1.cell(row, 2)
    c.value = band
    c.font = make_font("Arial", 9, bold=True, color=color)
    c.fill = make_fill(bg)
    c.alignment = make_align("left", "center")

    # Count
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    c = ws1.cell(row, 3)
    c.value = count
    c.font = make_font("Arial", 9, bold=True, color=color)
    c.fill = make_fill(bg)
    c.alignment = make_align("center", "center")
    ws1.cell(row, 4).fill = make_fill(bg)

    # Visual bar
    bar_len = int((count / max_urg) * 12) if max_urg > 0 else 0
    ws1.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    c = ws1.cell(row, 5)
    c.value = "█" * bar_len if bar_len > 0 else "·"
    c.font = make_font("Arial", 9, color=color)
    c.fill = make_fill("FAFAFA")
    c.alignment = make_align("left", "center")
    ws1.cell(row, 6).fill = make_fill("FAFAFA")

# Top sectors
sec_tints = ["EBF3FB", "E8F5E9", "FFF8E1", "FFF3E0", "F3E5F5",
             "FFEBEE", "E0F7FA", "F9FBE7", "FCE4EC", "E8EAF6",
             "FFF9C4", "F1F8E9"]
total_sectors_all = sum(c for _, c in top_sectors)
top8 = top_sectors[:8]

for i, (sec, cnt) in enumerate(top8):
    row = urg_hdr + 1 + i
    ws1.row_dimensions[row].height = 14
    tint = sec_tints[i % len(sec_tints)]

    # Sector name (H:I)
    ws1.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
    c = ws1.cell(row, 8)
    c.value = sec
    c.font = make_font("Arial", 9)
    c.fill = make_fill(tint)
    c.alignment = make_align("left", "center")
    ws1.cell(row, 9).fill = make_fill(tint)

    # Count (J:K)
    ws1.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
    c = ws1.cell(row, 10)
    c.value = cnt
    c.font = make_font("Arial", 9, bold=True)
    c.fill = make_fill(tint)
    c.alignment = make_align("center", "center")
    ws1.cell(row, 11).fill = make_fill(tint)

    # Share (L:M)
    ws1.merge_cells(start_row=row, start_column=12, end_row=row, end_column=13)
    c = ws1.cell(row, 12)
    share = f"{cnt/total_sectors_all*100:.1f}%" if total_sectors_all > 0 else "—"
    c.value = share
    c.font = make_font("Arial", 9)
    c.fill = make_fill(tint)
    c.alignment = make_align("center", "center")
    ws1.cell(row, 13).fill = make_fill(tint)

# Footer
footer_row = urg_hdr + 1 + max(len(urgency_bands), len(top8)) + 1
ws1.row_dimensions[footer_row].height = 16
ws1.merge_cells(start_row=footer_row, start_column=2, end_row=footer_row, end_column=13)
c = ws1.cell(footer_row, 2)
c.value = f"Generated by TenderRadar Automated Intelligence System  ·  IDCG Consulting  ·  {TODAY.strftime('%d %B %Y')}"
c.font = make_font("Arial", 8, italic=True, color=GRAY_MD)
c.alignment = make_align("center", "center")

_log.info(f"  Sheet 1 done — rows up to ~{footer_row}")

# ─────────────────────────────────────────────
# SHEET 2: Live Pipeline
# ─────────────────────────────────────────────

ws2 = wb.create_sheet("🔥 Live Pipeline")
ws2.sheet_properties.tabColor = RED

# Column widths
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 38
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 13
ws2.column_dimensions["E"].width = 9
ws2.column_dimensions["F"].width = 8
ws2.column_dimensions["G"].width = 20
ws2.column_dimensions["H"].width = 18
ws2.column_dimensions["I"].width = 12
ws2.column_dimensions["J"].width = 30
ws2.column_dimensions["K"].width = 10
ws2.column_dimensions["L"].width = 2

# Row 1: Header
ws2.row_dimensions[1].height = 30
ws2.merge_cells("A1:L1")
c = ws2["A1"]
c.value = "  TenderRadar™  —  Live Opportunity Pipeline  |  Active & Closing Tenders Only"
c.font = make_font("Arial", 14, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")

# Row 2: Subtitle
ws2.row_dimensions[2].height = 16
ws2.merge_cells("A2:L2")
c = ws2["A2"]
c.value = (f"  {active_now} active opportunities across {portals_count} portals"
           f"  ·  Sorted by priority score  ·  {closing_7} closing within 7 days")
c.font = make_font("Arial", 9, italic=True, color=WHITE)
c.fill = make_fill(NAVY2)
c.alignment = make_align("left", "center")

# Row 3: Spacer
ws2.row_dimensions[3].height = 6

# Row 4: Column headers
ws2.row_dimensions[4].height = 18
pipe_headers = ["", "TITLE", "PORTAL", "DEADLINE", "DAYS LEFT",
                "SCORE", "SECTOR", "SERVICE TYPE", "STATUS",
                "OPPORTUNITY INSIGHT", "LINK", ""]
for col_i, hdr in enumerate(pipe_headers):
    col = col_i + 1
    c = ws2.cell(4, col)
    c.value = hdr
    c.font = make_font("Arial", 9, bold=True, color=WHITE)
    c.fill = make_fill(NAVY)
    c.alignment = make_align("center", "center")

# Apply autofilter
from openpyxl.worksheet.filters import AutoFilter
ws2.auto_filter.ref = "B4:K4"

# Freeze panes at row 5
ws2.freeze_panes = "B5"

# Data rows
for i, tender in enumerate(active_rows):
    row = 5 + i
    ws2.row_dimensions[row].height = 15
    rf = make_fill(WHITE) if i % 2 == 0 else make_fill(OFF_WHITE)
    bd = make_border(bottom="thin", color=GRAY_LT)

    def set_col(col, value, font=None, fill=None, align=None, border=None, hyperlink=None):
        c = ws2.cell(row, col)
        c.value = value
        c.font = font or make_font("Arial", 9)
        c.fill = fill or rf
        c.alignment = align or make_align("left", "center")
        c.border = border or bd
        if hyperlink:
            c.hyperlink = hyperlink
            c.style = "Hyperlink"
        return c

    # A: margin
    ws2.cell(row, 1).fill = rf

    # B: Title
    c = ws2.cell(row, 2)
    c.value = trunc(tender.get("Title") or "", 85)
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # C: Portal
    c = ws2.cell(row, 3)
    c.value = tender.get("Portal") or ""
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # D: Deadline
    c = ws2.cell(row, 4)
    c.value = tender.get("_deadline_fmt") or "—"
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # E: Days left
    c = ws2.cell(row, 5)
    dl = tender["_days_left"]
    c.value = dl if dl is not None else "—"
    if isinstance(dl, int) and dl <= 7:
        c.font = make_font("Arial", 9, bold=True, color=RED)
        c.fill = make_fill(RED_LT)
    elif isinstance(dl, int) and dl <= 30:
        c.font = make_font("Arial", 9, color=ORANGE)
        c.fill = make_fill(ORANGE_LT)
    elif isinstance(dl, int):
        c.font = make_font("Arial", 9, color=GREEN_DK)
        c.fill = rf
    else:
        c.font = make_font("Arial", 9, color=GRAY_MD)
        c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # F: Priority Score
    c = ws2.cell(row, 6)
    score = tender.get("Priority Score") or 0
    c.value = score
    if score >= 75:
        c.font = make_font("Arial", 9, bold=True, color="7D5A00")
        c.fill = make_fill("FFD700")
    elif score >= 50:
        c.font = make_font("Arial", 9, bold=True, color=NAVY)
        c.fill = make_fill(BLUE_HDR)
    else:
        c.font = make_font("Arial", 9, bold=True)
        c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # G: Sector
    c = ws2.cell(row, 7)
    sec_raw = tender.get("Sector") or ""
    secs = [s.strip() for s in sec_raw.split(",")][:2]
    c.value = ", ".join(secs) if secs else "—"
    c.font = make_font("Arial", 8)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # H: Service Type
    c = ws2.cell(row, 8)
    c.value = trunc(tender.get("Service Type") or "", 40)
    c.font = make_font("Arial", 8)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # I: Status badge
    c = ws2.cell(row, 9)
    st = tender["_status"]
    if st == "CRITICAL":
        c.value = "🔴 CRITICAL"
        c.font = make_font("Arial", 9, bold=True, color=RED)
        c.fill = make_fill(RED_LT)
    elif st == "URGENT":
        c.value = "🟡 URGENT"
        c.font = make_font("Arial", 9, bold=True, color=ORANGE)
        c.fill = make_fill(ORANGE_LT)
    else:
        c.value = "🟢 ACTIVE"
        c.font = make_font("Arial", 9, bold=True, color=GREEN_DK)
        c.fill = make_fill(GREEN_LT)
    c.alignment = make_align("center", "center")
    c.border = bd

    # J: Opportunity Insight
    c = ws2.cell(row, 10)
    c.value = trunc(tender.get("Opportunity Insight") or "", 80)
    c.font = make_font("Arial", 8, italic=True)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # K: URL Link
    c = ws2.cell(row, 11)
    url = tender.get("Tender URL") or ""
    if url:
        c.value = "🔗 Open"
        c.hyperlink = url
        c.font = make_font("Arial", 9, color="2E75B6")
        c.style = "Hyperlink"
    else:
        c.value = "—"
        c.font = make_font("Arial", 9, color=GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # L: margin
    ws2.cell(row, 12).fill = rf

_log.info(f"  Sheet 2 done — {len(active_rows)} active rows")

# ─────────────────────────────────────────────
# SHEET 3: Portal Health
# ─────────────────────────────────────────────

ws3 = wb.create_sheet("📊 Portal Health")
ws3.sheet_properties.tabColor = GREEN_DK

# Column widths
ws3.column_dimensions["A"].width = 2
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 8
ws3.column_dimensions["E"].width = 8
ws3.column_dimensions["F"].width = 8
ws3.column_dimensions["G"].width = 8
ws3.column_dimensions["H"].width = 8
ws3.column_dimensions["I"].width = 8
ws3.column_dimensions["J"].width = 10
ws3.column_dimensions["K"].width = 10
ws3.column_dimensions["L"].width = 10
ws3.column_dimensions["M"].width = 16
ws3.column_dimensions["N"].width = 2

# Row 1
ws3.row_dimensions[1].height = 30
ws3.merge_cells("A1:N1")
c = ws3["A1"]
c.value = "  TenderRadar™  —  Portal Health & Coverage Report"
c.font = make_font("Arial", 14, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")

# Row 2
ws3.row_dimensions[2].height = 16
ws3.merge_cells("A2:N2")
c = ws3["A2"]
c.value = f"  {portals_count} portals monitored  ·  Showing per-portal intelligence metrics  ·  As of {TODAY.strftime('%d %B %Y')}"
c.font = make_font("Arial", 9, italic=True, color=WHITE)
c.fill = make_fill(NAVY2)
c.alignment = make_align("left", "center")

# Row 3: Spacer
ws3.row_dimensions[3].height = 8

# Row 4: Section header
ws3.row_dimensions[4].height = 22
ws3.merge_cells("B4:M4")
c = ws3["B4"]
c.value = "  PORTAL INTELLIGENCE SCORECARDS"
c.font = make_font("Arial", 13, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")
for col in range(2, 14):
    ws3.cell(4, col).fill = make_fill(NAVY)

# Row 5: Column headers
ws3.row_dimensions[5].height = 18
ph_headers = [
    "PORTAL", "GROUP", "TOTAL", "ACTIVE", "CRITICAL",
    "URGENT", "EXPIRED", "UNKNOWN", "LIVE %", "HIGH PRIO",
    "AVG SCORE", "LAST SCRAPED"
]
for i, hdr in enumerate(ph_headers):
    col = i + 2  # starts at B=2
    c = ws3.cell(5, col)
    c.value = hdr
    c.font = make_font("Arial", 9, bold=True, color=WHITE)
    c.fill = make_fill(NAVY)
    c.alignment = make_align("center", "center")
    c.border = make_border(bottom="medium", color=WHITE)

# Rows 6-23: Portal data
for i, (portal, ps) in enumerate(portals_sorted):
    row = 6 + i
    ws3.row_dimensions[row].height = 15
    rf = make_fill(WHITE) if i % 2 == 0 else make_fill(OFF_WHITE)
    bd = make_border(bottom="thin", color=GRAY_LT)

    total = ps["total"]
    active = ps["active"]
    critical = ps["critical"]
    urgent = ps["urgent"]
    expired = ps["expired"]
    unknown = ps["unknown"]
    live_pct = (active / total * 100) if total > 0 else 0
    group = PORTAL_GROUPS.get(portal, "International")

    # Highlight stale portals
    if active == 0 and total > 10:
        rf = make_fill(YELLOW_LT)

    c = ws3.cell(row, 2)
    c.value = portal
    c.font = make_font("Arial", 9, bold=True)
    c.fill = rf
    c.alignment = make_align("left", "center")
    c.border = bd

    c = ws3.cell(row, 3)
    c.value = group
    c.font = make_font("Arial", 8, color=GRAY_DK)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 4)
    c.value = total
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 5)
    c.value = active
    c.font = make_font("Arial", 9, color=GREEN_DK if active > 0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 6)
    c.value = critical
    c.font = make_font("Arial", 9, bold=critical>0, color=RED if critical>0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 7)
    c.value = urgent
    c.font = make_font("Arial", 9, color=ORANGE if urgent>0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 8)
    c.value = expired
    c.font = make_font("Arial", 9, color=GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 9)
    c.value = unknown
    c.font = make_font("Arial", 9, italic=True, color=GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 10)
    c.value = f"{live_pct:.0f}%"
    if live_pct >= 50:
        c.font = make_font("Arial", 9, bold=True, color=GREEN_DK)
    elif live_pct >= 10:
        c.font = make_font("Arial", 9, bold=True, color=ORANGE)
    else:
        c.font = make_font("Arial", 9, bold=True, color=RED)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 11)
    c.value = ps["high_prio"]
    c.font = make_font("Arial", 9, bold=ps["high_prio"]>0, color=GOLD if ps["high_prio"]>0 else GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 12)
    c.value = ps["avg_score"] if ps["avg_score"] > 0 else "—"
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    c = ws3.cell(row, 13)
    ls = ps.get("last_scraped") or ""
    c.value = str(ls)[:16] if ls else "—"
    c.font = make_font("Arial", 8, color=GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

n_portals = len(portals_sorted)
spacer3_row = 6 + n_portals

ws3.row_dimensions[spacer3_row].height = 10

# Portal groups summary
grp_hdr_row = spacer3_row + 1
ws3.row_dimensions[grp_hdr_row].height = 22
ws3.merge_cells(start_row=grp_hdr_row, start_column=2, end_row=grp_hdr_row, end_column=13)
c = ws3.cell(grp_hdr_row, 2)
c.value = "  PORTAL GROUPS SUMMARY"
c.font = make_font("Arial", 13, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")
for col in range(2, 14):
    ws3.cell(grp_hdr_row, col).fill = make_fill(NAVY)

# Group summary headers
grp_col_hdr = grp_hdr_row + 1
ws3.row_dimensions[grp_col_hdr].height = 16
for i, hdr in enumerate(["GROUP", "PORTALS", "TOTAL TENDERS", "ACTIVE", "LIVE %"]):
    col = i + 2
    c = ws3.cell(grp_col_hdr, col)
    c.value = hdr
    c.font = make_font("Arial", 9, bold=True, color=NAVY)
    c.fill = make_fill(BLUE_HDR)
    c.alignment = make_align("center", "center")

# Compute group stats
group_stats = defaultdict(lambda: {"portal_count": 0, "total": 0, "active": 0})
for portal, ps in portals_sorted:
    g = PORTAL_GROUPS.get(portal, "International")
    group_stats[g]["portal_count"] += 1
    group_stats[g]["total"] += ps["total"]
    group_stats[g]["active"] += ps["active"]

groups_order = ["India Central", "India State", "International", "NGO/Civil Society"]
grp_colors = [BLUE_LIGHT, GREEN_LT, ORANGE_LT, PURPLE_LT]

for i, grp in enumerate(groups_order):
    row = grp_col_hdr + 1 + i
    ws3.row_dimensions[row].height = 15
    gs = group_stats.get(grp, {"portal_count": 0, "total": 0, "active": 0})
    rf = make_fill(grp_colors[i])
    live_pct = (gs["active"] / gs["total"] * 100) if gs["total"] > 0 else 0
    bd = make_border(bottom="thin", color=GRAY_LT)

    for col, val in enumerate([grp, gs["portal_count"], gs["total"], gs["active"], f"{live_pct:.0f}%"], start=2):
        c = ws3.cell(row, col)
        c.value = val
        c.font = make_font("Arial", 9, bold=(col == 2))
        c.fill = rf
        c.alignment = make_align("center" if col > 2 else "left", "center")
        c.border = bd

_log.info(f"  Sheet 3 done — {n_portals} portals + group summary")

# ─────────────────────────────────────────────
# SHEET 4: All Tenders
# ─────────────────────────────────────────────

ws4 = wb.create_sheet("📋 All Tenders")
ws4.sheet_properties.tabColor = GRAY_DK

# Column widths
ws4.column_dimensions["A"].width = 2
ws4.column_dimensions["B"].width = 14
ws4.column_dimensions["C"].width = 40
ws4.column_dimensions["D"].width = 20
ws4.column_dimensions["E"].width = 14
ws4.column_dimensions["F"].width = 8
ws4.column_dimensions["G"].width = 10
ws4.column_dimensions["H"].width = 8
ws4.column_dimensions["I"].width = 10
ws4.column_dimensions["J"].width = 20
ws4.column_dimensions["K"].width = 18
ws4.column_dimensions["L"].width = 35
ws4.column_dimensions["M"].width = 10
ws4.column_dimensions["N"].width = 2

# Row 1
ws4.row_dimensions[1].height = 28
ws4.merge_cells("A1:N1")
c = ws4["A1"]
c.value = f"  TenderRadar™  —  Complete Tender Database  |  {total_tenders} Records"
c.font = make_font("Arial", 14, bold=True, color=WHITE)
c.fill = make_fill(NAVY)
c.alignment = make_align("left", "center")

# Row 2
ws4.row_dimensions[2].height = 16
ws4.merge_cells("A2:N2")
c = ws4["A2"]
c.value = "  Includes all portals, statuses, and time periods  ·  Use filters to explore"
c.font = make_font("Arial", 9, italic=True, color=WHITE)
c.fill = make_fill(NAVY2)
c.alignment = make_align("left", "center")

# Row 3: Spacer
ws4.row_dimensions[3].height = 6

# Row 4: Column headers
ws4.row_dimensions[4].height = 18
all_headers = [
    "", "PORTAL", "TITLE", "ORGANIZATION", "DEADLINE",
    "DAYS LEFT", "STATUS", "PRIORITY SCORE", "REL. SCORE",
    "SECTOR", "SERVICE TYPE", "OPPORTUNITY INSIGHT", "LINK", ""
]
for i, hdr in enumerate(all_headers):
    col = i + 1
    c = ws4.cell(4, col)
    c.value = hdr
    c.font = make_font("Arial", 9, bold=True, color=WHITE)
    c.fill = make_fill(NAVY)
    c.alignment = make_align("center", "center")

# AutoFilter and freeze
ws4.auto_filter.ref = "B4:M4"
ws4.freeze_panes = "B5"

# Data rows (all 741)
all_rows_sorted = sorted(rows, key=lambda r: (-(r.get("Priority Score") or 0)))
for i, tender in enumerate(all_rows_sorted):
    row = 5 + i
    ws4.row_dimensions[row].height = 14
    rf = make_fill(WHITE) if i % 2 == 0 else make_fill(OFF_WHITE)
    bd = make_border(bottom="thin", color=GRAY_LT)

    # Margin
    ws4.cell(row, 1).fill = rf
    ws4.cell(row, 14).fill = rf

    # B: Portal
    c = ws4.cell(row, 2)
    c.value = tender.get("Portal") or ""
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("left", "center")
    c.border = bd

    # C: Title
    c = ws4.cell(row, 3)
    c.value = trunc(tender.get("Title") or "", 100)
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # D: Organization
    c = ws4.cell(row, 4)
    c.value = trunc(tender.get("Organization") or "", 40)
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("left", "center")
    c.border = bd

    # E: Deadline
    c = ws4.cell(row, 5)
    c.value = tender.get("_deadline_fmt") or "—"
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # F: Days left
    c = ws4.cell(row, 6)
    dl = tender["_days_left"]
    c.value = dl if dl is not None else "—"
    if isinstance(dl, int) and dl <= 7:
        c.font = make_font("Arial", 9, bold=True, color=RED)
        c.fill = make_fill(RED_LT)
    elif isinstance(dl, int) and dl <= 30:
        c.font = make_font("Arial", 9, color=ORANGE)
        c.fill = rf
    elif isinstance(dl, int) and dl < 0:
        c.font = make_font("Arial", 9, color=GRAY_MD)
        c.fill = rf
    else:
        c.font = make_font("Arial", 9)
        c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # G: Status
    c = ws4.cell(row, 7)
    st = tender["_status"]
    status_map = {
        "CRITICAL": ("🔴 CRITICAL", RED,     RED_LT),
        "URGENT":   ("🟡 URGENT",   ORANGE,  ORANGE_LT),
        "ACTIVE":   ("🟢 ACTIVE",   GREEN_DK, GREEN_LT),
        "EXPIRED":  ("⏰ EXPIRED",  GRAY_DK, "F5F5F5"),
        "UNKNOWN":  ("❓ UNKNOWN",  GRAY_MD, "FAFAFA"),
    }
    label, color, bg = status_map.get(st, ("—", GRAY_MD, WHITE))
    c.value = label
    c.font = make_font("Arial", 8, bold=True, color=color)
    c.fill = make_fill(bg)
    c.alignment = make_align("center", "center")
    c.border = bd

    # H: Priority Score
    c = ws4.cell(row, 8)
    score = tender.get("Priority Score") or 0
    c.value = score
    if score >= 75:
        c.font = make_font("Arial", 9, bold=True, color="7D5A00")
        c.fill = make_fill("FFD700")
    elif score >= 50:
        c.font = make_font("Arial", 9, bold=True, color=NAVY)
        c.fill = make_fill(BLUE_HDR)
    else:
        c.font = make_font("Arial", 9)
        c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # I: Relevance Score
    c = ws4.cell(row, 9)
    c.value = tender.get("Relevance Score") or 0
    c.font = make_font("Arial", 9)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

    # J: Sector
    c = ws4.cell(row, 10)
    c.value = trunc(tender.get("Sector") or "", 40)
    c.font = make_font("Arial", 8)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # K: Service Type
    c = ws4.cell(row, 11)
    c.value = trunc(tender.get("Service Type") or "", 45)
    c.font = make_font("Arial", 8)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # L: Opportunity Insight
    c = ws4.cell(row, 12)
    c.value = trunc(tender.get("Opportunity Insight") or "", 90)
    c.font = make_font("Arial", 8, italic=True)
    c.fill = rf
    c.alignment = make_align("left", "center", wrap=True)
    c.border = bd

    # M: URL
    c = ws4.cell(row, 13)
    url = tender.get("Tender URL") or ""
    if url:
        c.value = "🔗 Link"
        c.hyperlink = url
        c.font = make_font("Arial", 9, color="2E75B6")
        c.style = "Hyperlink"
    else:
        c.value = "—"
        c.font = make_font("Arial", 9, color=GRAY_MD)
    c.fill = rf
    c.alignment = make_align("center", "center")
    c.border = bd

_log.info(f"  Sheet 4 done — {len(all_rows_sorted)} rows")

# ─────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────

_log.info(f"\nSaving to {OUTPUT}...")
wb.save(OUTPUT)
_log.info("✓ Saved successfully!")

# Print summary
_log.info("\n=== SHEET SUMMARY ===")
_log.info(f"Sheet 1 (Command Center):  KPIs + Top {len(top_opps)} opps + {len(portals_sorted)} portals + urgency/sectors")
_log.info(f"Sheet 2 (Live Pipeline):   {len(active_rows)} active tenders")
_log.info(f"Sheet 3 (Portal Health):   {n_portals} portals + group summary")
_log.info(f"Sheet 4 (All Tenders):     {len(all_rows_sorted)} total tenders")
