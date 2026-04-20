#!/usr/bin/env python3
# =============================================================================
# scripts/validate_system.py — TenderRadar Pre-Deployment Validation
#
# Usage:
#   python3 scripts/validate_system.py            # all checks
#   python3 scripts/validate_system.py --quick    # skip slow checks
#   python3 scripts/validate_system.py --api      # API health only
#
# Exit codes:
#   0  — all checks passed
#   1  — one or more checks failed
#   2  — critical failure (DB or config missing)
# =============================================================================

from __future__ import annotations

import argparse
import importlib
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, List, Optional, Tuple

# ── Add project root to path ──────────────────────────────────────────────────
BASE_DIR = os.path.expanduser("~/tender_system")
sys.path.insert(0, BASE_DIR)
os.chdir(BASE_DIR)

# ── Colour helpers (no deps) ──────────────────────────────────────────────────
_GREEN  = "\033[92m"
_YELLOW = "\033[93m"
_RED    = "\033[91m"
_CYAN   = "\033[96m"
_BOLD   = "\033[1m"
_RESET  = "\033[0m"

def ok(msg: str)   -> str: return f"{_GREEN}  ✓  {_RESET}{msg}"
def warn(msg: str) -> str: return f"{_YELLOW}  ⚠  {_RESET}{msg}"
def fail(msg: str) -> str: return f"{_RED}  ✗  {_RESET}{msg}"
def info(msg: str) -> str: return f"{_CYAN}  ·  {_RESET}{msg}"
def hdr(msg: str)  -> str: return f"\n{_BOLD}{_CYAN}{'─'*60}{_RESET}\n{_BOLD}  {msg}{_RESET}\n{'─'*60}"


# =============================================================================
# Result accumulator
# =============================================================================

class ValidationReport:
    def __init__(self):
        self.results: List[Tuple[str, str, str]] = []   # (status, section, message)
        self.start   = time.time()

    def add(self, status: str, section: str, msg: str, detail: str = ""):
        full = f"{msg}  {detail}".strip()
        self.results.append((status, section, full))
        symbol = {"PASS": ok, "WARN": warn, "FAIL": fail, "INFO": info}[status]
        print(symbol(full))

    def summary(self):
        elapsed = time.time() - self.start
        passes  = sum(1 for s, _, _ in self.results if s == "PASS")
        warns   = sum(1 for s, _, _ in self.results if s == "WARN")
        fails   = sum(1 for s, _, _ in self.results if s == "FAIL")
        total   = passes + warns + fails
        print(f"\n{'='*60}")
        print(f"{_BOLD}Validation complete in {elapsed:.1f}s{_RESET}")
        print(f"  {_GREEN}{passes} passed{_RESET}  "
              f"{_YELLOW}{warns} warnings{_RESET}  "
              f"{_RED}{fails} failed{_RESET}  "
              f"({total} checks)")
        print("="*60)
        return fails


rep = ValidationReport()


# =============================================================================
# SECTION 1 — Configuration
# =============================================================================

def check_config():
    print(hdr("1. Configuration"))
    try:
        import config as cfg
        rep.add("PASS", "config", "config.py imports OK")
    except Exception as e:
        rep.add("FAIL", "config", f"config.py import failed: {e}")
        return

    for attr in ("OUTPUT_DIR", "LOG_FILE", "UNIFIED_EXCEL_PATH",
                 "NOTIFICATIONS_ENABLED"):
        val = getattr(cfg, attr, None)
        if val is None:
            rep.add("WARN", "config", f"config.{attr} is not defined")
        else:
            rep.add("PASS", "config", f"config.{attr}", f"= {str(val)[:60]}")

    # OUTPUT_DIR must exist
    out_dir = getattr(cfg, "OUTPUT_DIR", None)
    if out_dir:
        if os.path.isdir(out_dir):
            rep.add("PASS", "config", f"OUTPUT_DIR exists: {out_dir}")
        else:
            try:
                os.makedirs(out_dir, exist_ok=True)
                rep.add("WARN", "config", f"OUTPUT_DIR created (was missing): {out_dir}")
            except Exception as e:
                rep.add("FAIL", "config", f"OUTPUT_DIR missing and cannot create: {e}")


# =============================================================================
# SECTION 2 — Database
# =============================================================================

def check_database():
    print(hdr("2. Database"))
    try:
        from database.db import init_db, get_connection, get_stats
    except Exception as e:
        rep.add("FAIL", "db", f"db.py import failed: {e}")
        return

    # Connection test
    try:
        conn = get_connection()
        conn.close()
        rep.add("PASS", "db", "MySQL connection OK")
    except Exception as e:
        rep.add("FAIL", "db", f"MySQL connection failed: {e}")
        return  # no point continuing

    # init_db (idempotent)
    try:
        init_db()
        rep.add("PASS", "db", "init_db() ran without error")
    except Exception as e:
        rep.add("FAIL", "db", f"init_db() failed: {e}")

    # Check tables exist
    expected_tables = ["seen_tenders", "tenders", "tender_hashes"]
    try:
        conn = get_connection()
        cur  = conn.cursor()
        cur.execute("SHOW TABLES")
        existing = {row[0] for row in cur.fetchall()}
        conn.close()
        for tbl in expected_tables:
            if tbl in existing:
                rep.add("PASS", "db", f"Table `{tbl}` exists")
            else:
                rep.add("WARN", "db", f"Table `{tbl}` NOT found — may need migration")
    except Exception as e:
        rep.add("WARN", "db", f"Could not check tables: {e}")

    # get_stats
    try:
        stats = get_stats()
        total = sum(stats.values())
        rep.add("PASS", "db", f"get_stats() OK — {total} total seen_tenders across {len(stats)} portals")
    except Exception as e:
        rep.add("WARN", "db", f"get_stats() failed: {e}")

    # DRY_RUN flag
    try:
        import database.db as _db
        if hasattr(_db, "DRY_RUN"):
            rep.add("PASS", "db", f"DRY_RUN flag present (currently={_db.DRY_RUN})")
        else:
            rep.add("FAIL", "db", "DRY_RUN flag missing from db.py")
    except Exception:
        pass


# =============================================================================
# SECTION 3 — Pipeline registry and imports
# =============================================================================

def check_pipeline():
    print(hdr("3. Pipeline Registry"))
    try:
        from core.registry import all_jobs
        jobs = all_jobs()
        rep.add("PASS", "pipeline", f"pipeline.registry imported — {len(jobs)} jobs registered")
    except Exception as e:
        rep.add("FAIL", "pipeline", f"pipeline.registry import failed: {e}")
        return

    # Check each job has required attrs
    for job in jobs:
        for attr in ("flag", "label", "module", "fn"):
            if not hasattr(job, attr):
                rep.add("WARN", "pipeline", f"Job missing attr `{attr}`: {getattr(job, 'flag', '?')}")

    # Runner
    try:
        from core.runner import JobRunner
        rep.add("PASS", "pipeline", "pipeline.runner.JobRunner imports OK")
    except Exception as e:
        rep.add("FAIL", "pipeline", f"pipeline.runner import failed: {e}")

    # Reporter
    try:
        from core.reporter import RunReporter
        rep.add("PASS", "pipeline", "pipeline.reporter.RunReporter imports OK")
    except Exception as e:
        rep.add("FAIL", "pipeline", f"pipeline.reporter import failed: {e}")

    # Excel exporter
    try:
        from exporters.excel_exporter import write_unified_excel
        rep.add("PASS", "pipeline", "pipeline.excel_exporter.write_unified_excel imports OK")
    except Exception as e:
        rep.add("FAIL", "pipeline", f"pipeline.excel_exporter import failed: {e}")


# =============================================================================
# SECTION 4 — Intelligence layer
# =============================================================================

def check_intelligence():
    print(hdr("4. Intelligence Layer"))

    # Classifier
    try:
        from intelligence.classifier import classify_tender
        result = classify_tender(
            "Evaluation of health and WASH programme in rural India",
            "Seeking consulting firm for mid-term evaluation and capacity building"
        )
        sectors  = result.sectors
        services = result.service_types
        rep.add("PASS", "intelligence",
                f"classify_tender() OK — sectors={sectors}, services={services}")
    except Exception as e:
        rep.add("FAIL", "intelligence", f"classify_tender() failed: {e}")

    # Normalizer
    try:
        from intelligence.normalizer import normalize_tender
        sample = {
            "title": "Test Tender",
            "organization": "Test Org",
            "deadline": "2026-06-30",
            "url": "https://example.com/tender/1",
        }
        norm = normalize_tender(sample, tender_id="test-001", source_portal="test")
        rep.add("PASS", "intelligence", f"normalize_tender() OK — id={norm.tender_id}")
    except Exception as e:
        rep.add("WARN", "intelligence", f"normalize_tender() skipped/failed: {e}")

    # Deduplicator
    try:
        from intelligence.deduplicator import check_duplicate, DedupResult
        rep.add("PASS", "intelligence", "intelligence.deduplicator imports OK")
    except Exception as e:
        rep.add("WARN", "intelligence", f"intelligence.deduplicator import failed: {e}")

    # intelligence_layer process_batch (import only — don't run full batch)
    try:
        from intelligence.intelligence_layer import process_batch
        rep.add("PASS", "intelligence", "intelligence_layer.process_batch imports OK")
    except Exception as e:
        rep.add("WARN", "intelligence", f"intelligence_layer import failed: {e}")


# =============================================================================
# SECTION 5 — Excel outputs
# =============================================================================

def check_excel():
    print(hdr("5. Excel Outputs"))
    try:
        import config as cfg
        unified_path = getattr(cfg, "UNIFIED_EXCEL_PATH", None)
        portal_excels_dir = getattr(cfg, "PORTAL_EXCELS_DIR", None)
    except Exception as e:
        rep.add("FAIL", "excel", f"Cannot read config: {e}")
        return

    if not unified_path:
        rep.add("WARN", "excel", "UNIFIED_EXCEL_PATH not set in config")
    elif os.path.isfile(unified_path):
        size_kb = os.path.getsize(unified_path) // 1024
        mtime   = datetime.fromtimestamp(os.path.getmtime(unified_path))
        rep.add("PASS", "excel",
                f"Unified Excel exists: {unified_path}",
                f"({size_kb} KB, last modified {mtime:%Y-%m-%d %H:%M})")

        # Validate column headers
        try:
            from openpyxl import load_workbook
            wb  = load_workbook(unified_path, read_only=True, data_only=True)
            ws  = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            required = {"Portal", "Title", "Organization", "Country",
                        "Deadline", "Sectors", "Service Types",
                        "Relevance", "Detail Link", "Scraped Date"}
            missing  = required - set(headers)
            if missing:
                rep.add("FAIL", "excel", f"Missing columns in unified Excel: {missing}")
            else:
                row_count = ws.max_row - 1  # subtract header
                rep.add("PASS", "excel",
                        f"All required columns present, {row_count} data rows")
            wb.close()
        except Exception as e:
            rep.add("WARN", "excel", f"Could not validate unified Excel columns: {e}")
    else:
        rep.add("WARN", "excel",
                f"Unified Excel not yet generated: {unified_path}",
                "(run main.py once to create it)")

    # Check per-portal Excel files (all pipelines write to portal_excels/)
    if portal_excels_dir and os.path.isdir(portal_excels_dir):
        xlsx_files = list(Path(portal_excels_dir).glob("*.xlsx"))
        if xlsx_files:
            rep.add("PASS", "excel",
                    f"{len(xlsx_files)} per-portal Excel file(s) in portal_excels/")
            for xf in sorted(xlsx_files, key=os.path.getmtime, reverse=True)[:5]:
                sz = os.path.getsize(xf) // 1024
                mt = datetime.fromtimestamp(os.path.getmtime(xf))
                rep.add("INFO", "excel", f"  {xf.name}", f"({sz} KB, {mt:%Y-%m-%d %H:%M})")
        else:
            rep.add("WARN", "excel",
                    "No per-portal Excel files found in portal_excels/",
                    "(run a scraper to generate them)")
    else:
        rep.add("WARN", "excel", "PORTAL_EXCELS_DIR missing — cannot check per-portal Excel")


# =============================================================================
# SECTION 6 — Logging and monitoring
# =============================================================================

def check_logging():
    print(hdr("6. Logging & Monitoring"))
    try:
        import config as cfg
        log_file = getattr(cfg, "LOG_FILE", None)
    except Exception:
        log_file = None

    if log_file:
        if os.path.isfile(log_file):
            size = os.path.getsize(log_file)
            rep.add("PASS", "logging", f"run.log exists ({size:,} bytes): {log_file}")
        else:
            rep.add("WARN", "logging",
                    f"run.log not yet created: {log_file}",
                    "(normal on first run)")
    else:
        rep.add("WARN", "logging", "LOG_FILE not set in config")

    # Structured log
    structured_log = os.path.join(BASE_DIR, "monitoring", "tenderradar.log")
    if os.path.isfile(structured_log):
        size = os.path.getsize(structured_log)
        rep.add("PASS", "logging", f"Structured log exists ({size:,} bytes)")
    else:
        rep.add("INFO", "logging",
                "Structured log not yet created (created on first run)")

    # Logger setup
    try:
        from monitoring.logs import setup_logging
        rep.add("PASS", "logging", "pipeline.logger.setup_logging imports OK")
    except Exception as e:
        rep.add("WARN", "logging", f"pipeline.logger import failed: {e}")


# =============================================================================
# SECTION 7 — Notifier
# =============================================================================

def check_notifier():
    print(hdr("7. Notifier"))
    try:
        from notifier import notify_all, send_rich_alert
        rep.add("PASS", "notifier", "notifier.py imports OK")
    except Exception as e:
        rep.add("WARN", "notifier", f"notifier.py import failed: {e}")
        return

    try:
        import config as cfg
        enabled = getattr(cfg, "NOTIFICATIONS_ENABLED", None)
        if enabled is True:
            rep.add("PASS", "notifier", "NOTIFICATIONS_ENABLED = True")
        elif enabled is False:
            rep.add("WARN", "notifier", "NOTIFICATIONS_ENABLED = False (alerts silenced)")
        else:
            rep.add("WARN", "notifier", "NOTIFICATIONS_ENABLED not set in config")

        # Check Telegram token without sending
        tg_token = getattr(cfg, "TELEGRAM_BOT_TOKEN", "") or ""
        if tg_token and tg_token != "YOUR_BOT_TOKEN":
            rep.add("PASS", "notifier", f"TELEGRAM_BOT_TOKEN configured (…{tg_token[-6:]})")
        else:
            rep.add("WARN", "notifier", "TELEGRAM_BOT_TOKEN not configured")
    except Exception as e:
        rep.add("WARN", "notifier", f"Notifier config check failed: {e}")


# =============================================================================
# SECTION 8 — FastAPI backend (optional — skip if not running)
# =============================================================================

def check_api(quick: bool = False):
    print(hdr("8. FastAPI Backend"))
    if quick:
        rep.add("INFO", "api", "Skipped (--quick mode)")
        return

    import urllib.request, urllib.error

    api_url = os.environ.get("API_URL", "http://localhost:8000")

    def _get(path: str) -> Optional[dict]:
        try:
            req = urllib.request.urlopen(f"{api_url}{path}", timeout=5)
            return json.loads(req.read())
        except urllib.error.URLError:
            return None
        except Exception:
            return None

    health = _get("/health")
    if health is None:
        rep.add("WARN", "api",
                f"FastAPI not reachable at {api_url}",
                "(start with: uvicorn api.main:app --reload)")
        return

    rep.add("PASS", "api", f"FastAPI /health OK at {api_url}")

    # Stats endpoint
    stats = _get("/api/v1/stats")
    if stats:
        rep.add("PASS", "api",
                f"/api/v1/stats OK — total_tenders={stats.get('total_tenders', '?')}")
    else:
        rep.add("WARN", "api", "/api/v1/stats not reachable")

    # Portals endpoint
    portals = _get("/api/v1/portals")
    if portals:
        count = len(portals) if isinstance(portals, list) else portals.get("count", "?")
        rep.add("PASS", "api", f"/api/v1/portals OK — {count} portals")
    else:
        rep.add("WARN", "api", "/api/v1/portals not reachable")

    # Tenders list
    tenders = _get("/api/v1/tenders?page=1&page_size=5")
    if tenders:
        rep.add("PASS", "api",
                f"/api/v1/tenders OK — total={tenders.get('total', '?')}")
    else:
        rep.add("WARN", "api", "/api/v1/tenders not reachable")


# =============================================================================
# SECTION 9 — Dry-run smoke test
# =============================================================================

def check_dry_run(quick: bool = False):
    print(hdr("9. Dry-Run Mode"))
    if quick:
        rep.add("INFO", "dry_run", "Skipped (--quick mode)")
        return

    # Verify DRY_RUN flag mechanics
    try:
        import database.db as _db
        original = _db.DRY_RUN
        _db.DRY_RUN = True

        # mark_as_seen should be a no-op
        result = _db.mark_as_seen("__validate_dry_run__", "test_source")
        _db.DRY_RUN = original

        if result is True:
            rep.add("PASS", "dry_run", "mark_as_seen() correctly no-ops under DRY_RUN")
        else:
            rep.add("WARN", "dry_run",
                    f"mark_as_seen() under DRY_RUN returned unexpected: {result}")
    except Exception as e:
        rep.add("WARN", "dry_run", f"DRY_RUN mark_as_seen test failed: {e}")

    # Check argparse flag wiring in main.py
    main_src = Path(BASE_DIR) / "main.py"
    if main_src.exists():
        src = main_src.read_text()
        checks = [
            ("--dry-run arg",    "--dry-run"         in src),
            ("DRY_RUN = True",   "DRY_RUN = True"    in src),
            ("write_unified_excel call", "write_unified_excel" in src),
        ]
        for label, found in checks:
            if found:
                rep.add("PASS", "dry_run", f"main.py: {label} present")
            else:
                rep.add("FAIL", "dry_run", f"main.py: {label} NOT found")
    else:
        rep.add("FAIL", "dry_run", "main.py not found")


# =============================================================================
# SECTION 10 — Python dependencies
# =============================================================================

def check_dependencies():
    print(hdr("10. Python Dependencies"))
    deps = [
        ("requests",      "requests"),
        ("bs4",           "beautifulsoup4"),
        ("openpyxl",      "openpyxl"),
        ("mysql.connector","mysql-connector-python"),
        ("pdfplumber",    "pdfplumber"),
        ("selenium",      "selenium"),
        ("fastapi",       "fastapi"),
        ("uvicorn",       "uvicorn"),
        ("pydantic",      "pydantic"),
    ]
    for module, pip_name in deps:
        spec = importlib.util.find_spec(module.split(".")[0])
        if spec is not None:
            try:
                mod = importlib.import_module(module.split(".")[0])
                ver = getattr(mod, "__version__", "?")
                rep.add("PASS", "deps", f"{pip_name} installed", f"v{ver}")
            except Exception:
                rep.add("PASS", "deps", f"{pip_name} importable")
        else:
            rep.add("WARN", "deps", f"{pip_name} NOT installed",
                    f"(pip install {pip_name})")

    # Optional: chromadb
    spec = importlib.util.find_spec("chromadb")
    if spec:
        rep.add("PASS", "deps", "chromadb installed (semantic dedup enabled)")
    else:
        rep.add("INFO", "deps",
                "chromadb not installed (semantic dedup disabled — hash-only dedup active)")


# =============================================================================
# ENTRY POINT
# =============================================================================

import importlib.util   # ensure available for check_dependencies

def main():
    parser = argparse.ArgumentParser(description="TenderRadar Pre-Deployment Validator")
    parser.add_argument("--quick", action="store_true",
                        help="Skip slow checks (API ping, dry-run test)")
    parser.add_argument("--api", action="store_true",
                        help="Run only the API health check")
    args = parser.parse_args()

    print(f"\n{_BOLD}{_CYAN}TenderRadar — Pre-Deployment Validation{_RESET}")
    print(f"  Run at : {datetime.now():%Y-%m-%d %H:%M:%S}")
    print(f"  Base   : {BASE_DIR}")

    if args.api:
        check_api(quick=False)
    else:
        check_config()
        check_database()
        check_pipeline()
        check_intelligence()
        check_excel()
        check_logging()
        check_notifier()
        check_api(quick=args.quick)
        check_dry_run(quick=args.quick)
        check_dependencies()

    fails = rep.summary()
    sys.exit(0 if fails == 0 else 1)


if __name__ == "__main__":
    main()
