"""
Tender Evidence Packaging Layer — TenderRadar
============================================

Creates structured, reusable evidence packages for NEW shortlisted tenders.
Each package is stored under:

    output/tender_packages/YYYY-MM-DD/<sanitized_title_or_id>/
        metadata.json       — machine-readable tender facts
        summary.txt         — analyst-readable one-pager
        evidence.txt        — evidence state + document notes
        documents/          — downloaded attachments (if any)

Only processes:
    - Is New == YES
    - AI_Suggested_Label in {Relevant, Borderline}
    - Tenders not already packaged today (dedup via run_state.json)

Evidence states (deterministic):
    SIGNAL_ONLY      — URL only; no page content, no docs
    PAGE_ONLY        — page content present, no doc attachments
    PARTIAL_PACKAGE  — some docs found but package incomplete
    FULL_PACKAGE     — full content + doc attachments available

No semantic search. No email sending. No Excel modification.
"""

from __future__ import annotations

import json
import logging
import os
import re
import time
import urllib.parse
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("tenderradar.evidence_packager")

# ── Paths ──────────────────────────────────────────────────────────────────────
_PROJECT_ROOT    = Path(__file__).resolve().parent.parent
_OUTPUT_DIR      = _PROJECT_ROOT / "output"
_PACKAGES_ROOT   = _OUTPUT_DIR / "tender_packages"
_WORKBOOK_PATH   = _OUTPUT_DIR / "Tender_Monitor_Master.xlsx"

# ── Constants ──────────────────────────────────────────────────────────────────
_SHORTLIST_LABELS = {"Relevant", "Borderline"}
_MAX_SAFE_DIRNAME = 60    # chars
_DOWNLOAD_TIMEOUT = 15    # seconds per document
_MAX_DOCS         = 5     # max attachments to attempt per tender

# ── CAPSTAT client tiers (for IDCG fit reasoning) ─────────────────────────────
_CLIENT_TIER1 = {
    "world bank", "undp", "giz", "unicef", "fao", "wfp", "adb", "afdb",
    "ifc", "kfw", "kreditanstalt", "afd", "jica", "usaid", "fcdo", "dfid",
    "european commission", "european union", "ifad", "ilo", "mcc",
}
_CLIENT_TIER2_PAT = re.compile(
    r"\b(IUCN|Winrock|TNC|Room to Read|Save the Children|Tata Trust|"
    r"Hans Foundation|MicroSave|MSC|CUTS|CLASP|WRI|SELCO|Oxfam|CARE|"
    r"Plan International|British Council|Reliance Foundation|"
    r"Leadership for Equity|Genesis Analytics|Tanager|I4DI)\b",
    re.IGNORECASE,
)
_GEO_STRONG = re.compile(
    r"\b(india|rajasthan|gujarat|maharashtra|karnataka|kerala|tamil ?nadu|"
    r"andhra|telangana|odisha|bihar|jharkhand|uttarakhand|himachal|"
    r"chhattisgarh|madhya pradesh|uttar pradesh|tripura|assam|meghalaya|"
    r"nagaland|afghanistan|tajikistan|bangladesh|sri lanka|nepal)\b",
    re.IGNORECASE,
)
_GEO_WEAK = re.compile(
    r"\b(china|tanzania|kenya|nigeria|ethiopia|ghana|latin america|"
    r"sub-saharan|sub saharan|south africa|colombia|peru|brazil|cambodia|"
    r"vietnam|laos|indonesia)\b",
    re.IGNORECASE,
)
_SERVICE_STRONG = re.compile(
    r"\b(evaluation|M&E|MEL|baseline|midline|endline|TPM|IVA|"
    r"independent verification|third.?party|impact assessment)\b",
    re.IGNORECASE,
)
_SERVICE_MODERATE = re.compile(
    r"\b(advisory|technical assistance|capacity building|research|survey|"
    r"feasibility|diagnostic|policy|governance|institutional)\b",
    re.IGNORECASE,
)
_SECTOR_STRONG = re.compile(
    r"\b(education|health|nutrition|agriculture|forestry|climate|"
    r"environment|rural|livelihoods?|energy|renewable|gender|"
    r"governance|MSME|skills?|social protection|water|sanitation)\b",
    re.IGNORECASE,
)
_DOC_URL_PAT = re.compile(
    r"https?://\S+\.(?:pdf|docx?|xlsx?|zip|rar)\b",
    re.IGNORECASE,
)

# Shared maturity module (imported lazily to avoid circular imports at module load)
def _get_maturity_module():
    from intelligence.maturity import classify_evidence, maturity_label, maturity_digest_note, recommended_action
    return classify_evidence, maturity_label, maturity_digest_note, recommended_action


# ══════════════════════════════════════════════════════════════════════════════
# 1. WORKBOOK READER
# ══════════════════════════════════════════════════════════════════════════════

def _load_new_shortlisted(workbook_path: Path) -> List[Dict[str, Any]]:
    """
    Load all rows from master Excel where:
      Is New == YES
      AI_Suggested_Label in {Relevant, Borderline}
    Returns list of row dicts with normalized keys.
    """
    try:
        from openpyxl import load_workbook as _load_wb
    except ImportError:
        logger.error("[packager] openpyxl not available")
        return []

    try:
        wb = _load_wb(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        logger.error("[packager] Cannot open workbook: %s", exc)
        return []

    sheet_name = "All Tenders" if "All Tenders" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    rows = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        def _get(key: str, default: str = "") -> str:
            idx = col.get(key)
            if idx is None:
                return default
            v = row_vals[idx]
            return str(v).strip() if v is not None else default

        is_new   = _get("Is New").upper()
        ai_label = _get("AI_Suggested_Label")

        if is_new != "YES" or ai_label not in _SHORTLIST_LABELS:
            continue

        rows.append({
            "title":          _get("Title"),
            "portal":         _get("Portal"),
            "organization":   _get("Organization"),
            "country":        _get("Country"),
            "deadline":       _get("Deadline"),
            "priority_score": _get("Priority Score", "0"),
            "relevance_score": _get("Relevance Score", "0"),
            "sector":         _get("Sector"),
            "service_type":   _get("Service Type"),
            "ai_label":       ai_label,
            "human_label":    _get("Human_Label"),
            "label_reason":   _get("Label_Reason"),
            "tender_url":     _get("Tender URL"),
            "tender_id":      _get("Tender ID"),
            "scraped_date":   _get("Scraped Date"),
            "deep_scope":     _get("Deep Scope"),
            "ai_summary":     _get("AI Summary"),
            "evaluation_criteria": _get("Evaluation Criteria"),
            "relevance_reason":    _get("Relevance Reason"),
            "opportunity_insight": _get("Opportunity Insight"),
            "cross_sources":       _get("Cross Sources"),
            "evidence_state":      _get("Evidence_State"),
            "opportunity_maturity": _get("Opportunity_Maturity"),
            "recommended_action":  _get("Recommended_Action"),
            "scoring_note":        _get("Relevance Reason"),  # alias
        })
    wb.close()
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# 2. EVIDENCE STATE CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════

def _classify_evidence(row: Dict[str, Any]) -> Tuple[str, str, int]:
    """
    Returns (evidence_state, opportunity_maturity_label, doc_url_count).

    Delegates to intelligence.maturity for canonical classification.
    opportunity_maturity_label is the business-facing label:
      Signal First / Partial Package / Full Package
    """
    classify_evidence, maturity_label, _, _ = _get_maturity_module()
    evidence_state, doc_url_count = classify_evidence(row)
    return evidence_state, maturity_label(evidence_state), doc_url_count


def _fetch_rich_packager_data(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Rehydrate shortlisted workbook rows from DB so packager uses the same evidence
    layer as Excel maturity, rather than only the visible workbook columns.
    """
    ids = []
    urls = []
    for row in rows:
        tid = str(row.get("tender_id") or "").strip()
        url = str(row.get("tender_url") or "").strip()
        if tid:
            ids.append(tid)
        if url:
            urls.append(url)
    ids = list(dict.fromkeys(ids))
    urls = list(dict.fromkeys(urls))
    if not ids and not urls:
        return {}

    try:
        from database.db import get_connection

        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        clauses = []
        params: List[str] = []
        if ids:
            clauses.append("t.tender_id IN (" + ", ".join(["%s"] * len(ids)) + ")")
            params.extend(ids)
        if urls:
            clauses.append("t.url IN (" + ", ".join(["%s"] * len(urls)) + ")")
            params.extend(urls)
        cur.execute(
            f"""
            SELECT
                t.tender_id,
                t.url,
                t.description,
                t.deep_description,
                t.deep_scope,
                t.deep_eval_criteria,
                t.deep_pdf_text,
                t.deep_document_links,
                t.deep_ai_summary
            FROM tenders t
            WHERE {" OR ".join(clauses)}
            """,
            params,
        )
        out: Dict[str, Dict[str, Any]] = {}
        for record in (cur.fetchall() or []):
            tid = str(record.get("tender_id") or "").strip()
            url = str(record.get("url") or "").strip().lower()
            if tid:
                out[tid] = record
            if url:
                out[url] = record
        cur.close()
        conn.close()
        return out
    except Exception as exc:
        logger.debug("[packager] DB evidence rehydrate skipped: %s", exc)
        return {}


def _extract_doc_urls(row: Dict[str, Any]) -> List[str]:
    """Extract all document URLs from content fields."""
    all_text = " ".join([
        str(row.get("deep_scope") or ""),
        str(row.get("deep_description") or ""),
        str(row.get("deep_pdf_text") or ""),
        str(row.get("deep_ai_summary") or row.get("ai_summary") or ""),
        str(row.get("cross_sources") or ""),
        str(row.get("evaluation_criteria") or ""),
    ])
    urls = list(_DOC_URL_PAT.findall(all_text))
    raw_links = row.get("deep_document_links")
    if isinstance(raw_links, str) and raw_links.strip():
        try:
            raw_links = json.loads(raw_links)
            if isinstance(raw_links, str):
                raw_links = json.loads(raw_links)
        except Exception:
            raw_links = []
    if isinstance(raw_links, list):
        for item in raw_links:
            if isinstance(item, dict) and item.get("url"):
                urls.append(str(item["url"]).strip())
    return list(dict.fromkeys(urls))  # deduplicated, ordered


# ══════════════════════════════════════════════════════════════════════════════
# 3. IDCG FIT REASONING
# ══════════════════════════════════════════════════════════════════════════════

def _compute_fit_signals(row: Dict[str, Any]) -> Dict[str, str]:
    """
    Returns:
      technical_fit:      High / Medium / Low
      delivery_feasibility: High / Medium / Low
      fit_reason:         short plain-text explanation
    """
    title   = str(row.get("title") or "")
    sector  = str(row.get("sector") or "")
    stype   = str(row.get("service_type") or "")
    org     = str(row.get("organization") or "")
    country = str(row.get("country") or "")
    note    = str(row.get("scoring_note") or "")

    combined = title + " " + sector + " " + stype + " " + note

    # Technical fit
    strong_svc  = bool(_SERVICE_STRONG.search(combined))
    moderate_svc = bool(_SERVICE_MODERATE.search(combined))
    strong_sector = bool(_SECTOR_STRONG.search(combined))

    if strong_svc and strong_sector:
        technical_fit = "High"
    elif strong_svc or (moderate_svc and strong_sector):
        technical_fit = "Medium"
    else:
        technical_fit = "Low"

    # Delivery feasibility
    org_lower = org.lower()
    is_t1 = any(t in org_lower for t in _CLIENT_TIER1)
    is_t2 = bool(_CLIENT_TIER2_PAT.search(org))
    geo_text = country + " " + title
    is_strong_geo = bool(_GEO_STRONG.search(geo_text))
    is_weak_geo   = bool(_GEO_WEAK.search(geo_text))

    if is_strong_geo and (is_t1 or is_t2):
        delivery_feasibility = "High"
    elif is_strong_geo or (is_t1 and not is_weak_geo):
        delivery_feasibility = "Medium"
    elif is_weak_geo:
        delivery_feasibility = "Low"
    else:
        delivery_feasibility = "Medium"  # unknown geo → cautious medium

    # Build fit reason
    reason_parts = []
    if strong_svc:
        reason_parts.append("core IDCG service area (evaluation/M&E/TA)")
    elif moderate_svc:
        reason_parts.append("adjacent IDCG service area (advisory/research)")
    if strong_sector:
        reason_parts.append("sector match (education/health/environment/agriculture)")
    if is_t1:
        reason_parts.append(f"Tier-1 donor client ({org})")
    elif is_t2:
        reason_parts.append(f"known IDCG partner ({org})")
    if is_strong_geo:
        reason_parts.append("geography within IDCG delivery network")
    elif is_weak_geo:
        reason_parts.append("geography outside IDCG core network (delivery risk)")

    if "CAPSTAT" in note or "IDCG service match" in note:
        reason_parts.append("CAPSTAT service match confirmed")
    if "ML strong" in note:
        reason_parts.append("shadow ML model: strong IDCG fit signal")
    elif "ML borderline" in note:
        reason_parts.append("shadow ML model: borderline IDCG fit signal")

    fit_reason = "; ".join(reason_parts) if reason_parts else "limited structured signals available"

    return {
        "technical_fit":       technical_fit,
        "delivery_feasibility": delivery_feasibility,
        "fit_reason":          fit_reason,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 4. SAFE PATH HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _safe_dirname(title: str, tender_id: str = "") -> str:
    """
    Create a safe, ASCII-only directory name from the tender title (or fall back to ID).
    Strips non-ASCII, special chars, truncates.
    """
    if title:
        # Encode to ASCII dropping non-ASCII bytes (removes â, ã, Marathi, Hindi chars)
        safe = title.encode("ascii", "ignore").decode("ascii")
        safe = re.sub(r"[^\w\s-]", "", safe)
        safe = re.sub(r"[\s]+", "_", safe).strip("_")
        safe = safe[:_MAX_SAFE_DIRNAME]
    else:
        safe = ""

    if not safe:
        safe = re.sub(r"[^\w-]", "_", tender_id)[:_MAX_SAFE_DIRNAME] or "tender"

    return safe


def _unique_package_dir(run_dir: Path, title: str, tender_id: str) -> Path:
    """Return a unique sub-directory path inside run_dir."""
    base = _safe_dirname(title, tender_id)
    pkg_dir = run_dir / base
    suffix = 1
    while pkg_dir.exists():
        pkg_dir = run_dir / f"{base}_{suffix}"
        suffix += 1
    return pkg_dir


# ══════════════════════════════════════════════════════════════════════════════
# 5. DOCUMENT DOWNLOADER
# ══════════════════════════════════════════════════════════════════════════════

def _try_download_doc(url: str, dest_dir: Path) -> Optional[str]:
    """
    Attempt to download a document URL into dest_dir.
    Returns filename on success, None on failure.
    Non-fatal: always swallows exceptions.
    """
    try:
        import urllib.request
        parsed = urllib.parse.urlparse(url)
        fname  = os.path.basename(parsed.path) or "document"
        # Sanitize filename
        fname = re.sub(r"[^\w.\-]", "_", fname)[:80]
        if not fname or fname == "_":
            fname = "document"
        dest = dest_dir / fname

        # Avoid overwriting if same name downloaded twice
        stem = Path(fname).stem
        ext  = Path(fname).suffix
        idx  = 1
        while dest.exists():
            dest = dest_dir / f"{stem}_{idx}{ext}"
            idx += 1

        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 TenderRadar/1.0 (document retrieval)"
        })
        with urllib.request.urlopen(req, timeout=_DOWNLOAD_TIMEOUT) as resp:
            content = resp.read()

        if len(content) < 100:
            return None  # Likely an error page / redirect body

        dest.write_bytes(content)
        return dest.name
    except Exception as exc:
        logger.debug("[packager] doc download failed for %s: %s", url, exc)
        return None


# ══════════════════════════════════════════════════════════════════════════════
# 6. FILE WRITERS
# ══════════════════════════════════════════════════════════════════════════════

def _write_metadata_json(
    pkg_dir: Path,
    row: Dict[str, Any],
    evidence_state: str,
    opportunity_maturity: str,
    doc_url_count: int,
    fit_signals: Dict[str, str],
    downloaded_files: List[str],
) -> None:
    meta = {
        "title":                row.get("title", ""),
        "portal":               row.get("portal", ""),
        "organization":         row.get("organization", ""),
        "country":              row.get("country", ""),
        "deadline":             row.get("deadline", ""),
        "priority_score":       row.get("priority_score", "0"),
        "relevance_score":      row.get("relevance_score", "0"),
        "sector":               row.get("sector", ""),
        "service_type":         row.get("service_type", ""),
        "AI_Suggested_Label":   row.get("ai_label", ""),
        "Human_Label":          row.get("human_label", ""),
        "tender_url":           row.get("tender_url", ""),
        "scraped_date":         row.get("scraped_date", ""),
        "evidence_state":       evidence_state,
        "opportunity_maturity": opportunity_maturity,
        "doc_url_count":        doc_url_count,
        "downloaded_files":     downloaded_files,
        "technical_fit":        fit_signals["technical_fit"],
        "delivery_feasibility": fit_signals["delivery_feasibility"],
        "packaged_at":          datetime.now(timezone.utc).isoformat(),
    }
    (pkg_dir / "metadata.json").write_text(
        json.dumps(meta, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _write_summary_txt(
    pkg_dir: Path,
    row: Dict[str, Any],
    evidence_state: str,
    fit_signals: Dict[str, str],
    opportunity_maturity: str = "",
) -> None:
    title   = row.get("title", "")
    org     = row.get("organization", "") or "Unknown"
    country = row.get("country", "") or "Unknown"
    deadline = row.get("deadline", "") or "Not specified"
    ps      = row.get("priority_score", "0")
    rs      = row.get("relevance_score", "0")
    ai_lbl  = row.get("ai_label", "")
    summary = str(row.get("ai_summary") or "").strip()
    rel_reason = str(row.get("relevance_reason") or "").strip()
    scoring_note = str(row.get("scoring_note") or "").strip()
    opportunity_insight = str(row.get("opportunity_insight") or "").strip()

    _, mat_label_fn, _, rec_action_fn = _get_maturity_module()
    mat_label = opportunity_maturity or mat_label_fn(evidence_state)
    rec_action = rec_action_fn(evidence_state)

    lines = [
        "=" * 70,
        "IDCG TENDER SUMMARY",
        "=" * 70,
        "",
        f"Title        : {title}",
        f"Organization : {org}",
        f"Country      : {country}",
        f"Deadline     : {deadline}",
        f"Portal       : {row.get('portal', '')}",
        f"Sector       : {row.get('sector', '')}",
        f"Service Type : {row.get('service_type', '')}",
        "",
        "-" * 70,
        "SCORING",
        "-" * 70,
        f"Priority Score      : {ps}",
        f"Relevance Score     : {rs}",
        f"AI Label            : {ai_lbl}",
        f"Human Label         : {row.get('human_label', '') or 'Not yet reviewed'}",
        f"Evidence State      : {evidence_state}",
        f"Opportunity Maturity: {mat_label}",
        f"Recommended Action  : {rec_action}",
        "",
        "-" * 70,
        "IDCG FIT ASSESSMENT",
        "-" * 70,
        f"Technical Fit        : {fit_signals['technical_fit']}",
        f"Delivery Feasibility : {fit_signals['delivery_feasibility']}",
        f"Why IDCG fit         : {fit_signals['fit_reason']}",
        "",
    ]

    if summary:
        lines += [
            "-" * 70,
            "AI SUMMARY",
            "-" * 70,
            summary,
            "",
        ]
    else:
        lines += [
            "-" * 70,
            "AI SUMMARY",
            "-" * 70,
            "(No AI summary available — signal-only listing)",
            "",
        ]

    if rel_reason:
        lines += [
            "-" * 70,
            "RELEVANCE EXPLANATION",
            "-" * 70,
            rel_reason,
            "",
        ]

    if scoring_note and scoring_note != rel_reason:
        lines += [
            "-" * 70,
            "SCORING NOTE",
            "-" * 70,
            scoring_note,
            "",
        ]

    if opportunity_insight:
        lines += [
            "-" * 70,
            "OPPORTUNITY INSIGHT",
            "-" * 70,
            opportunity_insight,
            "",
        ]

    lines += [
        "=" * 70,
        f"Tender URL: {row.get('tender_url', '')}",
        f"Packaged  : {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        "=" * 70,
    ]

    (pkg_dir / "summary.txt").write_text("\n".join(lines), encoding="utf-8")


def _write_evidence_txt(
    pkg_dir: Path,
    row: Dict[str, Any],
    evidence_state: str,
    opportunity_maturity: str,
    doc_urls: List[str],
    downloaded_files: List[str],
) -> None:
    _EVIDENCE_NOTES = {
        "SIGNAL_ONLY":      "Signal-only listing — no TOR or page content available yet. "
                            "Monitor the portal URL for document publication.",
        "PAGE_ONLY":        "Page-rich listing — content extracted from portal page, "
                            "but no document attachments found.",
        "PARTIAL_PACKAGE":  "Partial package — document URLs found in page content, "
                            "but full TOR/annexures may be missing.",
        "FULL_PACKAGE":     "Full package available — rich content and document "
                            "attachments present.",
    }

    _, mat_label_fn, mat_digest_fn, rec_action_fn = _get_maturity_module()
    # opportunity_maturity is the business-facing label passed in from the caller
    mat_label = opportunity_maturity or mat_label_fn(evidence_state)
    rec_action = rec_action_fn(evidence_state)
    digest_note = mat_digest_fn(evidence_state)

    lines = [
        "=" * 70,
        "EVIDENCE REPORT",
        "=" * 70,
        "",
        f"Evidence State      : {evidence_state}",
        f"Opportunity Maturity: {mat_label}",
        f"Recommended Action  : {rec_action} — {digest_note}",
        f"Document URLs Found : {len(doc_urls)}",
        f"Files Downloaded    : {len(downloaded_files)}",
        "",
        "-" * 70,
        "EVIDENCE NOTE",
        "-" * 70,
        _EVIDENCE_NOTES.get(evidence_state, "Evidence state unknown."),
        "",
    ]

    if doc_urls:
        lines += [
            "-" * 70,
            "DOCUMENT URLS",
            "-" * 70,
        ]
        for i, url in enumerate(doc_urls, 1):
            lines.append(f"  [{i}] {url}")
        lines.append("")

    if downloaded_files:
        lines += [
            "-" * 70,
            "DOWNLOADED FILES",
            "-" * 70,
        ]
        for f in downloaded_files:
            lines.append(f"  - {f}")
        lines.append("")
    else:
        lines += [
            "-" * 70,
            "DOWNLOADED FILES",
            "-" * 70,
            "  (No files downloaded — see documents/ folder)",
            "",
        ]

    if row.get("deep_scope"):
        scope_preview = str(row["deep_scope"])[:500]
        lines += [
            "-" * 70,
            "DEEP SCOPE PREVIEW (first 500 chars)",
            "-" * 70,
            scope_preview,
            "...",
            "",
        ]

    lines += [
        "=" * 70,
        f"Portal URL: {row.get('tender_url', '')}",
        "=" * 70,
    ]

    (pkg_dir / "evidence.txt").write_text("\n".join(lines), encoding="utf-8")


# ══════════════════════════════════════════════════════════════════════════════
# 7. RUN-STATE (DEDUPLICATION)
# ══════════════════════════════════════════════════════════════════════════════

def _load_run_state(run_dir: Path) -> Dict[str, Any]:
    state_path = run_dir / "run_state.json"
    if state_path.exists():
        try:
            return json.loads(state_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"packaged_urls": [], "packaged_titles": []}


def _save_run_state(run_dir: Path, state: Dict[str, Any]) -> None:
    (run_dir / "run_state.json").write_text(
        json.dumps(state, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def _already_packaged(row: Dict[str, Any], state: Dict[str, Any]) -> bool:
    url   = row.get("tender_url", "")
    title = row.get("title", "")
    return url in state.get("packaged_urls", []) or (
        title and title in state.get("packaged_titles", [])
    )


def _mark_packaged(row: Dict[str, Any], state: Dict[str, Any]) -> None:
    url   = row.get("tender_url", "")
    title = row.get("title", "")
    if url and url not in state["packaged_urls"]:
        state["packaged_urls"].append(url)
    if title and title not in state["packaged_titles"]:
        state["packaged_titles"].append(title)


# ══════════════════════════════════════════════════════════════════════════════
# 8. MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def package_new_shortlisted(
    workbook_path: Optional[Path] = None,
    run_date: Optional[str] = None,
    download_docs: bool = True,
    dry_run: bool = False,
) -> Dict[str, Any]:
    """
    Main entry point for the evidence packaging layer.

    Args:
        workbook_path:  Path to master Excel (defaults to output/Tender_Monitor_Master.xlsx)
        run_date:       YYYY-MM-DD string for output folder (defaults to today)
        download_docs:  Whether to attempt document downloads (default True)
        dry_run:        If True, classify and log but do not write any files

    Returns dict with:
        packaged_count:   int
        skipped_count:    int (already packaged)
        run_dir:          str (path to today's package folder)
        evidence_summary: dict (counts per evidence state)
        errors:           list of non-fatal error strings
    """
    wb_path  = workbook_path or _WORKBOOK_PATH
    date_str = run_date or date.today().isoformat()
    run_dir  = _PACKAGES_ROOT / date_str

    logger.info("[packager] Starting evidence packaging — run_dir: %s", run_dir)

    # Load rows
    rows = _load_new_shortlisted(wb_path)
    if not rows:
        logger.info("[packager] No new shortlisted tenders found in workbook.")
        return {
            "packaged_count": 0, "skipped_count": 0,
            "run_dir": str(run_dir), "evidence_summary": {}, "errors": [],
        }

    logger.info("[packager] Found %d new shortlisted tender(s) to process", len(rows))
    _rich_by_key = _fetch_rich_packager_data(rows)

    if dry_run:
        logger.info("[packager] DRY-RUN mode — no files will be written")
    else:
        run_dir.mkdir(parents=True, exist_ok=True)

    # Load dedup state
    state = _load_run_state(run_dir) if not dry_run else {"packaged_urls": [], "packaged_titles": []}

    packaged_count = 0
    skipped_count  = 0
    evidence_summary: Dict[str, int] = {}
    errors: List[str] = []

    for row in rows:
        title = row.get("title", "")
        key_tid = str(row.get("tender_id") or "").strip()
        key_url = str(row.get("tender_url") or "").strip().lower()
        rich = (_rich_by_key.get(key_tid) if key_tid else None) or _rich_by_key.get(key_url, {})
        if rich:
            # DB-enriched evidence is authoritative when available; workbook values remain fallback.
            row["deep_scope"] = row.get("deep_scope") or str(rich.get("deep_scope") or rich.get("deep_description") or rich.get("deep_pdf_text") or "")
            row["deep_description"] = str(rich.get("deep_description") or "")
            row["deep_pdf_text"] = str(rich.get("deep_pdf_text") or "")
            row["ai_summary"] = row.get("ai_summary") or str(rich.get("deep_ai_summary") or "")
            row["evaluation_criteria"] = row.get("evaluation_criteria") or str(rich.get("deep_eval_criteria") or "")
            row["deep_document_links"] = rich.get("deep_document_links")

        # Skip already-packaged tenders (same-day dedup)
        if _already_packaged(row, state):
            logger.debug("[packager] Skipping already-packaged: %s", title[:60])
            skipped_count += 1
            continue

        try:
            # Prefer workbook maturity if already computed from the shared exporter logic.
            evidence_state = str(row.get("evidence_state") or "").strip()
            opportunity_maturity = str(row.get("opportunity_maturity") or "").strip()
            if evidence_state and opportunity_maturity:
                doc_url_count = len(_extract_doc_urls(row))
            else:
                evidence_state, opportunity_maturity, doc_url_count = _classify_evidence(row)
            doc_urls = _extract_doc_urls(row)[:_MAX_DOCS]

            # Compute IDCG fit signals
            fit_signals = _compute_fit_signals(row)

            # Log classification
            logger.info(
                "[packager] %s | %s | fit=%s/delivery=%s | docs=%d | %s",
                evidence_state, row.get("ai_label", ""),
                fit_signals["technical_fit"], fit_signals["delivery_feasibility"],
                len(doc_urls), title[:55],
            )

            # Track evidence summary
            evidence_summary[evidence_state] = evidence_summary.get(evidence_state, 0) + 1

            if dry_run:
                packaged_count += 1
                continue

            # Create package directory
            tender_id = row.get("tender_url", "")[-20:] or str(packaged_count)
            pkg_dir = _unique_package_dir(run_dir, title, tender_id)
            pkg_dir.mkdir(parents=True, exist_ok=True)

            # Create documents/ subfolder
            docs_dir = pkg_dir / "documents"
            docs_dir.mkdir(exist_ok=True)

            # Download attachments
            downloaded_files: List[str] = []
            if download_docs and doc_urls:
                for url in doc_urls:
                    fname = _try_download_doc(url, docs_dir)
                    if fname:
                        downloaded_files.append(fname)
                        logger.debug("[packager] Downloaded: %s", fname)

            # Write package files
            _write_metadata_json(
                pkg_dir, row, evidence_state, opportunity_maturity,
                doc_url_count, fit_signals, downloaded_files,
            )
            _write_summary_txt(pkg_dir, row, evidence_state, fit_signals, opportunity_maturity)
            _write_evidence_txt(
                pkg_dir, row, evidence_state, opportunity_maturity,
                doc_urls, downloaded_files,
            )

            # Mark as packaged
            _mark_packaged(row, state)
            packaged_count += 1

        except Exception as exc:
            err = f"Failed to package '{title[:50]}': {exc}"
            logger.warning("[packager] %s", err)
            errors.append(err)

    # Save dedup state
    if not dry_run and run_dir.exists():
        _save_run_state(run_dir, state)

    result = {
        "packaged_count":  packaged_count,
        "skipped_count":   skipped_count,
        "run_dir":         str(run_dir),
        "evidence_summary": evidence_summary,
        "errors":          errors,
    }
    logger.info(
        "[packager] Done — packaged=%d skipped=%d errors=%d | %s",
        packaged_count, skipped_count, len(errors), evidence_summary,
    )
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 9. CLI / STANDALONE RUNNER
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse
    logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")

    ap = argparse.ArgumentParser(description="TenderRadar Evidence Packager")
    ap.add_argument("--workbook", type=str, default="", help="Path to master workbook")
    ap.add_argument("--date", type=str, default="", help="Run date YYYY-MM-DD (default: today)")
    ap.add_argument("--no-download", action="store_true", help="Skip document downloads")
    ap.add_argument("--dry-run", action="store_true", help="Classify only, no file writes")
    args = ap.parse_args()

    result = package_new_shortlisted(
        workbook_path=Path(args.workbook) if args.workbook else None,
        run_date=args.date or None,
        download_docs=not args.no_download,
        dry_run=args.dry_run,
    )

    print(f"\nPackaging complete:")
    print(f"  Packaged : {result['packaged_count']}")
    print(f"  Skipped  : {result['skipped_count']}")
    print(f"  Errors   : {len(result['errors'])}")
    print(f"  Run dir  : {result['run_dir']}")
    print(f"  Evidence : {result['evidence_summary']}")
    if result["errors"]:
        print("\nErrors:")
        for e in result["errors"]:
            print(f"  - {e}")
