from __future__ import annotations

from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config.config import OPPORTUNITY_SIGNALS_EXCEL_PATH

_COLUMNS = [
    ("Source", 24),
    ("Title", 60),
    ("Organization", 24),
    ("Geography", 18),
    ("Sector", 24),
    ("Signal Stage", 16),
    ("Consulting Signal", 16),
    ("Confidence Score", 14),
    ("Summary", 70),
    ("URL", 55),
    ("Recommended Action", 24),
    ("Published Date", 14),
    ("Captured At", 20),
]


def export_opportunity_signals_excel(rows: List[Dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunity Signals"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append([name for name, _ in _COLUMNS])
    for col_idx, (name, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in rows:
        ws.append([
            row.get("source", ""),
            row.get("title", ""),
            row.get("organization", ""),
            row.get("geography", ""),
            row.get("sector", ""),
            row.get("signal_stage", ""),
            "Yes" if int(row.get("consulting_signal") or 0) == 1 else "No",
            int(row.get("confidence_score") or 0),
            row.get("summary", ""),
            row.get("url", ""),
            row.get("recommended_action", ""),
            row.get("published_date", ""),
            row.get("captured_at", ""),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    meta = wb.create_sheet("Run Summary")
    meta["A1"] = "Generated"
    meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta["A2"] = "Rows"
    meta["B2"] = len(rows)

    wb.save(OPPORTUNITY_SIGNALS_EXCEL_PATH)
    return OPPORTUNITY_SIGNALS_EXCEL_PATH
