# =============================================================================
# exporters/excel_feedback_sync.py — Excel-based Feedback Loop Sync
#
# The user fills ONE column in Tender_Monitor_Master.xlsx:
#   "My Decision"  →  Bid / No Bid / Review Later
#
# This module reads that value and syncs it to the bid_pipeline DB table
# so TenderRadar can learn from real bid decisions over time.
# Outcome tracking (Won/Lost) happens via the frontend API, not Excel.
#
# Column mapping (Excel → DB)
# ────────────────────────────────────────────────────────────────────────────
#   My Decision = Bid          → bid_decision = "bid"
#   My Decision = No Bid       → bid_decision = "no_bid"
#   My Decision = Review Later → bid_decision = "review_later"
#
# NOTE: "Outcome" column was removed from Excel (2026-04-02). The OUTCOME_MAP
# and out_ci logic below still run for backward compat with old Excel files.
#
# Public API
# ────────────────────────────────────────────────────────────────────────────
#   sync_excel_feedback(excel_path=None)  → dict  (Task 3 — main sync)
#   compute_feedback_metrics()            → dict  (Task 4 — basic metrics)
#   print_feedback_summary(metrics=None)  → None  (Task 5 — log output)
# =============================================================================

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

logger = logging.getLogger("tenderradar.excel_feedback_sync")

# ── Column name constants ─────────────────────────────────────────────────────
_COL_URL      = "Tender URL"
_COL_DECISION = "My Decision"
_COL_OUTCOME  = "Outcome"
_COL_TITLE    = "Title"
_COL_PORTAL   = "Portal"
_COL_TID      = "Tender ID"
_COL_HUMAN_LABEL = "Human_Label"
_COL_LABEL_REASON = "Label_Reason"
_COL_TRAINING_APPROVED = "Training_Approved"
_COL_ACTION_LABEL = "Action_Label"

# ── Value maps ────────────────────────────────────────────────────────────────
DECISION_MAP: Dict[str, str] = {
    "bid": "bid",
    "no bid": "no_bid",
    "review later": "review_later",
    # backward-compatible legacy labels
    "no": "no_bid",
    "later": "review_later",
}

OUTCOME_MAP: Dict[str, str] = {
    "won": "won",
    "lost": "lost",
    "no submission": "no_submission",
    "pending": "pending",
    # backward-compatible legacy label
    "na": "pending",
}

TRAINING_APPROVED_MAP: Dict[str, str] = {
    "yes": "Yes",
    "no": "No",
}

HUMAN_LABEL_TO_SCORE: Dict[str, int] = {
    "relevant": 2,
    "borderline": 1,
    "not relevant": 0,
}


def _norm_label(value: str) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


# =============================================================================
# SECTION 1 — Read feedback from Excel  (Task 1)
# =============================================================================

def read_excel_feedback(excel_path: Optional[str] = None) -> List[Dict]:
    """
    Open Tender_Monitor_Master.xlsx and extract rows where the user has
    filled "My Decision" or "Outcome".

    Returns a list of dicts, each with:
        url          : str  (from "Tender URL" column)
        title        : str  (for logging)
        portal       : str  (for logging)
        raw_decision : str  (raw cell value, e.g. "BID")
        raw_outcome  : str  (raw cell value, e.g. "WON")
        decision     : str  (mapped: "bid" / "no_bid" / "pending" / "")
        outcome      : str  (mapped: "won" / "lost" / "")
    """
    if excel_path is None:
        try:
            from config.config import UNIFIED_EXCEL_PATH
            excel_path = UNIFIED_EXCEL_PATH
        except ImportError:
            excel_path = os.path.join(
                os.path.dirname(__file__), "..", "output", "Tender_Monitor_Master.xlsx"
            )
        excel_path = os.path.normpath(excel_path)

    if not os.path.exists(excel_path):
        logger.warning("[feedback_sync] Excel not found: %s", excel_path)
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("[feedback_sync] openpyxl not available — cannot read Excel")
        return []

    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active

        # Locate header columns by name (case-insensitive)
        header_row = next(ws.iter_rows(max_row=1), None)
        if not header_row:
            wb.close()
            return []
        headers = [str(c.value or "").strip() for c in header_row]

        def _find_col(name: str) -> Optional[int]:
            nl = name.lower()
            for i, h in enumerate(headers):
                if h.lower() == nl:
                    return i
            return None

        url_ci  = _find_col(_COL_URL)
        dec_ci  = _find_col(_COL_DECISION)
        out_ci  = _find_col(_COL_OUTCOME)
        ttl_ci  = _find_col(_COL_TITLE)
        por_ci  = _find_col(_COL_PORTAL)

        if dec_ci is None and out_ci is None:
            logger.info(
                "[feedback_sync] Excel has no '%s' or '%s' columns — "
                "nothing to sync. Fill these columns and re-run.",
                _COL_DECISION, _COL_OUTCOME,
            )
            wb.close()
            return []

        rows: List[Dict] = []
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            raw_dec = str(row_vals[dec_ci] or "").strip() if dec_ci is not None else ""
            raw_out = str(row_vals[out_ci] or "").strip() if out_ci is not None else ""

            # Skip rows without any user input
            if not raw_dec and not raw_out:
                continue
            if _norm_label(raw_dec) == "na" and (not raw_out or _norm_label(raw_out) == "na"):
                continue

            url    = str(row_vals[url_ci] or "").strip() if url_ci is not None else ""
            title  = str(row_vals[ttl_ci] or "").strip() if ttl_ci is not None else ""
            portal = str(row_vals[por_ci] or "").strip() if por_ci is not None else ""

            decision = DECISION_MAP.get(_norm_label(raw_dec), "")
            outcome  = OUTCOME_MAP.get(_norm_label(raw_out), "")

            rows.append({
                "url":          url,
                "title":        title[:100],
                "portal":       portal,
                "raw_decision": raw_dec,
                "raw_outcome":  raw_out,
                "decision":     decision,
                "outcome":      outcome,
            })

        wb.close()
        logger.info(
            "[feedback_sync] Read %d decision(s) from %s",
            len(rows), os.path.basename(excel_path),
        )
        return rows

    except Exception as exc:
        logger.error("[feedback_sync] Failed to read Excel: %s", exc)
        return []


# =============================================================================
# SECTION 2 — DB helpers  (Task 2)
# =============================================================================

def _url_to_tender_id(cur, url: str) -> Optional[str]:
    """
    Resolve a Tender URL to a tender_id from seen_tenders.
    Tries exact match first; falls back to URL without trailing slash.
    Returns None if not found.
    """
    if not url:
        return None

    # Exact match
    cur.execute(
        "SELECT tender_id FROM seen_tenders WHERE url = %s LIMIT 1", (url,)
    )
    row = cur.fetchone()
    if row:
        return row[0] if isinstance(row, tuple) else row.get("tender_id")

    # Fallback: strip trailing slash and retry
    url_stripped = url.rstrip("/")
    if url_stripped != url:
        cur.execute(
            "SELECT tender_id FROM seen_tenders WHERE url = %s LIMIT 1",
            (url_stripped,),
        )
        row = cur.fetchone()
        if row:
            return row[0] if isinstance(row, tuple) else row.get("tender_id")

    return None


def _set_bid_decision(conn, tender_id: str, bid_decision: str) -> bool:
    """
    Directly update bid_decision in bid_pipeline.
    Used for BID (without final outcome yet) and LATER decisions.
    """
    if bid_decision not in ("bid", "no_bid", "review_later"):
        logger.warning("[feedback_sync] Invalid bid_decision ignored: %s", bid_decision)
        return False
    try:
        cur = conn.cursor()
        cur.execute(
            "UPDATE bid_pipeline "
            "SET bid_decision = %s, updated_at = NOW() "
            "WHERE tender_id = %s",
            (bid_decision, tender_id),
        )
        affected = cur.rowcount
        conn.commit()
        cur.close()
        return affected > 0
    except Exception as exc:
        logger.warning("[feedback_sync] _set_bid_decision failed: %s", exc)
        return False


def _ensure_tender_labels_table(conn) -> None:
    cur = conn.cursor()
    try:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS tender_labels (
                id BIGINT PRIMARY KEY AUTO_INCREMENT,
                url VARCHAR(2048) NOT NULL,
                tender_id VARCHAR(255) NULL,
                human_label VARCHAR(32) NOT NULL,
                label_score TINYINT NOT NULL,
                label_reason TEXT NULL,
                action_label VARCHAR(64) NULL,
                training_approved TINYINT(1) NOT NULL DEFAULT 1,
                reviewed_at DATETIME NOT NULL,
                source_sheet VARCHAR(64) NOT NULL DEFAULT 'All Tenders',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_tender_labels_url (url(512))
            )
            """
        )
        conn.commit()
    finally:
        cur.close()


def ingest_approved_labels(excel_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Read master workbook (All Tenders) and ingest rows with Training_Approved=Yes
    into tender_labels. Safe and idempotent via URL upsert.
    """
    if excel_path is None:
        try:
            from config.config import UNIFIED_EXCEL_PATH
            excel_path = UNIFIED_EXCEL_PATH
        except ImportError:
            excel_path = os.path.join(
                os.path.dirname(__file__), "..", "output", "Tender_Monitor_Master.xlsx"
            )
        excel_path = os.path.normpath(excel_path)

    if not os.path.exists(excel_path):
        return {"ingested": 0, "skipped": 0, "errors": 0, "note": "Master workbook not found."}

    try:
        from openpyxl import load_workbook
        from database.db import get_connection
    except Exception as exc:
        return {"ingested": 0, "skipped": 0, "errors": 1, "note": f"Import error: {exc}"}

    wb = None
    conn = None
    ingested = 0
    skipped = 0
    errors = 0
    reviewed_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["All Tenders"] if "All Tenders" in wb.sheetnames else wb.active

        header_row = next(ws.iter_rows(max_row=1), None)
        if not header_row:
            return {"ingested": 0, "skipped": 0, "errors": 0, "note": "No headers in master workbook."}
        headers = [str(c.value or "").strip() for c in header_row]

        def _find_col(name: str) -> Optional[int]:
            nl = name.lower()
            for i, h in enumerate(headers):
                if h.lower() == nl:
                    return i
            return None

        url_ci = _find_col(_COL_URL)
        tid_ci = _find_col(_COL_TID)
        human_ci = _find_col(_COL_HUMAN_LABEL)
        reason_ci = _find_col(_COL_LABEL_REASON)
        train_ci = _find_col(_COL_TRAINING_APPROVED)
        action_ci = _find_col(_COL_ACTION_LABEL)
        if None in (url_ci, human_ci, train_ci):
            return {"ingested": 0, "skipped": 0, "errors": 0, "note": "Labeling columns not found; nothing to ingest."}

        conn = get_connection()
        _ensure_tender_labels_table(conn)
        cur = conn.cursor()

        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            approved_raw = _norm_label(str(row_vals[train_ci] or ""))
            if TRAINING_APPROVED_MAP.get(approved_raw, "") != "Yes":
                continue

            url = str(row_vals[url_ci] or "").strip()
            human_raw = _norm_label(str(row_vals[human_ci] or ""))
            if not url or human_raw not in HUMAN_LABEL_TO_SCORE:
                skipped += 1
                continue

            tender_id = str(row_vals[tid_ci] or "").strip() if tid_ci is not None else ""
            if not tender_id:
                tender_id = _url_to_tender_id(cur, url) or ""
            label_reason = str(row_vals[reason_ci] or "").strip() if reason_ci is not None else ""
            action_label = str(row_vals[action_ci] or "").strip() if action_ci is not None else ""

            try:
                cur.execute(
                    """
                    INSERT INTO tender_labels (
                        url, tender_id, human_label, label_score, label_reason,
                        action_label, training_approved, reviewed_at, source_sheet
                    ) VALUES (%s, %s, %s, %s, %s, %s, 1, %s, 'All Tenders')
                    ON DUPLICATE KEY UPDATE
                        tender_id = VALUES(tender_id),
                        human_label = VALUES(human_label),
                        label_score = VALUES(label_score),
                        label_reason = VALUES(label_reason),
                        action_label = VALUES(action_label),
                        training_approved = VALUES(training_approved),
                        reviewed_at = VALUES(reviewed_at)
                    """,
                    (
                        url,
                        tender_id or None,
                        human_raw.title() if human_raw != "not relevant" else "Not Relevant",
                        HUMAN_LABEL_TO_SCORE[human_raw],
                        label_reason or None,
                        action_label or None,
                        reviewed_at,
                    ),
                )
                ingested += 1
            except Exception as exc:
                logger.warning("[feedback_sync] Failed to ingest label for %s: %s", url[:120], exc)
                errors += 1

        conn.commit()
        cur.close()
        note = f"Ingested {ingested} approved label(s)." + (f" Skipped {skipped}." if skipped else "")
        return {"ingested": ingested, "skipped": skipped, "errors": errors, "note": note}
    except Exception as exc:
        logger.warning("[feedback_sync] ingest_approved_labels failed: %s", exc)
        return {"ingested": ingested, "skipped": skipped, "errors": errors + 1, "note": f"Label ingestion failed: {exc}"}
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass


# =============================================================================
# SECTION 3 — Main sync function  (Task 3)
# =============================================================================

def sync_excel_feedback(excel_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Read "My Decision" / "Outcome" from the master Excel and sync to DB.

    Flow per row:
      1. Look up tender_id from seen_tenders via Tender URL
      2. ensure_pipeline_entry(tender_id)  — INSERT IGNORE into bid_pipeline
      3. Apply decision:
           Bid + Won/Lost        → record_outcome(outcome, "bid")
           Bid + Pending         → record_outcome("pending", "bid")
           Bid + blank outcome   → set bid_decision="bid"
           No Bid                → record_outcome("no_submission", "no_bid")
           Review Later          → set bid_decision="review_later"

    Returns summary dict:
        synced   : int  — rows successfully written to DB
        skipped  : int  — rows skipped (URL not found in seen_tenders)
        no_url   : int  — rows missing Tender URL
        total    : int  — total rows with decision/outcome filled
        errors   : int  — unexpected errors
    """
    feedback_rows = read_excel_feedback(excel_path)

    if not feedback_rows:
        return {
            "synced": 0, "skipped": 0, "no_url": 0,
            "total": 0, "errors": 0,
            "note": "No feedback rows found in Excel.",
        }

    try:
        from database.db import get_connection
        from pipeline.opportunity_pipeline import (
            ensure_pipeline_entry,
            record_outcome,
        )
    except ImportError as exc:
        logger.error("[feedback_sync] Import failed: %s", exc)
        return {
            "synced": 0, "skipped": 0, "no_url": 0,
            "total": len(feedback_rows), "errors": 1,
            "note": f"Import error: {exc}",
        }

    conn = get_connection()
    cur  = conn.cursor()

    synced  = 0
    skipped = 0
    no_url  = 0
    errors  = 0

    for row in feedback_rows:
        url      = row["url"]
        decision = row["decision"]
        outcome  = row["outcome"]
        label    = row.get("title") or url[:60]

        if not url:
            no_url += 1
            logger.debug("[feedback_sync] No URL for row: %s", label)
            continue

        # Step 1 — resolve tender_id
        try:
            tender_id = _url_to_tender_id(cur, url)
        except Exception as exc:
            logger.warning("[feedback_sync] URL lookup failed for '%s': %s", url, exc)
            errors += 1
            continue

        if not tender_id:
            skipped += 1
            logger.debug("[feedback_sync] tender_id not found for URL: %s", url[:80])
            continue

        # Step 2 — ensure in pipeline
        try:
            ensure_pipeline_entry(tender_id)
        except Exception as exc:
            logger.warning("[feedback_sync] ensure_pipeline_entry failed for %s: %s",
                           tender_id, exc)
            errors += 1
            continue

        # Step 3 — apply decision / outcome
        try:
            if not decision:
                # Only outcome is set (no decision col or empty decision)
                if outcome in ("won", "lost", "pending"):
                    record_outcome(tender_id, outcome, "bid")
                    synced += 1
                continue

            if decision == "no_bid":
                # No Bid without explicit outcome defaults to No Submission.
                _out = outcome if outcome in ("no_submission", "pending") else "no_submission"
                record_outcome(tender_id, _out, "no_bid")
                synced += 1

            elif decision == "bid":
                if outcome in ("won", "lost", "pending"):
                    # Bid placed AND outcome known — full record
                    record_outcome(tender_id, outcome, "bid")
                elif outcome == "no_submission":
                    # Inconsistent combo: a bid cannot have no-submission outcome.
                    logger.debug(
                        "[feedback_sync] Skipping inconsistent outcome for %s: %s + %s",
                        tender_id, decision, outcome
                    )
                else:
                    # Bid placed but result pending
                    _set_bid_decision(conn, tender_id, "bid")
                synced += 1

            elif decision == "review_later":
                # Review Later — preserve undecided state.
                if outcome == "pending":
                    record_outcome(tender_id, "pending", "review_later")
                else:
                    _set_bid_decision(conn, tender_id, "review_later")
                synced += 1

        except Exception as exc:
            logger.warning("[feedback_sync] Could not apply decision for %s: %s",
                           tender_id, exc)
            errors += 1

    cur.close()
    conn.close()

    summary = {
        "synced":  synced,
        "skipped": skipped,
        "no_url":  no_url,
        "total":   len(feedback_rows),
        "errors":  errors,
        "note":    (
            f"Synced {synced}/{len(feedback_rows)} decisions. "
            + (f"{skipped} URL(s) not matched. " if skipped else "")
            + (f"{errors} error(s)." if errors else "")
        ).strip(),
    }

    # Also ingest approved training labels from master workbook (non-fatal).
    try:
        label_sync = ingest_approved_labels(excel_path)
        summary["labels_ingested"] = int(label_sync.get("ingested", 0))
        summary["labels_skipped"] = int(label_sync.get("skipped", 0))
        summary["labels_errors"] = int(label_sync.get("errors", 0))
        if summary["labels_ingested"] > 0:
            summary["note"] = f"{summary['note']} Labels: {label_sync.get('note', '')}".strip()
    except Exception as exc:
        logger.warning("[feedback_sync] Approved label ingestion failed (non-fatal): %s", exc)
        summary["labels_ingested"] = 0
        summary["labels_skipped"] = 0
        summary["labels_errors"] = 1

    logger.info("[feedback_sync] Sync complete: %s", summary["note"])
    return summary


# =============================================================================
# SECTION 4 — Compute basic metrics  (Task 4)
# =============================================================================

def compute_feedback_metrics() -> Dict[str, Any]:
    """
    Query bid_pipeline for a lightweight feedback metrics snapshot.

    Returns:
        total_bid_now        : total tenders the model tagged BID_NOW
        bid_now_acted_on     : BID_NOW tenders where firm bid (bid_decision='bid')
        bid_now_ignored      : BID_NOW tenders where firm chose NO (bid_decision='no_bid')
        total_bids_placed    : all bids placed (any tier)
        total_wins           : outcome = 'won'
        total_losses         : outcome = 'lost'
        win_rate             : wins / (wins + losses)  — None if no data
        bid_now_act_rate     : bid_now_acted_on / total_bid_now  — None if no data
    """
    try:
        from database.db import get_connection
        conn = get_connection()
        cur  = conn.cursor()

        def _count(sql: str, args: tuple = ()) -> int:
            cur.execute(sql, args)
            row = cur.fetchone()
            return int(row[0]) if row else 0

        total_bid_now    = _count(
            "SELECT COUNT(*) FROM bid_pipeline WHERE model_decision_tag = 'BID_NOW'"
        )
        bid_now_acted_on = _count(
            "SELECT COUNT(*) FROM bid_pipeline "
            "WHERE model_decision_tag = 'BID_NOW' AND bid_decision = 'bid'"
        )
        bid_now_ignored  = _count(
            "SELECT COUNT(*) FROM bid_pipeline "
            "WHERE model_decision_tag = 'BID_NOW' "
            "AND bid_decision IN ('no_bid', 'review_later')"
        )
        total_bids       = _count(
            "SELECT COUNT(*) FROM bid_pipeline WHERE bid_decision = 'bid'"
        )
        total_wins       = _count(
            "SELECT COUNT(*) FROM bid_pipeline WHERE outcome = 'won'"
        )
        total_losses     = _count(
            "SELECT COUNT(*) FROM bid_pipeline WHERE outcome = 'lost'"
        )

        cur.close()
        conn.close()

        total_decided  = total_wins + total_losses
        win_rate       = round(total_wins / total_decided, 3) if total_decided > 0 else None
        act_rate       = round(bid_now_acted_on / total_bid_now, 3) if total_bid_now > 0 else None

        return {
            "total_bid_now":        total_bid_now,
            "bid_now_acted_on":     bid_now_acted_on,
            "bid_now_ignored":      bid_now_ignored,
            "total_bids_placed":    total_bids,
            "total_wins":           total_wins,
            "total_losses":         total_losses,
            "win_rate":             win_rate,
            "bid_now_act_rate":     act_rate,
        }

    except Exception as exc:
        logger.warning("[feedback_sync] compute_feedback_metrics failed: %s", exc)
        return {
            "total_bid_now": 0, "bid_now_acted_on": 0, "bid_now_ignored": 0,
            "total_bids_placed": 0, "total_wins": 0, "total_losses": 0,
            "win_rate": None, "bid_now_act_rate": None,
            "error": str(exc),
        }


# =============================================================================
# SECTION 5 — Log output  (Task 5)
# =============================================================================

def print_feedback_summary(metrics: Optional[Dict] = None) -> None:
    """
    Print a FEEDBACK SUMMARY block to stdout (captured by the logger).

    If metrics is None, computes fresh from DB.
    """
    if metrics is None:
        metrics = compute_feedback_metrics()

    def _pct(rate: Optional[float]) -> str:
        return f"{rate*100:.0f}%" if rate is not None else "—"

    total_bid_now   = metrics.get("total_bid_now",     0)
    bid_now_acted   = metrics.get("bid_now_acted_on",  0)
    bid_now_ignored = metrics.get("bid_now_ignored",   0)
    total_bids      = metrics.get("total_bids_placed", 0)
    total_wins      = metrics.get("total_wins",        0)
    total_losses    = metrics.get("total_losses",      0)
    win_rate        = metrics.get("win_rate")
    act_rate        = metrics.get("bid_now_act_rate")
    err             = metrics.get("error")

    W = 52
    print(f"\n{'─' * W}")
    print(f"  FEEDBACK SUMMARY")
    print(f"{'─' * W}")
    if err:
        print(f"  ⚠  Could not load metrics: {err}")
    else:
        print(f"  BID NOW tenders (model)  : {total_bid_now:>4}")
        print(f"  └─ Acted on (BID)        : {bid_now_acted:>4}  ({_pct(act_rate)})")
        print(f"  └─ Ignored (NO)          : {bid_now_ignored:>4}")
        print(f"  Total bids placed        : {total_bids:>4}")
        print(f"  Wins                     : {total_wins:>4}")
        print(f"  Losses                   : {total_losses:>4}")
        print(f"  Win rate (resolved bids) : {_pct(win_rate):>5}")
    print(f"{'─' * W}\n")
