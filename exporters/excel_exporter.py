# =============================================================================
# pipeline/excel_exporter.py — Unified Post-Run Excel Export
#
# Writes a single UNIFIED Excel file (Tender_Monitor_Master.xlsx) after each
# pipeline run containing ALL scraped tenders from ALL portals in one sheet,
# with the full set of required columns including sector and service type.
#
# This is separate from the per-portal Excel files (which each scraper writes
# itself and are preserved as-is). The unified file is the aggregate view used
# for human review and cross-portal analysis.
#
# Required columns (8 core + extras):
#   Title · Organization · Country · Deadline · Portal · URL ·
#   Sectors · Service Types ·
#   Relevance (keyword match) · Is New (this run) · Scraped Date
#
# Sector/service-type classification is performed by the ZERO-COST rule-based
# classify_tender() — does NOT need OpenAI. Runs on every row.
#
# Usage (from main.py):
#   from pipeline.excel_exporter import write_unified_excel
#   write_unified_excel(results, dry_run=args.dry_run)
# =============================================================================

from __future__ import annotations

import logging
import os
from datetime import datetime, timezone
from typing import Any, Dict, List


# ── Deadline formats tried in order (mirrors tender_intelligence._DEADLINE_FORMATS) ──
_EXPORT_DEADLINE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y",
    "%d %b, %Y", "%B %d, %Y", "%b %d, %Y", "%m/%d/%Y", "%Y/%m/%d",
    "%d %b %Y %H:%M", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M",
)


def _deadline_is_expired(deadline_str: str) -> bool:
    """Return True if the deadline string parses to a date strictly in the past."""
    raw = str(deadline_str or "").strip()
    if not raw or raw in ("", "N/A", "None", "TBD", "nan"):
        return False   # unknown deadline → do not drop
    raw_short = raw[:20]
    today = datetime.now(timezone.utc).date()
    for fmt in _EXPORT_DEADLINE_FORMATS:
        try:
            dl_date = datetime.strptime(raw_short, fmt).date()
            return dl_date < today
        except ValueError:
            continue
    return False   # unparseable → do not drop

logger = logging.getLogger("tenderradar.excel_exporter")


# ── Column schema for the unified sheet ──────────────────────────────────────
# (column_name, column_width)
# Phase 10: added Relevance Score (0-100) and Relevance Reason columns.
# "Sector" / "Service Type" / "Tender URL" are the canonical names.
UNIFIED_COLUMNS: List[tuple] = [
    ("Portal",               18),
    ("Title",                55),
    ("Organization",         28),
    ("Country",              15),
    ("Deadline",             15),
    ("Priority Score",       13),   # 0-100 composite opportunity priority (NEW)
    ("Relevance Score",      13),   # 0-100 keyword fit score
    ("Sector",               28),
    ("Service Type",         28),
    ("Cross Sources",        18),
    ("Relevance",            30),
    ("Relevance Reason",     45),   # one-sentence keyword explanation
    ("Opportunity Insight",  50),   # strategic insight text (NEW)
    ("Deep Scope",           60),
    ("Evaluation Criteria",  45),
    ("AI Summary",           45),
    ("Is New",                8),
    ("Tender URL",           55),
    ("Scraped Date",         16),
    ("My Decision",          15),   # user fills: Bid / No Bid / Review Later
    ("AI_Suggested_Label",   18),
    ("Human_Label",          15),
    ("Label_Reason",         36),
    ("Training_Approved",    16),
    ("Action_Label",         16),
    ("Label_Status",         16),
    # ── Opportunity Maturity Layer ────────────────────────────────────────────
    ("Evidence_State",       18),   # SIGNAL_ONLY / PAGE_ONLY / PARTIAL_PACKAGE / FULL_PACKAGE
    ("Opportunity_Maturity", 18),   # Signal First / Partial Package / Full Package
    ("Maturity_Summary",     45),   # one-line description of evidence depth
    ("Recommended_Action",   16),   # Monitor / Review / Prepare Bid
]

# Gold header fill for user-input columns (My Decision)
_USER_COL_HEADER_COLOR = "FFD966"   # amber-gold — signals "fill me in"
_USER_COL_HEADER_TEXT  = "7F4A00"   # dark brown text on gold

# Strict controlled feedback labels
DECISION_LABELS = ["Bid", "No Bid", "Review Later"]
HUMAN_LABELS = ["Relevant", "Borderline", "Not Relevant"]
TRAINING_APPROVED_LABELS = ["Yes", "No"]
ACTION_LABELS = ["Pursue", "Review Later", "Do Not Pursue"]

# ── 9-column standard for per-portal Excel files ──────────────────────────────
PORTAL_COLUMNS: List[tuple] = [
    ("Title",        60),
    ("Organization", 30),
    ("Country",      15),
    ("Sector",       25),
    ("Service Type", 25),
    ("Deadline",     16),
    ("Portal",       18),
    ("Tender URL",   55),
    ("Scraped Date", 16),
]

# ── Excel theme — same as BaseScraper and all Phase 1 pipelines ──────────────
_HDR_COLOR = "1F3864"   # dark navy header
_ALT_COLOR = "F5F8FF"   # even-row soft blue
_REL_COLOR = "E2EFDA"   # relevance cell soft green
_NEW_COLOR = "DDEEFF"   # light blue — new tender highlight

# ── Output quality gate (master Excel) ───────────────────────────────────────
# Keeps output actionable by suppressing obvious procurement-noise rows.
_OUTPUT_GATE_ENABLED = os.getenv("OUTPUT_QUALITY_GATE", "true").strip().lower() not in (
    "false", "0", "no", "off"
)
_MIN_PRIORITY = int(os.getenv("OUTPUT_MIN_PRIORITY", "20"))
_MIN_RELEVANCE = int(os.getenv("OUTPUT_MIN_RELEVANCE", "60"))
_MIN_NEW_PRIORITY = int(os.getenv("OUTPUT_MIN_NEW_PRIORITY", "50"))
_MIN_NEW_RELEVANCE = int(os.getenv("OUTPUT_MIN_NEW_RELEVANCE", "50"))

_POSITIVE_HINTS = (
    "consult", "evaluation", "assess", "baseline", "endline", "survey",
    "technical assistance", "ta ", "policy", "research", "capacity building",
    "monitoring", "m&e", "advisory", "study",
)
_NOISE_HINTS = (
    "supply", "procurement of", "construction", "civil work", "repair",
    "material", "equipment", "furniture", "poultry feed", "rcc nala",
    "drain", "pipe laying", "pump installation", "house keeping", "sweeper",
    "bitumen", "cement", "tile", "road work", "boundary wall",
)


def _extract_field(row: Dict, *aliases: str, default: str = "") -> str:
    """Try aliases in order; return first truthy value or default."""
    for k in aliases:
        v = row.get(k)
        if v and str(v).strip() not in ("", "N/A", "None", "nan"):
            return str(v).strip()
    return default


def _classify_row(row: Dict) -> tuple:
    """
    Wrapper — delegates to _classify_row_v2 (self-contained, no external imports).
    Kept for backward compatibility; new code should call _classify_row_v2 directly.
    """
    return _classify_row_v2(row)


# ── Lightweight label helpers (avoids importing Next.js constants) ────────────

_SECTOR_LABELS = {
    "health":                "Health",
    "education":             "Education",
    "environment":           "Environment",
    "agriculture":           "Agriculture",
    "water_sanitation":      "WASH",
    "urban_development":     "Urban",
    "energy":                "Energy",
    "governance":            "Governance",
    "gender_inclusion":      "Gender",
    "infrastructure":        "Infrastructure",
    "research":              "Research",
    "finance":               "Finance",
    "communications":        "Communications",
    "circular_economy":      "Circular Economy",
    "tourism":               "Tourism",
    "evaluation_monitoring": "M&E",
}

_SERVICE_LABELS = {
    "evaluation_monitoring": "Evaluation",
    "consulting_advisory":   "Consulting",
    "research_study":        "Research",
    "capacity_building":     "Capacity Building",
    "audit_finance":         "Audit",
    "communications_media":  "Communications",
    "project_management":    "PMC",
}


def _sector_label(slug: str) -> str:
    return _SECTOR_LABELS.get(slug, slug.replace("_", " ").title())


def _service_label(slug: str) -> str:
    return _SERVICE_LABELS.get(slug, slug.replace("_", " ").title())


def _normalize_feedback_label(column: str, value: str) -> str:
    """Normalize legacy feedback labels to strict dropdown values."""
    txt = " ".join(str(value or "").strip().lower().replace("_", " ").split())
    if not txt:
        return ""
    if column == "My Decision":
        mapping = {
            "bid": "Bid",
            "no": "No Bid",
            "no bid": "No Bid",
            "later": "Review Later",
            "review later": "Review Later",
            "pending": "Review Later",
        }
        return mapping.get(txt, "")
    return ""


def _to_int(v: Any) -> int:
    try:
        return int(float(v or 0))
    except (TypeError, ValueError):
        return 0


def _normalize_title_for_match(title: str) -> str:
    txt = " ".join(str(title or "").strip().lower().split())
    if not txt:
        return ""
    return "".join(ch for ch in txt if ch.isalnum() or ch.isspace()).strip()


def _suggest_ai_label(row: Dict[str, Any]) -> str:
    """
    Lightweight labeling using existing live scoring/flags only.
    """
    if _to_int(row.get("_is_consulting_relevant", 1)) == 0:
        return "Not Relevant"

    # IC / non-firm-eligible roles → always Not Relevant for labeling purposes.
    if _to_int(row.get("_is_firm_eligible", 1)) == 0:
        return "Not Relevant"

    priority = _to_int(row.get("Priority Score"))
    relevance = _to_int(row.get("Relevance Score"))
    low_conf = _to_int(row.get("_is_low_confidence", 0)) == 1
    client_fit = _to_int(row.get("_client_fit_score", 0))
    service_fit = _to_int(row.get("_service_fit_score", 0))
    consulting_conf = _to_int(row.get("_consulting_confidence_score", 0))
    procurement_penalty = _to_int(row.get("_procurement_penalty_score", 0))

    title = str(row.get("Title") or "").lower()
    reason = str(row.get("Relevance Reason") or "").lower()
    scoring_note = str(row.get("_scoring_note") or "").lower()
    org = str(row.get("Organization") or "").lower()
    service_type = str(row.get("Service Type") or "").lower()
    sector = str(row.get("Sector") or "").lower()
    text = " | ".join([title, reason, scoring_note, org, service_type, sector])

    # Catch IC noise that may not have been picked up by scoring layer (edge cases).
    # Checks title + scoring_note + relevance_reason for IC patterns.
    _ic_label_signals = (
        "individual consultant", "individual contractor", "individual expert",
        "short-term consultant", "national consultant", "international consultant",
        "external collaborator", "external expert", "ic role", "not firm eligible",
        # UNDP/UN vacancy title patterns (e.g. "IC - ", "National IC -", "; IC -")
        " ic - ", " ic–", "national ic", "local ic", "ic position", "- ic -",
    )
    if any(k in text for k in _ic_label_signals):
        return "Not Relevant"

    donor_terms = (
        "world bank", "undp", "ilo", "adb", "afdb", "afd", "fao", "unicef",
        "un women", "unops", "giz", "eu", "european commission", "dfid", "usaid",
    )
    strong_service_terms = (
        "evaluation", "m&e", "monitoring and evaluation", "assessment",
        "baseline", "endline", "advisory", "technical assistance", "policy",
        "institutional strengthening", "capacity building", "feasibility study",
    )
    borderline_service_terms = (
        "pmc", "project management consultant", "engineering advisory",
        "design supervision", "owner's engineer", "transaction advisory",
    )
    negative_terms = (
        "supply", "goods", "construction", "civil work", "manpower",
        "equipment", "furniture", "repair", "installation",
    )

    has_donor = any(k in text for k in donor_terms)
    has_strong_service = any(k in text for k in strong_service_terms)
    has_borderline_service = any(k in text for k in borderline_service_terms)
    has_negative_pattern = any(k in text for k in negative_terms)

    # Strong non-fit / procurement-like cases.
    if procurement_penalty >= 45:
        return "Not Relevant"
    if has_negative_pattern and not has_strong_service and procurement_penalty >= 20:
        return "Not Relevant"

    # Low-confidence rows should not default to Borderline.
    if low_conf:
        if (
            has_donor
            and has_strong_service
            and consulting_conf >= 55
            and procurement_penalty <= 20
            and (priority >= 45 or relevance >= 50)
        ):
            return "Borderline"
        return "Not Relevant"

    # Strong IDCG fit promotion.
    if (
        has_strong_service
        and (has_donor or client_fit >= 60)
        and (service_fit >= 55 or consulting_conf >= 60)
        and procurement_penalty <= 20
        and (priority >= 55 or relevance >= 60)
    ):
        return "Relevant"
    if (
        (priority >= 72 and relevance >= 62)
        or (priority >= 65 and relevance >= 68)
        or (service_fit >= 70 and consulting_conf >= 65 and procurement_penalty <= 20)
    ):
        return "Relevant"

    # Mixed but plausible consulting opportunities.
    if (
        (priority >= 45 and relevance >= 45)
        or (service_fit >= 45 and consulting_conf >= 45)
        or has_borderline_service
        or (has_strong_service and procurement_penalty <= 30)
    ):
        return "Borderline"

    return "Not Relevant"


def _passes_output_quality_gate(row: Dict[str, Any]) -> bool:
    """
    Final output gate for master Excel quality.
    This does not affect DB persistence; it only controls what reaches output.
    """
    if not _OUTPUT_GATE_ENABLED:
        return True

    # Keep explicit analyst decisions regardless of score.
    decision = str(row.get("My Decision") or "").strip()
    if decision in ("Bid", "Review Later"):
        return True

    # Defensive: drop tenders with a past deadline regardless of priority score.
    # This is defense-in-depth: the scoring pipeline already sets priority=0 for
    # expired deadlines, but this catches any case where that path was bypassed
    # (e.g., deadline_category = "unknown" but the raw string is in the past).
    deadline_str = str(row.get("Deadline") or "").strip()
    if deadline_str and _deadline_is_expired(deadline_str):
        return False

    title = str(row.get("Title") or "").lower()
    relevance_text = str(row.get("Relevance") or "").lower()
    service_text = str(row.get("Service Type") or "").lower()
    portal = str(row.get("Portal") or "")

    priority = _to_int(row.get("Priority Score"))
    relevance = _to_int(row.get("Relevance Score"))
    is_new = str(row.get("Is New") or "").strip().upper() == "YES"

    has_positive = any(k in title or k in relevance_text or k in service_text for k in _POSITIVE_HINTS)
    has_noise = any(k in title for k in _NOISE_HINTS)

    # Hard reject obvious non-consulting procurement rows unless strongly relevant.
    if has_noise and not has_positive and priority < 75 and relevance < 75:
        return False

    # Primary thresholds.
    if priority >= _MIN_PRIORITY and relevance >= _MIN_RELEVANCE:
        return True
    if priority >= 75 or relevance >= 80:
        return True
    if is_new and priority >= _MIN_NEW_PRIORITY and relevance >= _MIN_NEW_RELEVANCE and has_positive:
        return True

    # Strong international consulting portals can pass with moderate scores.
    if portal in ("World Bank", "UNDP Procurement", "ILO Procurement", "DTVP Germany", "European Commission (EC)", "TED EU"):
        if priority >= 45 and relevance >= 55 and has_positive:
            return True

    return False


# ── Core: classify using internal helpers (avoids lib_utils import issue) ────

def _classify_row_v2(row: Dict) -> tuple[str, str]:
    """Classify without external lib_utils dependency."""
    try:
        from intelligence.classifier import classify_tender
        title       = _extract_field(row, "title", "Title")
        description = _extract_field(row, "description", "Description", "Summary")
        cls         = classify_tender(title, description)
        sectors_str  = ", ".join(_sector_label(s)  for s in cls.sectors[:3])
        services_str = ", ".join(_service_label(s) for s in cls.service_types[:3])
        return sectors_str, services_str
    except Exception as exc:
        logger.debug(f"[excel_exporter] classify skipped: {exc}")
        return "", ""


# =============================================================================
# Opportunity data batch-fetch (priority_score + opportunity_insight)
# =============================================================================

def _resolve_tid(row: Dict) -> str:
    """Extract tender_id from a raw scraper row using common field name variants."""
    return str(
        row.get("tender_id") or row.get("id")
        or row.get("sol_num") or row.get("Bid Number") or ""
    ).strip()[:255]


def _fetch_opportunity_data(rows: List[Dict]) -> Dict[str, Dict]:
    """
    Batch-fetch priority_score and opportunity_insight from tender_structured_intel
    for the given raw scraper rows.

    Returns {tender_id → {"priority_score": int, "opportunity_insight": str}}.
    Returns {} on any DB error (graceful degradation: rows fall back to 0/empty).
    """
    ids = [_resolve_tid(r) for r in rows]
    ids = list({tid for tid in ids if tid})   # deduplicate
    if not ids:
        return {}

    try:
        from database.db import get_connection
        conn = get_connection()
        cur  = conn.cursor(dictionary=True)
        placeholders = ", ".join(["%s"] * len(ids))
        cur.execute(
            f"SELECT tender_id, priority_score, opportunity_insight, "
            f"client_fit_score, service_fit_score, consulting_confidence_score, procurement_penalty_score "
            f"FROM tender_structured_intel "
            f"WHERE tender_id IN ({placeholders})",
            ids,
        )
        result = {r["tender_id"]: r for r in (cur.fetchall() or [])}
        cur.close()
        conn.close()
        return result
    except Exception as exc:
        logger.debug(f"[excel_exporter] _fetch_opportunity_data failed: {exc}")
        return {}


def _fetch_rich_tender_data(rows: List[Dict]) -> Dict[str, Dict]:
    """
    Pull the richest currently known backend record for each tender.

    Uses `v_tender_full` plus grouped cross-source metadata so Excel can prefer
    merged DB knowledge over thinner raw scraper rows.
    """
    ids = [_resolve_tid(r) for r in rows]
    ids = list({tid for tid in ids if tid})
    urls = []
    for row in rows:
        url = (
            row.get("Tender URL")
            or row.get("url")
            or row.get("URL")
            or row.get("detail_url")
            or row.get("Detail Link")
            or row.get("link")
            or ""
        )
        url = str(url).strip()
        if url:
            urls.append(url)
    urls = list({url for url in urls if url})
    if not ids and not urls:
        return {}

    try:
        from database.db import get_connection

        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        clauses = []
        params: List[str] = []
        if ids:
            placeholders = ", ".join(["%s"] * len(ids))
            clauses.append(f"v.tender_id IN ({placeholders})")
            params.extend(ids)
        if urls:
            placeholders = ", ".join(["%s"] * len(urls))
            clauses.append(f"v.url IN ({placeholders})")
            params.extend(urls)
        cur.execute(
            f"""
            SELECT
                v.tender_id,
                v.title,
                v.organization,
                v.country,
                v.deadline,
                v.deadline_raw,
                v.sector,
                v.consulting_type,
                v.relevance_score,
                v.priority_score,
                v.opportunity_insight,
                v.is_consulting_relevant,
                v.is_low_confidence,
                v.scoring_note,
                v.description,
                v.deep_scope,
                v.deep_description,
                v.deep_pdf_text,
                v.deep_eval_criteria,
                COALESCE(v.deep_ai_summary, v.ai_summary, '') AS ai_summary,
                v.deep_document_links,
                v.url,
                COUNT(cs.id) AS cross_source_count,
                GROUP_CONCAT(DISTINCT COALESCE(cs.source_portal, cs.portal) ORDER BY COALESCE(cs.source_portal, cs.portal) SEPARATOR ', ') AS cross_source_portals
            FROM v_tender_full v
            LEFT JOIN tender_cross_sources cs ON v.tender_id = cs.tender_id
            WHERE {" OR ".join(clauses)}
            GROUP BY
                v.tender_id, v.title, v.organization, v.country, v.deadline,
                v.deadline_raw, v.sector, v.consulting_type, v.relevance_score,
                v.priority_score, v.opportunity_insight, v.is_consulting_relevant,
                v.is_low_confidence, v.scoring_note, v.description, v.deep_scope,
                v.deep_description, v.deep_pdf_text, v.deep_eval_criteria,
                COALESCE(v.deep_ai_summary, v.ai_summary, ''), v.deep_document_links, v.url
            """,
            params,
        )
        out: Dict[str, Dict] = {}
        for record in (cur.fetchall() or []):
            tid = str(record.get("tender_id") or "").strip()
            url = str(record.get("url") or "").strip().lower()
            if tid:
                out[tid] = record
            if url:
                out[url] = record
        cur.close()
        conn.close()
        return out
    except Exception as exc:
        logger.debug(f"[excel_exporter] _fetch_rich_tender_data failed: {exc}")
        return {}


# =============================================================================
# Main entry point
# =============================================================================

def write_unified_excel(
    results:  List[Any],   # list of pipeline.runner.JobResult
    dry_run:  bool = False,
) -> str:
    """
    Build and save the unified Master Excel from all job results.

    Args:
        results:  List of JobResult objects from pipeline.runner.JobRunner.run().
        dry_run:  If True, still writes the Excel (so output can be verified)
                  but logs that it's a dry-run.

    Returns:
        Path of the written Excel file, or "" on failure.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from config.config import UNIFIED_EXCEL_PATH

    if dry_run:
        logger.info("[excel_exporter] DRY-RUN: unified Excel will still be written "
                    "(DB writes are skipped, Excel is not)")

    # ── Import numeric scorer ─────────────────────────────────────────────────
    try:
        from intelligence.keywords import score_tender_numeric as _score_numeric
    except Exception as _e:
        logger.warning(f"[excel_exporter] Could not import score_tender_numeric: {_e}")
        def _score_numeric(title, description="", country=""):
            return 0, "Scorer unavailable."

    # ── Aggregate all rows from every job result ──────────────────────────────
    unified_rows: List[Dict] = []
    new_tender_ids: set       = set()

    for result in results:
        portal_label = getattr(result, "label", getattr(result, "flag", "unknown"))

        # Mark which tender IDs are new this run
        for t in (getattr(result, "new_tenders", []) or []):
            tid = t.get("url", "") or t.get("title", "")
            new_tender_ids.add(tid)

        for row in (getattr(result, "all_rows", []) or []):
            tender_id = _resolve_tid(row)
            title  = _extract_field(row, "title", "Title", "Tender Title", "Project")
            org    = _extract_field(row, "organization", "Organization",
                                    "Authority", "Department", "Ministry",
                                    "Client", "Agency", "Entity")
            country = _extract_field(row, "country", "Country", "Location", "Geography")
            if not country:
                # Infer India for domestic portals
                if any(tag in portal_label.lower()
                       for tag in ("gem", "cg", "ngobox", "devnet", "karnataka",
                                   "meghalaya", "sidbi", "icfre", "phfi", "jtds")):
                    country = "India"

            deadline = _extract_field(row, "deadline", "Deadline",
                                      "Bid Closing Date", "Last Date",
                                      "Response Deadline", "Closing Date",
                                      "Submission Deadline")
            url      = _extract_field(row, "url", "Detail Link", "link",
                                      "detail_url", "URL", "Link",
                                      "PDF / Link", "Source URL", "Tender URL",
                                      "detail_link", "pdf_link")
            relevance = _extract_field(row, "Relevance", "relevance",
                                       "keyword_relevance")
            description = _extract_field(row, "description", "Description",
                                         "Summary", "Brief Description")

            sectors_str, services_str = _classify_row_v2(row)

            # ── Numeric relevance score ────────────────────────────────────────
            rel_score, rel_reason = _score_numeric(title, description + " " + relevance, country)

            # Is this a new tender this run?
            is_new = "YES" if (url in new_tender_ids or title in new_tender_ids) else ""

            unified_rows.append({
                "Portal":            portal_label,
                "Title":             title[:400] if title else "",
                "Organization":      org[:200]   if org else "",
                "Country":           country,
                "Deadline":          deadline,
                "Relevance Score":   rel_score,
                "Sector":            sectors_str,
                "Service Type":      services_str,
                "Relevance":         relevance[:200] if relevance else "",
                "Relevance Reason":  rel_reason[:300],
                "Is New":            is_new,
                "Tender URL":        url[:2000]  if url else "",
                "Scraped Date":      datetime.now().strftime("%Y-%m-%d %H:%M"),
                "_tender_id":        tender_id,
                "_from_current_run": 1,
            })

    if not unified_rows:
        logger.warning("[excel_exporter] No rows to write — skipping unified Excel")
        return ""

    # ── Enrich with priority_score + opportunity_insight from DB ─────────────
    # Single batch SELECT for all current-run rows; falls back silently on error.
    _all_raw_rows = [row for result in results
                     for row in (getattr(result, "all_rows", []) or [])]
    _opp_data = _fetch_opportunity_data(_all_raw_rows)
    _rich_data = _fetch_rich_tender_data(_all_raw_rows)

    # Build a URL+title → tid map so we can look up by the same keys used
    # for is_new detection (URL is the most reliable cross-scraper identifier).
    _tid_by_url: Dict[str, str] = {}
    for raw in _all_raw_rows:
        tid = _resolve_tid(raw)
        if tid:
            u = (raw.get("url") or raw.get("URL") or raw.get("detail_url") or
                 raw.get("Detail Link") or raw.get("link") or "").strip()
            if u:
                _tid_by_url[u] = tid

    def _apply_rich_evidence(urow: Dict[str, Any], rich: Dict[str, Any]) -> None:
        """
        Hydrate a workbook row with the richest evidence available before maturity classification.

        Authoritative source:
          - Current-run rows: DB-enriched `v_tender_full` fields when present.
          - Carry-over rows: previous workbook columns first, then refreshed from DB by tender_id/URL.
        """
        if not rich:
            return

        urow["Organization"] = (urow.get("Organization") or rich.get("organization") or "")[:200]
        urow["Country"] = urow.get("Country") or str(rich.get("country") or "")
        urow["Deadline"] = urow.get("Deadline") or str(rich.get("deadline") or rich.get("deadline_raw") or "")
        if not urow.get("Sector") and rich.get("sector"):
            urow["Sector"] = str(rich.get("sector") or "")
        if not urow.get("Service Type") and rich.get("consulting_type"):
            urow["Service Type"] = str(rich.get("consulting_type") or "")
        if not urow.get("Relevance Score"):
            try:
                urow["Relevance Score"] = int(rich.get("relevance_score") or 0)
            except (ValueError, TypeError):
                pass
        rich_desc = str(rich.get("description") or "")
        if rich_desc and not urow.get("Relevance"):
            urow["Relevance"] = rich_desc[:200]
        urow["Cross Sources"] = str(rich.get("cross_source_count") or 0)
        if rich.get("cross_source_portals"):
            urow["Relevance Reason"] = (
                (urow.get("Relevance Reason") or "") + f" Sources: {rich.get('cross_source_portals')}."
            ).strip()[:300]

        deep_scope = str(rich.get("deep_scope") or "").strip()
        deep_description = str(rich.get("deep_description") or "").strip()
        deep_pdf_text = str(rich.get("deep_pdf_text") or "").strip()
        ai_summary = str(rich.get("ai_summary") or "").strip()

        if not urow.get("Deep Scope"):
            urow["Deep Scope"] = (deep_scope or deep_description or deep_pdf_text)[:1000]
        if not urow.get("Evaluation Criteria"):
            urow["Evaluation Criteria"] = str(rich.get("deep_eval_criteria") or "")[:800]
        if not urow.get("AI Summary"):
            urow["AI Summary"] = ai_summary[:500]

        # Hidden row-only evidence for the shared maturity classifier / packager parity.
        urow["_deep_scope"] = deep_scope
        urow["_deep_description"] = deep_description
        urow["_deep_pdf_text"] = deep_pdf_text
        urow["_deep_ai_summary"] = ai_summary
        urow["_deep_document_links"] = rich.get("deep_document_links")

    for urow in unified_rows:
        # Try URL-based lookup first, then row-level tender_id
        _u = urow.get("Tender URL", "")
        _tid = _tid_by_url.get(_u) or str(urow.get("_tender_id") or "")
        urow["_tender_id"] = _tid
        _opp = _opp_data.get(_tid, {}) if _tid else {}
        _rich = (_rich_data.get(_tid) if _tid else None) or _rich_data.get(str(_u).strip().lower(), {})
        try:
            urow["Priority Score"] = int((_opp.get("priority_score") or _rich.get("priority_score") or 0))
        except (ValueError, TypeError):
            urow["Priority Score"] = 0
        urow["Opportunity Insight"] = str(
            _opp.get("opportunity_insight") or _rich.get("opportunity_insight") or ""
        )[:500]
        urow["_is_consulting_relevant"] = _to_int(
            _rich.get("is_consulting_relevant", 1)
        )
        urow["_is_low_confidence"] = _to_int(_rich.get("is_low_confidence", 0))
        urow["_scoring_note"] = str(_rich.get("scoring_note") or "").strip()
        urow["_client_fit_score"] = _to_int(
            _opp.get("client_fit_score") if _opp.get("client_fit_score") is not None else _rich.get("client_fit_score", 0)
        )
        urow["_service_fit_score"] = _to_int(
            _opp.get("service_fit_score") if _opp.get("service_fit_score") is not None else _rich.get("service_fit_score", 0)
        )
        urow["_consulting_confidence_score"] = _to_int(
            _opp.get("consulting_confidence_score") if _opp.get("consulting_confidence_score") is not None else _rich.get("consulting_confidence_score", 0)
        )
        urow["_procurement_penalty_score"] = _to_int(
            _opp.get("procurement_penalty_score") if _opp.get("procurement_penalty_score") is not None else _rich.get("procurement_penalty_score", 0)
        )
        _apply_rich_evidence(urow, _rich)

        # ── Opportunity Maturity (computed after all rich fields are set) ──────
        # Uses DB-backed deep fields for current-run rows and workbook/DB carry-forward
        # fields for carry-over rows so all outputs share the same maturity inputs.
        try:
            from intelligence.maturity import classify_row as _classify_maturity
            _maturity = _classify_maturity({
                **urow,
                "deep_scope": urow.get("_deep_scope") or urow.get("Deep Scope"),
                "deep_description": urow.get("_deep_description") or urow.get("Deep Scope"),
                "deep_pdf_text": urow.get("_deep_pdf_text") or "",
                "deep_ai_summary": urow.get("_deep_ai_summary") or urow.get("AI Summary"),
                "deep_document_links": urow.get("_deep_document_links"),
                "evaluation_criteria": urow.get("Evaluation Criteria"),
                "cross_sources": urow.get("Cross Sources"),
                "description": urow.get("Relevance"),
            })
            urow["Evidence_State"]       = _maturity["evidence_state"]
            urow["Opportunity_Maturity"] = _maturity["opportunity_maturity"]
            urow["Maturity_Summary"]     = _maturity["maturity_summary"]
            urow["Recommended_Action"]   = _maturity["recommended_action"]
        except Exception as _me:
            logger.debug("[excel_exporter] maturity classify skipped: %s", _me)
            urow["Evidence_State"]       = "SIGNAL_ONLY"
            urow["Opportunity_Maturity"] = "Signal First"
            urow["Maturity_Summary"]     = ""
            urow["Recommended_Action"]   = "Monitor"

        if urow.get("_scoring_note"):
            base_reason = str(urow.get("Relevance Reason") or "").strip()
            urow["Relevance Reason"] = (
                urow["_scoring_note"]
                if not base_reason
                else f"{urow['_scoring_note']} | {base_reason}"
            )[:300]
        if urow.get("_is_low_confidence") == 1:
            urow["Relevance Reason"] = (
                f"[LOW CONFIDENCE] {str(urow.get('Relevance Reason') or '').strip()}"
            )[:300]

    # ── Dedup key helper ──────────────────────────────────────────────────────
    def _row_key(row: Dict) -> str:
        """
        Stable deduplication key for a unified row.
        Primary:  Tender URL (case-insensitive, stripped) — most reliable.
        Fallback: 'Portal|Title[:120]' when URL is absent (e.g. some domestic portals).
        Returns "" for rows that have neither — those are never deduplicated.
        """
        url = (row.get("Tender URL") or "").strip().lower()
        if url:
            return url
        portal = (row.get("Portal") or "").strip().lower()
        title  = (row.get("Title")  or "").strip().lower()[:120]
        if portal or title:
            return f"\x00{portal}|{title}"   # \x00 prefix marks fallback keys
        return ""

    # ── Previous-review cache (for label carry-forward) ──────────────────────
    _prev_feedback_by_url: Dict[str, Dict[str, str]] = {}
    _prev_feedback_by_tid: Dict[str, Dict[str, str]] = {}
    _prev_feedback_by_title: Dict[str, Dict[str, str]] = {}

    def _collect_prev_feedback(row_vals: tuple, headers: List[Any], idx_map: Dict[str, Any]) -> None:
        url_val = str(row_vals[idx_map["url_ci"]] or "").strip().lower() if idx_map["url_ci"] is not None else ""
        tid_val = str(row_vals[idx_map["tid_ci"]] or "").strip() if idx_map["tid_ci"] is not None else ""
        title_val = str(row_vals[idx_map["title_ci"]] or "").strip() if idx_map["title_ci"] is not None else ""
        title_key = _normalize_title_for_match(title_val)

        feedback = {
            "My Decision": _normalize_feedback_label(
                "My Decision",
                str(row_vals[idx_map["decision_ci"]] or "") if idx_map["decision_ci"] is not None else "",
            ),
            "Human_Label": str(row_vals[idx_map["human_label_ci"]] or "").strip() if idx_map["human_label_ci"] is not None else "",
            "Label_Reason": str(row_vals[idx_map["label_reason_ci"]] or "").strip() if idx_map["label_reason_ci"] is not None else "",
            "Training_Approved": str(row_vals[idx_map["training_ci"]] or "").strip() if idx_map["training_ci"] is not None else "",
            "Action_Label": str(row_vals[idx_map["action_ci"]] or "").strip() if idx_map["action_ci"] is not None else "",
            "AI_Suggested_Label": str(row_vals[idx_map["ai_label_ci"]] or "").strip() if idx_map["ai_label_ci"] is not None else "",
        }

        if url_val:
            _prev_feedback_by_url[url_val] = feedback
        if tid_val:
            _prev_feedback_by_tid[tid_val] = feedback
        if title_key:
            _prev_feedback_by_title[title_key] = feedback

    # ── Merge rows from portals NOT in this run (from previous unified master) ─
    # Strategy:
    #   • Portals IN current run  → use ONLY the freshly-scraped rows (full replace).
    #   • Portals NOT in current run → carry over all rows from the previous master,
    #     but skip any whose Tender URL already appears in the current rows
    #     (cross-portal duplicate defence).
    # The most-recent data always wins: current-run rows are indexed first.
    current_portals = {row["Portal"] for row in unified_rows}
    seen_keys: set = {k for r in unified_rows if (k := _row_key(r))}

    if os.path.exists(UNIFIED_EXCEL_PATH):
        try:
            from openpyxl import load_workbook as _load_wb
            wb_old = _load_wb(UNIFIED_EXCEL_PATH, read_only=True, data_only=True)
            ws_old = wb_old.active
            old_headers_row = next(ws_old.iter_rows(max_row=1), None)
            if old_headers_row:
                old_headers = [c.value for c in old_headers_row]
                valid_cols   = {c[0] for c in UNIFIED_COLUMNS}  # includes Priority Score + Opportunity Insight
                idx_map = {
                    "url_ci": next((i for i, h in enumerate(old_headers) if h == "Tender URL"), None),
                    "tid_ci": next((i for i, h in enumerate(old_headers) if h == "Tender ID"), None),
                    "title_ci": next((i for i, h in enumerate(old_headers) if h == "Title"), None),
                    "decision_ci": next((i for i, h in enumerate(old_headers) if h == "My Decision"), None),
                    "human_label_ci": next((i for i, h in enumerate(old_headers) if h == "Human_Label"), None),
                    "label_reason_ci": next((i for i, h in enumerate(old_headers) if h == "Label_Reason"), None),
                    "training_ci": next((i for i, h in enumerate(old_headers) if h == "Training_Approved"), None),
                    "action_ci": next((i for i, h in enumerate(old_headers) if h == "Action_Label"), None),
                    "ai_label_ci": next((i for i, h in enumerate(old_headers) if h == "AI_Suggested_Label"), None),
                }
                old_count    = 0
                dup_count    = 0
                for row_vals in ws_old.iter_rows(min_row=2, values_only=True):
                    row_dict = {h: (str(v) if v is not None else "")
                                for h, v in zip(old_headers, row_vals)
                                if h in valid_cols}
                    row_dict["_tender_id"] = str(row_vals[idx_map["tid_ci"]] or "").strip() if idx_map["tid_ci"] is not None else ""

                    # Collect previous review inputs for all rows.
                    _collect_prev_feedback(row_vals, old_headers, idx_map)

                    portal = row_dict.get("Portal", "")

                    # Skip: portal fully covered by current run
                    if portal in current_portals:
                        continue

                    # Skip: URL already present (cross-portal duplicate)
                    key = _row_key(row_dict)
                    if key and key in seen_keys:
                        dup_count += 1
                        continue

                    row_dict["Is New"] = ""   # clear — not new in this run
                    row_dict["_from_current_run"] = 0
                    # Re-score carry-over rows with current algorithm
                    if not row_dict.get("Relevance Score"):
                        _t = row_dict.get("Title", "")
                        _d = row_dict.get("Relevance", "")
                        _c = row_dict.get("Country", "")
                        _rs, _rr = _score_numeric(_t, _d, _c)
                        row_dict["Relevance Score"]  = _rs
                        row_dict["Relevance Reason"] = _rr
                    # Re-classify maturity for carry-over rows so all rows are consistent
                    if not row_dict.get("Evidence_State"):
                        try:
                            from intelligence.maturity import classify_row as _cm
                            _mat = _cm(row_dict)
                            row_dict["Evidence_State"]       = _mat["evidence_state"]
                            row_dict["Opportunity_Maturity"] = _mat["opportunity_maturity"]
                            row_dict["Maturity_Summary"]     = _mat["maturity_summary"]
                            row_dict["Recommended_Action"]   = _mat["recommended_action"]
                        except Exception:
                            row_dict["Evidence_State"]       = "SIGNAL_ONLY"
                            row_dict["Opportunity_Maturity"] = "Signal First"
                            row_dict["Maturity_Summary"]     = ""
                            row_dict["Recommended_Action"]   = "Monitor"
                    unified_rows.append(row_dict)
                    if key:
                        seen_keys.add(key)
                    old_count += 1

            wb_old.close()
            parts = [f"{old_count} rows from {len(set(r.get('Portal','') for r in unified_rows) - current_portals)} carry-over portals"]
            if dup_count:
                parts.append(f"{dup_count} cross-portal duplicates removed")
            logger.info(f"[excel_exporter] Merged: {', '.join(parts)}")
        except Exception as exc:
            logger.warning(f"[excel_exporter] Could not merge previous unified master: {exc}")

    # ── Back-fill decisions/labels for re-scraped rows (match URL → TID → title) ─
    _backfill_count = 0
    for urow in unified_rows:
        _u = (urow.get("Tender URL") or "").strip().lower()
        _tid = str(urow.get("_tender_id") or "").strip()
        _title_key = _normalize_title_for_match(urow.get("Title") or "")
        prev = (
            _prev_feedback_by_url.get(_u)
            or (_prev_feedback_by_tid.get(_tid) if _tid else None)
            or (_prev_feedback_by_title.get(_title_key) if _title_key else None)
        )
        if not prev:
            continue

        changed = False
        if not urow.get("My Decision") and prev.get("My Decision"):
            urow["My Decision"] = prev["My Decision"]
            changed = True
        for lbl_col in ("Human_Label", "Label_Reason", "Training_Approved", "Action_Label"):
            if not str(urow.get(lbl_col) or "").strip() and prev.get(lbl_col):
                urow[lbl_col] = prev[lbl_col]
                changed = True
        if changed:
            _backfill_count += 1

        urow["_prev_ai_suggested_label"] = prev.get("AI_Suggested_Label", "")
        urow["_has_prev_review"] = any(
            str(prev.get(col) or "").strip()
            for col in ("Human_Label", "Label_Reason", "Training_Approved", "Action_Label")
        )
    if _backfill_count:
        logger.info(
            "[excel_exporter] Back-filled prior decisions/labels for %d row(s)",
            _backfill_count,
        )

    # Re-hydrate all rows after carry-over merge so maturity uses the same DB-backed
    # evidence sources for both current-run and carry-over rows wherever possible.
    _all_rich_data = _fetch_rich_tender_data(unified_rows)
    for urow in unified_rows:
        _u = str(urow.get("Tender URL") or "").strip().lower()
        _tid = str(urow.get("_tender_id") or "").strip()
        _rich = (_all_rich_data.get(_tid) if _tid else None) or _all_rich_data.get(_u, {})
        _apply_rich_evidence(urow, _rich)
        try:
            from intelligence.maturity import classify_row as _classify_maturity
            _maturity = _classify_maturity({
                **urow,
                "deep_scope": urow.get("_deep_scope") or urow.get("Deep Scope"),
                "deep_description": urow.get("_deep_description") or urow.get("Deep Scope"),
                "deep_pdf_text": urow.get("_deep_pdf_text") or "",
                "deep_ai_summary": urow.get("_deep_ai_summary") or urow.get("AI Summary"),
                "deep_document_links": urow.get("_deep_document_links"),
                "evaluation_criteria": urow.get("Evaluation Criteria"),
                "cross_sources": urow.get("Cross Sources"),
                "description": urow.get("Relevance"),
            })
            urow["Evidence_State"]       = _maturity["evidence_state"]
            urow["Opportunity_Maturity"] = _maturity["opportunity_maturity"]
            urow["Maturity_Summary"]     = _maturity["maturity_summary"]
            urow["Recommended_Action"]   = _maturity["recommended_action"]
        except Exception as _me:
            logger.debug("[excel_exporter] maturity reclassify skipped: %s", _me)

    # ── Final quality gate (reduce non-consulting noise in output) ───────────
    _before_gate = len(unified_rows)
    _dropped_new = 0
    _gated_rows: List[Dict] = []
    for _r in unified_rows:
        # Hard export gate from DB relevance layer.
        if _to_int(_r.get("_is_consulting_relevant", 1)) == 0:
            continue
        _keep = _passes_output_quality_gate(_r)
        # Low-confidence rows are allowed only if new this run or explicitly
        # selected by analyst; they are demoted in sorting.
        if _keep and _to_int(_r.get("_is_low_confidence", 0)) == 1:
            _decision = str(_r.get("My Decision") or "").strip()
            _is_new_lc = str(_r.get("Is New") or "").strip().upper() == "YES"
            if _decision not in ("Bid", "Review Later") and not _is_new_lc:
                _keep = False
        if _keep:
            _gated_rows.append(_r)
        else:
            if str(_r.get("Is New") or "").strip().upper() == "YES":
                _dropped_new += 1
    unified_rows = _gated_rows
    if _OUTPUT_GATE_ENABLED:
        logger.info(
            "[excel_exporter] Output gate: kept %d/%d rows (dropped=%d, dropped_new=%d)",
            len(unified_rows),
            _before_gate,
            _before_gate - len(unified_rows),
            _dropped_new,
        )

    def _maturity_counts(rows: List[Dict[str, Any]]) -> Dict[str, int]:
        counts = {"Signal First": 0, "Partial Package": 0, "Full Package": 0}
        for row in rows:
            label = str(row.get("Opportunity_Maturity") or "").strip()
            if label in counts:
                counts[label] += 1
        return counts

    _all_maturity = _maturity_counts(unified_rows)
    _current_maturity = _maturity_counts([r for r in unified_rows if _to_int(r.get("_from_current_run", 0)) == 1])
    _carry_maturity = _maturity_counts([r for r in unified_rows if _to_int(r.get("_from_current_run", 0)) != 1])
    logger.info(
        "[excel_exporter] Maturity distribution — all: Signal First=%d, Partial Package=%d, Full Package=%d | "
        "current-run: Signal First=%d, Partial Package=%d, Full Package=%d | "
        "carry-over: Signal First=%d, Partial Package=%d, Full Package=%d",
        _all_maturity["Signal First"], _all_maturity["Partial Package"], _all_maturity["Full Package"],
        _current_maturity["Signal First"], _current_maturity["Partial Package"], _current_maturity["Full Package"],
        _carry_maturity["Signal First"], _carry_maturity["Partial Package"], _carry_maturity["Full Package"],
    )

    # ── Labeling fields (master-only) ────────────────────────────────────────
    for _r in unified_rows:
        ai_label = _suggest_ai_label(_r)
        _r["AI_Suggested_Label"] = ai_label
        _r["Human_Label"] = str(_r.get("Human_Label") or "").strip()
        _r["Label_Reason"] = str(_r.get("Label_Reason") or "").strip()
        _r["Training_Approved"] = str(_r.get("Training_Approved") or "").strip() or "No"
        _r["Action_Label"] = str(_r.get("Action_Label") or "").strip()

        _is_new = str(_r.get("Is New") or "").strip().upper() == "YES"
        _has_human_label = bool(_r["Human_Label"])
        _had_prev_review = bool(_r.get("_has_prev_review"))
        _prev_ai = str(_r.get("_prev_ai_suggested_label") or "").strip()
        _ai_changed = bool(_prev_ai and _prev_ai != ai_label)

        if _had_prev_review and _has_human_label and _ai_changed:
            _r["Label_Status"] = "Updated"
        elif _had_prev_review and _has_human_label:
            _r["Label_Status"] = "Carried Forward"
        elif _is_new:
            _r["Label_Status"] = "New"
        else:
            _r["Label_Status"] = "Unlabeled"

    # ── Sort: Priority Score DESC, Relevance Score DESC, Deadline ASC ──────────
    def _deadline_iso(row: Dict) -> str:
        """Normalise deadline string to YYYY-MM-DD for ascending sort."""
        dl = (row.get("Deadline") or "").strip()
        if not dl or dl in ("N/A", "None", "TBD", ""):
            return "9999-99-99"
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y",
                    "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
                    "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M"):
            try:
                from datetime import datetime as _dt
                return _dt.strptime(dl[:len(fmt)], fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return dl   # fallback: lexicographic

    def _sort_key(row: Dict) -> tuple:
        try:
            priority  = int(row.get("Priority Score",  0) or 0)
        except (ValueError, TypeError):
            priority  = 0
        try:
            relevance = int(row.get("Relevance Score", 0) or 0)
        except (ValueError, TypeError):
            relevance = 0
        low_confidence = _to_int(row.get("_is_low_confidence", 0))
        # Primary: low-confidence last · then Priority Score ↓ · Relevance Score ↓ · Deadline ↑
        return (low_confidence, -priority, -relevance, _deadline_iso(row))

    unified_rows.sort(key=_sort_key)
    logger.info(f"[excel_exporter] Writing {len(unified_rows)} rows to unified Excel "
                f"(sorted: Priority Score ↓, Relevance Score ↓, Deadline ↑)…")

    # ── Build workbook ────────────────────────────────────────────────────────
    DARK_BLUE = PatternFill("solid", fgColor=_HDR_COLOR)
    ALT_FILL  = PatternFill("solid", fgColor=_ALT_COLOR)
    REL_FILL  = PatternFill("solid", fgColor=_REL_COLOR)
    NEW_FILL  = PatternFill("solid", fgColor=_NEW_COLOR)
    WHITE     = PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    CELL_FONT = Font(name="Calibri", size=10)
    REL_FONT  = Font(name="Calibri", size=10, color="375623", bold=True)
    NO_FONT   = Font(name="Calibri", size=10, color="999999")
    LINK_FONT = Font(name="Calibri", size=10, color="1155CC", underline="single")
    NEW_FONT  = Font(name="Calibri", size=10, color="0F4C81", bold=True)
    THIN      = Border(
        left   = Side(style="thin"), right  = Side(style="thin"),
        top    = Side(style="thin"), bottom = Side(style="thin"),
    )

    col_names      = [c[0] for c in UNIFIED_COLUMNS]
    rel_idx        = col_names.index("Relevance") + 1
    link_idx       = col_names.index("Tender URL") + 1
    new_idx        = col_names.index("Is New") + 1
    score_idx      = col_names.index("Relevance Score") + 1
    reason_idx     = col_names.index("Relevance Reason") + 1
    priority_idx   = col_names.index("Priority Score") + 1       # NEW
    insight_idx    = col_names.index("Opportunity Insight") + 1  # NEW
    decision_idx   = col_names.index("My Decision") + 1          # user input
    ai_label_idx   = col_names.index("AI_Suggested_Label") + 1
    human_label_idx = col_names.index("Human_Label") + 1
    label_reason_idx = col_names.index("Label_Reason") + 1
    training_idx   = col_names.index("Training_Approved") + 1
    action_idx     = col_names.index("Action_Label") + 1
    label_status_idx = col_names.index("Label_Status") + 1

    # Score band fills  (background / text colour)
    from openpyxl.styles import PatternFill as _PF
    _SCORE_FILLS = {
        "high":   (_PF("solid", fgColor="D1FAE5"), "065F46"),  # green
        "medium": (_PF("solid", fgColor="FEF3C7"), "92400E"),  # amber
        "low":    (_PF("solid", fgColor="FEE2E2"), "991B1B"),  # red-light
        "none":   (_PF("solid", fgColor="F3F4F6"), "9CA3AF"),  # gray
    }

    def _score_band(score):
        if score >= 75: return "high"
        if score >= 50: return "medium"
        if score >= 25: return "low"
        return "none"

    wb = Workbook()

    # User-input column header style (gold — "please fill me in")
    USER_HDR_FILL = PatternFill("solid", fgColor=_USER_COL_HEADER_COLOR)
    USER_HDR_FONT = Font(name="Calibri", bold=True,
                         color=_USER_COL_HEADER_TEXT, size=11)
    _user_input_cols = {decision_idx, human_label_idx, label_reason_idx, training_idx, action_idx}

    def _render_master_sheet(ws, rows_for_sheet: List[Dict]) -> None:
        # Header
        for ci, (col_name, col_width) in enumerate(UNIFIED_COLUMNS, 1):
            c = ws.cell(1, ci, col_name)
            if ci in _user_input_cols:
                c.fill = USER_HDR_FILL
                c.font = USER_HDR_FONT
            else:
                c.fill = DARK_BLUE
                c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN
            ws.column_dimensions[get_column_letter(ci)].width = col_width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(UNIFIED_COLUMNS))}1"

        # Data rows
        for ri, row_data in enumerate(rows_for_sheet, 2):
            ws.row_dimensions[ri].height = 45
            is_new_row = row_data.get("Is New") == "YES"
            base_fill = NEW_FILL if is_new_row else (ALT_FILL if ri % 2 == 0 else WHITE)

            for ci, col_name in enumerate(col_names, 1):
                val = row_data.get(col_name, "")
                cell = ws.cell(ri, ci, val)
                cell.border = THIN
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if ci == priority_idx:
                    try:
                        s_val = int(val or 0)
                    except (ValueError, TypeError):
                        s_val = 0
                    band = _score_band(s_val)
                    sfill, scolor = _SCORE_FILLS[band]
                    cell.value = s_val
                    cell.fill = sfill
                    cell.font = Font(name="Calibri", size=10, bold=(s_val >= 50), color=scolor)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif ci == score_idx:
                    try:
                        s_val = int(val or 0)
                    except (ValueError, TypeError):
                        s_val = 0
                    band = _score_band(s_val)
                    sfill, scolor = _SCORE_FILLS[band]
                    cell.value = s_val
                    cell.fill = sfill
                    cell.font = Font(name="Calibri", size=10, bold=(s_val >= 50), color=scolor)
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif ci == insight_idx:
                    cell.fill = base_fill
                    cell.font = Font(name="Calibri", size=9, color="374151", italic=True)
                elif ci == reason_idx:
                    cell.fill = base_fill
                    cell.font = Font(name="Calibri", size=9, color="4B5563", italic=True)
                elif ci == link_idx and str(val).startswith("http"):
                    cell.hyperlink = val
                    cell.font = LINK_FONT
                    cell.fill = base_fill
                elif ci == rel_idx:
                    if val:
                        cell.fill = REL_FILL
                        cell.font = REL_FONT
                    else:
                        cell.fill = base_fill
                        cell.font = NO_FONT
                elif ci == new_idx and val == "YES":
                    cell.fill = NEW_FILL
                    cell.font = NEW_FONT
                elif ci == ai_label_idx:
                    cell.fill = base_fill
                    if val == "Relevant":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276221")
                    elif val == "Borderline":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="7F6000")
                    else:
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif ci == label_status_idx:
                    cell.fill = base_fill
                    if val == "Updated":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="1D4ED8")
                    elif val == "Carried Forward":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276221")
                    elif val == "New":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="7F6000")
                    else:
                        cell.font = Font(name="Calibri", size=10, color="6B7280")
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                elif ci == decision_idx:
                    cell.fill = base_fill
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    if val == "Bid":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276221")
                    elif val == "No Bid":
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                    elif val == "Review Later":
                        cell.font = Font(name="Calibri", size=10, italic=True, color="7F6000")
                    else:
                        cell.font = Font(name="Calibri", size=10, color="9CA3AF")
                else:
                    cell.fill = base_fill
                    cell.font = CELL_FONT

        # Strict data validations
        if len(rows_for_sheet) >= 1:
            last_row = len(rows_for_sheet) + 1
            dv_specs = [
                (decision_idx, DECISION_LABELS, "Invalid My Decision", "Select one of: Bid, No Bid, Review Later."),
                (human_label_idx, HUMAN_LABELS, "Invalid Human_Label", "Select one of: Relevant, Borderline, Not Relevant."),
                (training_idx, TRAINING_APPROVED_LABELS, "Invalid Training_Approved", "Select Yes or No."),
                (action_idx, ACTION_LABELS, "Invalid Action_Label", "Select one of: Pursue, Review Later, Do Not Pursue."),
            ]
            for col_idx, labels, title, error in dv_specs:
                dv = DataValidation(
                    type="list",
                    formula1=f"\"{','.join(labels)}\"",
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle=title,
                    error=error,
                )
                ws.add_data_validation(dv)
                col_letter = get_column_letter(col_idx)
                dv.add(f"{col_letter}2:{col_letter}{last_row}")

    ws = wb.active
    ws.title = "All Tenders"
    _render_master_sheet(ws, unified_rows)

    ws_new = wb.create_sheet("New Tenders")
    new_rows = [r for r in unified_rows if str(r.get("Is New") or "").strip().upper() == "YES"]
    _render_master_sheet(ws_new, new_rows)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.title = "Run Summary"

    # Count by portal
    portal_counts: Dict[str, int] = {}
    new_counts:    Dict[str, int] = {}
    for row_data in unified_rows:
        p = row_data["Portal"]
        portal_counts[p] = portal_counts.get(p, 0) + 1
        if row_data["Is New"] == "YES":
            new_counts[p] = new_counts.get(p, 0) + 1

    summary_headers = [
        ("Portal", 20), ("Total Tenders", 15), ("New This Run", 14)
    ]
    for ci, (h, w) in enumerate(summary_headers, 1):
        c = ws_summary.cell(1, ci, h)
        c.fill = DARK_BLUE; c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws_summary.column_dimensions[get_column_letter(ci)].width = w

    for ri, (portal, total) in enumerate(
        sorted(portal_counts.items(), key=lambda x: -x[1]), 2
    ):
        ws_summary.cell(ri, 1, portal).font  = CELL_FONT
        ws_summary.cell(ri, 2, total).font   = CELL_FONT
        ws_summary.cell(ri, 3, new_counts.get(portal, 0)).font = CELL_FONT

    # Totals row
    tr = len(portal_counts) + 2
    total_new = sum(new_counts.values())
    ws_summary.cell(tr, 1, "TOTAL").font = Font(name="Calibri", bold=True, size=10)
    ws_summary.cell(tr, 2, len(unified_rows)).font = Font(name="Calibri", bold=True, size=10)
    ws_summary.cell(tr, 3, total_new).font = Font(name="Calibri", bold=True, size=10)

    # ── Save ──────────────────────────────────────────────────────────────────
    try:
        wb.save(UNIFIED_EXCEL_PATH)
        size_kb = os.path.getsize(UNIFIED_EXCEL_PATH) // 1024
        logger.info(
            f"[excel_exporter] ✓ Unified Excel saved: {UNIFIED_EXCEL_PATH} "
            f"({len(unified_rows)} rows, {size_kb} KB)"
        )
    except Exception as exc:
        logger.error(f"[excel_exporter] Failed to save unified Excel: {exc}")
        return ""

    # Also write filtered portal-wise excels from the same gated unified rows so
    # output/portal_excels reflects the same relevance quality threshold.
    written: Dict[str, str] = {}
    try:
        written = write_portal_excels(unified_rows)
        logger.info(
            "[excel_exporter] Filtered portal excels refreshed: %d file(s)",
            len(written),
        )
    except Exception as exc:
        logger.warning(
            "[excel_exporter] write_portal_excels failed (non-fatal): %s", exc
        )

    _cleanup_output_artifacts(UNIFIED_EXCEL_PATH, written)
    return UNIFIED_EXCEL_PATH


# =============================================================================
# Per-portal standardised Excel writer
# =============================================================================

# Map portal labels (lowercase, as they appear in the "Portal" column) to the
# canonical per-portal Excel filenames written to PORTAL_EXCELS_DIR.
# Both short keys ("gem") and full label strings ("gem bidplus") are listed so
# that any portal label variant resolves correctly.
_PORTAL_FILENAME_MAP: Dict[str, str] = {
    # World Bank
    "worldbank":              "WorldBank_Tenders_Master.xlsx",
    "world bank":             "WorldBank_Tenders_Master.xlsx",
    # GeM
    "gem":                    "GeM_BidPlus_Tenders_Master.xlsx",
    "gem bidplus":            "GeM_BidPlus_Tenders_Master.xlsx",
    # DevNet
    "devnet":                 "DevNet_India_Tenders_Master.xlsx",
    "devnet india":           "DevNet_India_Tenders_Master.xlsx",
    # CG eProcurement
    "cg":                     "CG_eProcurement_Tenders_Master.xlsx",
    "cg eprocurement":        "CG_eProcurement_Tenders_Master.xlsx",
    # GIZ
    "giz":                    "GIZ_India_Tenders_Master.xlsx",
    "giz india":              "GIZ_India_Tenders_Master.xlsx",
    # UNDP
    "undp":                   "UNDP_Procurement_Tenders_Master.xlsx",
    "undp procurement":       "UNDP_Procurement_Tenders_Master.xlsx",
    # Meghalaya
    "meghalaya":              "Meghalaya_MBDA_Tenders_Master.xlsx",
    "meghalaya mbda":         "Meghalaya_MBDA_Tenders_Master.xlsx",
    # NGO Box
    "ngobox":                 "NGO_Box_Tenders_Master.xlsx",
    "ngo box":                "NGO_Box_Tenders_Master.xlsx",
    # IUCN
    "iucn":                   "IUCN_Procurement_Tenders_Master.xlsx",
    "iucn procurement":       "IUCN_Procurement_Tenders_Master.xlsx",
    # Welthungerhilfe
    "whh":                    "Welthungerhilfe_Tenders_Master.xlsx",
    "welthungerhilfe":        "Welthungerhilfe_Tenders_Master.xlsx",
    # UNGM
    "ungm":                   "UNGM_Tenders_Master.xlsx",
    # SIDBI
    "sidbi":                  "SIDBI_Tenders_Master.xlsx",
    # NIC State Portals
    "nic":                    "NIC_States_Tenders_Master.xlsx",
    "nic state portals":      "NIC_States_Tenders_Master.xlsx",
    # DTVP Germany
    "dtvp":                   "DTVP_Germany_Tenders_Master.xlsx",
    "dtvp germany":           "DTVP_Germany_Tenders_Master.xlsx",
    # TED EU
    "ted":                    "TED_EU_Tenders_Master.xlsx",
    "ted eu":                 "TED_EU_Tenders_Master.xlsx",
    # TANEPS Tanzania
    "taneps":                 "TANEPS_Tanzania_Tenders_Master.xlsx",
    "taneps tanzania":        "TANEPS_Tanzania_Tenders_Master.xlsx",
    # AfDB
    "afdb":                   "AfDB_Consultants_Tenders_Master.xlsx",
    "afdb consultants":       "AfDB_Consultants_Tenders_Master.xlsx",
    "african development bank": "AfDB_Consultants_Tenders_Master.xlsx",
    # SAM.gov
    "sam":                    "SAM_Tenders_Master.xlsx",
    "sam opportunities":      "SAM_Tenders_Master.xlsx",
    # Karnataka
    "karnataka":              "Karnataka_eProcure_Tenders_Master.xlsx",
    "karnataka eprocure":     "Karnataka_eProcure_Tenders_Master.xlsx",
    # USAID
    "usaid":                  "USAID_Tenders_Master.xlsx",
    # ICFRE
    "icfre":                  "ICFRE_Tenders_Master.xlsx",
    "icfre tenders":          "ICFRE_Tenders_Master.xlsx",
    # JTDS Jharkhand
    "jtds":                   "JTDS_Jharkhand_Tenders_Master.xlsx",
    "jtds jharkhand":         "JTDS_Jharkhand_Tenders_Master.xlsx",
    # PHFI
    "phfi":                   "PHFI_Tenders_Master.xlsx",
    # AFD France
    "afd":                    "AFD_France_Tenders_Master.xlsx",
    "afd france":             "AFD_France_Tenders_Master.xlsx",
    # European Commission
    "ec":                     "European_Commission_EC_Tenders_Master.xlsx",
    "european commission (ec)": "European_Commission_EC_Tenders_Master.xlsx",
    # ILO Procurement
    "ilo":                    "ILO_Procurement_Tenders_Master.xlsx",
    "ilo procurement":        "ILO_Procurement_Tenders_Master.xlsx",
    # High-volume state portals
    "maharashtra tenders":    "Maharashtra_Tenders_Master.xlsx",
    "up etenders":            "UP_eTenders_Master.xlsx",
}


def _portal_filename(portal_label: str) -> str:
    """Derive an output filename from a portal label."""
    lc = portal_label.lower().strip()
    if lc in _PORTAL_FILENAME_MAP:
        return _PORTAL_FILENAME_MAP[lc]
    # Fallback: sanitise label into a filename
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in portal_label)
    return f"{safe}_Tenders_Master.xlsx"


def write_portal_excels(
    unified_rows: List[Dict],
    output_dir:   str = None,
) -> Dict[str, str]:
    """
    Write one standardised Excel file per portal using the 9-column PORTAL_COLUMNS
    schema.  These overwrite the legacy per-portal files with a uniform layout.

    Args:
        unified_rows: rows already built by write_unified_excel() (or any list of
                      dicts with at least "Portal" key).
        output_dir:   directory for output files; defaults to config.OUTPUT_DIR.

    Returns:
        Dict mapping portal_label → written file path.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from config.config import PORTAL_EXCELS_DIR as _PORTAL_DIR

    if output_dir is None:
        output_dir = _PORTAL_DIR
    os.makedirs(output_dir, exist_ok=True)

    if not unified_rows:
        logger.warning("[excel_exporter] write_portal_excels: no rows to process")
        return {}

    # ── Styles (shared across all portal files) ───────────────────────────────
    DARK_BLUE = PatternFill("solid", fgColor=_HDR_COLOR)
    ALT_FILL  = PatternFill("solid", fgColor=_ALT_COLOR)
    WHITE     = PatternFill("solid", fgColor="FFFFFF")
    REL_FILL  = PatternFill("solid", fgColor=_REL_COLOR)
    NEW_FILL  = PatternFill("solid", fgColor=_NEW_COLOR)
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    CELL_FONT = Font(name="Calibri", size=10)
    REL_FONT  = Font(name="Calibri", size=10, color="375623", bold=True)
    LINK_FONT = Font(name="Calibri", size=10, color="1155CC", underline="single")
    NEW_FONT  = Font(name="Calibri", size=10, color="0F4C81", bold=True)
    THIN      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    col_names  = [c[0] for c in PORTAL_COLUMNS]
    url_idx    = col_names.index("Tender URL") + 1
    rel_idx    = col_names.index("Sector") + 1     # highlight sector column
    scraped_idx = col_names.index("Scraped Date") + 1

    # ── Group rows by portal ──────────────────────────────────────────────────
    by_portal: Dict[str, List[Dict]] = {}
    for row in unified_rows:
        p = row.get("Portal", "unknown")
        by_portal.setdefault(p, []).append(row)

    written: Dict[str, str] = {}

    def _dl_key(row: Dict) -> str:
        """Deadline sort key for portal files (YYYY-MM-DD or original string)."""
        dl = (row.get("Deadline") or "").strip()
        if not dl or dl in ("N/A", "None", "TBD"):
            return "9999-99-99"
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y",
                    "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
                    "%d-%m-%Y %H:%M", "%d/%m/%Y %H:%M"):
            try:
                from datetime import datetime as _dt
                return _dt.strptime(dl[:len(fmt)], fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
        return dl

    for portal_label, rows in by_portal.items():
        rows = sorted(rows, key=_dl_key)   # sort by deadline ascending
        filename = _portal_filename(portal_label)
        out_path = os.path.join(output_dir, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = portal_label[:31]   # Excel sheet names max 31 chars

        # Header row
        for ci, (col_name, col_width) in enumerate(PORTAL_COLUMNS, 1):
            c = ws.cell(1, ci, col_name)
            c.fill      = DARK_BLUE
            c.font      = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border    = THIN
            ws.column_dimensions[get_column_letter(ci)].width = col_width
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(PORTAL_COLUMNS))}1"

        # Data rows
        for ri, row_data in enumerate(rows, 2):
            ws.row_dimensions[ri].height = 40
            is_new_row = row_data.get("Is New") == "YES"
            base_fill  = NEW_FILL if is_new_row else (ALT_FILL if ri % 2 == 0
                                                       else WHITE)

            for ci, col_name in enumerate(col_names, 1):
                val  = row_data.get(col_name, "")
                cell = ws.cell(ri, ci, val)
                cell.border    = THIN
                cell.alignment = Alignment(wrap_text=True, vertical="top")

                if ci == url_idx and str(val).startswith("http"):
                    cell.hyperlink = val
                    cell.font      = LINK_FONT
                    cell.fill      = base_fill
                elif col_name == "Sector" and val:
                    cell.fill = REL_FILL
                    cell.font = REL_FONT
                elif col_name == "Is New" and val == "YES":
                    cell.fill = NEW_FILL
                    cell.font = NEW_FONT
                else:
                    cell.fill = base_fill
                    cell.font = CELL_FONT

        try:
            wb.save(out_path)
            size_kb = os.path.getsize(out_path) // 1024
            logger.info(
                f"[excel_exporter] ✓ Portal Excel: {filename} "
                f"({len(rows)} rows, {size_kb} KB)"
            )
            written[portal_label] = out_path
        except Exception as exc:
            logger.error(f"[excel_exporter] Failed to save {filename}: {exc}")

    logger.info(f"[excel_exporter] Per-portal Excels written: {len(written)} files")
    return written


def _cleanup_output_artifacts(unified_excel_path: str, written_portals: Dict[str, str]) -> None:
    """
    Keep output folders clean:
      - One canonical master in output/
      - Only freshly regenerated portal excels in output/portal_excels
      - No stale preview/temp/cache artifacts
    """
    try:
        from config.config import OUTPUT_DIR, PORTAL_EXCELS_DIR
    except Exception as exc:
        logger.warning("[excel_exporter] Output cleanup skipped (config import): %s", exc)
        return

    keep_root = {os.path.abspath(unified_excel_path)}
    removed_root = 0

    # Explicit stale artifact from earlier manual flows.
    preview_path = os.path.join(OUTPUT_DIR, "Tender_Monitor_Master_filtered_preview.xlsx")
    if os.path.exists(preview_path):
        try:
            os.remove(preview_path)
            removed_root += 1
        except Exception as exc:
            logger.warning("[excel_exporter] Could not remove preview file %s: %s", preview_path, exc)

    # Keep only canonical master .xlsx in output root.
    for name in os.listdir(OUTPUT_DIR):
        path = os.path.join(OUTPUT_DIR, name)
        if not os.path.isfile(path):
            continue
        lower = name.lower()
        if lower.startswith("~$") and lower.endswith(".xlsx"):
            try:
                os.remove(path)
                removed_root += 1
            except Exception as exc:
                logger.warning("[excel_exporter] Could not remove temp file %s: %s", path, exc)
            continue
        if lower.endswith(".xlsx") and os.path.abspath(path) not in keep_root:
            try:
                os.remove(path)
                removed_root += 1
            except Exception as exc:
                logger.warning("[excel_exporter] Could not remove non-canonical root excel %s: %s", path, exc)

    removed_portal = 0
    if written_portals:
        keep_portal = {os.path.abspath(p) for p in written_portals.values()}
        for name in os.listdir(PORTAL_EXCELS_DIR):
            path = os.path.join(PORTAL_EXCELS_DIR, name)
            if not os.path.isfile(path):
                continue
            lower = name.lower()
            remove = False
            if lower.endswith(".xlsx") and os.path.abspath(path) not in keep_portal:
                remove = True
            if lower.startswith("~$") and lower.endswith(".xlsx"):
                remove = True
            if lower.endswith(".json") or lower == ".ds_store":
                remove = True
            if remove:
                try:
                    os.remove(path)
                    removed_portal += 1
                except Exception as exc:
                    logger.warning("[excel_exporter] Could not remove stale portal artifact %s: %s", path, exc)
    else:
        logger.warning(
            "[excel_exporter] Skipping portal artifact cleanup (no regenerated portal files)"
        )

    if removed_root or removed_portal:
        logger.info(
            "[excel_exporter] Output cleanup removed root=%d, portal=%d artifacts",
            removed_root,
            removed_portal,
        )
