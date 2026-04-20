"""
Manual daily email digest for TenderRadar.

Reads the canonical master workbook plus same-day evidence packages and sends
one concise daily update email with:
  - summary counts for new tenders
  - top tenders in the body
  - master workbook attached
  - optional ZIP of top relevant evidence packages

This module is intentionally separate from the main 6-hour scraper flow so the
digest can be triggered manually or safely auto-invoked from main.py.
"""

from __future__ import annotations

import json
import logging
import os
import smtplib
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from config.config import (
    OUTPUT_DIR,
    SMTP_FROM_EMAIL,
    SMTP_FROM_NAME,
    UNIFIED_EXCEL_PATH,
    email_configured,
    get_email_config_errors,
    NOTIFY_EMAIL_TO,
    SMTP_HOST,
    SMTP_PASSWORD,
    SMTP_PORT,
    SMTP_USE_TLS,
    SMTP_USERNAME,
)

logger = logging.getLogger("tenderradar.daily_digest")

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_PACKAGES_ROOT = _PROJECT_ROOT / "output" / "tender_packages"
_NOTIFICATIONS_DIR = _PROJECT_ROOT / "output" / "notifications"
_STATE_PATH = _NOTIFICATIONS_DIR / "daily_email_state.json"
_MAX_TOP_TENDERS = 10
_DEFAULT_TOP_TENDERS = 5
_DEFAULT_MAX_PACKAGES = 5


@dataclass
class DigestRow:
    title: str
    organization: str
    country: str
    is_new: str
    ai_label: str
    human_label: str
    relevance_reason: str
    priority_score: float
    relevance_score: float
    tender_url: str
    deadline: str
    evidence_state: str = ""
    opportunity_maturity: str = ""
    recommended_action: str = ""


_EVIDENCE_STATES = ("SIGNAL_ONLY", "PAGE_ONLY", "PARTIAL_PACKAGE", "FULL_PACKAGE")


def _smtp_enabled() -> bool:
    return email_configured()


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _clean_text(value: Any) -> str:
    text = str(value or "").strip()
    return " ".join(text.split())


def _display_text(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    if "â" in text or "\x80" in text or "\x93" in text or "\x94" in text:
        try:
            text = text.encode("latin1").decode("utf-8")
        except Exception:
            pass
    return " ".join(text.split())


def _reason_short(reason: str, limit: int = 160) -> str:
    reason = _display_text(reason)
    if not reason:
        return "Aligned with IDCG's consulting focus based on the master workbook assessment."
    if len(reason) <= limit:
        return reason
    return reason[: limit - 3].rstrip() + "..."


def _business_reason(reason: str) -> str:
    reason = _reason_short(reason, limit=135)
    replacements = {
        "Medium relevance": "Moderate IDCG fit",
        "High relevance": "Strong IDCG fit",
        "Low relevance": "Lower-confidence fit",
        "CAP-STAT": "CAPSTAT",
        "firm-fit": "firm fit",
    }
    for old, new in replacements.items():
        reason = reason.replace(old, new)
    return reason


def _display_label(row: DigestRow) -> str:
    return row.human_label or row.ai_label or "Unlabeled"


def _format_score(score: float) -> str:
    if score <= 0:
        return "-"
    return str(int(round(score)))


def _summarize_evidence_mix(rows: Sequence[DigestRow]) -> Dict[str, int]:
    counts = {state: 0 for state in _EVIDENCE_STATES}
    for row in rows:
        if row.evidence_state in counts:
            counts[row.evidence_state] += 1
    return counts


def _maturity_display(row: DigestRow) -> str:
    """Return the business-facing maturity label for display, falling back gracefully."""
    if row.opportunity_maturity:
        return row.opportunity_maturity
    # Derive from evidence state if maturity column is absent (older workbooks)
    _fallback = {
        "SIGNAL_ONLY": "Signal First",
        "PAGE_ONLY": "Partial Package",
        "PARTIAL_PACKAGE": "Partial Package",
        "FULL_PACKAGE": "Full Package",
    }
    return _fallback.get(row.evidence_state, "Signal First")


def _best_opportunity_line(top_relevant: Optional[DigestRow]) -> str:
    if not top_relevant:
        return ""
    maturity = _maturity_display(top_relevant)
    action = top_relevant.recommended_action or "Review"
    return (
        f"Best opportunity today: {top_relevant.title} — "
        f"{_business_reason(top_relevant.relevance_reason)} "
        f"[{maturity} | {action}]"
    )


def _build_caveats(rows: Sequence[DigestRow], evidence_mix: Dict[str, int]) -> List[str]:
    notes: List[str] = []
    signal_only = evidence_mix.get("SIGNAL_ONLY", 0)
    full_package = evidence_mix.get("FULL_PACKAGE", 0)
    borderline = sum(1 for row in rows if _display_label(row).lower() == "borderline")
    if signal_only:
        plural = "opportunities are" if signal_only != 1 else "opportunity is"
        notes.append(f"{signal_only} {plural} signal-only and may publish TOR or full documents later.")
    if full_package == 0:
        notes.append("No full-package tenders were available today.")
    if borderline:
        plural = "opportunities remain" if borderline != 1 else "opportunity remains"
        notes.append(f"{borderline} borderline {plural} under review due to fit or delivery-feasibility uncertainty.")
    if not notes:
        notes.append("No material caveats for today's shortlisted opportunities.")
    return notes[:3]


def _load_digest_rows(workbook_path: Path) -> List[DigestRow]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to read Tender_Monitor_Master.xlsx")

    if not workbook_path.exists():
        raise FileNotFoundError(f"Master workbook not found: {workbook_path}")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb["All Tenders"] if "All Tenders" in wb.sheetnames else wb[wb.sheetnames[0]]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {str(h).strip(): i for i, h in enumerate(headers) if h}

        def _row_get(row_vals: Sequence[Any], key: str, default: Any = "") -> Any:
            idx = col.get(key)
            return default if idx is None else row_vals[idx]

        rows: List[DigestRow] = []
        for row_vals in ws.iter_rows(min_row=2, values_only=True):
            is_new = _clean_text(_row_get(row_vals, "Is New")).upper()
            if is_new != "YES":
                continue
            rows.append(
                DigestRow(
                    title=_display_text(_row_get(row_vals, "Title")),
                    organization=_display_text(_row_get(row_vals, "Organization")),
                    country=_display_text(_row_get(row_vals, "Country")),
                    is_new=is_new,
                    ai_label=_display_text(_row_get(row_vals, "AI_Suggested_Label")),
                    human_label=_display_text(_row_get(row_vals, "Human_Label")),
                    relevance_reason=_display_text(_row_get(row_vals, "Relevance Reason")),
                    priority_score=_to_float(_row_get(row_vals, "Priority Score")),
                    relevance_score=_to_float(_row_get(row_vals, "Relevance Score")),
                    tender_url=_display_text(_row_get(row_vals, "Tender URL")),
                    deadline=_display_text(_row_get(row_vals, "Deadline")),
                    evidence_state=_display_text(_row_get(row_vals, "Evidence_State")),
                    opportunity_maturity=_display_text(_row_get(row_vals, "Opportunity_Maturity")),
                    recommended_action=_display_text(_row_get(row_vals, "Recommended_Action")),
                )
            )
        return rows
    finally:
        wb.close()


def _sort_key(row: DigestRow) -> tuple:
    return (
        row.priority_score,
        row.relevance_score,
        1 if row.ai_label == "Relevant" else 0,
        row.title.lower(),
    )


def _build_email_body(
    report_date: str,
    total_new: int,
    relevant_count: int,
    borderline_count: int,
    top_rows: Sequence[DigestRow],
    package_count: int,
    evidence_mix: Dict[str, int],
    best_opportunity: str,
    caveats: Sequence[str],
) -> str:
    lines = [
        f"TenderRadar Daily Update — {report_date}",
        "",
        "Summary",
        f"- New tenders: {total_new}",
        f"- Relevant: {relevant_count}",
        f"- Borderline: {borderline_count}",
        f"- Packages attached: {package_count}",
        (
            "- Evidence mix: "
            f"SIGNAL_ONLY {evidence_mix['SIGNAL_ONLY']} | "
            f"PAGE_ONLY {evidence_mix['PAGE_ONLY']} | "
            f"PARTIAL_PACKAGE {evidence_mix['PARTIAL_PACKAGE']} | "
            f"FULL_PACKAGE {evidence_mix['FULL_PACKAGE']}"
        ),
        "",
    ]

    if best_opportunity:
        lines.extend([best_opportunity, ""])

    lines.append("Top opportunities")

    if not top_rows:
        lines.append("- No new tenders met the shortlist criteria in the master workbook.")
    else:
        _MATURITY_DIGEST_NOTES = {
            "Signal First":    "Monitor for TOR publication",
            "Partial Package": "Review content — docs may be pending",
            "Full Package":    "Bid-ready package — begin preparation",
        }
        for idx, row in enumerate(top_rows, start=1):
            org = row.organization or "Unknown organization"
            country = row.country or "Country not listed"
            label = _display_label(row)
            maturity = _maturity_display(row)
            action = row.recommended_action or "Review"
            maturity_note = _MATURITY_DIGEST_NOTES.get(maturity, "")
            why = _business_reason(row.relevance_reason)
            lines.append(f"{idx}. {row.title}")
            lines.append(
                f"   {org} | {country} | {label} | Priority {_format_score(row.priority_score)}"
            )
            lines.append(f"   Maturity: {maturity} — {maturity_note} | Action: {action}")
            lines.append(f"   Why relevant: {why}")

    lines.extend(["", "Notes"])
    for note in caveats:
        lines.append(f"- {note}")

    lines.extend(
        [
            "",
            "Attachments",
            "- Attached: Tender_Monitor_Master.xlsx",
            (
                f"- Attached: ZIP of top relevant packages ({package_count})"
                if package_count > 0
                else "- Attached: No package ZIP today"
            ),
        ]
    )
    return "\n".join(lines)


def _load_state() -> Dict[str, Any]:
    if not _STATE_PATH.exists():
        return {"sent_reports": []}
    try:
        return json.loads(_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        logger.warning("[daily_digest] Could not parse state file; starting fresh")
        return {"sent_reports": []}


def _save_state(state: Dict[str, Any]) -> None:
    _NOTIFICATIONS_DIR.mkdir(parents=True, exist_ok=True)
    _STATE_PATH.write_text(json.dumps(state, indent=2, ensure_ascii=True), encoding="utf-8")


def _recipients_key(recipients: Sequence[str]) -> str:
    return ",".join(sorted({r.strip().lower() for r in recipients if r.strip()}))


def _already_sent(report_date: str, recipients: Sequence[str]) -> bool:
    state = _load_state()
    key = _recipients_key(recipients)
    for item in state.get("sent_reports", []):
        if item.get("report_date") == report_date and item.get("recipients") == key:
            return True
    return False


def _mark_sent(report_date: str, recipients: Sequence[str], subject: str, attachments: Sequence[str]) -> None:
    state = _load_state()
    state.setdefault("sent_reports", []).append(
        {
            "report_date": report_date,
            "recipients": _recipients_key(recipients),
            "subject": subject,
            "attachments": [os.path.basename(p) for p in attachments],
            "sent_at": datetime.now().isoformat(timespec="seconds"),
        }
    )
    _save_state(state)


def _index_package_dirs(run_dir: Path) -> Dict[str, Path]:
    indexed: Dict[str, Path] = {}
    if not run_dir.exists():
        return indexed

    for child in sorted(run_dir.iterdir()):
        if not child.is_dir():
            continue
        meta_path = child / "metadata.json"
        if not meta_path.exists():
            continue
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception as exc:
            logger.warning("[daily_digest] Failed to read package metadata %s: %s", meta_path, exc)
            continue

        for key in (
            _clean_text(meta.get("tender_url")),
            _clean_text(meta.get("title")).lower(),
        ):
            if key:
                indexed[key] = child
    return indexed


def _load_package_metadata(run_dir: Path) -> Dict[str, Dict[str, Any]]:
    indexed: Dict[str, Dict[str, Any]] = {}
    if not run_dir.exists():
        return indexed

    for child in sorted(run_dir.iterdir()):
        if not child.is_dir():
            continue
        meta_path = child / "metadata.json"
        if not meta_path.exists():
            continue
        try:
            meta = json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception as exc:
            logger.warning("[daily_digest] Failed to read package metadata %s: %s", meta_path, exc)
            continue
        meta["_package_dir"] = str(child)
        for key in (
            _display_text(meta.get("tender_url")),
            _display_text(meta.get("title")).lower(),
        ):
            if key:
                indexed[key] = meta
    return indexed


def _build_packages_zip(report_date: str, package_dirs: Iterable[Path]) -> Optional[Path]:
    package_dirs = [p for p in package_dirs if p.exists() and p.is_dir()]
    if not package_dirs:
        return None

    _NOTIFICATIONS_DIR.mkdir(parents=True, exist_ok=True)
    zip_path = _NOTIFICATIONS_DIR / f"tenderradar_top_packages_{report_date}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for pkg_dir in package_dirs:
            for path in sorted(pkg_dir.rglob("*")):
                if path.is_file():
                    arcname = str(Path(pkg_dir.name) / path.relative_to(pkg_dir))
                    zf.write(path, arcname=arcname)
    return zip_path


def _send_email(subject: str, body: str, attachment_paths: Sequence[Path]) -> bool:
    if not _smtp_enabled():
        logger.warning(
            "[daily_digest] Email skipped; missing config: %s",
            ", ".join(get_email_config_errors()) or "unknown",
        )
        return False

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = formataddr((SMTP_FROM_NAME, SMTP_FROM_EMAIL))
    msg["To"] = ", ".join(NOTIFY_EMAIL_TO)
    msg.set_content(body)

    for attachment_path in attachment_paths:
        if not attachment_path or not attachment_path.exists():
            continue
        with open(attachment_path, "rb") as fh:
            data = fh.read()
        subtype = "zip" if attachment_path.suffix.lower() == ".zip" else "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        msg.add_attachment(
            data,
            maintype="application",
            subtype=subtype,
            filename=attachment_path.name,
        )

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=40) as server:
            if SMTP_USE_TLS:
                server.starttls()
            if SMTP_USERNAME:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)
        logger.info("[daily_digest] Email sent to %s", ", ".join(NOTIFY_EMAIL_TO))
        return True
    except Exception as exc:
        logger.warning("[daily_digest] send failed: %s", exc)
        return False


def send_daily_digest(
    workbook_path: Optional[Path] = None,
    report_date: Optional[str] = None,
    top_tenders_limit: int = _DEFAULT_TOP_TENDERS,
    max_package_count: int = _DEFAULT_MAX_PACKAGES,
    include_packages: bool = True,
    dry_run: bool = False,
    force: bool = False,
) -> Dict[str, Any]:
    """
    Build and optionally send the manual daily TenderRadar digest.
    """
    workbook = Path(workbook_path or UNIFIED_EXCEL_PATH)
    report_date = report_date or date.today().isoformat()
    top_tenders_limit = max(1, min(int(top_tenders_limit), _MAX_TOP_TENDERS))
    max_package_count = max(0, min(int(max_package_count), _MAX_TOP_TENDERS))

    rows = _load_digest_rows(workbook)
    sorted_rows = sorted(rows, key=_sort_key, reverse=True)

    relevant_rows = [r for r in sorted_rows if r.ai_label == "Relevant"]
    borderline_rows = [r for r in sorted_rows if r.ai_label == "Borderline"]
    shortlist_rows = [r for r in sorted_rows if r.ai_label in {"Relevant", "Borderline"}]

    package_run_dir = _PACKAGES_ROOT / report_date
    package_index = _index_package_dirs(package_run_dir)
    package_meta_index = _load_package_metadata(package_run_dir)

    for row in shortlist_rows:
        # Workbook columns are authoritative (set by excel_exporter); only fall back
        # to package metadata if the workbook columns are absent (older workbooks).
        if not row.evidence_state:
            meta = package_meta_index.get(row.tender_url) or package_meta_index.get(row.title.lower())
            if meta:
                row.evidence_state = _display_text(meta.get("evidence_state")) or ""

    top_rows = shortlist_rows[: max(5, min(top_tenders_limit, 7))] if shortlist_rows else sorted_rows[:top_tenders_limit]

    if not sorted_rows:
        logger.info("[daily_digest] No new tenders in workbook for %s; skipping digest", report_date)
        return {
            "ok": True,
            "duplicate_blocked": False,
            "skipped_no_new": True,
            "dry_run": dry_run,
            "report_date": report_date,
            "subject": f"TenderRadar Daily Update — {report_date} | 0 New | 0 Relevant",
            "body": "",
            "total_new": 0,
            "relevant_count": 0,
            "borderline_count": 0,
            "top_tenders": [],
            "attachments": [str(workbook)] if workbook.exists() else [],
            "package_dirs": [],
            "errors": [],
        }

    package_dirs: List[Path] = []
    packages_zip: Optional[Path] = None
    if include_packages and max_package_count > 0:
        for row in relevant_rows:
            if len(package_dirs) >= max_package_count:
                break
            candidate = package_index.get(row.tender_url) or package_index.get(row.title.lower())
            if candidate and candidate not in package_dirs:
                package_dirs.append(candidate)
        packages_zip = _build_packages_zip(report_date, package_dirs)

    evidence_mix = _summarize_evidence_mix(shortlist_rows)
    best_opportunity = _best_opportunity_line(relevant_rows[0] if relevant_rows else None)
    caveats = _build_caveats(shortlist_rows, evidence_mix)

    subject = (
        f"TenderRadar Daily Update — {report_date} | "
        f"{len(sorted_rows)} New | {len(relevant_rows)} Relevant"
    )
    body = _build_email_body(
        report_date=report_date,
        total_new=len(sorted_rows),
        relevant_count=len(relevant_rows),
        borderline_count=len(borderline_rows),
        top_rows=top_rows,
        package_count=len(package_dirs),
        evidence_mix=evidence_mix,
        best_opportunity=best_opportunity,
        caveats=caveats,
    )

    attachments: List[Path] = [workbook]
    if packages_zip:
        attachments.append(packages_zip)

    if not dry_run and not force and _already_sent(report_date, NOTIFY_EMAIL_TO):
        logger.warning("[daily_digest] Duplicate daily digest blocked for %s", report_date)
        return {
            "ok": False,
            "duplicate_blocked": True,
            "dry_run": False,
            "report_date": report_date,
            "subject": subject,
            "body": body,
            "total_new": len(sorted_rows),
            "relevant_count": len(relevant_rows),
            "borderline_count": len(borderline_rows),
            "attachments": [str(p) for p in attachments if p.exists()],
            "package_dirs": [str(p) for p in package_dirs],
        }

    ok = True
    if dry_run:
        logger.info("[daily_digest] DRY-RUN — no email sent")
    else:
        ok = _send_email(subject=subject, body=body, attachment_paths=attachments)
        if ok:
            _mark_sent(report_date, NOTIFY_EMAIL_TO, subject, [str(p) for p in attachments if p.exists()])

    result = {
        "ok": ok,
        "duplicate_blocked": False,
        "skipped_no_new": False,
        "dry_run": dry_run,
        "report_date": report_date,
        "subject": subject,
        "body": body,
        "total_new": len(sorted_rows),
        "relevant_count": len(relevant_rows),
        "borderline_count": len(borderline_rows),
        "top_tenders": [
            {
                "title": row.title,
                "organization": row.organization,
                "country": row.country,
                "label": _display_label(row),
                "priority_score": _format_score(row.priority_score),
                "evidence_state": row.evidence_state or "",
                "opportunity_maturity": _maturity_display(row),
                "recommended_action": row.recommended_action or "Review",
                "why_relevant": _business_reason(row.relevance_reason),
            }
            for row in top_rows
        ],
        "attachments": [str(p) for p in attachments if p.exists()],
        "package_dirs": [str(p) for p in package_dirs],
        "errors": [],
    }
    logger.info(
        "[daily_digest] Prepared digest | new=%d relevant=%d borderline=%d packages=%d dry_run=%s ok=%s",
        result["total_new"],
        result["relevant_count"],
        result["borderline_count"],
        len(package_dirs),
        dry_run,
        ok,
    )
    return result
