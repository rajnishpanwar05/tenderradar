# =============================================================================
# usaid_pipeline.py — USAID WorkWithUSAID Sub-Opportunities
#
# Source  : https://www.workwithusaid.org/sub-opportunities
# Method  : Selenium headless Chrome (React SPA — content loads via JS)
#
# What this captures:
#   Sub-contracts and sub-grants posted by USAID implementing partners.
#   These are RFPs/RFAs issued by prime contractors looking for sub-contractors
#   in program countries including India, South Asia, and Africa.
#   Highly relevant for IDCG as a sub-contractor / technical partner.
#
# No login required. No CAPTCHA.
# =============================================================================

import os, re, time
import requests

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import sys
sys.path.insert(0, os.path.expanduser("~/tender_system"))

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

# ── Registry metadata (override static defaults) ─────────────────────────────
SCRAPER_META = {
    "flag":        "usaid",
    "label":       "USAID",
    "group":       "selenium",
    "timeout":     240,
    "max_retries": 0,
    "auto":        True,   # primary path is SAM.gov API (reliable); partner2peer as fallback
}

# ── Constants ─────────────────────────────────────────────────────────────────
USAID_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "USAID_SubOpp_Master.xlsx")

# workwithusaid.org is frequently down. Updated priority order (2026-04):
#   1. SAM.gov API (most reliable — same USAID opportunities via federal API)
#   2. partner2peer.usaid.gov (USAID's own sub-award portal)
#   3. workwithusaid.org (original URL, often unreachable)
PORTAL_URLS = [
    "https://partner2peer.usaid.gov/opportunities",
    "https://www.workwithusaid.org/sub-opportunities",
    "https://www.workwithusaid.org/",
]
PORTAL_URL   = PORTAL_URLS[0]   # default; overridden in run() to first live URL

# SAM.gov API for USAID opportunities (free public API — same data, more reliable)
SAM_API_URL = "https://api.sam.gov/opportunities/v2/search"
SAM_USAID_QUERIES = [
    "USAID India consulting",
    "USAID South Asia evaluation",
    "USAID Africa monitoring evaluation",
    "USAID capacity building advisory",
]
PAGE_WAIT    = 20
SCROLL_PAUSE = 2.0
MAX_SCROLLS  = 15   # infinite-scroll guard

# ── Excel styles ───────────────────────────────────────────────────────────────
DARK_BLUE  = PatternFill("solid", fgColor="1F3864")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
WHITE      = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT  = Font(name="Calibri", size=10)
THIN       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Title",          60),
    ("Organization",   35),
    ("Country",        20),
    ("Type",           18),
    ("Posted",         14),
    ("Deadline",       16),
    ("Link",           50),
    ("Relevance",      30),
]


def _init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "USAID Sub-Opportunities"
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        c = ws.cell(1, col_idx, col_name)
        c.fill      = DARK_BLUE
        c.font      = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = col_width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLUMNS)).column_letter}1"
    return wb, ws


def _save_excel(rows: list):
    if os.path.exists(USAID_EXCEL_PATH):
        wb = load_workbook(USAID_EXCEL_PATH)
        ws = wb.active
    else:
        wb, ws = _init_excel()

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.add(str(row[0]))

    for r in rows:
        key = r.get("title", "")[:80]
        if key in existing:
            continue
        row_idx = ws.max_row + 1
        fill    = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
        ws.append([
            r["title"], r["org"], r["country"], r["opp_type"],
            r["posted"], r["deadline"], r["link"], r["relevance"],
        ])
        for col_i in range(1, len(COLUMNS) + 1):
            c = ws.cell(row_idx, col_i)
            c.font      = CELL_FONT
            c.fill      = fill
            c.border    = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        existing.add(key)

    wb.save(USAID_EXCEL_PATH)


def _build_driver():
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0.0.0 Safari/537.36")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


def _parse_cards(soup: BeautifulSoup) -> list:
    """
    Parse opportunity cards/rows from the rendered React page.
    Tries multiple selector strategies to handle different layouts.
    """
    results = []

    # Strategy 1: article / card elements
    cards = (soup.find_all("article") or
             soup.find_all("div", class_=re.compile(r"card|opportunity|listing|result", re.I)))

    for card in cards:
        title_el = (card.find(["h2", "h3", "h4"]) or
                    card.find("a") or
                    card.find(class_=re.compile(r"title|name", re.I)))
        if not title_el:
            continue
        title = title_el.get_text(strip=True)
        if not title or len(title) < 5:
            continue

        link = ""
        a_tag = card.find("a", href=True)
        if a_tag:
            href = a_tag["href"]
            link = href if href.startswith("http") else f"https://www.workwithusaid.org{href}"

        # Extract metadata from card text
        card_text  = card.get_text(" | ", strip=True)
        org        = ""
        country    = ""
        opp_type   = ""
        posted     = ""
        deadline   = ""

        # Look for labelled spans/divs
        for label_el in card.find_all(class_=re.compile(r"label|meta|info|detail", re.I)):
            label_text = label_el.get_text(strip=True).lower()
            sibling    = label_el.find_next_sibling()
            value      = sibling.get_text(strip=True) if sibling else ""
            if "organ" in label_text:  org      = value
            if "country" in label_text: country = value
            if "type" in label_text:   opp_type = value
            if "posted" in label_text: posted   = value
            if "dead" in label_text or "close" in label_text: deadline = value

        results.append({
            "title"   : title,
            "org"     : org,
            "country" : country,
            "opp_type": opp_type,
            "posted"  : posted,
            "deadline": deadline,
            "link"    : link,
        })

    # Strategy 2: table rows (if the page uses a table layout)
    if not results:
        for table in soup.find_all("table"):
            hdrs_text = table.get_text()
            if not any(k in hdrs_text for k in ("Title", "Organization", "Country", "Type")):
                continue
            for tr in table.find_all("tr")[1:]:
                tds = tr.find_all("td")
                if len(tds) < 2:
                    continue
                title = tds[0].get_text(strip=True)
                if not title:
                    continue
                org      = tds[1].get_text(strip=True) if len(tds) > 1 else ""
                country  = tds[2].get_text(strip=True) if len(tds) > 2 else ""
                opp_type = tds[3].get_text(strip=True) if len(tds) > 3 else ""
                deadline = tds[4].get_text(strip=True) if len(tds) > 4 else ""
                a_tag    = tds[0].find("a", href=True)
                link     = a_tag["href"] if a_tag else ""
                results.append({
                    "title"   : title,
                    "org"     : org,
                    "country" : country,
                    "opp_type": opp_type,
                    "posted"  : "",
                    "deadline": deadline,
                    "link"    : link,
                })

    return results


def _find_live_url(driver) -> str:
    """
    Try each PORTAL_URL until one loads successfully (200-like response).
    Returns the first URL that loads content, or empty string if all fail.
    """
    for url in PORTAL_URLS:
        try:
            print(f"  USAID: Trying {url} ...")
            driver.get(url)
            # Wait briefly to see if page loads
            time.sleep(4)
            # Check if page has meaningful content (not just an error page)
            src = driver.page_source
            if len(src) > 5000 and not any(k in src.lower() for k in (
                "404", "not found", "service unavailable", "err_", "connection refused",
            )):
                print(f"  USAID: Live at {url}")
                return url
            print(f"  USAID: {url} loaded but looks empty/error — trying next")
        except Exception as e:
            print(f"  USAID: {url} failed: {str(e)[:60]}")
    return ""


def _try_sam_api(api_key: str = "") -> list:
    """
    Pull USAID opportunities from SAM.gov API as primary data source.
    Requires SAM_API_KEY from config. Returns list of card-format dicts.
    Falls back to empty list so Selenium path still tries.
    """
    try:
        from config.config import SAM_API_KEY as _key
        key = api_key or _key
        if not key or key in ("", "YOUR_SAM_API_KEY"):
            return []

        from datetime import datetime, timedelta
        today     = datetime.utcnow()
        date_to   = today.strftime("%m/%d/%Y")
        date_from = (today - timedelta(days=45)).strftime("%m/%d/%Y")

        def _append_results(payload: dict) -> None:
            for opp in (payload.get("opportunitiesData") or []):
                nid = (opp.get("noticeId") or "").strip()
                if not nid or nid in seen_ids:
                    continue
                seen_ids.add(nid)
                title = (opp.get("title") or "").strip()
                if not title:
                    continue
                agency = (opp.get("fullParentPathName") or opp.get("organizationName") or "")
                if "USAID" not in agency.upper() and "USAID" not in title.upper():
                    continue
                results.append({
                    "title": title,
                    "org": agency,
                    "country": "",
                    "opp_type": (opp.get("type") or "").upper(),
                    "posted": (opp.get("postedDate") or "")[:10],
                    "deadline": (opp.get("responseDeadLine") or "")[:10],
                    "link": (opp.get("uiLink") or f"https://sam.gov/opp/{nid}/view"),
                })

        results = []
        seen_ids: set = set()
        queries = SAM_USAID_QUERIES
        for query in queries:
            try:
                resp = requests.get(
                    SAM_API_URL,
                    params={
                        "api_key": key,
                        "title": query,
                        "organizationName": "USAID",
                        "ptype": ["p", "r", "s", "o", "k", "i"],
                        "limit": 50,
                        "postedFrom": date_from,
                        "postedTo": date_to,
                        "status": "active",
                    },
                    headers={"User-Agent": "TenderMonitor/3.0"},
                    timeout=20,
                )
                if resp.status_code != 200:
                    body_snip = (resp.text or "")[:400].replace("\n", " ")
                    print(f"  USAID: SAM API HTTP {resp.status_code} for '{query}' body='{body_snip}'")
                    continue
                _append_results(resp.json())
                time.sleep(0.5)
            except Exception:
                continue
        # Broad fallback if nothing returned from focused queries
        if not results:
            for query in ("USAID", "USAID consulting", "development consulting"):
                try:
                    resp = requests.get(
                        SAM_API_URL,
                        params={
                            "api_key": key,
                            "title": query,
                            "organizationName": "USAID",
                            "ptype": ["p", "r", "s", "o", "k", "i"],
                            "limit": 50,
                            "postedFrom": date_from,
                            "postedTo": date_to,
                            "status": "active",
                        },
                        headers={"User-Agent": "TenderMonitor/3.0"},
                        timeout=20,
                    )
                    if resp.status_code != 200:
                        body_snip = (resp.text or "")[:400].replace("\n", " ")
                        print(f"  USAID: SAM API HTTP {resp.status_code} for '{query}' body='{body_snip}'")
                        continue
                    _append_results(resp.json())
                    time.sleep(0.5)
                except Exception:
                    continue
        if results:
            print(f"  USAID: SAM API returned {len(results)} USAID opportunities")
        return results
    except Exception as exc:
        print(f"  USAID: SAM API failed: {exc}")
        return []


def run():
    all_rows    = []
    new_tenders = []
    seen_titles = set()

    # ── Priority 1: SAM.gov API (most reliable — no Selenium needed) ───────────
    sam_cards = _try_sam_api()
    if sam_cards:
        print(f"  USAID: SAM API returned {len(sam_cards)} cards")
        for c in sam_cards:
            title = c["title"]
            key   = title[:80]
            if key in seen_titles:
                continue
            seen_titles.add(key)
            relevance = score_relevance(title, "")
            if not relevance and not title_is_relevant(title):
                continue
            slug      = re.sub(r"[^A-Za-z0-9]", "_", title)[:60]
            tender_id = f"USAID_{slug}"
            c["relevance"] = relevance
            all_rows.append(c)
            if check_if_new(tender_id):
                mark_as_seen(tender_id, title, "USAID", c.get("link") or "")
                new_tenders.append(c)
        if all_rows:
            _save_excel(all_rows)
            print(f"USAID (SAM API) done — {len(all_rows)} relevant, {len(new_tenders)} NEW")
            return new_tenders, all_rows

    # ── Priority 2: Selenium on partner2peer / workwithusaid ──────────────────
    print("  USAID: SAM API unavailable — trying Selenium on portal URLs...")
    driver = None
    try:
        driver = _build_driver()
    except Exception as exc:
        print(f"  USAID: Selenium init failed: {exc}")
        print("  USAID: Skipping this run without marking a failure.")
        return [], []
    try:
        live_url = _find_live_url(driver)
        if not live_url:
            print("  USAID: All portal URLs unreachable — skipping this run (not unstable).")
            print("  USAID: Add SAM_API_KEY to config/.env for reliable USAID data.")
            return [], []

        print(f"  USAID WorkWithUSAID: Loading from {live_url}...")
        driver.get(live_url)

        wait = WebDriverWait(driver, PAGE_WAIT)
        # Wait for any content to appear
        try:
            wait.until(lambda d: (
                len(d.find_elements(By.TAG_NAME, "a")) > 15 or
                len(d.find_elements(By.CSS_SELECTOR,
                    '[class*="opportunit"], [class*="listing"], '
                    '[class*="result"], [data-testid]')) > 0
            ))
        except Exception:
            pass
        time.sleep(6)   # React hydration settle

        # ── Scroll to load all lazy-loaded items ──────────────────────────────
        last_height = driver.execute_script("return document.body.scrollHeight")
        for _ in range(MAX_SCROLLS):
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(SCROLL_PAUSE)
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height

        soup  = BeautifulSoup(driver.page_source, "html.parser")
        cards = _parse_cards(soup)
        print(f"  USAID: {len(cards)} opportunities parsed")

        for c in cards:
            title = c["title"]
            key   = title[:80]
            if key in seen_titles:
                continue
            seen_titles.add(key)

            relevance = score_relevance(title, "")
            if not relevance and not title_is_relevant(title):
                continue

            slug      = re.sub(r"[^A-Za-z0-9]", "_", title)[:60]
            tender_id = f"USAID_{slug}"
            c["relevance"] = relevance
            c["tender_id"] = tender_id
            all_rows.append(c)

            if check_if_new(tender_id):
                mark_as_seen(tender_id, title, "USAID", c.get("link") or PORTAL_URL)
                new_tenders.append(c)

    except Exception as e:
        print(f"  USAID ERROR: {e}")
    finally:
        if driver:
            driver.quit()

    _save_excel(all_rows)
    print(f"USAID done — {len(all_rows)} relevant, {len(new_tenders)} NEW")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, all_ = run()
    print(f"\n✅  USAID: {len(new)} new / {len(all_)} total")
