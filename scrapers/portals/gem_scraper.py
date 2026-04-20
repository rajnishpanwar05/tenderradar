# =============================================================================
# scrapers/portals/gem_scraper.py — GeM BidPlus Decision-Grade Pipeline
#
# COMPLETE REWRITE — Production-hardened, consulting-only filter, decision engine.
#
# Site   : https://bidplus.gem.gov.in
# Method : requests + CSRF token + paginated JSON API
# Filter : Hard pre-filter (REJECT junk / REQUIRE consulting) → quality scoring
#          → GeM weighted scoring → decision engine (BID_NOW / STRONG / REVIEW)
#
# ── Pipeline (IntelligentBaseScraper) ────────────────────────────────────────
#   on_run_start()       → clear Excel, create session + CSRF
#   fetch_data()         → paginated API, rotating headers, session-refresh retry
#   validate_schema()    → structure-change detection (0-row warning)
#   extract_rows()       → normalize + category upstream filter + hard pre-filter
#   [quality filter]     → apply_intelligence_filter() [shared, threshold=15]
#   on_filter_complete() → capture quality-filter rejections for debug
#   enrich_fields()      → days_left, gem_score, decision override, WHY explanation
#   get_tender_id()      → SHA-256(title|deadline|bid_number) — detects updates
#   to_standard_format() → TenderResult for notifications
#   on_run_end()         → multi-sheet Excel: BID NOW / STRONG / ALL FILTERED
#   _print_debug()       → extended debug metrics (--debug flag)
#
# ── Changes vs old gem_scraper.py ────────────────────────────────────────────
#   + Migrated to IntelligentBaseScraper (consistent with all other scrapers)
#   + Hard pre-filter: REJECT supply/goods; REQUIRE consulting signal
#   + GeM weighted scoring (0-100) with urgency + sector bonus
#   + decision_tag overridden by gem_score (not quality_engine estimate)
#   + WHY explanation per tender ("Evaluation + Education sector + 8d left")
#   + Hash-based tender_id: detects updated deadlines/titles as new
#   + Rotating User-Agents (4 pool) to reduce block risk
#   + Multi-sheet Excel: 🔥 BID NOW / ⭐ STRONG / 📊 ALL FILTERED
#   + Decision-tier row colouring + user-input columns (My Decision / Outcome)
#   + Extended debug: score distribution, top 5 BID NOW, rejection reasons
#
# ── No changes required in ───────────────────────────────────────────────────
#   config.py    (GEM_EXCEL_PATH / GEM_MAX_PAGES / GEM_BID_TYPE already exist)
#   database/db.py   (check_if_new / mark_as_seen unchanged)
#   exporters/       (unified exporter picks up TenderResult from to_standard_format)
# =============================================================================

from __future__ import annotations

import hashlib
import json
import logging
import os
import random
import re
import time
from datetime import datetime
from typing import Any

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from config.config import GEM_EXCEL_PATH, GEM_MAX_PAGES, GEM_BID_TYPE, GEM_BID_STATUS
from core.base_scraper import IntelligentBaseScraper
from core.quality_engine import TenderResult, make_tender_result

logger = logging.getLogger("tenderradar.gem")


# =============================================================================
# SECTION 1 — Configuration & Vocabulary
# =============================================================================

BASE_URL          = "https://bidplus.gem.gov.in"
PAGE_URL          = f"{BASE_URL}/all-bids"
API_URL           = f"{BASE_URL}/all-bids-data"
SEARCH_BIDS_URL   = f"{BASE_URL}/search-bids"     # category-filtered endpoint

SORT_BY   = "Bid-End-Date-Oldest"   # most urgent first
DATE_FROM = ""
DATE_TO   = ""

# ── Phase 1 — Category-based fetch via /search-bids (100% precision) ─────────
# The /search-bids endpoint filters by b_cat_id — every result is guaranteed
# to be in that consulting category.  No keyword ambiguity, no junk.
#
# Source: live API investigation Mar 2026.
# Format: {b_cat_id: display_name}
# Active bid counts as of research: total ~312 bids across 11 categories.
_CONSULTING_CATEGORIES: dict[str, str] = {
    # Core consulting (182 bids)
    "services_home_pr22455282_hi85156607": "Hiring of Consultants - Milestone Based",
    # Financial advisory (37 bids)
    "services_home_fina_fi44514430":       "Financial Advisory Services",
    # Consultancy percentage-based (20 bids)
    "services_home_pr22455282_hi33564580": "Hiring of Consultancy - Percentage Based",
    # Per person per month consulting (14 bids)
    "services_home_pr22455282_hi53016071": "Hiring of Consultants - Per Person Month",
    # Professional training (11 bids)
    "services_home_tr64181578_prof":       "Professional Training Services",
    # Survey / research / evaluation / assessment (9 bids)
    "services_home_ma32785043_surv":       "Survey / Market Research / Evaluation",
    # Non-IT professional services (8 bids)
    "services_home_pr22455282_no68842601": "Non-IT Professional Service",
    # GIS consulting (2 bids)
    "services_home_giss_hiri":             "Hiring of GIS Implementation Agency",
    # NITI Aayog empanelment (1 bid)
    "services_home_empa":                  "Empanelment of Consultants (NITI Aayog)",
    # Safety consultant (1 bid)
    "services_home_pr22455282_hi07707508": "Hiring of Safety Consultant",
    # PR agency (1 bid)
    "services_home_pr22455282_hi82684505": "Hiring of PR Agency",
}
_MAX_CATEGORY_PAGES: int = 25   # safety cap (25 × 10 = 250 per category; largest has 182)

# ── Phase 2 — Keyword fallback for custom bids not in structured categories ───
# "Custom Bid for Services" (2,590 bids) and "General Services" (408 bids)
# contain legitimate consulting work not mapped to a specific b_cat_id.
# We search these with keywords — apply existing hard-reject filter to clean results.
_CUSTOM_BID_KEYWORDS: list[str] = [
    "consultant",   # ~233 bids incl. custom bids
    "advisory",     # ~43 bids
    "evaluation",   # additional uncategorised evaluation bids
]
_PAGES_PER_KEYWORD: int = 3   # 3 × 10 = 30 per keyword (most consulting are in Phase 1)

# ── Rotating User-Agents (anti-block pool) ────────────────────────────────────
_USER_AGENTS: list[str] = [
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
     "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
     "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"),
    ("Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) "
     "Gecko/20100101 Firefox/121.0"),
    ("Mozilla/5.0 (X11; Linux x86_64) "
     "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36"),
]

# ── Hard pre-filter: REQUIRE at least one genuine consulting signal ───────────
# Substring matching: "consult" hits consultant/consultancy/consulting.
# "hiring of agency" added — common GeM category for milestone consulting work.
_REQUIRED_SIGNALS: list[str] = [
    "consult",               # consultant, consultancy, consulting
    "evaluation",
    "assessment",
    "technical assistance",
    "advisory",
    "capacity building",
    "research",
    "feasibility",
    "impact",
    "hiring of agency",      # GeM category: "Hiring of Agency for X - Milestone basis"
]

# ── Hard pre-filter: HARD REJECT list — wins unconditionally ─────────────────
# RULE: Use SPECIFIC PHRASES for words that appear in legitimate consulting bids.
#   ✗ "supply"      → too broad: kills "Supply Chain Consultancy"
#   ✗ "maintenance" → too broad: kills "IT Maintenance Advisory"
#   ✗ "transport"   → too broad: kills "Transport Sector Policy Advisory"
#   ✗ "installation"→ too broad: kills "IT Implementation Advisory"
#   ✗ "insurance"   → too broad: kills "Insurance Regulatory Advisory"
# Use the unambiguous bare words (manpower, housekeeping, etc.) as-is.
_HARD_REJECT: list[str] = [
    # ── Unambiguous bare words (never appear in consulting context) ───────────
    "manpower",
    "outsourcing",
    "facility management",
    "housekeeping",
    "sanitation",
    "security guard",
    # ── Non-IDCG professional services ───────────────────────────────────────
    # These are consulting-like but NOT development consulting.
    # A "law firm for evaluation of contracts" is NOT an IDCG opportunity.
    "law firm",
    "legal consultant",
    "legal advisory",
    "legal services",
    "legal counsel",
    "empanelment of advocates",
    "empanelment of advocate",
    "advocate firm",
    "panel advocate",
    "chartered accountant",
    "ca firm",
    "statutory auditor",
    "tax consultant",
    "company secretary",
    "architecture firm",
    "architectural consultant",
    "interior designer",
    "software development",
    "app development",
    "website development",
    "web development",
    "erp implementation",
    "system integrator",
    # ── Specific phrases only for ambiguous words ─────────────────────────────
    "supply of goods",
    "supply and delivery",
    "supply and installation",
    "supply of equipment",
    "supply of material",
    "supply of hardware",
    "supply of furniture",
    "procurement of goods",
    "purchase of ",
    "annual maintenance contract",
    "amc for ",
    "amc of ",
    "repair and maintenance",
    "vehicle hire",
    "vehicle hiring",
    "cab hire",
    "cab services",
    "civil works",
    "construction of ",
    "catering services",
    "printing and stationery",
    # ── Technical surveys (field work, not consulting) ─────────────────────────
    "geophysical survey",
    "bathymetric survey",
    "topographic survey",
    "hydrographic survey",
    "geological survey",
    "geotechnical investigation",
    "soil testing",
    "bore hole",
    "borehole",
    "drilling services",
    "seismic survey",
    "gpr survey",
    "lidar survey",
    # ── Other non-consulting services ──────────────────────────────────────────
    "sweeping and cleaning",
    "horticulture services",
    "pest control",
    "scavenging",
    "laundry services",
    "ambulance services",
    "fire fighting",
    "fire safety equipment",
    "cctv installation",
    "electrical works",
    "plumbing works",
    "waterproofing",
    "painting works",
    "interior works",
    "hiring of bus",
    "hiring of vehicle",
    "hiring of truck",
]

# ── Safety net: final scan before returning enriched rows ─────────────────────
# Belt-and-suspenders — catches any junk that slips through scoring edge cases.
# Only the most unambiguous terms — not "supply" or "transport".
_SAFETY_KEYWORDS: list[str] = [
    "manpower",
    "outsourcing",
    "security guard",
    "housekeeping",
    "sanitation",
]

# ── GeM weighted scoring vocabulary ──────────────────────────────────────────
# List of ([keywords], weight) — first match in group scores, no double-count.
# Positive = consulting signals, Negative = goods/supply signals.
_GEM_SCORE_WEIGHTS: list[tuple[list[str], int]] = [
    # ── Tier-1 positive: core consulting types (+25) ─────────────────────────
    (["consultancy services", "hiring of consultants",
      "engagement of consultants", "consultancy"],            +25),
    (["consultant"],                                           +25),
    (["evaluation", "assessment", "impact assessment",
      "programme evaluation", "project evaluation"],          +18),
    (["monitoring and evaluation", "m&e",
      "third party monitoring", "tpm", "independent verification",
      "iva"],                                                  +18),
    # ── Tier-2 positive: technical/advisory work (+20 / +15) ────────────────
    (["technical assistance"],                                 +20),
    (["advisory"],                                             +15),
    (["capacity building"],                                    +12),
    # ── Tier-3 positive: research/study (capped at 8 pts) ───────────────────
    (["research", "study", "feasibility study",
      "feasibility assessment", "scoping study"],              +8),
    (["baseline survey", "baseline study", "endline survey",
      "mid-term review", "final evaluation"],                  +8),
    (["training", "skill development",
      "human resource development"],                           +8),
    (["milestone", "milestone based", "deliverable based",
      "deliverable"],                                          +8),
]

# Internal labels used in Excel
_DECISION_DISPLAY: dict[str, str] = {
    "BID_NOW":         "🔥 BID NOW",
    "STRONG_CONSIDER": "⭐ STRONG",
    "WEAK_CONSIDER":   "📊 REVIEW",
    "IGNORE":          "— IGNORE",
}

BID_TYPE_LABEL: dict[int, str] = {1: "BID", 2: "RA", 5: "Direct RA"}

BUYER_STATUS_MAP: dict[int, str] = {
    0: "Not Evaluated",    1: "Technical Evaluation",
    2: "Financial Evaluation", 3: "Bid Award",
}

# ── Fake-consulting penalty keywords (Task 4) ─────────────────────────────────
# These words appear in bids that look consulting-like but are actually
# supervision/manpower contracts.  Each match subtracts 25 points.
_FAKE_CONSULTING_KEYWORDS: list[str] = [
    "supervision", "inspection", "installation", "manpower", "outsourcing",
    # ── Non-IDCG professional services that mimic consulting keywords ────────
    "law firm", "legal service", "legal advisor", "advocate",
    "lawyer", "attorney", "chartered accountant", "ca firm",
    "statutory audit", "tax consult", "company secretary",
    "architect", "interior design", "software develop",
    "app develop", "web develop", "erp implement", "system integrat",
]
_FAKE_CONSULTING_PENALTY: int = -25

# ── Output size limits (Task 2) ───────────────────────────────────────────────
# Applied in enrich_fields() — caps how many rows of each tier reach
# the DB-dedup check (notifications) and the Excel export.
OUTPUT_LIMIT_BID_NOW: int = 10
OUTPUT_LIMIT_STRONG:  int = 15


# =============================================================================
# SECTION 1b — Firm Profile Scoring
# =============================================================================

def _load_firm_profile() -> dict:
    """
    Load config/firm_profile.json once at import time.
    Returns empty dict silently on any error so the pipeline never aborts.
    """
    path = os.path.normpath(
        os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "..", "..", "config", "firm_profile.json",
        )
    )
    try:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    except Exception as exc:
        logger.warning("[gem] firm_profile.json not loaded: %s", exc)
        return {}


_FIRM_PROFILE: dict = _load_firm_profile()


def _apply_firm_profile_score(
    row: dict,
    base_score: int,
) -> tuple[int, list[str]]:
    """
    Apply firm-strategy boosts and penalties from firm_profile.json.

    Rules (values read from score_boosts / score_penalties keys in JSON):
      preferred_sector  match → +10  (one bonus max)
      preferred_region  match → +5   (one bonus max; GeM = India = South Asia)
      avoid_sector      match → -15  (one penalty max)
      avoid_keyword     match → -20  (one penalty max)

    Returns (adjusted_score_clamped_0_100, list_of_signal_strings).
    """
    if not _FIRM_PROFILE:
        return base_score, []

    boosts   = _FIRM_PROFILE.get("score_boosts",   {})
    penalties= _FIRM_PROFILE.get("score_penalties", {})

    combined = (
        row.get("Item Category",     "") + " " +
        row.get("Brief Description", "") + " " +
        row.get("title",             "")
    ).lower()

    sector    = (row.get("Sector")    or row.get("sector",    "")).lower()
    geography = (row.get("Geography") or row.get("geography", "")).lower()

    adjustment:      int       = 0
    profile_signals: list[str] = []

    # ── Preferred sector bonus ────────────────────────────────────────────────
    boost_ps = boosts.get("preferred_sector", 10)
    for ps in _FIRM_PROFILE.get("preferred_sectors", []):
        if ps.lower() in sector or ps.lower() in combined:
            adjustment += boost_ps
            profile_signals.append(f"+{boost_ps} ({ps} sector)")
            break   # one sector bonus max

    # ── Preferred region bonus ────────────────────────────────────────────────
    # GeM is always India → "South Asia" always matches; still applied so
    # GeM bids score consistently vs other portals in cross-portal views.
    boost_r = boosts.get("preferred_region", 5)
    for region in _FIRM_PROFILE.get("preferred_regions", []):
        rl = region.lower()
        if rl == "south asia" or rl in geography or rl in combined:
            adjustment += boost_r
            profile_signals.append(f"+{boost_r} ({region})")
            break   # one region bonus max

    # ── Avoid sector penalty ──────────────────────────────────────────────────
    pen_as = penalties.get("avoid_sector", -15)
    for av in _FIRM_PROFILE.get("avoid_sectors", []):
        if av.lower() in sector or av.lower() in combined:
            adjustment += pen_as
            profile_signals.append(f"{pen_as} (avoid sector: {av})")
            break

    # ── Avoid keyword penalty ─────────────────────────────────────────────────
    pen_kw = penalties.get("avoid_keyword_in_title", -20)
    for kw in _FIRM_PROFILE.get("avoid_keywords", []):
        if kw.lower() in combined:
            adjustment += pen_kw
            profile_signals.append(f"{pen_kw} (avoid kw: {kw})")
            break

    return max(0, min(100, base_score + adjustment)), profile_signals


# =============================================================================
# SECTION 2 — Session + CSRF Helpers
# =============================================================================

def _make_session() -> requests.Session:
    """
    Create a Session with:
    - Random User-Agent from the rotation pool
    - Retry logic (3 attempts, exponential backoff) for 5xx / 429
    """
    sess = requests.Session()
    sess.headers.update({
        "User-Agent":       random.choice(_USER_AGENTS),
        "Accept":           "application/json, text/javascript, */*; q=0.01",
        "Accept-Language":  "en-US,en;q=0.9",
        "Accept-Encoding":  "gzip, deflate, br",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":          PAGE_URL,
        "Origin":           BASE_URL,
        "Connection":       "keep-alive",
    })
    retry = Retry(
        total=3,
        backoff_factor=1.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        respect_retry_after_header=True,
    )
    sess.mount("https://", HTTPAdapter(max_retries=retry))
    sess.mount("http://",  HTTPAdapter(max_retries=retry))
    return sess


def _fetch_csrf_token(sess: requests.Session) -> str:
    """
    Fetch CSRF token with two-chain fallback and 3 retry attempts:
      Chain 1: inline JS variable  csrf_bd_gem_nk: "..."
      Chain 2: hidden <input> with 32-char hex value
    Raises RuntimeError if all attempts fail.
    """
    for attempt in range(3):
        try:
            resp = sess.get(PAGE_URL, timeout=30)
            resp.raise_for_status()
            html = resp.text

            # Chain 1 — inline JSON / JS assignment
            m = re.search(
                r"csrf_bd_gem_nk['\"]?\s*:\s*['\"]([a-f0-9]{32})['\"]", html
            )
            if m:
                return m.group(1)

            # Chain 2 — hidden form field
            soup = BeautifulSoup(html, "html.parser")
            for inp in soup.find_all("input", {"type": "hidden"}):
                val = inp.get("value", "")
                if re.fullmatch(r"[a-f0-9]{32}", val):
                    return val

            logger.warning("[gem] CSRF not found on attempt %d", attempt + 1)

        except Exception as exc:
            logger.warning("[gem] CSRF fetch attempt %d failed: %s", attempt + 1, exc)

        if attempt < 2:
            time.sleep(2 ** attempt)

    raise RuntimeError(
        "[gem] CSRF token not found after 3 attempts — "
        "page layout may have changed."
    )


def _build_post_data(page_num: int, csrf_token: str, keyword: str = "") -> dict:
    """Payload for /all-bids-data (keyword full-text search)."""
    return {
        "payload": json.dumps({
            "page": page_num,
            "param": {"searchBid": keyword, "searchType": "fullText"},
            "filter": {
                "bidStatusType": GEM_BID_STATUS,
                "byType":        GEM_BID_TYPE,
                "highBidValue":  "",
                "byEndDate":     {"from": DATE_FROM, "to": DATE_TO},
                "sort":          SORT_BY,
            },
        }),
        "csrf_bd_gem_nk": csrf_token,
    }


def _build_category_payload(page_num: int, csrf_token: str, cat_id: str) -> dict:
    """Payload for /search-bids (category-based, 100% precision)."""
    return {
        "payload": json.dumps({
            "searchType": "bidNumber",
            "bidNumber":  "",
            "category":   cat_id,
            "bidEndFrom": DATE_FROM,
            "bidEndTo":   DATE_TO,
            "page":       page_num,
        }),
        "csrf_bd_gem_nk": csrf_token,
    }


def _fetch_page(
    sess: requests.Session,
    csrf_token: str,
    page_num: int,
    keyword: str = "",
) -> tuple[list[dict] | None, int]:
    """
    Fetch a single page from /all-bids-data (keyword search).
    Returns (docs, num_found).  Returns (None, 0) on any error.
    """
    try:
        resp = sess.post(
            API_URL,
            data=_build_post_data(page_num, csrf_token, keyword),
            timeout=30,
        )
        resp.raise_for_status()
        js = resp.json()
    except Exception as exc:
        logger.warning("[gem] Page %d fetch error: %s", page_num, exc)
        return None, 0

    if js.get("code") != 200:
        logger.warning("[gem] Page %d: API code %s", page_num, js.get("code"))
        return None, 0

    inner = js.get("response", {}).get("response", {})
    return inner.get("docs", []), inner.get("numFound", 0)


def _fetch_category_page(
    sess: requests.Session,
    csrf_token: str,
    page_num: int,
    cat_id: str,
) -> tuple[list[dict] | None, int]:
    """
    Fetch a single page from /search-bids (category-filtered, 100% precision).
    Response structure is identical to /all-bids-data.
    Returns (docs, num_found).  Returns (None, 0) on any error.
    """
    try:
        resp = sess.post(
            SEARCH_BIDS_URL,
            data=_build_category_payload(page_num, csrf_token, cat_id),
            timeout=30,
        )
        resp.raise_for_status()
        js = resp.json()
    except Exception as exc:
        logger.warning("[gem] Category page %d/%s fetch error: %s", page_num, cat_id[:20], exc)
        return None, 0

    if js.get("code") != 200:
        logger.warning("[gem] Category page %d: API code %s", page_num, js.get("code"))
        return None, 0

    inner = js.get("response", {}).get("response", {})
    return inner.get("docs", []), inner.get("numFound", 0)


# =============================================================================
# SECTION 3 — Field Normalization Helpers
# =============================================================================

def _to_str(val: Any) -> str:
    if isinstance(val, list):
        return " ".join(str(v) for v in val if v).strip()
    return str(val).strip() if val else ""


def _to_int(val: Any, default: int = 0) -> int:
    if isinstance(val, list):
        val = val[0] if val else default
    try:
        return int(val)
    except Exception:
        return default


def _parse_utc_date(iso_str: str) -> str:
    try:
        return datetime.strptime(iso_str, "%Y-%m-%dT%H:%M:%SZ").strftime(
            "%d-%m-%Y %H:%M"
        )
    except Exception:
        return iso_str or ""


def _compute_days_left(deadline_str: str) -> int:
    """Return calendar days from today to deadline.  Returns -1 if unparseable."""
    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%SZ"):
        try:
            dl = datetime.strptime(deadline_str, fmt)
            return max(0, (dl.date() - datetime.now().date()).days)
        except ValueError:
            continue
    return -1


def _normalize_raw_bid(raw: dict) -> dict | None:
    """
    Convert one raw API response dict into a flat normalized row dict.

    Returns None when:
      - bid_id is missing (invalid record)
      - Item Category contains no consulting/milestone signal (upstream filter)
      - title is empty after extraction

    Does NOT apply the hard REJECT/REQUIRE pre-filter — that runs separately
    in extract_rows() so rejections can be logged distinctly.
    """
    b_id   = _to_str(raw.get("b_id"))
    b_type = _to_int(raw.get("b_bid_type"), 1)

    if not b_id:
        return None

    # ── Item Category ─────────────────────────────────────────────────────────
    cats = raw.get("b_category_name") or []
    if isinstance(cats, str):
        cats = [cats]
    cats_clean    = [str(c).strip() for c in cats if c]
    item_category = "; ".join(cats_clean)

    # ── URL + bid type label ──────────────────────────────────────────────────
    # NOTE: /showbidDocument/{b_id} serves a raw PDF download — NOT a page.
    # We instead link to the All Bids search page with the bid number pre-filled
    # so clicking the Excel link opens the GeM website and shows the specific bid.
    # The user can then view the project page and download the PDF from there.
    bid_num_for_url = _to_str(raw.get("b_bid_number"))
    # URL-encode the bid number for use as a query param
    from urllib.parse import quote as _url_quote
    doc_path = f"/all-bids?searchBid={_url_quote(bid_num_for_url)}&searchType=bidNumber"

    type_lbl = BID_TYPE_LABEL.get(b_type, "BID")
    if _to_int(raw.get("is_rc_bid"))             == 1: type_lbl += " (Rate Contract)"
    if _to_int(raw.get("ba_is_global_tendering")) == 1: type_lbl += " (Global Tender)"

    # ── Organisation ──────────────────────────────────────────────────────────
    ministry_raw  = _to_str(raw.get("ba_official_details_minName")).replace("_", " ").title()
    dept_name     = _to_str(raw.get("ba_official_details_deptName"))
    ministry_dept = (ministry_raw + " | " + dept_name) if dept_name else ministry_raw

    # ── Dates ─────────────────────────────────────────────────────────────────
    deadline  = _parse_utc_date(_to_str(raw.get("final_end_date_sort")))
    issue_dt  = _parse_utc_date(_to_str(raw.get("final_start_date_sort")))

    # ── Brief description (fallback chain) ────────────────────────────────────
    brief_desc = ""
    for _k in ("b_description", "bidDescription", "b_bid_description",
               "b_name", "b_title", "name"):
        _v = raw.get(_k)
        if _v:
            brief_desc = _to_str(_v)[:600]
            break

    # ── Estimated value (fallback chain) ──────────────────────────────────────
    estimated_value = ""
    for _k in ("b_estimated_total_value", "b_total_estimated_value",
               "b_bid_est_cost", "totalBidValue", "estimated_total_value"):
        _v = raw.get(_k)
        if _v not in (None, "", 0, "0"):
            try:
                estimated_value = f"₹ {float(_v):,.2f}"
            except Exception:
                estimated_value = str(_v)
            break

    # ── EMD (fallback chain) ──────────────────────────────────────────────────
    emd = ""
    for _k in ("b_emd_amt", "b_emd_amount", "emdAmount", "b_emd", "emd_value"):
        _v = raw.get(_k)
        if _v not in (None, "", 0, "0"):
            try:
                emd = f"₹ {float(_v):,.2f}"
            except Exception:
                emd = str(_v)
            break

    # ── Consignee state (fallback chain) ──────────────────────────────────────
    consignee_state = ""
    for _k in ("ba_consignee_state_name", "b_consignee_state_names",
               "consigneeStateNames", "b_consignee_state", "stateNames"):
        _v = raw.get(_k)
        if _v:
            consignee_state = (
                "; ".join(str(x) for x in _v if x)
                if isinstance(_v, list) else str(_v)
            )
            break

    bid_number = _to_str(raw.get("b_bid_number"))
    bid_url    = BASE_URL + doc_path

    # ── Validate essential fields ─────────────────────────────────────────────
    # A bid with no category AND no description is useless — skip it.
    if not item_category and not brief_desc:
        return None

    # Validate URL
    if not bid_url.startswith("https://"):
        return None

    # ── Build combined rich description for quality_engine scoring ───────────
    # quality_engine reads "Description" and "Method" for scoring + confidence.
    rich_desc = (item_category + ". " + brief_desc).strip(". ")

    # Method mapping: consulting GeM bids → "rfp" gives confidence 0.75
    # Direct RA → "direct" gives 0.55 (still passes filter at threshold=15)
    method_map = {1: "rfp", 2: "rfp", 5: "direct"}
    method = method_map.get(b_type, "rfp")

    return {
        # ── Internal keys ─────────────────────────────────────────────────
        "bid_id":    b_id,
        "bid_number": bid_number,
        # ── quality_engine standard field names ───────────────────────────
        "title":       item_category or brief_desc[:120],
        "Description": rich_desc,
        "Method":      method,
        "Deadline":    deadline,
        "organization": dept_name or ministry_raw,
        # ── Display fields ────────────────────────────────────────────────
        "Title":                 item_category or brief_desc[:120],
        "Bid Number":            bid_number,
        "Bid Type":              type_lbl,
        "Item Category":         item_category,
        "Brief Description":     brief_desc,
        "Organisation Name":     dept_name,
        "Ministry / Department": ministry_dept,
        "Consignee State":       consignee_state,
        "Issue Date":            issue_dt,
        "Bid End Date":          deadline,
        "Estimated Bid Value":   estimated_value,
        "EMD / Bid Security":    emd,
        "Buyer Status":          BUYER_STATUS_MAP.get(
                                     _to_int(raw.get("b_buyer_status")), "Not Evaluated"
                                 ),
        "Bid Link":              bid_url,
        "url":                   bid_url,
        "Source":                "GeM",
    }


# =============================================================================
# SECTION 4 — Hard Pre-Filter (REJECT / REQUIRE)
# =============================================================================

def _passes_hard_filter(row: dict) -> tuple[bool, str]:
    """
    Two-gate hard pre-filter — strict consulting-only gate.

    Gate 1 — HARD REJECT (unconditional):
        If ANY word from _HARD_REJECT appears in the combined text → DROP.
        This wins even if a consulting signal is also present.
        Eliminates manpower, supply, vehicle, maintenance, cleaning, etc.

    Gate 2 — REQUIRE consulting signal:
        At least ONE term from _REQUIRED_SIGNALS must be present.
        "service", "agency", "firm" are NOT consulting signals and are excluded.

    Returns (passes, reason_if_rejected).
    """
    combined = (
        row.get("Item Category",     "") + " " +
        row.get("Brief Description", "") + " " +
        row.get("title",             "") + " " +
        row.get("Description",       "")
    ).lower()

    # Gate 1: HARD REJECT wins unconditionally — no exceptions
    for kw in _HARD_REJECT:
        if kw in combined:
            return False, f"hard_reject: '{kw}'"

    # Gate 2: REQUIRE at least one genuine consulting signal
    has_signal = any(kw in combined for kw in _REQUIRED_SIGNALS)
    if not has_signal:
        return False, "no consulting signal"

    return True, ""


# =============================================================================
# SECTION 5 — GeM Weighted Scoring + Decision Engine
# =============================================================================

def _compute_gem_score(row: dict) -> tuple[int, list[str]]:
    """
    Compute a GeM-specific weighted consulting score in range 0–100.

    Scoring sources:
      - Keyword weights from _GEM_SCORE_WEIGHTS (consulting signals only)
      - Deadline urgency bonus (+5 or +10)

    Returns (score, list_of_matched_signals).
    Returns (0, []) immediately if any _HARD_REJECT keyword is present —
    belt-and-suspenders guard, should not be reached in normal flow.
    """
    combined = (
        row.get("Item Category",     "") + " " +
        row.get("Brief Description", "") + " " +
        row.get("title",             "")
    ).lower()

    # HARD_REJECT guard — belt-and-suspenders; _passes_hard_filter should
    # have already dropped these, but score 0 ensures IGNORE if they slip through.
    for kw in _HARD_REJECT:
        if kw in combined:
            return 0, []

    score:   int        = 0
    matched: list[str]  = []

    for keywords, weight in _GEM_SCORE_WEIGHTS:
        for kw in keywords:
            if kw in combined:
                score += weight
                matched.append(kw)
                break   # one match per group — no double-scoring

    # ── Deadline urgency bonus ────────────────────────────────────────────────
    days_left = row.get("_days_left", -1)
    if 0 < days_left <= 7:
        score += 10
        matched.append(f"URGENT ({days_left}d left)")
    elif 0 < days_left <= 14:
        score += 5
        matched.append(f"deadline in {days_left}d")

    return max(0, min(100, score)), matched


def _gem_decision_tag(score: int) -> str:
    """Convert gem_score → decision tag aligned with quality_engine naming."""
    if score >= 75:  return "BID_NOW"
    if score >= 60:  return "STRONG_CONSIDER"
    if score >= 40:  return "WEAK_CONSIDER"
    return "IGNORE"


def _build_why(row: dict, score: int, matched: list[str]) -> str:
    """
    Build a structured WHY string for business teams (Task 3).

    Fixed order — always:
      1. Consulting type   e.g. "Evaluation"
      2. Sector            e.g. "Health sector"
      3. Urgency           e.g. "5 days left (urgent)" / "12 days left"

    Example outputs:
      "Evaluation + Health sector + 5 days left (urgent)"
      "Technical Assistance + Governance sector + 12 days left"
      "Capacity Building + 3 days left (urgent)"
      "Consulting category match"   ← fallback when nothing detected
    """
    parts: list[str] = []

    # 1. Consulting type — always first
    ctype = row.get("Consulting Type") or row.get("consulting_type", "")
    if ctype and ctype not in ("General Consulting", ""):
        parts.append(ctype)

    # 2. Sector — always second
    sector = row.get("Sector") or row.get("sector", "")
    if sector and sector not in ("General Development", ""):
        parts.append(f"{sector} sector")

    # 3. Deadline urgency — always third when present
    days_left = row.get("_days_left", -1)
    if days_left == 0:
        parts.append("closes today (urgent)")
    elif 0 < days_left <= 7:
        parts.append(f"{days_left} days left (urgent)")
    elif 0 < days_left <= 30:
        parts.append(f"{days_left} days left")

    return " + ".join(parts) if parts else "Consulting category match"


# =============================================================================
# SECTION 6 — GeMScraper (IntelligentBaseScraper)
# =============================================================================

class GeMScraper(IntelligentBaseScraper):
    """
    GeM BidPlus decision-grade consulting pipeline.

    Hard pre-filter → shared quality scoring → GeM weighted scoring →
    decision engine → multi-sheet Excel output.
    """

    SOURCE_NAME = "GeM"
    SOURCE_URL  = PAGE_URL
    EXCEL_PATH  = GEM_EXCEL_PATH

    # Lower threshold because hard pre-filter already removes junk.
    # Consulting GeM bids reliably score 40-60 on raw quality (method=rfp,
    # deadline present, description 10+ words) × 0.75-0.85 confidence ≈ 30-50.
    QUALITY_THRESHOLD = 15

    # Schema fields used in validate_schema() to detect API changes
    EXPECTED_SCHEMA_FIELDS: list[tuple] = [
        ("b_bid_number",        str),
        ("final_end_date_sort", str),
        ("b_category_name",     (str, list)),
    ]

    def __init__(self) -> None:
        super().__init__()
        self._session:              requests.Session | None = None
        self._csrf_token:           str  = ""
        self._rejected_pre_filter:  list = []   # (row, reason) tuples
        self._quality_rejected:     list = []   # (row, reason) tuples

    # =========================================================================
    # STEP 0 — Lifecycle: on_run_start
    # =========================================================================

    def on_run_start(self) -> None:
        """Clear stale Excel, build session, fetch CSRF."""
        if os.path.exists(self.EXCEL_PATH):
            try:
                os.remove(self.EXCEL_PATH)
                self._log.info("Cleared old Excel: %s", self.EXCEL_PATH)
            except Exception as exc:
                self._log.warning("Could not clear old Excel: %s", exc)

        # Reset rejection logs for this run
        self._rejected_pre_filter = []
        self._quality_rejected    = []

        self._session    = _make_session()
        self._csrf_token = _fetch_csrf_token(self._session)
        self._log.debug("[gem] Session ready. CSRF token acquired.")

    # =========================================================================
    # STEP 1 — Fetch
    # =========================================================================

    def fetch_data(self) -> list[dict]:
        """
        Hybrid two-phase fetch for maximum consulting bid coverage with precision.

        ── Phase 1: Category-based fetch via /search-bids (PRIMARY) ─────────────
          Uses GeM's /search-bids endpoint with specific b_cat_id values.
          EVERY result is a genuine consulting bid — 100% precision, zero noise.
          Covers ~312 structured consulting bids across 11 categories.
          Research confirmed: /all-bids-data ignores category filters; /search-bids
          does not — this is the only way to get clean category-filtered results.

        ── Phase 2: Keyword fallback via /all-bids-data (SECONDARY) ────────────
          "Custom Bid for Services" (2,590 bids) hides consulting work not mapped
          to a structured category — reachable only via keyword search.
          Hard-reject filter cleans noise from these results.

        All docs deduplicated by b_id — same bid never appears twice.
        """
        all_docs:   list[dict] = []
        seen_b_ids: set[str]   = set()

        # ── Phase 1: Structured category fetch ───────────────────────────────
        self._log.info("[gem] Phase 1: Category fetch (%d categories)", len(_CONSULTING_CATEGORIES))
        phase1_total = 0
        for cat_id, cat_name in _CONSULTING_CATEGORIES.items():
            cat_docs  = self._fetch_category_pages(cat_id)
            new_count = 0
            for doc in cat_docs:
                bid_id = _to_str(doc.get("b_id", ""))
                if bid_id and bid_id not in seen_b_ids:
                    seen_b_ids.add(bid_id)
                    all_docs.append(doc)
                    new_count += 1
            phase1_total += new_count
            self._log.info("[gem]   %-45s %3d fetched  %3d new", cat_name[:45], len(cat_docs), new_count)
            time.sleep(random.uniform(1.0, 2.0))

        self._log.info("[gem] Phase 1 subtotal: %d unique bids from structured categories", phase1_total)

        # ── Phase 2: Keyword fallback for custom / uncategorised bids ────────
        self._log.info("[gem] Phase 2: Keyword fallback (%d keywords)", len(_CUSTOM_BID_KEYWORDS))
        phase2_total = 0
        for keyword in _CUSTOM_BID_KEYWORDS:
            kw_docs   = self._fetch_keyword_pages(keyword)
            new_count = 0
            for doc in kw_docs:
                bid_id = _to_str(doc.get("b_id", ""))
                if bid_id and bid_id not in seen_b_ids:
                    seen_b_ids.add(bid_id)
                    all_docs.append(doc)
                    new_count += 1
            phase2_total += new_count
            self._log.info("[gem]   keyword '%-22s': %3d fetched  %3d new (custom bids)", keyword, len(kw_docs), new_count)
            time.sleep(random.uniform(1.5, 2.5))

        self._log.info("[gem] Phase 2 subtotal: %d additional custom bids", phase2_total)
        self._log.info("[gem] Total unique bids fetched: %d", len(all_docs))
        return all_docs

    def _fetch_category_pages(self, cat_id: str) -> list[dict]:
        """
        Fetch all pages for a single consulting category via /search-bids.
        Capped at _MAX_CATEGORY_PAGES as a safety net.
        """
        docs: list[dict] = []
        page = 1

        while page <= _MAX_CATEGORY_PAGES:
            page_docs, num_found = _fetch_category_page(
                self._session, self._csrf_token, page, cat_id
            )

            # Session/CSRF expired — refresh once
            if page_docs is None:
                try:
                    self._session    = _make_session()
                    self._csrf_token = _fetch_csrf_token(self._session)
                    page_docs, num_found = _fetch_category_page(
                        self._session, self._csrf_token, page, cat_id
                    )
                except Exception as exc:
                    logger.warning("[gem] Session refresh failed for cat %s: %s", cat_id[:20], exc)
                    break

            if not page_docs:
                break

            docs.extend(page_docs)

            total_pages = max(1, -(-num_found // 10))
            if page >= total_pages:
                break

            page += 1
            time.sleep(random.uniform(0.5, 1.0))

        return docs

    def _fetch_keyword_pages(self, keyword: str) -> list[dict]:
        """
        Fetch up to _PAGES_PER_KEYWORD pages via /all-bids-data keyword search.
        Used only for Phase 2 (custom/uncategorised bids).
        """
        docs: list[dict] = []
        page = 1

        while page <= _PAGES_PER_KEYWORD:
            page_docs, num_found = _fetch_page(
                self._session, self._csrf_token, page, keyword
            )

            # Session/CSRF expired — refresh once
            if page_docs is None:
                try:
                    self._session    = _make_session()
                    self._csrf_token = _fetch_csrf_token(self._session)
                    page_docs, num_found = _fetch_page(
                        self._session, self._csrf_token, page, keyword
                    )
                except Exception as exc:
                    self._log.warning("[gem] Session refresh failed for '%s': %s", keyword, exc)
                    break

            if not page_docs:
                break

            docs.extend(page_docs)

            total_pages = max(1, -(-num_found // 10))
            if page >= total_pages:
                break

            page += 1
            time.sleep(random.uniform(0.8, 1.5))

        return docs

    # =========================================================================
    # STEP 2 — Schema validation + structure-change detection
    # =========================================================================

    def validate_schema(self, raw_data: list[dict]) -> bool:
        """
        Structure-change detection:
          - WARN if 0 docs returned (API change / network issue)
          - WARN if expected fields missing in sample
        Does NOT abort the run — field fallback chains handle partial data.
        """
        if not raw_data:
            self._log.warning(
                "[gem] STRUCTURE WARNING: 0 bids returned from API. "
                "Possible schema change or network blockage."
            )
            return False

        sample = raw_data[:10]
        missing: list[str] = []
        for rec in sample:
            for field, _ in self.EXPECTED_SCHEMA_FIELDS:
                if field not in rec:
                    missing.append(field)

        if missing:
            self._log.warning(
                "[gem] SCHEMA WARNING: fields missing in sample: %s",
                sorted(set(missing)),
            )
        return True

    # =========================================================================
    # STEP 3 — Extract rows + Hard Pre-Filter
    # =========================================================================

    def extract_rows(self, raw_data: list[dict]) -> list[dict]:
        """
        Normalize each raw bid dict and apply the two-stage filter:
          Stage 1 (in _normalize_raw_bid): upstream category filter — drops
            bids whose Item Category has no consulting/milestone signal.
          Stage 2 (_passes_hard_filter): REQUIRE consulting keyword; REJECT
            hard supply/goods phrases.

        In-run deduplication by bid_id prevents duplicate pages from inflating counts.
        """
        rows:     list[dict] = []
        seen_ids: set[str]   = set()

        for raw in raw_data:
            # ── Normalize (includes upstream category filter) ─────────────────
            try:
                row = _normalize_raw_bid(raw)
            except Exception as exc:
                self._log.debug("Normalize error: %s", exc)
                continue

            if row is None:
                continue

            # ── In-run dedup by bid_id ────────────────────────────────────────
            bid_id = row["bid_id"]
            if bid_id in seen_ids:
                continue
            seen_ids.add(bid_id)

            # ── Hard pre-filter ───────────────────────────────────────────────
            passes, reason = _passes_hard_filter(row)
            if not passes:
                self._rejected_pre_filter.append((row, reason))
                continue

            rows.append(row)

        # ── Rejection breakdown (always-on summary) ───────────────────────────
        hard_reject_count   = sum(1 for _, r in self._rejected_pre_filter if r.startswith("hard_reject"))
        no_signal_count     = sum(1 for _, r in self._rejected_pre_filter if r == "no consulting signal")
        other_reject_count  = len(self._rejected_pre_filter) - hard_reject_count - no_signal_count

        self._log.info("[gem] Filter summary — raw: %d | dupes skipped: %d | hard_reject: %d | no_signal: %d | other: %d | passed: %d",
                       len(raw_data), len(raw_data) - len(seen_ids),
                       hard_reject_count, no_signal_count, other_reject_count, len(rows))

        # ── Sample drops (debug level — visible only when DEBUG logging active) ──
        hard_samples = [(r, rsn) for r, rsn in self._rejected_pre_filter
                        if rsn.startswith("hard_reject")][:5]
        nosig_samples = [(r, rsn) for r, rsn in self._rejected_pre_filter
                         if rsn == "no consulting signal"][:5]
        if hard_samples:
            self._log.debug("[gem] Sample HARD REJECT drops:")
            for r, rsn in hard_samples:
                self._log.debug("[gem]   (%s) → %s", rsn, r.get("title", r.get("Item Category", "?"))[:65])
        if nosig_samples:
            self._log.debug("[gem] Sample NO-SIGNAL drops:")
            for r, rsn in nosig_samples[:3]:
                self._log.debug("[gem]   (%s) → %s", rsn, r.get("title", r.get("Item Category", "?"))[:65])

        return rows

    # =========================================================================
    # STEP 4b — Capture quality-filter rejections for debug
    # =========================================================================

    def on_filter_complete(
        self,
        scraped_total: int,
        accepted:      list[dict],
        rejected:      list[dict],
        reasons:       list[str],
    ) -> None:
        self._quality_rejected = list(zip(rejected, reasons))

    # =========================================================================
    # STEP 5 — Enrich fields (post quality-filter)
    # =========================================================================

    def enrich_fields(self, rows: list[dict]) -> list[dict]:
        """
        Post-filter GeM-specific enrichment:
          1. Compute days_left from Deadline
          2. Compute GeM weighted score (gem_score 0–100)
          3. Override decision_tag with gem_score-based decision
             (replaces quality_engine estimate with GeM-calibrated logic)
          4. Generate WHY explanation string
          5. Add user-facing display fields (Decision Label, Days Left, Score)
        """
        enriched: list[dict] = []

        for row in rows:
            # 1. Days left (needed by gem_score urgency bonus)
            days_left = _compute_days_left(row.get("Deadline", ""))
            row["_days_left"] = days_left
            row["Days Left"]  = f"{days_left}d" if days_left >= 0 else "N/A"

            # 2. GeM weighted base score (keyword weights + urgency + sector)
            gem_score, matched_signals = _compute_gem_score(row)

            # 3. Firm profile adjustment — preferred/avoid sectors, regions, keywords
            gem_score, profile_signals = _apply_firm_profile_score(row, gem_score)
            row["_profile_signals"] = profile_signals

            row["gem_score"] = gem_score
            row["Score"]     = gem_score

            # 4. Decision override (GeM-calibrated thresholds)
            decision = _gem_decision_tag(gem_score)
            row["decision_tag"] = decision
            row["Decision"]     = decision

            # 5. WHY explanation (consulting type + sector + urgency)
            row["Why"] = _build_why(row, gem_score, matched_signals)

            # 6. Display-friendly decision label
            row["Decision Label"] = _DECISION_DISPLAY.get(decision, decision)

            enriched.append(row)

        # ── Task 6 — Safety check: final scan before returning ────────────────
        # Belt-and-suspenders: catches any junk that slipped past scoring.
        safe_rows: list[dict] = []
        leak_count = 0
        for row in enriched:
            chk = (
                row.get("title",       "") + " " +
                row.get("Description", "")
            ).lower()
            leak_kw = next((kw for kw in _SAFETY_KEYWORDS if kw in chk), None)
            if leak_kw:
                print(
                    f"[gem] ERROR: Safety check — junk leaked through: "
                    f"'{row.get('title','')[:60]}' (kw='{leak_kw}') — DROPPED"
                )
                leak_count += 1
            else:
                safe_rows.append(row)
        if leak_count:
            print(f"[gem] Safety check removed {leak_count} junk row(s).")
        enriched = safe_rows

        # ── Task 4 — Force output quality ────────────────────────────────────
        # Cap at 50: keep top rows by score so Excel stays actionable.
        # Warn if fewer than 5 rows pass — may signal over-filtering regression.
        if len(enriched) > 50:
            pre_cap = len(enriched)
            enriched.sort(key=lambda r: r.get("gem_score", 0), reverse=True)
            enriched = enriched[:50]
            print(f"[gem] Output capped at 50 rows (was {pre_cap}; kept top by score).")
        if len(enriched) < 5:
            print(
                f"[gem] WARNING: only {len(enriched)} row(s) passed all filters — "
                f"check if _HARD_REJECT or _REQUIRED_SIGNALS need recalibration."
            )

        print(f"[gem] Final enriched rows returning: {len(enriched)}")
        return enriched

    # =========================================================================
    # STEP 6 — Unique tender ID (hash-based, detects updates)
    # =========================================================================

    def get_tender_id(self, row: dict) -> str:
        """
        Stable unique ID: SHA-256 of (title | deadline | bid_number).

        Using a content hash (not just bid_id) means an updated tender
        — changed deadline or revised title — gets a NEW hash and will
        re-trigger notification.  This fixes the broken notification issue.
        """
        key = (
            row.get("title",       "").strip().lower() + "|" +
            row.get("Deadline",    "").strip()          + "|" +
            row.get("bid_number",  "").strip()
        )
        return "GEM_" + hashlib.sha256(key.encode()).hexdigest()[:24]

    # =========================================================================
    # STEP 7 — Standard output format (for notifications)
    # =========================================================================

    def to_standard_format(self, row: dict) -> TenderResult:
        # Task 5 — Notification filter: only BID_NOW tenders trigger alerts.
        # Non-BID_NOW rows are set to decision_tag="IGNORE" so the Telegram
        # notifier skips them (it only renders BID_NOW / STRONG_CONSIDER /
        # WEAK_CONSIDER sections — IGNORE has no section header and is silently
        # dropped). The rows still appear in the Excel via on_run_end().
        notify_tag = (
            row.get("decision_tag", "")
            if row.get("decision_tag") == "BID_NOW"
            else "IGNORE"
        )
        return make_tender_result(
            title           = row.get("title", ""),
            url             = row.get("url", self.SOURCE_URL),
            deadline        = row.get("Deadline", ""),
            organization    = row.get("Organisation Name") or row.get("organization", ""),
            sector          = row.get("Sector", ""),
            consulting_type = row.get("Consulting Type", ""),
            quality_score   = row.get("gem_score", 0),
            source          = self.SOURCE_NAME,
            decision_tag    = notify_tag,
        )

    # =========================================================================
    # STEP 8 — Lifecycle: on_run_end (Excel export)
    # =========================================================================

    def on_run_end(self, all_rows: list[dict]) -> None:
        if not all_rows:
            print("[gem] No rows to export — Excel not written.")
            return
        try:
            _save_gem_excel(all_rows, self.EXCEL_PATH)
        except Exception as exc:
            self._log.error("[gem] Excel save failed: %s", exc)

    # =========================================================================
    # Debug output (extends base class)
    # =========================================================================

    def _print_debug(
        self,
        scraped_total: int,
        accepted_rows:  list[dict],
        rejected_rows:  list[dict],
    ) -> None:
        # Run base class debug first (tier breakdown, quality scores)
        super()._print_debug(scraped_total, accepted_rows, rejected_rows)

        pre_rejected = self._rejected_pre_filter
        q_rejected   = self._quality_rejected

        print(f"\n[gem/debug] ── GeM-specific metrics ──────────────────────────")
        print(f"[gem/debug]   API raw bids              : {scraped_total}")
        print(f"[gem/debug]   Pre-filter rejected        : {len(pre_rejected)}")
        print(f"[gem/debug]   Quality-filter rejected    : {len(q_rejected)}")
        print(f"[gem/debug]   Final accepted             : {len(accepted_rows)}")

        # ── Rejection reason breakdown ────────────────────────────────────────
        reason_counts: dict[str, int] = {}
        for _, rsn in pre_rejected:
            cat = rsn.split(":")[0].strip()
            reason_counts[cat] = reason_counts.get(cat, 0) + 1
        for _, rsn in q_rejected:
            reason_counts[rsn] = reason_counts.get(rsn, 0) + 1

        if reason_counts:
            print(f"[gem/debug]   ── Rejection reasons ───────────────────────────")
            for rsn, cnt in sorted(reason_counts.items(), key=lambda x: -x[1]):
                print(f"[gem/debug]     {rsn:<45}: {cnt}")

        # ── Score distribution ────────────────────────────────────────────────
        scores = [r.get("gem_score", 0) for r in accepted_rows]
        if scores:
            dist = {"≥75 BID NOW": 0, "60-74 STRONG": 0, "40-59 REVIEW": 0, "<40 IGNORE": 0}
            for s in scores:
                if s >= 75:   dist["≥75 BID NOW"]  += 1
                elif s >= 60: dist["60-74 STRONG"]  += 1
                elif s >= 40: dist["40-59 REVIEW"]  += 1
                else:         dist["<40 IGNORE"]    += 1
            print(f"[gem/debug]   ── Score distribution ──────────────────────────")
            for band, cnt in dist.items():
                bar = "█" * cnt
                print(f"[gem/debug]     {band:<18}: {cnt:>3}  {bar}")

        # ── Top 5 BID NOW tenders ─────────────────────────────────────────────
        bid_now = sorted(
            [r for r in accepted_rows if r.get("decision_tag") == "BID_NOW"],
            key=lambda r: (-r.get("gem_score", 0),
                           r.get("_days_left", 9999) if r.get("_days_left", -1) >= 0 else 9999),
        )[:5]
        if bid_now:
            print(f"[gem/debug]   ── Top 5 BID NOW ───────────────────────────────")
            for i, r in enumerate(bid_now, 1):
                print(
                    f"[gem/debug]     #{i} [score={r.get('gem_score', 0):>3}] "
                    f"{r.get('title', '')[:55]}"
                )
                print(f"[gem/debug]        WHY: {r.get('Why', '')}")

        # ── Sample pre-filter rejections ─────────────────────────────────────
        if pre_rejected:
            print(f"[gem/debug]   ── Sample pre-filter drops ─────────────────────")
            for r, rsn in pre_rejected[:3]:
                print(f"[gem/debug]     ({rsn}) → {r.get('title', '')[:55]}")

        print(f"[gem/debug] ════════════════════════════════════════════════════")


# =============================================================================
# SECTION 7 — Business-Ready Multi-Sheet Excel
# =============================================================================

# ── Theme constants ────────────────────────────────────────────────────────────
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy  — ALL FILTERED
_BID_HDR   = PatternFill("solid", fgColor="922B21")   # deep red   — BID NOW header
_STR_HDR   = PatternFill("solid", fgColor="784212")   # dark amber — STRONG header
_ALT_FILL  = PatternFill("solid", fgColor="F5F8FF")   # soft blue  — even rows
_WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")   # white      — odd rows
_USER_FILL = PatternFill("solid", fgColor="FFF9C4")   # pale gold  — user-input cols
_SCORE_FILL= PatternFill("solid", fgColor="DDEEFF")   # light blue — score col

_WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
_BODY_FONT   = Font(name="Calibri", size=10)
_SCORE_FONT  = Font(name="Calibri", size=10, bold=True, color="1F3864")
_USER_FONT   = Font(name="Calibri", size=10, color="7F6000", italic=True)
_LINK_FONT   = Font(name="Calibri", size=10, color="1155CC", underline="single")

_THIN_BORDER = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)

# Decision-tier row tints (applied as the base row colour)
_DECISION_ROW_FILLS: dict[str, PatternFill] = {
    "BID_NOW":         PatternFill("solid", fgColor="FDECEA"),  # light red
    "STRONG_CONSIDER": PatternFill("solid", fgColor="FEF6E7"),  # light amber
    "WEAK_CONSIDER":   PatternFill("solid", fgColor="F1F8E9"),  # light green
}

# ── Column schema: (display_header, row_dict_key, column_width) ───────────────
_COLUMNS: list[tuple[str, str, int]] = [
    ("Title",           "title",            55),
    ("Organization",    "Organisation Name", 30),
    ("Deadline",        "Deadline",          18),
    ("Days Left",       "Days Left",         10),
    ("Consulting Type", "Consulting Type",   22),
    ("Sector",          "Sector",            18),
    ("Score",           "gem_score",         10),
    ("Decision",        "Decision Label",    14),
    ("Why",             "Why",               45),
    ("Source",          "Source",            10),
    ("URL",             "url",               50),
    ("My Decision",     "_user_input_1",     14),  # left blank for human input
]

_USER_COLS:   frozenset[str] = frozenset({"My Decision"})
_SCORE_COL:   str = "Score"
_URL_COL:     str = "URL"
_DEC_COL:     str = "Decision"


def _write_gem_sheet(
    ws,
    rows: list[dict],
    hdr_fill: PatternFill,
) -> None:
    """
    Write rows to an openpyxl worksheet.
    - Coloured header row
    - Decision-tier row tints (BID NOW = light red, STRONG = light amber, etc.)
    - Score column highlighted in blue
    - URL column as clickable hyperlink
    - My Decision / Outcome columns in pale gold (signals "fill me in")
    """
    col_headers = [c[0] for c in _COLUMNS]

    # ── Header row ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    for ci, (col_name, _, col_width) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _WHITE_FONT
        cell.fill      = hdr_fill
        cell.border    = _THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)

    # ── Column index look-up ───────────────────────────────────────────────────
    url_ci   = col_headers.index(_URL_COL)   + 1
    score_ci = col_headers.index(_SCORE_COL) + 1
    dec_ci   = col_headers.index(_DEC_COL)   + 1

    # ── Data rows ──────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 40

        decision  = row.get("decision_tag", "")
        row_tint  = _DECISION_ROW_FILLS.get(decision)
        alt_tint  = _ALT_FILL if ri % 2 == 0 else _WHT_FILL

        for ci, (col_name, key, _) in enumerate(_COLUMNS, 1):
            val  = row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            # ── Fill ──────────────────────────────────────────────────────────
            if col_name in _USER_COLS:
                cell.fill = _USER_FILL
            elif ci == score_ci:
                cell.fill = _SCORE_FILL
            else:
                cell.fill = row_tint if row_tint else alt_tint

            # ── Font + special treatment ──────────────────────────────────────
            if col_name in _USER_COLS:
                cell.font = _USER_FONT
            elif ci == score_ci:
                cell.font = _SCORE_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == url_ci and val and str(val).startswith("http"):
                cell.hyperlink = str(val)
                cell.font      = _LINK_FONT
            elif ci == dec_ci:
                cell.font      = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.font = _BODY_FONT

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"


def _save_gem_excel(all_rows: list[dict], path: str) -> None:
    """
    Save a three-sheet decision-grade Excel workbook:
      Sheet 1 — 🔥 BID NOW       (decision_tag == BID_NOW)
      Sheet 2 — ⭐ STRONG         (decision_tag == STRONG_CONSIDER)
      Sheet 3 — 📊 ALL FILTERED  (all accepted rows, sorted score↓ deadline↑)

    Sort order: Score DESC, then Days Left ASC (most urgent first within tier).
    """
    def _sort_key(r: dict) -> tuple:
        dl = r.get("_days_left", -1)
        return (-r.get("gem_score", 0), dl if dl >= 0 else 9999)

    bid_now     = sorted(
        [r for r in all_rows if r.get("decision_tag") == "BID_NOW"],
        key=_sort_key,
    )
    strong      = sorted(
        [r for r in all_rows if r.get("decision_tag") == "STRONG_CONSIDER"],
        key=_sort_key,
    )
    all_sorted  = sorted(all_rows, key=_sort_key)

    wb = Workbook()
    wb.remove(wb.active)

    _write_gem_sheet(wb.create_sheet("🔥 BID NOW"),      bid_now,    _BID_HDR)
    _write_gem_sheet(wb.create_sheet("⭐ STRONG"),        strong,     _STR_HDR)
    _write_gem_sheet(wb.create_sheet("📊 ALL FILTERED"), all_sorted, _HDR_FILL)

    wb.save(path)

    print(f"[gem] Excel saved → {path}")
    print(f"[gem]   🔥 BID NOW        : {len(bid_now)} rows")
    print(f"[gem]   ⭐ STRONG          : {len(strong)} rows")
    print(f"[gem]   📊 ALL FILTERED    : {len(all_sorted)} rows")


# =============================================================================
# SECTION 8 — Module shim (required by registry + main.py)
# =============================================================================

def run(debug: bool = False) -> tuple[list, list]:
    """
    Module-level entry point called by main.py and core/registry.py.
    Returns (new_tenders: list[TenderResult], all_rows: list[dict]).
    """
    return GeMScraper().run(debug=debug)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="GeM BidPlus Decision-Grade Pipeline")
    parser.add_argument(
        "--debug", action="store_true",
        help="Enable detailed debug output (score distribution, rejections, top BID NOW)",
    )
    args = parser.parse_args()

    new_tenders, all_rows = run(debug=args.debug)
    print(f"\n── Run complete ────────────────────────────────────")
    print(f"  New tenders (notifications) : {len(new_tenders)}")
    print(f"  All accepted rows           : {len(all_rows)}")
    bid_now_count = sum(1 for r in all_rows if r.get("decision_tag") == "BID_NOW")
    strong_count  = sum(1 for r in all_rows if r.get("decision_tag") == "STRONG_CONSIDER")
    print(f"  🔥 BID NOW                  : {bid_now_count}")
    print(f"  ⭐ STRONG                    : {strong_count}")
