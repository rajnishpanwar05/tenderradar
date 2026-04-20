# =============================================================================
# meghalaya_pipeline.py — Meghalaya Basin Development Authority Tenders
#
# Site   : https://mbda.gov.in/tenders
# Login  : None required (public site)
# CAPTCHA: None
# Method : requests + BeautifulSoup (Drupal site, server-side rendered)
#
# Filter strategy:
#   • Skip pure goods/procurement tenders (RFQ for equipment, furniture, etc.)
#   • Keep: RFP, EOI, Empanelment, Consultancy, Agency selection tenders
#   • Score with shared IDCG keyword bank
#   • Notify only if relevance score ≥ 1 category (geo-specific, lower bar than UNDP)
#
# Run with: python3 main.py --meghalaya
# Included in default run automatically.
# =============================================================================

import os
import re
import time

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://mbda.gov.in"
LIST_URL    = f"{BASE_URL}/tenders"
MAX_PAGES   = 10    # 20 tenders/page × 10 pages = 200 tenders max
DELAY       = 0.8   # seconds between requests
# mbda.gov.in is a slow government server — use higher timeouts + retry
CONNECT_TIMEOUT = 12   # seconds to establish connection
READ_TIMEOUT    = 45   # seconds to read response (server is slow)
MAX_RETRIES     = 3    # retry count on connection/read errors

MEGHALAYA_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "Meghalaya_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── Goods/procurement title patterns to SKIP ──────────────────────────────────
# These are supply/equipment tenders — not relevant to IDCG
SKIP_TITLE_PATTERNS = [
    "procurement of office",
    "procurement of furniture",
    "procurement of equipment",
    "procurement of vehicle",
    "procurement of computer",
    "procurement of stationery",
    "supply of",
    "purchase of",
    "rfq for procurement",
    "request for bids (rfb) for procurement",
    "request for bids for procurement",
    "rfb for procurement",
    "rate contract for",
    "empanelment of vendors",
    "empanelment of suppliers",
    "empanelment of printers",
    "annual maintenance contract",
    "amc for",
]

# ── Excel styles ───────────────────────────────────────────────────────────────
# Note: PDF Links column removed per Phase 8 data output requirements.
# PDFs are not downloaded locally. Detail Link gives direct access to source page
# which contains the tender documents. PDF URLs remain in Description text.
MASTER_COLUMNS = [
    ("Title",        70),
    ("Date Posted",  18),
    ("Description",  65),
    ("Relevance",    40),
    ("Detail Link",  55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1B4F3A")   # dark green — Meghalaya theme
ALT_FILL       = PatternFill("solid", fgColor="F0FFF4")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="C6E0C4"),
    right=Side(style="thin", color="C6E0C4"),
    top=Side(style="thin", color="C6E0C4"),
    bottom=Side(style="thin", color="C6E0C4"),
)


# =============================================================================
# Helpers
# =============================================================================

def _is_goods_tender(title: str) -> bool:
    """Return True if this is a supply/goods/equipment tender — skip for IDCG."""
    t = title.lower().strip()
    return any(t.startswith(pat) or pat in t for pat in SKIP_TITLE_PATTERNS)


def _fetch_listing_page(session: requests.Session, page: int) -> list:
    """
    Fetch one page of tender listings. Returns list of (title, date, detail_url).
    Retries on connection/read timeout — mbda.gov.in is a slow government server.
    """
    url = f"{LIST_URL}?page={page}"
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, headers=HEADERS,
                            timeout=(CONNECT_TIMEOUT, READ_TIMEOUT))
            r.raise_for_status()
            break
        except Exception as e:
            last_err = e
            if attempt < MAX_RETRIES:
                wait = attempt * 3
                print(f"[meghalaya] Page {page} attempt {attempt} failed: {e} — retrying in {wait}s",
                      flush=True)
                time.sleep(wait)
    else:
        print(f"[meghalaya] Listing page {page} failed after {MAX_RETRIES} attempts: {last_err}",
              flush=True)
        return []

    soup = BeautifulSoup(r.text, "html.parser")

    # NOTE: #maincontent is an empty skip-anchor div on this Drupal site.
    # Tier 1: Drupal 7 — li.views-row containing div.views-field-php
    all_li = soup.find_all(
        "li",
        class_=lambda c: c and "views-row" in c,
    )
    tender_li = [
        li for li in all_li
        if li.find("div", class_=lambda c: c and "views-field-php" in c)
    ]

    # Tier 2: broader Drupal/CMS selectors (Drupal 8/9/10 class names)
    if not tender_li:
        candidates = (
            soup.find_all("article") or
            soup.find_all("li", class_=lambda c: c and any(
                k in c for k in ("views", "tender", "node", "item"))) or
            soup.find_all("div", class_=lambda c: c and any(
                k in c for k in ("views-row", "tender-item",
                                 "node--type-tender", "field-content")))
        )
        tender_li = [el for el in candidates
                     if el.find("a", href=True) and len(el.get_text(strip=True)) > 20]

    # Tier 3: bare anchor links to /node/ or /tender paths
    if not tender_li:
        tender_li = [
            a.parent for a in soup.find_all(
                "a", href=re.compile(r"/node/\d+|/tender", re.I))
            if len(a.get_text(strip=True)) > 8
        ]

    results = []
    for li in tender_li:
        a = li.find("a", href=True)
        if not a:
            continue
        title = a.get_text(strip=True)
        if not title or len(title) < 8:
            continue

        href = a["href"]
        detail_url = (f"{BASE_URL}{href}"
                      if href.startswith("/") else href)

        # Date: Drupal 7 span → HTML5 <time> → regex fallback
        date_span = (
            li.find("span", class_="date-display-single") or
            li.find("time")
        )
        if date_span:
            date_text = (date_span.get("datetime", "") or
                         date_span.get_text(strip=True))
        else:
            li_text = li.get_text(" ", strip=True)
            date_m  = re.search(
                r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{4}"
                r"|\d{1,2}\s+\w+\s+\d{4}"
                r"|\w+\s+\d{1,2},?\s+\d{4})",
                li_text,
            )
            date_text = date_m.group(1) if date_m else ""

        results.append((title, date_text, detail_url))

    return results


def _fetch_detail(session: requests.Session, url: str) -> tuple:
    """
    Fetch detail page. Returns (description, pdf_links_str).
    pdf_links_str is newline-separated list of full PDF URLs (not downloaded — links only).
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, headers=HEADERS,
                            timeout=(CONNECT_TIMEOUT, READ_TIMEOUT))
            r.raise_for_status()
            break
        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(attempt * 2)
            else:
                print(f"[meghalaya]   Detail error ({url}): {e}", flush=True)
                return "", ""
    try:
        soup = BeautifulSoup(r.text, "html.parser")

        # Remove nav/header/footer noise
        for tag in soup.find_all(["nav", "header", "footer", "script", "style"]):
            tag.decompose()

        # Extract PDF links
        pdf_links = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if ".pdf" in href.lower() or ".doc" in href.lower():
                full = (f"{BASE_URL}{href}"
                        if href.startswith("/") else href)
                if full not in pdf_links:
                    pdf_links.append(full)

        # Extract main description text
        description = ""
        for sel in ["div.content", "div#content", "main", "article", "div.field-items"]:
            el = soup.select_one(sel)
            if el:
                text = el.get_text(" ", strip=True)
                if len(text) > 80:
                    description = text[:2000]
                    break

        if not description:
            description = soup.get_text(" ", strip=True)[:2000]

        return description, "\n".join(pdf_links)

    except Exception as e:
        print(f"[meghalaya]   Detail error ({url}): {e}", flush=True)
        return "", ""


# =============================================================================
# Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MBDA Tenders"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes  = ws.cell(row=2, column=1)
    relevance_idx    = col_names.index("Relevance") + 1
    link_idx         = col_names.index("Detail Link") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill:
                        cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill:
                    cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(MEGHALAYA_EXCEL_PATH)
    print(f"[meghalaya] Excel saved: {MEGHALAYA_EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[meghalaya] Meghalaya Basin Dev Authority Pipeline starting...", flush=True)
    print("[meghalaya] Site: mbda.gov.in/tenders | No login required", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []

    if os.path.exists(MEGHALAYA_EXCEL_PATH):
        try:
            os.remove(MEGHALAYA_EXCEL_PATH)
        except Exception:
            pass

    session = requests.Session()

    # ── Fix 3: Connectivity pre-check — fail fast if site is unreachable ──────
    # mbda.gov.in is a slow government server. A quick HEAD request with a short
    # timeout tells us immediately if the site is down, rather than burning 3 full
    # retry cycles (each up to 45s) per listing page before giving up.
    reachable = False
    try:
        probe = session.head(LIST_URL, headers=HEADERS, timeout=(12, 20))
        probe.raise_for_status()
        reachable = True
        print("[meghalaya] Site reachable (HEAD) — proceeding.", flush=True)
    except requests.exceptions.ReadTimeout:
        try:
            probe2 = session.get(LIST_URL, headers=HEADERS,
                                 timeout=(12, 15), stream=True)
            probe2.close()
            reachable = True
            print("[meghalaya] HEAD timed out but GET connected — proceeding.", flush=True)
        except Exception as e2:
            print(f"[meghalaya] Site unreachable after HEAD+GET probe ({e2}) — aborting early.",
                  flush=True)
    except Exception as e:
        print(f"[meghalaya] Site unreachable ({e}) — aborting early.", flush=True)
    if not reachable:
        return [], []

    # ── Scrape all listing pages ───────────────────────────────────────────────
    all_listings = []
    for page in range(MAX_PAGES):
        items = _fetch_listing_page(session, page)
        if not items:
            print(f"[meghalaya] No items on page {page} — stopping.", flush=True)
            break
        print(f"[meghalaya] Page {page}: {len(items)} tenders", flush=True)
        all_listings.extend(items)
        time.sleep(DELAY)

    print(f"[meghalaya] Total listings: {len(all_listings)}", flush=True)

    # ── Filter 1: goods/equipment tenders ─────────────────────────────────────
    service_listings = [
        (title, date, url) for title, date, url in all_listings
        if not _is_goods_tender(title)
    ]
    skipped_goods = len(all_listings) - len(service_listings)

    # ── Fix 2: Within-run URL dedup — pagination sometimes overlaps ───────────
    # Build a deduplicated list while preserving first-seen order.
    seen_urls: set = set()
    unique_listings = []
    for item in service_listings:
        if item[2] not in seen_urls:
            seen_urls.add(item[2])
            unique_listings.append(item)
    skipped_dup = len(service_listings) - len(unique_listings)

    print(
        f"[meghalaya] After filters: {len(unique_listings)} unique service tenders "
        f"(skipped {skipped_goods} goods, {skipped_dup} pagination duplicates)",
        flush=True,
    )

    # ── Fix 1: Lazy detail fetching — only fetch pages for NEW tenders ─────────
    # For seen tenders we already have their data in the DB / previous Excel.
    # Fetching the detail page again adds nothing — skip the HTTP call entirely.
    # On a steady-state run (0 new tenders) this reduces runtime from ~90 min
    # to under 2 minutes.
    new_count = skipped_count = 0
    print("[meghalaya] Processing tenders (detail fetch only for new ones)...", flush=True)

    for i, (title, date, detail_url) in enumerate(unique_listings, 1):
        slug      = detail_url.replace(BASE_URL, "").strip("/")
        tender_id = f"MBDA/{slug[:100]}"

        if not check_if_new(tender_id):
            # Already seen — score from title alone, skip HTTP
            relevance = score_relevance(title, "")
            row = {
                "Title":       title,
                "Date Posted": date,
                "Description": "",
                "Relevance":   relevance,
                "Detail Link": detail_url,
            }
            all_rows.append(row)
            skipped_count += 1
            if i % 20 == 0:
                print(f"[meghalaya]   ... {i}/{len(unique_listings)} seen (fast-skip)", flush=True)
            continue

        # New tender — fetch full detail page
        print(f"[meghalaya]   [{i:>2}/{len(unique_listings)}] NEW → {title[:65]}", flush=True)
        description, _pdf_links = _fetch_detail(session, detail_url)
        time.sleep(DELAY)

        relevance = score_relevance(title, description)
        row = {
            "Title":       title,
            "Date Posted": date,
            "Description": description[:1500] if description else "",
            "Relevance":   relevance,
            "Detail Link": detail_url,
        }
        all_rows.append(row)

        mark_as_seen(
            tender_id,
            title=title,
            source_site="MBDA",
            url=detail_url,
        )
        new_count += 1

        if relevance:
            new_tenders.append({
                "title":    title,
                "deadline": date,
                "value":    "",
                "url":      detail_url,
            })
            print(f"[meghalaya]   → ALERT | {relevance[:60]}", flush=True)
        else:
            print(f"[meghalaya]   → saved (no keyword match)", flush=True)

    # ── Save Excel ─────────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(
        f"\n[meghalaya] Done — {len(all_rows)} tenders "
        f"({new_count} new, {skipped_count} fast-skipped), "
        f"{len(new_tenders)} NEW+relevant, {relevant} total relevant",
        flush=True,
    )
    return new_tenders, all_rows
