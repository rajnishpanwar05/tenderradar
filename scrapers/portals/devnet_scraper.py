# =============================================================================
# devnet_pipeline.py — DevNet Jobs India RFP/Tender Pipeline
#
# What this does:
#   1. Fetches the RFP listing page from devnetjobsindia.org
#   2. Parses all visible RFP/Tender entries (title, org, location, deadline)
#   3. Simulates ASP.NET __doPostBack() calls to fetch each detail page
#   4. Extracts full description + document links
#   5. Scores each listing for relevance against firm expertise
#   6. Deduplicates via MySQL (only NEW tenders trigger notifications)
#   7. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import DEVNET_EXCEL_PATH, LOG_FILE
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import FIRM_EXPERTISE

# ── Constants ─────────────────────────────────────────────────────────────────
BASE_URL    = "https://devnetjobsindia.org"
LISTING_URL = "https://devnetjobsindia.org/rfp_assignments.aspx"
HEADERS     = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer":         LISTING_URL,
    "Accept-Language": "en-US,en;q=0.9",
}
DELAY = 1.5   # seconds between requests — be polite to the server


# =============================================================================
# SECTION 1 — Relevance Scoring  (FIRM_EXPERTISE imported from keywords.py)
# =============================================================================

def _score(title: str, description: str):
    """Return (score, matched_categories_string)."""
    text    = (title + " " + description).lower()
    matched = [cat for cat, kws in FIRM_EXPERTISE.items()
               if any(kw in text for kw in kws)]
    score   = len(matched)
    return score, (", ".join(matched) if matched else "")


# =============================================================================
# SECTION 2 — ASP.NET Session + Page Fetching
# =============================================================================

def _get_aspnet_fields(soup: BeautifulSoup) -> dict:
    fields = {}
    for fid in ("__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION",
                "__VIEWSTATEENCRYPTED", "__SCROLLPOSITIONX", "__SCROLLPOSITIONY"):
        tag = soup.find("input", {"id": fid})
        if tag:
            fields[fid] = tag.get("value", "")
    return fields


def _fetch_listing(session: requests.Session) -> BeautifulSoup:
    r = session.get(LISTING_URL, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")


def _postback(session: requests.Session, aspnet_fields: dict,
              event_target: str):
    form_data = {**aspnet_fields, "__EVENTTARGET": event_target, "__EVENTARGUMENT": ""}
    try:
        r = session.post(LISTING_URL, data=form_data, headers=HEADERS, timeout=30)
        r.raise_for_status()
        return BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"    [devnet] Postback error for {event_target}: {e}")
        return None


# =============================================================================
# SECTION 3 — Parse Listing + Detail Pages
# =============================================================================

def _parse_listings(soup: BeautifulSoup) -> list[dict]:
    entries = []
    grid = soup.find("table", id=re.compile(r"grdJobs", re.I))
    if not grid:
        grid = soup.find("table")
    if not grid:
        print("[devnet] WARNING: Could not locate the RFP grid table.")
        return entries

    for row in grid.find_all("tr"):
        link = row.find("a", href=re.compile(r"__doPostBack", re.I))
        if not link:
            continue
        href  = link.get("href", "")
        match = re.search(r"__doPostBack\('([^']+)'", href)
        if not match:
            continue
        event_target = match.group(1)
        title        = link.get_text(strip=True)

        # Skip entries with no title — some DevNet rows have an anchor tag
        # but no visible text (image-only links, empty anchors, etc.)
        if not title or len(title) < 5:
            continue

        row_text     = row.get_text(" | ", strip=True)

        org_match      = re.search(r"\|\s*(.+?)\s*\|\s*Location:", row_text)
        loc_match      = re.search(r"Location:\s*(.+?)\s*\|\s*Apply by:", row_text)
        deadline_match = re.search(r"Apply by:\s*(\d{1,2}\s+\w+\s+\d{4})", row_text)

        entries.append({
            "Title":        title,
            "EventTarget":  event_target,
            "TenderID":     f"DEVNET/{event_target.replace('$', '_')}",
            "Organisation": org_match.group(1).strip()      if org_match      else "",
            "Location":     loc_match.group(1).strip()      if loc_match      else "",
            "Deadline":     deadline_match.group(1).strip() if deadline_match else "",
        })
    return entries


def _parse_detail(soup: BeautifulSoup) -> dict:
    detail = {"Description": "", "Document Links": ""}
    for selector in [
        {"id":    re.compile(r"pnlDetail|divDetail|JobDetail|rfpDetail", re.I)},
        {"class": re.compile(r"job.?detail|rfp.?detail|content.?area",  re.I)},
    ]:
        panel = soup.find(attrs=selector)
        if panel:
            detail["Description"] = panel.get_text(" ", strip=True)
            doc_links = []
            for a in panel.find_all("a", href=True):
                href = a["href"]
                if any(ext in href.lower() for ext in (".pdf", ".doc", ".docx", ".zip")):
                    full = href if href.startswith("http") else BASE_URL + "/" + href.lstrip("/")
                    doc_links.append(full)
            detail["Document Links"] = "\n".join(doc_links)
            break

    if not detail["Description"]:
        content = soup.find("div", {"id": re.compile(r"ContentPlaceHolder1", re.I)})
        if content:
            full_text = content.get_text(" ", strip=True)
            parts     = re.split(r"Requests for Proposals and Tenders", full_text, flags=re.I)
            detail["Description"] = parts[-1][:2000].strip() if len(parts) > 1 else ""

    return detail


# =============================================================================
# SECTION 4 — Excel Writer
# =============================================================================

MASTER_COLUMNS = [
    ("Title",           60),
    ("Organisation",    35),
    ("Location",        25),
    ("Deadline",        18),
    ("Description",     70),
    ("Document Links",  45),
    ("Relevance",       40),
    ("Source URL",      45),
]

HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
HIGH_FILL      = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


def _save_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DevNet RFPs"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes    = ws.cell(row=2, column=1)
    relevance_idx      = col_names.index("Relevance") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if col_name == "Source URL" and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font      = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = HIGH_FILL
                    cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt_fill:
                        cell.fill = alt_fill
            else:
                cell.font = BODY_FONT
                if alt_fill:
                    cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(DEVNET_EXCEL_PATH)
    print(f"[devnet] Excel saved: {DEVNET_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 5 — Main run() — called by main.py
# =============================================================================

def run() -> tuple:
    """
    Run the DevNet RFP pipeline.
    Returns:
        new_tenders — list of dicts for Telegram notification (new only)
        all_rows    — all fetched rows
    """
    print("\n" + "=" * 65)
    print("[devnet] DevNet Jobs India — RFP Pipeline starting...")
    print("=" * 65)

    new_tenders = []
    all_rows    = []

    # ── Clear old Excel for a fresh run ───────────────────────────────────────
    if os.path.exists(DEVNET_EXCEL_PATH):
        try:
            os.remove(DEVNET_EXCEL_PATH)
            print(f"[devnet] Cleared old Excel: {DEVNET_EXCEL_PATH}")
        except Exception as e:
            print(f"[devnet] Could not delete old Excel: {e}")

    # ── Fetch listing page ─────────────────────────────────────────────────────
    print("[devnet] Fetching listing page...")
    try:
        session       = requests.Session()
        session.headers.update(HEADERS)
        soup          = _fetch_listing(session)
        aspnet_fields = _get_aspnet_fields(soup)
        if not aspnet_fields.get("__VIEWSTATE"):
            print("[devnet] WARNING: No VIEWSTATE found — postbacks may fail.")
    except Exception as e:
        print(f"[devnet] ERROR fetching listing page: {e}")
        return new_tenders, all_rows

    # ── Parse all RFP rows ─────────────────────────────────────────────────────
    entries = _parse_listings(soup)
    print(f"[devnet] Found {len(entries)} RFP listings")
    if not entries:
        print("[devnet] No listings found — page structure may have changed.")
        return new_tenders, all_rows

    # ── Fetch details + score + deduplicate ───────────────────────────────────
    print("[devnet] Fetching details for each listing...")
    for i, entry in enumerate(entries, 1):
        print(f"[devnet]   [{i:>3}/{len(entries)}] {entry['Title'][:60]}...")

        detail_soup   = _postback(session, aspnet_fields, entry["EventTarget"])
        description   = ""
        document_links = ""

        if detail_soup:
            parsed         = _parse_detail(detail_soup)
            description    = parsed.get("Description",   "")
            document_links = parsed.get("Document Links", "")
            new_fields     = _get_aspnet_fields(detail_soup)
            if new_fields.get("__VIEWSTATE"):
                aspnet_fields = new_fields

        score, relevance = _score(entry["Title"], description)

        row = {
            "Title":          entry["Title"],
            "Organisation":   entry["Organisation"],
            "Location":       entry["Location"],
            "Deadline":       entry["Deadline"],
            "Description":    description[:1500] if description else "",
            "Document Links": document_links,
            "Relevance":      relevance,
            "Source URL":     LISTING_URL,
        }
        all_rows.append(row)

        # ── DB deduplication ──────────────────────────────────────────────────
        tender_id = entry["TenderID"]
        if check_if_new(tender_id):
            mark_as_seen(tender_id, title=entry["Title"], source_site="DevNet", url=LISTING_URL)
            new_tenders.append({
                "title":    entry["Title"],
                "deadline": entry["Deadline"],
                "value":    "See listing",
                "url":      LISTING_URL,
            })
            print(f"           → NEW | Relevance: {relevance or '—'}")
        else:
            print(f"           → seen | Relevance: {relevance or '—'}")

        time.sleep(DELAY)

    # ── Save Excel ─────────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)

    relevant_count = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[devnet] Done — {len(all_rows)} listings, {len(new_tenders)} NEW, "
          f"{relevant_count} relevant")

    return new_tenders, all_rows
