# =============================================================================
# iucn_pipeline.py — IUCN Procurement Tenders Pipeline
#
# Site   : https://iucn.org/procurement/currently-running-tenders
# Method : requests + BeautifulSoup (simple server-rendered HTML tables)
# Login  : None required
# CAPTCHA: None
#
# Table structure (2 tables on page):
#   Table 0 — "currently open" tenders          ← we scrape this
#   Table 1 — "under evaluation" (past deadline) ← skipped
#
# Columns per row:
#   [0] Submission Deadline
#   [1] RfP Title and Link  (href = PDF or detail page)
#   [2] IUCN Office (lead)
#   [3] Country of Performance
#   [4] Expected Contract Duration
#   [5] Estimated Contract Value
#
# Filter strategy:
#   • Take ALL open tenders (IUCN only posts consulting/RFP opportunities)
#   • Score relevance — India/South Asia or international M&E work gets high score
#   • Notify on any new tender regardless of relevance (volume is manageable ~95/run)
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://iucn.org"
LISTING_URL = f"{BASE_URL}/procurement/currently-running-tenders"
DELAY       = 0.3   # polite crawl — all data is on one page, this is for detail fetches

IUCN_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "IUCN_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
}

# India and South Asia terms for priority flagging
INDIA_SA_TERMS = [
    "india", "indian", "south asia", "south asian", "new delhi", "delhi",
    "mumbai", "bangalore", "bengaluru", "nepal", "bangladesh", "sri lanka",
    "pakistan", "bhutan", "maldives",
]


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("RfP Title",    65),
    ("IUCN Office",  30),
    ("Country",      25),
    ("Deadline",     18),
    ("Duration",     22),
    ("Budget",       22),
    ("Relevance",    40),
    ("India/SA",     10),
    ("PDF / Link",   55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
INDIA_FILL     = PatternFill("solid", fgColor="BDD7EE")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
INDIA_FONT     = Font(name="Calibri", size=10, color="1F3864", bold=True)
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse open tenders table
# =============================================================================

def _is_india_sa(text: str) -> bool:
    t = text.lower()
    return any(term in t for term in INDIA_SA_TERMS)

def _parse_open_tenders(soup: BeautifulSoup) -> list:
    """Parse Table 0 (currently open tenders). Returns list of dicts."""
    entries = []
    tables = soup.find_all('table')
    if not tables:
        print("[iucn] WARNING: No tables found on page.")
        return entries

    # Table 0 is always the "currently open" table
    open_table = tables[0]
    rows = open_table.find_all('tr')

    # Row 0 is the header row — skip it
    for row in rows[1:]:
        cells = row.find_all('td')
        if len(cells) < 2:
            continue

        deadline = cells[0].get_text(strip=True) if len(cells) > 0 else ""
        office   = cells[2].get_text(strip=True) if len(cells) > 2 else ""
        country  = cells[3].get_text(strip=True) if len(cells) > 3 else ""
        duration = cells[4].get_text(strip=True) if len(cells) > 4 else ""
        budget   = cells[5].get_text(strip=True) if len(cells) > 5 else ""

        # Title and link from col 1
        title_cell = cells[1]
        title_link = title_cell.find('a')
        title      = title_cell.get_text(strip=True)
        link       = ""
        if title_link:
            href = title_link.get('href', '').strip()
            if href:
                link = href if href.startswith('http') else BASE_URL + href
            # Title text might be the link text or the full cell text
            title = title_link.get_text(strip=True) or title

        if not title or len(title) < 5:
            continue

        is_india = _is_india_sa(country + " " + office + " " + title)
        tender_id = f"IUCN_{re.sub(r'[^a-zA-Z0-9]', '_', (link or title)[:80])}"

        entries.append({
            "RfP Title":   title,
            "IUCN Office": office,
            "Country":     country,
            "Deadline":    deadline,
            "Duration":    duration,
            "Budget":      budget,
            "PDF / Link":  link,
            "is_india":    is_india,
            "tender_id":   tender_id,
        })

    return entries


# =============================================================================
# SECTION 2 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    # Sort: India/SA first, then by deadline
    rows_sorted = sorted(rows, key=lambda r: (0 if r.get("is_india") else 1, r.get("Deadline", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "IUCN Open Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx    = col_names.index("Relevance")  + 1
    india_idx  = col_names.index("India/SA")   + 1
    link_idx   = col_names.index("PDF / Link") + 1

    for ri, row_data in enumerate(rows_sorted, 2):
        ws.row_dimensions[ri].height = 45
        alt       = ALT_FILL if ri % 2 == 0 else None
        is_india  = row_data.get("is_india", False)

        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
                if alt: cell.fill = alt
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            elif ci == india_idx:
                if is_india:
                    cell.value = "YES"
                    cell.fill  = INDIA_FILL
                    cell.font  = INDIA_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.value = ""
                    cell.font  = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(IUCN_EXCEL_PATH)
    print(f"[iucn] Excel saved: {IUCN_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 3 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the IUCN Procurement pipeline.
    Returns:
        new_tenders — list of dicts for notification (new only)
        all_rows    — all open tenders
    """
    print("\n" + "=" * 65)
    print("[iucn] IUCN Procurement Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(IUCN_EXCEL_PATH):
        try:
            os.remove(IUCN_EXCEL_PATH)
            print(f"[iucn] Cleared old Excel")
        except Exception:
            pass

    # ── Fetch page ─────────────────────────────────────────────────────────────
    try:
        r = requests.get(LISTING_URL, headers=HEADERS, timeout=25)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
    except Exception as e:
        print(f"[iucn] ERROR fetching page: {e}")
        return [], []

    entries = _parse_open_tenders(soup)
    print(f"[iucn] Open tenders found: {len(entries)}")

    if not entries:
        print("[iucn] No open tenders — page structure may have changed.")
        return [], []

    india_count = sum(1 for e in entries if e["is_india"])
    print(f"[iucn]   India/South Asia: {india_count} | Other: {len(entries) - india_count}")

    # ── Score relevance and deduplicate ───────────────────────────────────────
    for i, entry in enumerate(entries, 1):
        title     = entry["RfP Title"]
        relevance = score_relevance(title, entry["Country"] + " " + entry["IUCN Office"])

        row = {
            "RfP Title":   title,
            "IUCN Office": entry["IUCN Office"],
            "Country":     entry["Country"],
            "Deadline":    entry["Deadline"],
            "Duration":    entry["Duration"],
            "Budget":      entry["Budget"],
            "Relevance":   relevance,
            "India/SA":    "YES" if entry["is_india"] else "",
            "PDF / Link":  entry["PDF / Link"],
            "is_india":    entry["is_india"],
        }
        all_rows.append(row)

        tid = entry["tender_id"]
        if check_if_new(tid):
            mark_as_seen(tid, title=title, source_site="IUCN",
                         url=entry["PDF / Link"] or LISTING_URL)
            new_tenders.append({
                "title":    title,
                "deadline": entry["Deadline"],
                "value":    entry["Budget"] or "See listing",
                "url":      entry["PDF / Link"] or LISTING_URL,
            })

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"[iucn]   Relevant: {relevant} | NEW: {len(new_tenders)}")

    if all_rows:
        _save_excel(all_rows)

    print(f"\n[iucn] Done — {len(all_rows)} open tenders, {len(new_tenders)} NEW")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
