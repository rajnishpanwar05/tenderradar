from __future__ import annotations

# =============================================================================
# adb_scraper.py — Asian Development Bank (ADB) Procurement Notices  v2
#
# Source  : https://www.adb.org/projects/tenders
# Method  : POST to Drupal views/ajax endpoint (confirmed by API research)
#           → returns JSON array with "insert" command containing HTML fragment
#           → parse with BeautifulSoup
#
# Verified endpoint (stable since 2022, Drupal 9/10):
#   POST https://www.adb.org/views/ajax
#   Content-Type: application/x-www-form-urlencoded
#   Form fields: view_name, view_display_id, page
#
# Note: ADB has no public REST API for live tenders.
#   - selfservice.adb.org → authenticated Oracle CMS (EOI submission only)
#   - data.adb.org        → awarded contracts history only (CSV/Excel)
#   - views/ajax          → the ONLY way to get live open tenders programmatically
#
# Targets: TA projects, consulting services, evaluation/advisory notices
#
# SCRAPER_META (auto-registered via core/registry.py):
SCRAPER_META = {
    "flag":        "adb",
    "label":       "ADB (Asian Dev Bank)",
    "group":       "requests",
    "timeout":     240,
    "max_retries": 2,
    "auto":        True,
}
# =============================================================================

import json
import os
import random
import re
import time
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Constants ─────────────────────────────────────────────────────────────────
ADB_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "ADB_Tenders_Master.xlsx")
CACHE_PATH     = os.path.join(os.path.dirname(ADB_EXCEL_PATH), ".adb_cache.json")

SOURCE_NAME  = "ADB"
BASE_URL     = "https://www.adb.org"
AJAX_URL     = "https://www.adb.org/views/ajax"
TENDERS_URL  = "https://www.adb.org/projects/tenders"
MAX_PAGES    = 10    # 10 pages × ~20 items ≈ 200 notices
PAGE_PAUSE   = 2.0   # polite delay between pages

# View configurations for different tender types
_VIEW_CONFIGS = [
    # Main tenders/procurement page
    {
        "view_name":       "tenders",
        "view_display_id": "page_1",
        "label":           "Active Tenders",
    },
    # Technical assistance projects
    {
        "view_name":       "projects_procurement",
        "view_display_id": "block_procurement",
        "label":           "TA/Procurement",
    },
]

_MAX_RETRIES     = 2
_RETRY_BASE_WAIT = 2

_USER_AGENTS = [
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]


# ── Excel styles ──────────────────────────────────────────────────────────────
COLUMNS = [
    ("Notice ID",   16),
    ("Title",       62),
    ("Country",     16),
    ("Type",        18),
    ("Deadline",    13),
    ("Relevance",   38),
    ("Link",        55),
]
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_ALT_FILL  = PatternFill("solid", fgColor="F5F8FF")
_REL_FILL  = PatternFill("solid", fgColor="E2EFDA")
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_LINK_FONT = Font(name="Calibri", size=10, color="1155CC", underline="single")
_HIGH_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),  right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin",  color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


def _save_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ADB Tenders"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    col_names = [c[0] for c in COLUMNS]
    for ci, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    rel_idx  = col_names.index("Relevance") + 1
    link_idx = col_names.index("Link")      + 1

    for ri, row_data in enumerate(rows, 2):
        alt = _ALT_FILL if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 40
        vals = [
            row_data.get("notice_id", ""),
            row_data.get("title", ""),
            row_data.get("country", ""),
            row_data.get("notice_type", ""),
            row_data.get("deadline", ""),
            row_data.get("relevance", ""),
            row_data.get("link", ""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = _LINK_FONT
            elif ci == rel_idx:
                if val:
                    cell.fill = _REL_FILL
                    cell.font = _HIGH_FONT
                else:
                    cell.font = _BODY_FONT
                    if alt:
                        cell.fill = alt
            else:
                cell.font = _BODY_FONT
                if alt:
                    cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(ADB_EXCEL_PATH)
    print(f"[adb] Excel saved: {ADB_EXCEL_PATH}  ({len(rows)} rows)")


# ── Cache helpers ─────────────────────────────────────────────────────────────

def _load_cache() -> list[dict]:
    try:
        if os.path.exists(CACHE_PATH):
            with open(CACHE_PATH, encoding="utf-8") as f:
                data = json.load(f)
            age_h = (time.time() - data.get("saved_at", 0)) / 3600
            if age_h < 12:
                print(f"[adb] Using cache ({age_h:.1f}h old, {len(data['tenders'])} items)")
                return data["tenders"]
    except Exception:
        pass
    return []


def _save_cache(tenders: list[dict]) -> None:
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump({"saved_at": time.time(), "tenders": tenders}, f)
    except Exception:
        pass


# ── Session + HTTP helper ─────────────────────────────────────────────────────

def _make_session() -> requests.Session:
    """Create a session that looks like a real browser to ADB's Drupal site."""
    s = requests.Session()
    s.headers.update({
        "User-Agent":      random.choice(_USER_AGENTS),
        "Accept":          "application/json, text/javascript, */*; q=0.01",
        "Accept-Language": "en-US,en;q=0.9",
        "X-Requested-With": "XMLHttpRequest",   # Drupal AJAX check
        "Referer":         TENDERS_URL,
        "Origin":          BASE_URL,
    })
    return s


def _post_ajax(session: requests.Session, form_data: dict) -> Optional[list]:
    """
    POST to ADB's Drupal views/ajax endpoint with retry.
    Returns parsed JSON array or None on failure.
    """
    wait = _RETRY_BASE_WAIT
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            r = session.post(AJAX_URL, data=form_data, timeout=30)
            if r.status_code == 200:
                try:
                    return r.json()
                except ValueError:
                    # Not JSON — possibly a redirect or HTML error page
                    print(f"[adb]   Non-JSON response (len={len(r.text)})")
                    return None

            if r.status_code == 429:
                retry_after = int(r.headers.get("Retry-After", wait))
                print(f"[adb]   Rate-limited — waiting {retry_after}s")
                time.sleep(retry_after)
                wait *= 2
                continue

            if r.status_code == 403 and "just a moment" in r.text.lower():
                print("[adb]   Blocked by Cloudflare challenge — live AJAX path unavailable from this environment")
                return None

            if r.status_code >= 500:
                print(f"[adb]   Server error {r.status_code} — retrying in {wait}s")
                time.sleep(wait)
                wait *= 2
                continue

            print(f"[adb]   HTTP {r.status_code}")
            return None

        except (requests.Timeout, requests.ConnectionError) as exc:
            print(f"[adb]   Network error (attempt {attempt}/{_MAX_RETRIES}): {exc}")
            time.sleep(wait)
            wait *= 2
        except Exception as exc:
            print(f"[adb]   Unexpected error: {exc}")
            return None

    return None


# ── HTML parser ───────────────────────────────────────────────────────────────

def _extract_html_from_ajax(ajax_response: list) -> str:
    """
    ADB's Drupal views/ajax returns a JSON array of commands.
    Find the 'insert' command that contains the tender listing HTML.
    """
    for cmd in (ajax_response or []):
        if not isinstance(cmd, dict):
            continue
        if cmd.get("command") == "insert":
            html_data = cmd.get("data", "")
            if html_data and len(html_data) > 100:
                return html_data
    return ""


def _parse_tenders_from_html(html: str) -> list[dict]:
    """
    Parse ADB tender items from the HTML fragment returned by Drupal views/ajax.
    ADB tender rows are in <div class="views-row"> or <article>.
    """
    soup = BeautifulSoup(html, "html.parser")
    tenders = []

    # Try multiple selector patterns (ADB redesigns periodically)
    items = (
        soup.select("div.views-row")
        or soup.select("article.views-row")
        or soup.select("tr.odd, tr.even")
        or soup.select("li.views-row")
    )

    for item in items:
        # Title + link
        title_el = (
            item.select_one("span.field-content a")
            or item.select_one("h3 a, h4 a, h2 a")
            or item.select_one("td.views-field-title a")
            or item.select_one("a[href*='/projects/']")
        )
        if not title_el:
            continue

        title = title_el.get_text(strip=True)
        href  = title_el.get("href", "")
        if href.startswith("/"):
            href = BASE_URL + href
        if not title or not href:
            continue

        # Notice ID from URL (format: /projects/12345-001/...)
        nid_match = re.search(r"/projects/([A-Z0-9]+-\d{3}|\d+)", href, re.IGNORECASE)
        notice_id = nid_match.group(1) if nid_match else re.sub(r"[^\w]", "_", title[:30])

        # Country, type, deadline from labeled fields
        country     = ""
        notice_type = ""
        deadline    = ""

        for field in item.select("div.field, span.field-content, td"):
            text = field.get_text(separator=" ", strip=True)

            # Heuristic extraction based on common ADB field labels
            parent_class = " ".join(field.get("class", []))
            if "country" in parent_class.lower():
                country = text[:80]
            elif "type" in parent_class.lower() or "category" in parent_class.lower():
                notice_type = text[:60]
            elif "deadline" in parent_class.lower() or "closing" in parent_class.lower():
                deadline = text[:20]

        # Fallback: look for date-like strings in the item text
        if not deadline:
            date_match = re.search(
                r"\b(\d{1,2}[\s\-/]\w{3}[\s\-/]\d{4}|\d{4}-\d{2}-\d{2})\b",
                item.get_text()
            )
            if date_match:
                deadline = date_match.group(0)

        tenders.append({
            "notice_id":   notice_id,
            "title":       title,
            "link":        href,
            "country":     country,
            "notice_type": notice_type,
            "deadline":    deadline,
        })

    return tenders


# ── Main scrape loop ──────────────────────────────────────────────────────────

def _fetch_all_tenders() -> list[dict]:
    """
    Paginate through ADB tenders via Drupal views/ajax.
    Falls through multiple view configurations in case one returns empty.
    """
    session  = _make_session()
    all_tenders: list[dict] = []
    seen_ids: set[str] = set()

    # First: load the tenders page to get Drupal session/CSRF tokens if needed
    try:
        session.get(TENDERS_URL, timeout=20)
        time.sleep(1.0)
    except Exception:
        pass

    for view_cfg in _VIEW_CONFIGS:
        print(f"\n[adb]   View: {view_cfg['label']}...")
        view_found = False

        for page in range(MAX_PAGES):
            form_data = {
                "view_name":       view_cfg["view_name"],
                "view_display_id": view_cfg["view_display_id"],
                "view_args":       "",
                "view_path":       "/projects/tenders",
                "view_base_path":  "projects/tenders",
                "view_dom_id":     "1",
                "pager_element":   "0",
                "page":            str(page),
                "_drupal_ajax":    "1",
                "ajax_page_state[theme]":         "adb",
                "ajax_page_state[theme_token]":   "",
            }

            ajax_resp = _post_ajax(session, form_data)
            if ajax_resp is None:
                print(f"[adb]   Page {page}: AJAX failed — stopping this view")
                break

            html = _extract_html_from_ajax(ajax_resp)
            if not html:
                print(f"[adb]   Page {page}: no HTML in AJAX response — stopping")
                break

            items = _parse_tenders_from_html(html)
            if not items:
                print(f"[adb]   Page {page}: no items parsed — stopping")
                break

            view_found = True
            new_count = 0
            for t in items:
                nid = t["notice_id"]
                if nid not in seen_ids:
                    seen_ids.add(nid)
                    all_tenders.append(t)
                    new_count += 1

            print(f"[adb]   Page {page}: {len(items)} items ({new_count} new)")

            # Check for "no more pages" signal in the AJAX response
            has_next = any(
                isinstance(cmd, dict) and cmd.get("command") == "insert"
                and "pager" in str(cmd.get("selector", ""))
                and "next" in str(cmd.get("data", ""))
                for cmd in ajax_resp
            )
            if not has_next and page > 0:
                break

            time.sleep(PAGE_PAUSE + random.uniform(0, 1.0))

        if not view_found:
            print(f"[adb]   View '{view_cfg['label']}' returned nothing — trying next")
        time.sleep(2.0)

    return all_tenders


# ── Fallback: scrape HTML page directly ──────────────────────────────────────

def _fallback_html_scrape() -> list[dict]:
    """
    Direct GET fallback in case Drupal views/ajax returns empty.
    Loads the tenders page as a regular browser would and parses HTML.
    """
    print("[adb] Trying HTML fallback scrape...")
    tenders = []
    seen_ids: set[str] = set()

    headers = {
        "User-Agent":      random.choice(_USER_AGENTS),
        "Accept":          "text/html,application/xhtml+xml,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }

    for page in range(1, MAX_PAGES + 1):
        params = {"page": page - 1}
        wait = _RETRY_BASE_WAIT
        resp = None
        for attempt in range(3):
            try:
                resp = requests.get(TENDERS_URL, params=params,
                                    headers=headers, timeout=30)
                if resp.status_code == 200:
                    break
            except Exception as exc:
                print(f"[adb]   Fallback attempt {attempt+1}: {exc}")
                time.sleep(wait)
                wait *= 2
            resp = None

        if resp is None or resp.status_code != 200:
            break

        if "just a moment" in resp.text.lower():
            print("[adb]   HTML fallback blocked by Cloudflare challenge")
            break

        items = _parse_tenders_from_html(resp.text)
        if not items:
            break

        for t in items:
            nid = t["notice_id"]
            if nid not in seen_ids:
                seen_ids.add(nid)
                tenders.append(t)

        print(f"[adb]   Fallback page {page}: {len(items)} items")
        time.sleep(PAGE_PAUSE)

    return tenders


# ── Main ──────────────────────────────────────────────────────────────────────

def _selenium_scrape() -> list[dict]:
    """
    Selenium-based scrape to bypass Cloudflare protection.
    Loads the ADB tenders page in a headless browser and extracts items.
    """
    print("[adb] Attempting Selenium scrape to bypass Cloudflare...")
    tenders = []
    seen_ids: set[str] = set()
    driver = None
    try:
        opts = Options()
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument(f"user-agent={random.choice(_USER_AGENTS)}")

        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=opts)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })

        for page in range(MAX_PAGES):
            url = f"{TENDERS_URL}?page={page}"
            driver.get(url)
            time.sleep(3.0 + random.uniform(0, 2))

            html = driver.page_source
            if "just a moment" in html.lower() or len(html) < 500:
                print(f"[adb]   Selenium page {page}: still challenged or empty")
                break

            items = _parse_tenders_from_html(html)
            if not items:
                print(f"[adb]   Selenium page {page}: no items — stopping")
                break

            new_count = 0
            for t in items:
                nid = t["notice_id"]
                if nid not in seen_ids:
                    seen_ids.add(nid)
                    tenders.append(t)
                    new_count += 1

            print(f"[adb]   Selenium page {page}: {len(items)} items ({new_count} new)")
            if new_count == 0 and page > 0:
                break
            time.sleep(PAGE_PAUSE + random.uniform(0, 1.0))

    except Exception as exc:
        print(f"[adb]   Selenium scrape failed: {exc}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    return tenders


def run() -> tuple:
    """
    Run the ADB procurement scraper.
    Returns (new_tenders: list, all_rows: list).
    """
    print("\n[adb] ADB Procurement Scraper v2 (Drupal views/ajax + Selenium fallback)")

    # 1. Try Drupal views/ajax endpoint (primary)
    raw_tenders = _fetch_all_tenders()

    # 2. Fall back to Selenium (bypasses Cloudflare) if AJAX blocked
    if not raw_tenders:
        raw_tenders = _selenium_scrape()

    # 3. Fall back to direct HTML scraping
    if not raw_tenders:
        raw_tenders = _fallback_html_scrape()

    # 4. Fall back to cache if all live methods fail
    if not raw_tenders:
        print("[adb] Live methods returned nothing — trying cache...")
        raw_tenders = _load_cache()
        if not raw_tenders:
            print("[adb] No cache — skipping run.")
            return [], []

    if raw_tenders:
        _save_cache(raw_tenders)

    print(f"[adb] Total unique notices: {len(raw_tenders)}")

    # ── Filter + dedup via DB ─────────────────────────────────────────────────
    new_tenders: list[dict] = []
    all_rows:    list[dict] = []

    for t in raw_tenders:
        title = (t.get("title") or "").strip()
        if not title:
            continue

        relevance = score_relevance(title, "")
        row = {**t, "relevance": relevance}
        all_rows.append(row)

        tid = f"ADB_{t.get('notice_id', '')}"
        if check_if_new(tid):
            mark_as_seen(tid, title, SOURCE_NAME, t.get("link", ""))
            new_tenders.append({
                "title":    title,
                "deadline": t.get("deadline", ""),
                "value":    t.get("country", ""),
                "url":      t.get("link", ""),
            })

    relevant = sum(1 for r in all_rows if r.get("relevance"))
    print(f"[adb] {len(all_rows)} rows ({relevant} relevant, {len(new_tenders)} NEW)")

    if all_rows:
        _save_excel(all_rows)
    else:
        print("[adb] No rows to save.")

    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nResult: {len(rows)} total, {len(new)} new")
