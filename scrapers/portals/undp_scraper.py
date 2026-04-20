# =============================================================================
# undp_pipeline.py — UNDP Procurement Notices Pipeline
#
# Site   : https://procurement-notices.undp.org/
# Mode   : Option B — scrape ALL global notices, filter India-relevant ones
# Method : requests + BeautifulSoup (no Selenium, no CAPTCHA)
#
# India filter:
#   • Ref No contains "-IND-"  (UNDP-IND-XXXXX)
#   • Office/Country contains "INDIA" or "UNDP-IND"
#   • Title mentions India / Indian states / cities
#
# Run with: python3 main.py --undp
# Included in the default run automatically.
# =============================================================================

import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

SCRAPER_META = {
    "flag": "undp",
    "label": "UNDP Procurement",
    "group": "requests",
    "timeout": 900,
    "max_retries": 1,
    "auto": True,
}

# Max concurrent detail-page fetches — keeps UNDP response time ~100s regardless
# of how many relevant notices are found (vs. 600-1000s sequential).
_DETAIL_MAX_WORKERS = 10

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL   = "https://procurement-notices.undp.org"
SEARCH_URL = f"{BASE_URL}/search.cfm"
DETAIL_URL = f"{BASE_URL}/view_negotiation.cfm"

UNDP_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "UNDP_Tenders_Master.xlsx")
DELAY           = 1.0   # seconds between detail page requests

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         SEARCH_URL,
}

# Processes worth notifying IDCG about (consulting / advisory / expert work)
NOTIFY_PROCESSES = ["rfp", "ic ", "individual", "itb", "eoi", "rfi", "itp"]

# Processes to SKIP for notifications (goods, logistics, simple services)
SKIP_PROCESSES   = ["rfq", "request for quotation"]

# ── Geographic + topical relevance config ─────────────────────────────────────
# IDCG works across India, South Asia and international development programmes.
# We keep any UNDP notice that matches a country signal OR a topic signal.

# UN country codes in Ref No (UNDP-IND-XXXXX → IND) that are always kept
_KEEP_REF_CODES = {
    # South Asia
    "IND", "BGD", "NPL", "PAK", "LKA", "BTN", "AFG", "MDV",
    # East Africa
    "ETH", "KEN", "TZA", "UGA", "RWA", "MOZ", "ZMB", "ZWE",
    # West + Central Africa
    "GHA", "NGA", "SEN", "CIV", "CMR",
    # Southern Africa
    "ZAF", "MWI", "NAM", "BWA",
    # Multi-country / regional
    "GLO", "RAS", "RAF", "SAP",
}

# Office/country field substrings that always qualify
_KEEP_OFFICE_LOWER = {
    "india", "undp-ind", "bangladesh", "nepal", "pakistan", "sri lanka",
    "bhutan", "afghanistan", "maldives",
    "ethiopia", "kenya", "tanzania", "uganda", "rwanda", "mozambique",
    "zambia", "zimbabwe", "ghana", "nigeria", "senegal", "south africa",
    "africa", "south asia", "asia", "global",
}

# IDCG core topics — used to qualify global/multi-country notices by topic
_IDCG_TOPIC_KEYWORDS = [
    "evaluation", "monitoring", "assessment", "baseline", "survey",
    "m&e", "endline", "mid-term", "midterm", "review", "tpm",
    "capacity building", "capacity development", "training",
    "research", "study", "consultancy", "consultant", "advisory",
    "governance", "social", "gender", "health", "education",
    "environment", "climate", "wash", "nutrition", "livelihoods",
    "rural development", "community development", "programme evaluation",
]

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Ref No",       22),
    ("Title",        65),
    ("Office",       30),
    ("Process",      18),
    ("Deadline",     22),
    ("Posted",       15),
    ("Description",  70),
    ("Relevance",    40),
    ("Detail Link",  55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Helpers
# =============================================================================

def _is_relevant(ref_no: str, office: str, title: str) -> bool:
    """
    Return True for notices relevant to IDCG's geographic + topical scope.

    Rules (in priority order):
      1. Ref No country code in _KEEP_REF_CODES  → always keep
      2. Office/Country matches _KEEP_OFFICE_LOWER → always keep
      3. Title mentions Africa/Asia geography + IDCG topic → keep
      4. Process is consulting/advisory + IDCG topic in title → keep globally
    """
    # Rule 1: Ref No country code
    ref_up = ref_no.upper()
    m = re.search(r"UNDP-([A-Z]{3})-", ref_up)
    if m and m.group(1) in _KEEP_REF_CODES:
        return True
    if "-IND-" in ref_up:  # legacy check
        return True

    # Rule 2: Office/country field
    office_lower = office.lower()
    if any(kw in office_lower for kw in _KEEP_OFFICE_LOWER):
        return True

    # Rule 3 & 4: Title-based — require IDCG topic keyword
    title_lower = title.lower()
    has_idcg_topic = any(kw in title_lower for kw in _IDCG_TOPIC_KEYWORDS)
    if has_idcg_topic:
        return True

    return False


# Keep legacy alias for backward compatibility
def _is_india_relevant(ref_no: str, office: str, title: str) -> bool:
    return _is_relevant(ref_no, office, title)


def _is_notifiable_process(process: str) -> bool:
    """
    Return True if this process type is worth alerting IDCG about.
    RFQ (Request for Quotation) = goods/logistics → skip notifications.
    """
    p = process.lower()
    if any(skip in p for skip in SKIP_PROCESSES):
        return False
    return True


def _fetch_all_notices(session: requests.Session) -> list:
    """Fetch search.cfm and parse every notice row. Returns list of dicts."""
    print("[undp] Fetching full listing (all notices on one page)...", flush=True)
    raw_html = None
    for attempt in (1, 2, 3):
        try:
            r = session.get(SEARCH_URL, headers=HEADERS, timeout=180)
            r.raise_for_status()
            raw_html = r.text
            break
        except Exception as e:
            if attempt == 3:
                print(f"[undp] ERROR fetching search page: {e}", flush=True)
                return []
            print("[undp] Retry fetching search page after error…", flush=True)
            time.sleep(2)
    soup = BeautifulSoup(raw_html, "html.parser")

    rows    = soup.find_all("a", class_=re.compile(r"vacanciesTableLink"))
    notices = []

    for row in rows:
        href = row.get("href", "")
        if not href or "nego_id" not in href:
            continue

        m = re.search(r"nego_id=(\d+)", href)
        if not m:
            continue
        nego_id = m.group(1)

        # Parse label → value cells inside each row
        fields = {}
        for cell in row.find_all("div", class_=re.compile(r"vacanciesTable__cell$")):
            lbl = cell.find("div", class_=re.compile(r"vacanciesTable__cell__label"))
            val = cell.find("span")
            if lbl and val:
                fields[lbl.get_text(strip=True).lower()] = val.get_text(" ", strip=True)

        title = fields.get("title", "")
        if not title or len(title) < 5:
            continue

        detail_link = (f"{BASE_URL}/{href}"
                       if not href.startswith("http") else href)

        notices.append({
            "nego_id":     nego_id,
            "Ref No":      fields.get("ref no", ""),
            "Title":       title,
            "Office":      fields.get("undp office/country", ""),
            "Process":     fields.get("process", ""),
            "Deadline":    fields.get("deadline", ""),
            "Posted":      fields.get("posted", ""),
            "Detail Link": detail_link,
        })

    return notices


def _fetch_detail(session: requests.Session, nego_id: str) -> str:
    """Fetch the detail page and return the main description text (≤2000 chars)."""
    try:
        url = f"{DETAIL_URL}?nego_id={nego_id}"
        r   = session.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Remove nav / header / footer noise
        for tag in soup.find_all(["nav", "header", "footer", "script", "style"]):
            tag.decompose()

        # Try known content containers first
        for sel in ["div.grid-container", "main", "div#content", "div.content-area"]:
            el = soup.select_one(sel)
            if el:
                text = el.get_text(" ", strip=True)
                if len(text) > 150:
                    return text[:2000]

        # Fallback: whole body
        return soup.get_text(" ", strip=True)[:2000]

    except Exception as e:
        print(f"[undp]   Detail error (nego_id={nego_id}): {e}", flush=True)
        return ""


# =============================================================================
# Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "UNDP Procurement Notices"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    relevance_idx   = col_names.index("Relevance") + 1
    link_idx        = col_names.index("Detail Link") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill:
                        cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill:
                    cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(UNDP_EXCEL_PATH)
    print(f"[undp] Excel saved: {UNDP_EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[undp] UNDP Procurement Notices Pipeline starting...", flush=True)
    print("[undp] Mode: Global scrape → IDCG-relevant filter (India + South Asia + Africa + topic-match)", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []

    # Clear old Excel
    if os.path.exists(UNDP_EXCEL_PATH):
        try:
            os.remove(UNDP_EXCEL_PATH)
        except Exception:
            pass

    session = requests.Session()

    # ── Fetch all notices ──────────────────────────────────────────────────────
    try:
        all_notices = _fetch_all_notices(session)
    except Exception as e:
        print(f"[undp] ERROR fetching listing: {e}", flush=True)
        return new_tenders, all_rows

    print(f"[undp] Total notices fetched: {len(all_notices)}", flush=True)

    # ── Filter IDCG-relevant ───────────────────────────────────────────────────
    india_notices = [
        n for n in all_notices
        if _is_relevant(n["Ref No"], n["Office"], n["Title"])
    ]
    print(f"[undp] IDCG-relevant: {len(india_notices)} / {len(all_notices)}", flush=True)

    if not india_notices:
        print("[undp] No relevant notices found.", flush=True)
        return new_tenders, all_rows

    # ── Fetch details concurrently (avoids sequential timeout on large batches) ──
    total = len(india_notices)
    print(
        f"[undp] Fetching {total} detail pages "
        f"(max {_DETAIL_MAX_WORKERS} concurrent)...",
        flush=True,
    )

    # Fetch all detail pages in parallel; preserve original index for ordering.
    descriptions: list = [None] * total
    with ThreadPoolExecutor(max_workers=_DETAIL_MAX_WORKERS) as pool:
        future_to_idx = {
            pool.submit(_fetch_detail, session, notice["nego_id"]): idx
            for idx, notice in enumerate(india_notices)
        }
        done = 0
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            done += 1
            try:
                descriptions[idx] = future.result() or ""
            except Exception:
                descriptions[idx] = ""
            notice = india_notices[idx]
            print(
                f"[undp]   [{done:>3}/{total}] {notice['Title'][:60]}",
                flush=True,
            )

    # ── Score + dedup (sequential — no I/O) ──────────────────────────────────
    for idx, notice in enumerate(india_notices):
        description = descriptions[idx] or ""
        relevance   = score_relevance(notice["Title"], description)

        row = {
            "Ref No":      notice["Ref No"],
            "Title":       notice["Title"],
            "Office":      notice["Office"],
            "Process":     notice["Process"],
            "Deadline":    notice["Deadline"],
            "Posted":      notice["Posted"],
            "Description": description[:1500] if description else "",
            "Relevance":   relevance,
            "Detail Link": notice["Detail Link"],
        }
        all_rows.append(row)

        ref       = notice["Ref No"].strip()
        tender_id = f"UNDP/{ref}" if ref else f"UNDP/{notice['Title'][:60]}"

        rel_count  = len([r for r in relevance.split(",") if r.strip()]) if relevance else 0
        notifiable = (rel_count >= 2 and _is_notifiable_process(notice["Process"]))

        if check_if_new(tender_id):
            mark_as_seen(
                tender_id,
                title=notice["Title"],
                source_site="UNDP",
                url=notice["Detail Link"],
            )
            if notifiable:
                new_tenders.append({
                    "title":    notice["Title"],
                    "deadline": notice["Deadline"],
                    "value":    notice["Process"],
                    "url":      notice["Detail Link"],
                })
                print(f"[undp]   → NEW + ALERT | {relevance[:60]}", flush=True)
            else:
                reason = (
                    "RFQ/goods" if not _is_notifiable_process(notice["Process"])
                    else f"only {rel_count} keyword match"
                )
                print(f"[undp]   → NEW (no alert — {reason}) | {relevance or '—'}", flush=True)
        else:
            print(f"[undp]   → seen | {relevance or '—'}", flush=True)

    # ── Save Excel ─────────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(
        f"\n[undp] Done — {len(all_rows)} India notices, "
        f"{len(new_tenders)} NEW, {relevant} relevant",
        flush=True,
    )
    return new_tenders, all_rows
