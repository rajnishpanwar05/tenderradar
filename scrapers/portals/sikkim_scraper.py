# =============================================================================
# sikkim_pipeline.py — Sikkim eProcurement Portal Pipeline
#
# Site: https://sikkimtender.gov.in
# Uses Selenium in VISIBLE (non-headless) Chrome — CAPTCHA must be solved manually.
#
# ⚠️  MANUAL STEP REQUIRED:
#     When the browser window opens, solve the CAPTCHA and click Search.
#     The script then takes over and scrapes all pages automatically.
#
# Run with: python3 main.py --sikkim
# NOT included in the default run (none_selected) because of CAPTCHA dependency.
# =============================================================================

import os
import re
import time
import base64
import requests

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL          = "https://sikkimtender.gov.in/nicgep/app"
ACTIVE_TENDERS    = f"{BASE_URL}?page=FrontEndLatestActiveTenders&service=page"
BY_CLASSIFICATION = f"{BASE_URL}?page=FrontEndTendersByClassification&service=page"

SIKKIM_EXCEL_PATH    = os.path.join(PORTAL_EXCELS_DIR, "Sikkim_Tenders_Master.xlsx")
CAPTCHA_WAIT_SECONDS = 120   # seconds for user to solve CAPTCHA
ACTION_DELAY         = 2.0
MAX_PAGES            = 100
CAPTCHA_API_KEY = os.environ.get("CAPTCHA_API_KEY") or os.environ.get("TWOCAPTCHA_API_KEY", "")
SIKKIM_AUTO_CAPTCHA = str(os.environ.get("SIKKIM_AUTO_CAPTCHA", "")).strip().lower() in {"1", "true", "yes", "on"}
CAPTCHA_IN_URL = "https://2captcha.com/in.php"
CAPTCHA_RES_URL = "https://2captcha.com/res.php"
CAPTCHA_POLL_INTERVAL = 5
CAPTCHA_POLL_MAX = 20

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Tender ID",    22),
    ("Title",        65),
    ("Organisation", 35),
    ("Category",     18),
    ("Tender Type",  15),
    ("Closing Date", 20),
    ("Published",    18),
    ("Relevance",    40),
    ("Detail Link",  55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Browser helpers
# =============================================================================

def _create_driver():
    """Visible Chrome — required for CAPTCHA solve."""
    options = Options()
    # NOT headless — user must see browser window to solve CAPTCHA
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,900")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def _wait_for_captcha_solve(driver, timeout=CAPTCHA_WAIT_SECONDS):
    print(f"\n[sikkim] ⏳ Waiting up to {timeout}s for CAPTCHA solve...", flush=True)
    print("[sikkim]    → Look at the Chrome window", flush=True)
    print("[sikkim]    → Type the CAPTCHA text and click Search", flush=True)
    print("[sikkim]    → Return here — scraping starts automatically", flush=True)

    start = time.time()
    while time.time() - start < timeout:
        page_text = driver.page_source.lower()
        if any(kw in page_text for kw in
               ["tender id", "tender title", "closing date", "organisation name",
                "published date", "no tender found"]):
            print("[sikkim] ✅ CAPTCHA solved — results loaded!", flush=True)
            return True
        time.sleep(1.5)

    print("[sikkim] ❌ Timed out waiting for CAPTCHA.", flush=True)
    return False


def _two_captcha_submit_image(image_b64: str) -> str:
    if not CAPTCHA_API_KEY or not image_b64:
        return ""
    try:
        resp = requests.post(
            CAPTCHA_IN_URL,
            data={"key": CAPTCHA_API_KEY, "method": "base64", "body": image_b64, "json": 1},
            timeout=60,
        )
        data = resp.json() if resp.ok else {}
        if not data or data.get("status") != 1:
            return ""
        return str(data.get("request") or "")
    except Exception:
        return ""


def _two_captcha_poll(captcha_id: str) -> str:
    if not captcha_id:
        return ""
    for _ in range(CAPTCHA_POLL_MAX):
        try:
            time.sleep(CAPTCHA_POLL_INTERVAL)
            resp = requests.get(
                CAPTCHA_RES_URL,
                params={"key": CAPTCHA_API_KEY, "action": "get", "id": captcha_id, "json": 1},
                timeout=30,
            )
            data = resp.json() if resp.ok else {}
            if not data:
                continue
            if data.get("status") == 1:
                return str(data.get("request") or "")
            if data.get("request") not in ("CAPTCHA_NOT_READY", "CAPCHA_NOT_READY"):
                return ""
        except Exception:
            continue
    return ""


def _try_auto_solve_captcha(driver) -> bool:
    if not SIKKIM_AUTO_CAPTCHA or not CAPTCHA_API_KEY:
        return False
    try:
        images = driver.find_elements(By.CSS_SELECTOR, "img[src*='captcha'], img[id*='captcha'], img[class*='captcha']")
        if not images:
            return False
        img = next((i for i in images if i.is_displayed()), images[0])
        png = img.screenshot_as_png
        if not png:
            return False

        captcha_id = _two_captcha_submit_image(base64.b64encode(png).decode("ascii"))
        answer = _two_captcha_poll(captcha_id)
        if not answer:
            return False

        inputs = driver.find_elements(By.CSS_SELECTOR, "input[id*='captcha'], input[name*='captcha'], input[placeholder*='captcha']")
        if not inputs:
            inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
        target = next((i for i in inputs if i.is_displayed() and i.is_enabled()), None)
        if not target:
            return False
        target.clear()
        target.send_keys(answer)

        clicked = False
        for xp in [
            "//input[@type='submit']",
            "//button[@type='submit']",
            "//button[contains(translate(normalize-space(.),'SEARCHSUBMIT','searchsubmit'),'search')]",
            "//input[contains(translate(@value,'SEARCHSUBMIT','searchsubmit'),'search')]",
        ]:
            try:
                btn = driver.find_element(By.XPATH, xp)
                if btn.is_displayed() and btn.is_enabled():
                    driver.execute_script("arguments[0].click();", btn)
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            target.submit()

        for _ in range(18):
            page_text = driver.page_source.lower()
            if any(kw in page_text for kw in
                   ["tender id", "tender title", "closing date", "organisation name",
                    "published date", "no tender found"]):
                print("[sikkim] ✅ auto-captcha solved", flush=True)
                return True
            time.sleep(1.5)
    except Exception:
        return False
    return False


# =============================================================================
# Parsing
# =============================================================================

def _parse_tender_table(html):
    soup    = BeautifulSoup(html, "html.parser")
    tenders = []

    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if len(rows) < 2:
            continue

        header_row  = rows[0]
        headers     = [th.get_text(strip=True).lower()
                       for th in header_row.find_all(["th", "td"])]
        header_str  = " ".join(headers)

        if not any(kw in header_str for kw in
                   ["tender id", "tender title", "closing date", "organisation"]):
            continue

        print(f"[sikkim]   Found NIC tender table with {len(rows)-1} rows", flush=True)

        col = {}
        for i, h in enumerate(headers):
            if "tender id"    in h:                    col.setdefault("tender_id", i)
            if "tender title" in h or "title" in h:    col.setdefault("title", i)
            if "organisation" in h or "dept" in h:     col.setdefault("org", i)
            if "category"     in h:                    col.setdefault("category", i)
            if "type"         in h and "tender" in h:  col.setdefault("type", i)
            if "closing"      in h:                    col.setdefault("closing", i)
            if "published"    in h or "posted" in h:   col.setdefault("published", i)
            if "opening"      in h:                    col.setdefault("opening", i)

        def get_cell(row_cells, key):
            idx = col.get(key)
            if idx is not None and idx < len(row_cells):
                return row_cells[idx].get_text(" ", strip=True)
            return ""

        for row in rows[1:]:
            cells = row.find_all(["td", "th"])
            if len(cells) < 3:
                continue
            if len(row.get_text(" ", strip=True)) < 5:
                continue

            detail_link = ""
            for a in row.find_all("a", href=True):
                href = a["href"]
                if any(kw in href.lower() for kw in ["viewtender", "view", "detail", "tender"]):
                    detail_link = (href if href.startswith("http")
                                   else "https://sikkimtender.gov.in" + href)
                    break

            title = get_cell(cells, "title")
            if not title or len(title) < 3:
                title = max(cells, key=lambda c: len(c.get_text())).get_text(strip=True)

            entry = {
                "Tender ID":    get_cell(cells, "tender_id"),
                "Title":        title,
                "Organisation": get_cell(cells, "org"),
                "Category":     get_cell(cells, "category"),
                "Tender Type":  get_cell(cells, "type"),
                "Closing Date": get_cell(cells, "closing"),
                "Published":    get_cell(cells, "published"),
                "Detail Link":  detail_link,
            }

            if not entry["Title"] or len(entry["Title"]) < 3:
                continue

            tenders.append(entry)

        if tenders:
            return tenders

    return tenders


def _get_next_page(driver):
    for xpath in [
        "//a[normalize-space()='Next']",
        "//a[contains(text(),'Next')]",
        "//a[contains(text(),'>')]",
        "//input[@value='Next']",
        "//li[contains(@class,'next')]/a",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            if el.is_displayed() and el.is_enabled():
                driver.execute_script("arguments[0].click();", el)
                time.sleep(ACTION_DELAY)
                return True
        except Exception:
            continue
    return False


# =============================================================================
# Excel writer
# =============================================================================

def _save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sikkim eProcurement Tenders"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes  = ws.cell(row=2, column=1)
    relevance_idx    = col_names.index("Relevance") + 1
    link_idx         = col_names.index("Detail Link") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill: cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill: cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(SIKKIM_EXCEL_PATH)
    print(f"[sikkim] Excel saved: {SIKKIM_EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[sikkim] Sikkim eProcurement Portal Pipeline starting...", flush=True)
    print("[sikkim] ⚠️  CAPTCHA handling: auto via 2captcha (if SIKKIM_AUTO_CAPTCHA=1), else manual", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []
    raw         = []

    # Clear old Excel
    if os.path.exists(SIKKIM_EXCEL_PATH):
        try:
            os.remove(SIKKIM_EXCEL_PATH)
        except Exception:
            pass

    driver = _create_driver()
    try:
        print("[sikkim] Opening portal...", flush=True)
        driver.get(ACTIVE_TENDERS)
        time.sleep(2)

        print("\n[sikkim] >>> ACTION REQUIRED:", flush=True)
        print("[sikkim]     A Chrome window has opened.", flush=True)
        print("[sikkim]     Solve the CAPTCHA and click Search.", flush=True)

        solved = _try_auto_solve_captcha(driver)
        if not solved and not _wait_for_captcha_solve(driver):
            # Fallback: try Classification page
            print("[sikkim] Trying classification page...", flush=True)
            driver.get(BY_CLASSIFICATION)
            time.sleep(2)
            solved = _try_auto_solve_captcha(driver)
            if not solved and not _wait_for_captcha_solve(driver):
                print("[sikkim] ❌ Could not pass CAPTCHA — aborting.", flush=True)
                return new_tenders, all_rows

        # Scrape all pages
        print("\n[sikkim] Scraping tender listings...", flush=True)
        page_num = 1
        while page_num <= MAX_PAGES:
            print(f"[sikkim]   Page {page_num}...", flush=True)
            html = driver.page_source

            if "no tender found" in html.lower() or "no record" in html.lower():
                print("[sikkim]   No tenders on this page.", flush=True)
                break

            page_tenders = _parse_tender_table(html)
            print(f"[sikkim]   Extracted {len(page_tenders)} tenders", flush=True)
            raw.extend(page_tenders)

            if not _get_next_page(driver):
                print("[sikkim] No more pages.", flush=True)
                break
            page_num += 1

    finally:
        driver.quit()
        print("[sikkim] Browser closed.", flush=True)

    if not raw:
        print("[sikkim] ⚠️  No tenders extracted. Sikkim portal is small — try again.", flush=True)
        return new_tenders, all_rows

    # Deduplicate by Tender ID (portal-level)
    seen_ids, unique = set(), []
    for r in raw:
        key = r.get("Tender ID") or r.get("Title", "")[:40]
        if key not in seen_ids:
            seen_ids.add(key)
            unique.append(r)

    # Score relevance + DB deduplication
    for t in unique:
        combined  = t.get("Title", "") + " " + t.get("Organisation", "") + " " + t.get("Category", "")
        relevance = score_relevance(combined)
        t["Relevance"] = relevance
        all_rows.append(t)

        tid_raw   = t.get("Tender ID", "").strip()
        tender_id = f"SIKKIM/{tid_raw}" if tid_raw else f"SIKKIM/{t['Title'][:60]}"
        detail_url = t.get("Detail Link", ACTIVE_TENDERS)

        if check_if_new(tender_id):
            mark_as_seen(tender_id, title=t["Title"], source_site="SIKKIM", url=detail_url)
            new_tenders.append({
                "title":    t["Title"],
                "deadline": t.get("Closing Date", ""),
                "value":    t.get("Category", ""),
                "url":      detail_url,
            })
            print(f"[sikkim]   → NEW: {t['Title'][:70]} | {relevance[:40] or '—'}", flush=True)

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(f"\n[sikkim] Done — {len(all_rows)} tenders, {len(new_tenders)} NEW, {relevant} relevant", flush=True)
    return new_tenders, all_rows
