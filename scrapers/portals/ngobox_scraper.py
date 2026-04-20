# =============================================================================
# ngobox_pipeline.py — NGO Box RFP/EOI India Pipeline
#
# Site   : https://ngobox.org/rfp_eoi_listing.php
# Method : requests + BeautifulSoup (server-side rendered HTML)
# Login  : None required
# CAPTCHA: None
#
# What this does:
#   1. Fetches the RFP/EOI listing page
#   2. Extracts each tender: title, organisation, deadline, detail URL
#   3. Visits detail page for full description
#   4. Scores relevance against IDCG expertise keywords
#   5. Deduplicates via MySQL — only NEW tenders trigger notification
#   6. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://ngobox.org"
LISTING_URL = f"{BASE_URL}/rfp_eoi_listing.php"
DELAY       = 1.2

NGOBOX_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "NGOBox_RFPs_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         BASE_URL,
}

# Patterns that flag pure goods/supply tenders — skip these
SKIP_TITLE_PATTERNS = [
    "supply of", "purchase of", "procurement of equipment", "procurement of vehicle",
    "procurement of computer", "procurement of furniture", "procurement of stationery",
    "rfb for procurement", "rate contract for", "annual maintenance contract",
    "amc for", "nursing home items", "empanelment of vendors", "empanelment of suppliers",
]

def _is_goods_tender(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in SKIP_TITLE_PATTERNS)


# ── Excel styles (dark blue — consistent with all other pipelines) ─────────────
MASTER_COLUMNS = [
    ("Title",          65),
    ("Organisation",   38),
    ("Deadline",       16),
    ("Description",    65),
    ("Thematic Area",  28),
    ("Relevance",      40),
    ("Detail Link",    55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse listing page
# =============================================================================

def _parse_listings(soup: BeautifulSoup) -> list:
    """
    NGO Box structure (AJAX-rendered): inside #sresult, each tender card has:
      - <a href="full_rfp_eoi_..."> → title
      - text node → organisation name
      - "Deadline:" label + separate text node with the date
    We anchor on each full_rfp_eoi link and extract card data from its container.
    """
    entries = []
    seen_hrefs = set()

    SKIP_WORDS = ['google', 'calendar', 'deadline', 'premium', 'featured',
                  'standard', 'add to', 'shares', 'login', 'post deadline']

    # Iterate over all tender detail links
    for anchor in soup.find_all('a', href=lambda h: h and 'full_rfp_eoi' in h):
        href = anchor.get('href', '').strip()
        if not href or href in seen_hrefs:
            continue
        seen_hrefs.add(href)

        title = anchor.get_text(strip=True)
        if not title or len(title) < 8:
            continue

        detail_url = (BASE_URL + '/' + href.lstrip('/')) if not href.startswith('http') else href

        # Walk up to the card container (up to 12 levels)
        container = anchor.parent
        for _ in range(12):
            if not container:
                break
            txt = container.get_text(' ', strip=True)
            if 'Deadline' in txt and len(txt) > len(title) + 10:
                break
            container = container.parent
        if not container:
            container = anchor.parent

        # Extract all non-trivial text lines from the container
        all_lines = [
            ln.strip()
            for ln in container.get_text('\n', strip=True).split('\n')
            if ln.strip() and len(ln.strip()) > 2
            and not any(sw in ln.strip().lower() for sw in SKIP_WORDS)
        ]

        # Deadline: find "Deadline:" label and grab the following date line
        deadline = ''
        for i, ln in enumerate(all_lines):
            if re.match(r'Deadline\s*:', ln, re.I):
                # Date might be on the same line or the next
                rest = re.sub(r'Deadline\s*:\s*', '', ln, flags=re.I).strip()
                if rest and re.search(r'\d', rest):
                    deadline = rest
                elif i + 1 < len(all_lines):
                    deadline = all_lines[i + 1]
                break
        # Fallback: look for a date pattern in all lines
        if not deadline:
            for ln in all_lines:
                if re.search(r'\d{1,2}\s+\w{3}\.?\s+\d{4}', ln):
                    deadline = ln
                    break

        # Organisation: first line in container after the title that isn't a date/deadline/skip
        org = ''
        for ln in all_lines:
            ll = ln.lower()
            if title[:30].lower() in ll:
                continue
            if re.search(r'\d{1,2}\s+\w{3}\.?\s+\d{4}', ln):
                continue
            if any(sw in ll for sw in ['deadline', 'google', 'calendar']):
                continue
            if len(ln) > 4:
                org = ln
                break

        # Safe tender ID from the slug portion of the href
        slug = href.split('full_rfp_eoi_')[-1] if 'full_rfp_eoi_' in href else href
        tender_id = f"NGOBOX_{re.sub(r'[^a-zA-Z0-9]', '_', slug)[:80]}"

        entries.append({
            "Title":      title,
            "Organisation": org,
            "Deadline":   deadline,
            "detail_url": detail_url,
            "tender_id":  tender_id,
        })

    return entries


# =============================================================================
# SECTION 2 — Fetch detail page
# =============================================================================

def _fetch_detail(session: requests.Session, url: str) -> dict:
    result = {"Description": "", "Thematic Area": ""}
    try:
        r = session.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')

        # Main description — try common content containers
        for selector in [
            ('div', re.compile(r'rfp.?detail|job.?detail|content.?main', re.I)),
            ('div', {'class': re.compile(r'col-md-8|main.?content|detail', re.I)}),
        ]:
            panel = soup.find(selector[0], selector[1])
            if panel and len(panel.get_text(strip=True)) > 100:
                result["Description"] = panel.get_text(' ', strip=True)[:1800]
                break

        if not result["Description"]:
            # Fallback: main content area
            main = soup.find('main') or soup.find('div', {'id': re.compile(r'content', re.I)})
            if main:
                result["Description"] = main.get_text(' ', strip=True)[:1800]

        # Thematic area tag
        for tag in soup.find_all(string=re.compile(r'Thematic Area|Category|Sector', re.I)):
            parent = tag.parent
            if parent:
                nxt = parent.find_next_sibling()
                if nxt:
                    result["Thematic Area"] = nxt.get_text(strip=True)[:120]
                    break

    except Exception as e:
        print(f"[ngobox]   Detail fetch error ({url[:60]}): {e}")
    return result


# =============================================================================
# SECTION 3 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "NGOBox RFPs"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")   + 1
    link_idx = col_names.index("Detail Link") + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 50
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(NGOBOX_EXCEL_PATH)
    print(f"[ngobox] Excel saved: {NGOBOX_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 3b — Selenium driver (AJAX content loader)
# =============================================================================

def _build_driver():
    """Headless Chrome — needed because NGO Box loads listings via AJAX."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


# =============================================================================
# SECTION 4 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the NGO Box RFP/EOI pipeline.
    Returns:
        new_tenders — list of dicts for notification (new only)
        all_rows    — all fetched rows
    """
    print("\n" + "=" * 65)
    print("[ngobox] NGO Box RFP/EOI Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(NGOBOX_EXCEL_PATH):
        try:
            os.remove(NGOBOX_EXCEL_PATH)
            print(f"[ngobox] Cleared old Excel")
        except Exception:
            pass

    session = requests.Session()
    session.headers.update({
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/120.0.0.0 Safari/537.36"),
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": BASE_URL,
    })

    # ── Fetch listing via Selenium (results in #sresult loaded by AJAX) ───────
    soup   = None
    driver = None
    try:
        driver = _build_driver()
        driver.get(LISTING_URL)
        # Wait for #sresult to be populated with tender links
        try:
            WebDriverWait(driver, 25).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "#sresult a[href*='full_rfp_eoi']")
                )
            )
        except Exception:
            # Fallback: wait for any "Deadline:" text to appear
            try:
                WebDriverWait(driver, 15).until(
                    lambda d: "Deadline:" in d.page_source
                )
            except Exception:
                pass
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        # Transfer cookies to requests.Session for detail page fetching
        for ck in driver.get_cookies():
            session.cookies.set(ck["name"], ck["value"])
    except Exception as e:
        print(f"[ngobox] ERROR fetching listing: {e}")
        return [], []
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    if not soup:
        print("[ngobox] Could not fetch listing page.")
        return [], []

    entries = _parse_listings(soup)
    print(f"[ngobox] Found {len(entries)} RFP/EOI listings")

    if not entries:
        print("[ngobox] No listings found — page structure may have changed.")
        return [], []

    # ── Process each entry ─────────────────────────────────────────────────────
    for i, entry in enumerate(entries, 1):
        title = entry["Title"]

        # Skip pure goods tenders
        if _is_goods_tender(title):
            print(f"[ngobox]   [{i:>2}/{len(entries)}] SKIP (goods): {title[:60]}")
            continue

        print(f"[ngobox]   [{i:>2}/{len(entries)}] {title[:65]}...")

        detail = _fetch_detail(session, entry["detail_url"])
        relevance = score_relevance(title, detail.get("Description", ""))

        row = {
            "Title":        title,
            "Organisation": entry["Organisation"],
            "Deadline":     entry["Deadline"],
            "Description":  detail.get("Description", "")[:1500],
            "Thematic Area":detail.get("Thematic Area", ""),
            "Relevance":    relevance,
            "Detail Link":  entry["detail_url"],
        }
        all_rows.append(row)

        # ── DB deduplication ──────────────────────────────────────────────────
        tid = entry["tender_id"]
        if check_if_new(tid):
            mark_as_seen(tid, title=title, source_site="NGOBox", url=entry["detail_url"])
            new_tenders.append({
                "title":    title,
                "deadline": entry["Deadline"],
                "value":    "See listing",
                "url":      entry["detail_url"],
            })
            print(f"           → NEW | Relevance: {relevance or '—'}")
        else:
            print(f"           → seen | Relevance: {relevance or '—'}")

        time.sleep(DELAY)

    # ── Save Excel ─────────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[ngobox] Done — {len(all_rows)} listings, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
