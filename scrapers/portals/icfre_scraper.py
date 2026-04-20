# =============================================================================
# icfre_pipeline.py — ICFRE Tenders Pipeline
#
# Site   : https://icfre.gov.in/en/tenders
# Method : requests + BeautifulSoup (server-rendered HTML table)
# Login  : None required
# CAPTCHA: None
#
# Page structure:
#   div.recruitment-content > table.rec-table
#   Columns: [Sl. No.] [Title + PDF link] [Documents] [Last Date of Submission]
#
# What this does:
#   1. Fetches the current tenders page (no pagination needed — ~15 items max)
#   2. Also fetches the tender-archive page for recently closed items
#   3. Parses each row: title, PDF link, deadline
#   4. Scores relevance against IDCG expertise
#   5. Deduplicates via MySQL — only NEW tenders trigger notification
#   6. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
#
# ICFRE = Indian Council of Forestry Research and Education
# Focus: forestry, ecology, environment, research — strong overlap with IDCG
# =============================================================================

import os, re, time
from datetime import datetime, timedelta
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL     = "https://icfre.gov.in"
CURRENT_URL  = f"{BASE_URL}/en/tenders"
ARCHIVE_URL  = f"{BASE_URL}/en/tender-archive"
DELAY        = 0.5
ARCHIVE_DAYS = 90   # ISSUE 6: skip archive entries older than this many days

ICFRE_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "ICFRE_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer":         BASE_URL,
}

# Skip pure goods/printing tenders
SKIP_PATTERNS = [
    "supply of", "purchase of", "printing of", "rate contract",
    "amc for", "annual maintenance", "vehicle", "furniture",
]

def _is_goods(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in SKIP_PATTERNS)


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",        65),
    ("Deadline",     28),
    ("Source",       12),
    ("Relevance",    40),
    ("PDF / Link",   55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse tender table
# =============================================================================

def _parse_tender_table(soup: BeautifulSoup, source_label: str) -> list:
    """
    Parse ICFRE's rec-table structure.
    Columns: Sl.No | Title (+ PDF link) | Documents | Last Date
    """
    entries = []

    # Primary selector
    table = (
        soup.find('table', class_='rec-table') or
        soup.find('table', class_=re.compile(r'rec|tender|listing', re.I)) or
        soup.find('div', class_='recruitment-content')
    )

    # If we got a div wrapper, find the table inside it
    if table and table.name != 'table':
        table = table.find('table')

    # Final fallback: any table on the page
    if not table:
        tables = soup.find_all('table')
        for t in tables:
            if 'title' in t.get_text().lower() or 'tender' in t.get_text().lower():
                table = t
                break

    if not table:
        print(f"[icfre]   No table found on {source_label} page")
        return entries

    rows = table.find_all('tr')
    for row in rows:
        cells = row.find_all(['td', 'th'])
        if not cells:
            continue

        # Skip header row
        if row.find('th') or any(c.get('data-label','').lower() == 'sl. no.' and
                                  not c.get_text(strip=True).isdigit() for c in cells):
            cell_texts = [c.get_text(strip=True).lower() for c in cells]
            if any('title' in ct or 'tender' in ct for ct in cell_texts):
                continue

        # Find title cell — has class rec-title or data-label="Title"
        title_cell = None
        for cell in cells:
            dlabel = cell.get('data-label', '').lower()
            cls    = ' '.join(cell.get('class', [])).lower()
            if 'title' in dlabel or 'rec-title' in cls or 'title' in cls:
                title_cell = cell
                break

        # Fallback: cell with longest text that contains a link
        if not title_cell:
            linked = [c for c in cells if c.find('a')]
            if linked:
                title_cell = max(linked, key=lambda c: len(c.get_text()))

        if not title_cell:
            continue

        # Extract title text and PDF link
        title    = title_cell.get_text(' ', strip=True)
        # Clean up "Updated - date" suffix
        title    = re.sub(r'\s*Updated\s*-\s*\d+.*', '', title).strip()
        # Clean up icon text
        title    = re.sub(r'^\s*[\u2022\u25b6\uf105\uf054]\s*', '', title).strip()
        if not title or len(title) < 8:
            continue

        # PDF link
        pdf_link = ""
        link_tag = title_cell.find('a', href=True)
        if link_tag:
            href = link_tag.get('href', '')
            if href:
                # Resolve relative paths like ../pdf/tender/tender577.pdf
                if href.startswith('http'):
                    pdf_link = href
                elif href.startswith('..'):
                    # e.g. ../pdf/tender/X.pdf → https://icfre.gov.in/pdf/tender/X.pdf
                    pdf_link = BASE_URL + '/' + href.lstrip('./')
                elif href.startswith('/'):
                    pdf_link = BASE_URL + href
                else:
                    pdf_link = BASE_URL + '/' + href.lstrip('/')
                # Ensure /en/ prefix is never in the path (wrong, causes 404)
                pdf_link = pdf_link.replace('/en/pdf/', '/pdf/')

        # Deadline — find cell with data-label matching "last date" or date pattern
        deadline = ""
        for cell in cells:
            dlabel = cell.get('data-label', '').lower()
            txt    = cell.get_text(strip=True)
            if 'date' in dlabel or 'deadline' in dlabel or 'submission' in dlabel:
                deadline = txt
                break
            if re.search(r'\d{1,2}\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s+\d{4}', txt, re.I):
                deadline = txt[:60]

        # Updated date (for dedup purposes)
        updated = ""
        span = title_cell.find('span')
        if span:
            updated = span.get_text(strip=True)

        slug      = re.sub(r'[^a-zA-Z0-9]', '_', title[:60])
        tender_id = f"ICFRE_{source_label.upper()[:3]}_{slug}"

        entries.append({
            "Title":     title,
            "Deadline":  deadline,
            "Source":    source_label,
            "PDF / Link": pdf_link or CURRENT_URL,  # fallback to listing page
            "url":        pdf_link or CURRENT_URL,
            "tender_id": tender_id,
            "_updated":  updated,
        })

    return entries


# =============================================================================
# SECTION 2 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ICFRE Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")   + 1
    link_idx = col_names.index("PDF / Link")  + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(ICFRE_EXCEL_PATH)
    print(f"[icfre] Excel saved: {ICFRE_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 3 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the ICFRE Tenders pipeline.
    Returns (new_tenders, all_rows).
    """
    print("\n" + "=" * 65)
    print("[icfre] ICFRE Tenders Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(ICFRE_EXCEL_PATH):
        try:
            os.remove(ICFRE_EXCEL_PATH)
            print("[icfre] Cleared old Excel")
        except Exception:
            pass

    session = requests.Session()
    session.headers.update(HEADERS)
    all_entries = []

    # ── Fetch current tenders ──────────────────────────────────────────────────
    for label, url in [("Current", CURRENT_URL), ("Archive", ARCHIVE_URL)]:
        try:
            print(f"[icfre]   Fetching {label} tenders: {url}")
            r = session.get(url, timeout=25)
            r.raise_for_status()
            soup    = BeautifulSoup(r.text, 'html.parser')
            entries = _parse_tender_table(soup, label)
            print(f"[icfre]   {label}: {len(entries)} tenders found")
            all_entries.extend(entries)
        except Exception as e:
            print(f"[icfre]   ERROR fetching {label}: {e}")
        time.sleep(DELAY)

    print(f"[icfre] Total raw tenders: {len(all_entries)}")

    if not all_entries:
        print("[icfre] No tenders found — page structure may have changed.")
        return [], []

    # ── Score relevance + deduplicate ──────────────────────────────────────────
    _archive_cutoff = datetime.now() - timedelta(days=ARCHIVE_DAYS)
    _date_formats   = ("%d %b %Y", "%d %B %Y", "%d/%m/%Y", "%Y-%m-%d",
                       "%d-%m-%Y", "%d %b, %Y")

    for i, entry in enumerate(all_entries, 1):
        title = entry["Title"]

        # ISSUE 6: skip archive entries older than ARCHIVE_DAYS days
        if entry.get("Source") == "Archive" and entry.get("Deadline"):
            parsed_dl = None
            for _fmt in _date_formats:
                try:
                    parsed_dl = datetime.strptime(entry["Deadline"].strip(), _fmt)
                    break
                except ValueError:
                    pass
            if parsed_dl is not None and parsed_dl < _archive_cutoff:
                print(f"[icfre]   [{i:>2}/{len(all_entries)}] SKIP (archive >90d): {title[:60]}")
                continue

        if _is_goods(title):
            print(f"[icfre]   [{i:>2}/{len(all_entries)}] SKIP (goods): {title[:60]}")
            continue

        relevance = score_relevance(title, "")

        row = {
            "Title":      title,
            "Deadline":   entry["Deadline"],
            "Source":     entry["Source"],
            "Relevance":  relevance,
            "PDF / Link": entry["PDF / Link"],
        }
        all_rows.append(row)

        tid = entry["tender_id"]
        if check_if_new(tid):
            mark_as_seen(tid, title=title, source_site="ICFRE",
                         url=entry["PDF / Link"] or CURRENT_URL)
            new_tenders.append({
                "title":    title,
                "deadline": entry["Deadline"],
                "value":    "See ICFRE",
                "url":      entry["PDF / Link"] or CURRENT_URL,
            })
            print(f"[icfre]   → NEW: {title[:70]} | {relevance or '—'}")
        else:
            print(f"[icfre]   → seen: {title[:70]}")

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[icfre] Done — {len(all_rows)} tenders, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
