# =============================================================================
# wb_early_pipeline.py — World Bank Decision-Grade Early Pipeline  (v2)
#
# PURPOSE
#   Not just a listing — a decision engine.
#   Answers: "What should we start preparing for RIGHT NOW?"
#
# WHAT'S NEW IN V2
#   Task 1 — Change tracking: last_signal_score, content_hash, per-run diff
#   Task 2 — Stage classification: pipeline / approved / active
#   Task 3 — Expected procurement window: approval + 90d → approval + 180d
#   Task 4 — Firm profile integration: boosts + penalties from firm_profile.json
#   Task 5 — Smart alerts: new-high / score-upgrade / stage-change
#   Task 6 — Excel: Stage, Tender Window, Firm Fit, Adjusted Score, user columns
#   Task 7 — Debug: new/updated/unchanged, avg score delta, top upgraded
#
# USAGE
#   python3 main.py --wb-early
#   python3 main.py --wb-early --debug
#   python3 scrapers/portals/wb_early_pipeline.py --debug
# =============================================================================

from __future__ import annotations

import hashlib
import json
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests

from config.config import WB_EXCEL_PATH
from database.db import get_connection

# =============================================================================
# REGISTRY META
# =============================================================================

SCRAPER_META = {
    "flag":        "wb_early",
    "label":       "World Bank Early Pipeline",
    "group":       "api",
    "timeout":     180,
    "max_retries": 1,
    "auto":        True,
}

# =============================================================================
# SECTION 1  —  Constants
# =============================================================================

_PROJECTS_API = "https://search.worldbank.org/api/v2/projects"

_EARLY_FL = (
    "id,project_name,display_title,lending_project_name,"
    "boardapprovaldate,closingdate,totalamt,lendprojectcost,"
    "borrower,impagency,countryshortname,countryname,regionname,"
    "status,url,projdesc,projectdescription,"
    "sector1,sector2,sector3,theme1,theme2,lendinginstr"
)

_HEADER_POOL = [
    {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
     "Accept": "application/json, text/plain, */*"},
    {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36",
     "Accept": "application/json"},
    {"User-Agent": "curl/8.4.0", "Accept": "*/*"},
]

_MAX_PROJECTS          = 500
_MIN_ALERT_SCORE       = 50    # minimum adjusted_score to trigger an alert
_UPGRADE_THRESHOLD     = 12    # score must rise by this much to trigger "upgraded" alert
_PROCUREMENT_LEAD_DAYS = 90    # start_estimate = approval + 90d
_PROCUREMENT_END_DAYS  = 270   # Task 3: end_estimate = approval + 270d (was 180d)
_MAX_PROJECT_AGE_DAYS  = 548   # 18 months — older projects are filtered out
_MIN_SIGNAL_SCORE      = 12    # Task 1: minimum signal_raw to qualify (score-based gate)
_INFRA_MIN_SIGNAL_SCORE = 18   # Task 1: infra projects need a higher bar to pass

# =============================================================================
# SECTION 2  —  Consulting Signal Keywords
# =============================================================================

_STRONG_SIGNALS: List[Tuple[str, int]] = [
    ("technical assistance",        20),
    ("capacity building",           18),
    ("advisory services",           18),
    ("monitoring and evaluation",   18),
    ("impact evaluation",           16),
    ("consultancy services",        16),
    ("institutional strengthening", 15),   # already present
    ("systems strengthening",       14),   # Task 2
    ("advisory",                    14),
    ("capacity development",        14),
    ("implementation support",      13),   # Task 2
    ("knowledge management",        12),
    ("policy advisory",             12),
    ("program evaluation",          12),
    ("program management",          12),   # Task 2
    ("project management support",  12),
    ("independent verification",    12),
    ("third party monitoring",      12),
    ("technical support",           11),   # Task 2
    ("consulting",                  10),
    ("evaluation",                  10),
    ("assessment",                   9),
    ("capacity",                     8),
    ("training",                     8),
    ("research",                     7),
    ("survey",                       7),
    ("feasibility",                  7),
    ("social audit",                 7),
    ("baseline",                     6),
    ("review",                       6),
    ("study",                        5),
]

_INFRA_SIGNALS: List[str] = [
    # Fix 4: expanded list — any of these trigger infra-check
    "road", "bridge", "waterway", "irrigation",
    "civil works", "construction", "procurement of goods",
    "equipment supply", "infrastructure works", "road works",
    "bridge construction", "dam construction",
]

# Task 1: binary REQUIRED_SIGNALS gate replaced by _MIN_SIGNAL_SCORE threshold above.

# =============================================================================
# SECTION 3  —  Base Scoring (recency + sector + geography + signal)
# =============================================================================

# Fix 5: rebalanced weights — consulting signal now dominates
_MAX_RECENCY = 25   # was 30 — recency still important but capped lower
_MAX_SECTOR  = 20   # was 30 — sector alone should not drive the score
_MAX_GEO     = 15   # was 20 — geography is a tiebreaker, not a driver
_MAX_SIGNAL  = 40   # was 20 — consulting signal is the primary quality gate

_SECTOR_MAP: Dict[str, int] = {
    "health": 10, "education": 10, "agriculture": 10, "environment": 10,
    "governance": 10, "water": 10, "social protection": 10,
    "energy": 8,  "urban": 8,  "rural": 8, "gender": 8,
    "climate": 8, "nutrition": 8, "livelihoods": 8, "livelihood": 8,
    "public sector": 7, "financial sector": 7, "digital": 6, "transport": 5,
}


def _recency_score(approval_date_str: Optional[str]) -> int:
    if not approval_date_str:
        return 0
    try:
        d = datetime.strptime(str(approval_date_str)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return 0
    delta = (date.today() - d).days
    if delta < 0:    return 15
    if delta <= 90:  return _MAX_RECENCY
    if delta <= 180: return 20
    if delta <= 365: return 10
    return 0


def _sector_score(sector_text: str) -> int:
    text = sector_text.lower()
    return min(_MAX_SECTOR, sum(p for k, p in _SECTOR_MAP.items() if k in text))


def _geo_score(country: str, region: str) -> int:
    # Fix 5: tiers adjusted to match new _MAX_GEO = 15
    text = (country + " " + region).lower()
    if "india" in text:             return _MAX_GEO        # 15
    if "south asia" in text:        return 12
    if any(c in text for c in ["bangladesh", "nepal", "sri lanka", "pakistan"]): return 10
    if any(c in text for c in ["africa", "kenya", "nigeria", "ethiopia",
                                "ghana", "tanzania", "uganda"]): return 8
    return 4


def _consulting_signal_strength(description: str) -> Tuple[int, List[str]]:
    text      = description.lower()
    is_infra  = sum(1 for kw in _INFRA_SIGNALS if kw in text) >= 2
    raw, found = 0, []
    for kw, pts in _STRONG_SIGNALS:
        if kw in text:
            raw += pts
            found.append(kw)
            if raw >= _MAX_SIGNAL:
                break
    score = min(_MAX_SIGNAL, raw)
    if is_infra and score < _MAX_SIGNAL:
        score = max(0, score - 5)
    return score, found


def compute_base_score(
    approval_date: Optional[str],
    sectors: str,
    country: str,
    region: str,
    description: str,
) -> Tuple[int, str]:
    """Recency + sector + geo + consulting signal → 0-100 base score."""
    r           = _recency_score(approval_date)
    s           = _sector_score(sectors)
    g           = _geo_score(country, region)
    cs, matched = _consulting_signal_strength(description)
    total       = min(100, r + s + g + cs)
    parts       = []
    if r:  parts.append(f"recency:{r}")
    if s:  parts.append(f"sector:{s}")
    if g:  parts.append(f"geo:{g}")
    if cs: parts.append(f"signal:{cs} ({', '.join(matched[:3])})")
    return total, " | ".join(parts) if parts else "no match"

# keep old name as alias for tests that import it
compute_early_signal_score = compute_base_score


# =============================================================================
# SECTION 4  —  Firm Profile Integration  (Task 4)
# =============================================================================

_PROFILE_PATH = Path(__file__).resolve().parents[2] / "config" / "firm_profile.json"
_FIRM_PROFILE: Optional[dict] = None


def _load_firm_profile() -> dict:
    global _FIRM_PROFILE
    if _FIRM_PROFILE is not None:
        return _FIRM_PROFILE
    try:
        with open(_PROFILE_PATH, encoding="utf-8") as fh:
            _FIRM_PROFILE = json.load(fh)
    except Exception:
        _FIRM_PROFILE = {}
    return _FIRM_PROFILE


def compute_firm_fit(
    sectors: str,
    country: str,
    region: str,
    description: str,
    consulting_signals: List[str],
) -> Tuple[int, str]:
    """
    Apply firm_profile.json boosts and penalties.
    Returns (net_fit_score, reason_string).

    World Bank is always a preferred_client → guaranteed +15 base boost.
    Additional boosts for preferred sectors, region, consulting type.
    Penalties for avoid_sectors and avoid_keywords.
    """
    profile = _load_firm_profile()
    boosts   = profile.get("score_boosts",   {})
    penalties = profile.get("score_penalties", {})
    text     = (sectors + " " + description + " " + country + " " + region).lower()
    desc_lc  = description.lower()
    parts    = []
    net      = 0

    # Preferred client — World Bank is always in the list
    client_boost = boosts.get("preferred_client", 15)
    net  += client_boost
    parts.append(f"client:+{client_boost}")

    # Preferred sectors
    pref_sectors = profile.get("preferred_sectors", [])
    sec_boost    = boosts.get("preferred_sector", 10)
    matches      = [s for s in pref_sectors if s.lower() in text][:2]  # cap at 2
    if matches:
        earned = len(matches) * sec_boost
        net   += earned
        parts.append(f"sectors:+{earned} ({', '.join(matches)})")

    # Preferred regions
    pref_regions  = profile.get("preferred_regions", [])
    region_boost  = boosts.get("preferred_region", 6)
    geo_text      = (country + " " + region).lower()
    region_match  = next((r for r in pref_regions if r.lower() in geo_text), None)
    if region_match:
        net  += region_boost
        parts.append(f"region:+{region_boost} ({region_match})")

    # Preferred consulting types
    pref_consult = profile.get("preferred_consulting_types", [])
    ct_boost     = boosts.get("preferred_consulting_type", 5)
    ct_matches   = [c for c in pref_consult
                    if c.lower() in desc_lc or c.lower() in " ".join(consulting_signals)][:2]
    if ct_matches:
        earned = len(ct_matches) * ct_boost
        net   += earned
        parts.append(f"consulting_type:+{earned} ({', '.join(ct_matches)})")

    # Avoid sectors — penalise
    avoid_secs   = profile.get("avoid_sectors", [])
    sec_penalty  = abs(penalties.get("avoid_sector", 15))
    avoid_match  = [s for s in avoid_secs if s.lower() in text]
    if avoid_match:
        earned = len(avoid_match) * sec_penalty
        net   -= earned
        parts.append(f"avoid_sector:-{earned} ({', '.join(avoid_match)})")

    # Avoid keywords in description — penalise
    avoid_kws    = profile.get("avoid_keywords", [])
    kw_penalty   = abs(penalties.get("avoid_keyword_in_title", 20))
    kw_matches   = [kw for kw in avoid_kws if kw.lower() in desc_lc]
    if kw_matches:
        earned = min(len(kw_matches) * kw_penalty, 40)  # cap penalty at 40
        net   -= earned
        parts.append(f"avoid_kw:-{earned}")

    reason = " | ".join(parts) if parts else "WB client: neutral"
    return net, reason


# =============================================================================
# SECTION 5  —  Project Stage Classification  (Task 2)
# =============================================================================

def _classify_stage(status: str, approval_date: Optional[str]) -> str:
    """
    Map WB API status → internal stage label.

    pipeline  — project concept / under preparation (no board approval yet)
    approved  — board approved, implementation not yet started
    active    — full implementation underway
    """
    s = status.strip().lower()
    if "pipeline" in s:
        return "pipeline"
    if "active" in s:
        return "active"
    if "approved" in s:
        return "approved"
    # Fallback: infer from approval date recency
    if approval_date:
        try:
            d     = datetime.strptime(str(approval_date)[:10], "%Y-%m-%d").date()
            delta = (date.today() - d).days
            if delta < 0:   return "approved"   # future-dated → just approved
            if delta <= 90: return "approved"   # very recent → likely just approved
        except (ValueError, TypeError):
            pass
    return "active"


# =============================================================================
# SECTION 6  —  Procurement Window  (Task 3)
# =============================================================================

def _procurement_window(approval_date: Optional[str]) -> Tuple[Optional[str], Optional[str], str]:
    """
    Compute (start_estimate, end_estimate, display_text).
    start = approval_date + 90 days
    end   = approval_date + 270 days  (Task 3: extended from 180d)

    Display: "Jun 2026 – Dec 2026"  or  "—" if no approval date.
    """
    if not approval_date:
        return None, None, "—"
    try:
        d     = datetime.strptime(str(approval_date)[:10], "%Y-%m-%d").date()
        start = d + timedelta(days=_PROCUREMENT_LEAD_DAYS)
        end   = d + timedelta(days=_PROCUREMENT_END_DAYS)
        label = f"{start.strftime('%b %Y')} – {end.strftime('%b %Y')}"
        return str(start), str(end), label
    except (ValueError, TypeError):
        return None, None, "—"


# =============================================================================
# SECTION 7  —  Change Tracking  (Task 1)
# =============================================================================

def _content_hash(consulting_signal: str, sector: str, description: str) -> str:
    """MD5 fingerprint of the three most volatile fields."""
    raw = f"{consulting_signal}|{sector}|{description[:500]}"
    return hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()


@dataclass
class ChangeResult:
    """Per-project change classification returned by _upsert_early_project."""
    is_new:        bool  = False
    is_upgraded:   bool  = False   # score increased by >= UPGRADE_THRESHOLD
    is_changed:    bool  = False   # content hash changed (description/signals/sector)
    stage_changed: bool  = False   # stage label flipped
    old_score:     int   = 0
    new_score:     int   = 0
    old_stage:     str   = ""
    new_stage:     str   = ""

    @property
    def score_delta(self) -> int:
        return self.new_score - self.old_score

    @property
    def alert_type(self) -> Optional[str]:
        """
        Returns the highest-priority alert type or None.
        Priority: new_high > stage_change > upgraded
        """
        if self.is_new and self.new_score >= _MIN_ALERT_SCORE:
            return "new_high"
        if self.stage_changed:
            return "stage_change"
        if self.is_upgraded:
            return "upgraded"
        return None


# =============================================================================
# SECTION 8  —  API Helpers
# =============================================================================

def _api_get(url: str, params: dict, retries: int = 3) -> Optional[dict]:
    for attempt in range(retries):
        headers = _HEADER_POOL[attempt % len(_HEADER_POOL)]
        if attempt > 0:
            time.sleep(3 * (2 ** attempt))
        try:
            r = requests.get(url, params=params, headers=headers,
                             timeout=45 if attempt == 2 else 30)
            if r.status_code == 200 and r.text.strip():
                return r.json()
        except Exception as exc:
            print(f"[wb_early] API attempt {attempt+1}/{retries}: {exc}")
    return None


def _resolve(record: dict, keys: List[str], fallback: str = "") -> str:
    for k in keys:
        v = record.get(k)
        if v and str(v).strip() not in ("", "0", "None", "null"):
            return str(v).strip()
    return fallback


def _extract_sector_text(record: dict) -> str:
    parts = []
    for f in ["sector1", "sector2", "sector3", "theme1", "theme2", "lendinginstr"]:
        val = record.get(f)
        if isinstance(val, dict):
            name = val.get("Name") or val.get("name") or ""
            if name: parts.append(str(name))
        elif val and str(val).strip():
            parts.append(str(val).strip())
    return " | ".join(p for p in parts if p)


def _extract_description(record: dict) -> str:
    for key in ["projdesc", "projectdescription", "project_name",
                "display_title", "lending_project_name"]:
        val = record.get(key)
        if val and isinstance(val, str) and len(val) > 20:
            return val.strip()
    return _resolve(record, ["project_name", "display_title"], fallback="")


# =============================================================================
# SECTION 9  —  Project Fetching
# =============================================================================

def _fetch_projects(debug: bool = False) -> List[dict]:
    offset, all_records = 0, []
    while True:
        params = {"format": "json", "rows": 200, "os": offset,
                  "countrycode": "IN", "fl": _EARLY_FL, "status": "Active"}
        data = _api_get(_PROJECTS_API, params)
        if data is None or "projects" not in data:
            if data is None:
                print("[wb_early] Projects API unavailable")
            break
        batch = list(data["projects"].values())
        if not batch:
            break
        all_records.extend(batch)
        offset += len(batch)
        if debug:
            print(f"[wb_early]   Fetched page: {len(batch)} records "
                  f"(total: {len(all_records)})")
        if offset >= _MAX_PROJECTS or len(batch) < 200:
            break
        time.sleep(0.3)

    # Pipeline-stage projects (not yet board-approved)
    pipe_data = _api_get(_PROJECTS_API, {
        "format": "json", "rows": 100, "os": 0,
        "countrycode": "IN", "fl": _EARLY_FL, "status": "Pipeline",
    })
    if pipe_data and "projects" in pipe_data:
        pipe_batch = list(pipe_data["projects"].values())
        all_records.extend(pipe_batch)
        if debug:
            print(f"[wb_early]   Pipeline stage: {len(pipe_batch)} records")
    return all_records


# =============================================================================
# SECTION 10  —  Row Builder
# =============================================================================

def _action_category(adjusted_score: int) -> str:
    """Task 4: map adjusted score to a clear action label."""
    if adjusted_score >= 80: return "BID SOON"
    if adjusted_score >= 60: return "TRACK"
    return "LOW"


def _build_early_row(record: dict,
                     drops: Optional[dict] = None) -> Optional[dict]:
    """
    Convert raw WB API record → enriched early-pipeline dict.
    Returns None if filtered out.
    drops — when provided (debug mode), incremented by filter reason:
            drops["age"], drops["signal"], drops["infra"]
    """
    pid = _resolve(record, ["id"])
    if not pid or not pid.startswith("P"):
        return None

    name        = _resolve(record, ["project_name", "display_title",
                                     "lending_project_name"], fallback=pid)
    country     = _resolve(record, ["countryshortname", "countryname"], fallback="India")
    region      = _resolve(record, ["regionname"], fallback="South Asia")
    status_raw  = _resolve(record, ["status"], fallback="Active")
    approval_dt = (_resolve(record, ["boardapprovaldate"]) or "")[:10] or None

    # Fix 1 — age filter: drop projects older than 18 months
    if approval_dt:
        try:
            appr_d = datetime.strptime(approval_dt, "%Y-%m-%d").date()
            if (date.today() - appr_d).days > _MAX_PROJECT_AGE_DAYS:
                if drops is not None: drops["age"] = drops.get("age", 0) + 1
                return None
        except (ValueError, TypeError):
            pass
    elif "pipeline" not in status_raw.lower():
        # No approval date and not a pipeline project → nothing actionable
        if drops is not None: drops["age"] = drops.get("age", 0) + 1
        return None

    sectors     = _extract_sector_text(record)
    description = _extract_description(record)

    url = _resolve(record, ["url"])
    if not url or not url.startswith("http"):
        url = (f"https://projects.worldbank.org/en/"
               f"projects-operations/project-detail/{pid}")

    # ── Consulting signal ──────────────────────────────────────────────────────
    full_text  = description + " " + name + " " + sectors
    full_lower = full_text.lower()

    signal_raw, matched = _consulting_signal_strength(full_text)

    # Task 1 — score-based gate (replaces binary REQUIRED_SIGNALS check)
    if signal_raw < _MIN_SIGNAL_SCORE:
        if drops is not None: drops["signal"] = drops.get("signal", 0) + 1
        return None

    # Fix 4 — infrastructure exclusion: infra projects need a higher signal bar
    has_infra = any(kw in full_lower for kw in _INFRA_SIGNALS)
    if has_infra and signal_raw < _INFRA_MIN_SIGNAL_SCORE:
        if drops is not None: drops["infra"] = drops.get("infra", 0) + 1
        return None

    consulting_signal = ", ".join(matched[:8])

    # ── Stage ─────────────────────────────────────────────────────────────────
    stage = _classify_stage(status_raw, approval_dt)

    # ── Procurement window ─────────────────────────────────────────────────────
    start_est, end_est, window_label = _procurement_window(approval_dt)

    # Fix 2 — skip projects whose procurement window has already closed
    if end_est:
        try:
            if datetime.strptime(end_est, "%Y-%m-%d").date() < date.today():
                if drops is not None: drops["age"] = drops.get("age", 0) + 1
                return None
        except (ValueError, TypeError):
            pass

    # ── Base early signal score ────────────────────────────────────────────────
    base_score, score_reason = compute_base_score(
        approval_date=approval_dt,
        sectors=sectors,
        country=country,
        region=region,
        description=full_text,
    )

    # ── Firm fit boost/penalty ─────────────────────────────────────────────────
    firm_fit, firm_reason = compute_firm_fit(
        sectors=sectors,
        country=country,
        region=region,
        description=description,
        consulting_signals=matched,
    )

    adjusted_score = min(100, max(0, base_score + firm_fit))

    # ── Content hash (for change detection) ───────────────────────────────────
    chash = _content_hash(consulting_signal, sectors, description)

    return {
        "action_category":       _action_category(adjusted_score),  # Task 4
        "project_id":            pid,
        "project_name":          name[:500],
        "country":               country[:100],
        "region":                region[:100],
        "sector":                sectors[:500],
        "approval_date":         approval_dt,
        "status":                status_raw[:50],
        "project_stage":         stage,
        "procurement_plan_flag": 1 if approval_dt else 0,
        "consulting_signal":     consulting_signal[:500],
        "early_signal_score":    base_score,
        "score_reason":          score_reason[:500],
        "firm_fit_score":        firm_fit,
        "firm_fit_reason":       firm_reason[:500],
        "adjusted_score":        adjusted_score,
        "start_estimate":        start_est,
        "end_estimate":          end_est,
        "window_label":          window_label,
        "project_url":           url[:1000],
        "description":           description[:2000],
        "content_hash":          chash,
    }


# =============================================================================
# SECTION 11  —  DB Upsert + Change Detection  (Task 1)
# =============================================================================

def _upsert_early_project(conn, row: dict) -> ChangeResult:
    """
    INSERT new or UPDATE existing project. Returns a ChangeResult.
    """
    cur = conn.cursor()
    res = ChangeResult(new_score=row["adjusted_score"],
                       new_stage=row["project_stage"])
    try:
        cur.execute(
            """SELECT id, adjusted_score, content_hash, project_stage
                 FROM world_bank_early_pipeline
                WHERE project_id = %s LIMIT 1""",
            (row["project_id"],)
        )
        existing = cur.fetchone()

        if existing is None:
            # ── INSERT ────────────────────────────────────────────────────────
            res.is_new = True
            cur.execute("""
                INSERT INTO world_bank_early_pipeline
                    (project_id, project_name, country, region, sector,
                     approval_date, status, project_stage,
                     procurement_plan_flag, consulting_signal,
                     early_signal_score, last_signal_score,
                     score_reason, firm_fit_score, firm_fit_reason,
                     adjusted_score, start_estimate, end_estimate,
                     content_hash, project_url, description,
                     first_seen, last_updated, notified)
                VALUES
                    (%s,%s,%s,%s,%s,
                     %s,%s,%s,
                     %s,%s,
                     %s,%s,
                     %s,%s,%s,
                     %s,%s,%s,
                     %s,%s,%s,
                     NOW(),NOW(),0)
            """, (
                row["project_id"],  row["project_name"],  row["country"],
                row["region"],      row["sector"],
                row.get("approval_date"), row["status"], row["project_stage"],
                row["procurement_plan_flag"], row["consulting_signal"],
                row["early_signal_score"],  0,   # last_signal_score=0 for new
                row["score_reason"], row["firm_fit_score"], row["firm_fit_reason"],
                row["adjusted_score"], row.get("start_estimate"), row.get("end_estimate"),
                row["content_hash"], row["project_url"], row.get("description",""),
            ))
        else:
            db_id, old_score, old_hash, old_stage = existing
            res.old_score  = old_score  or 0
            res.old_stage  = old_stage  or ""
            res.is_upgraded    = (row["adjusted_score"] - res.old_score) >= _UPGRADE_THRESHOLD
            res.is_changed     = (row["content_hash"] != (old_hash or ""))
            res.stage_changed  = (row["project_stage"] != res.old_stage and
                                  res.old_stage != "")

            # ── UPDATE ────────────────────────────────────────────────────────
            cur.execute("""
                UPDATE world_bank_early_pipeline
                   SET project_name       = %s,
                       sector             = %s,
                       status             = %s,
                       project_stage      = %s,
                       consulting_signal  = %s,
                       last_signal_score  = adjusted_score,
                       early_signal_score = %s,
                       score_reason       = %s,
                       firm_fit_score     = %s,
                       firm_fit_reason    = %s,
                       adjusted_score     = %s,
                       start_estimate     = %s,
                       end_estimate       = %s,
                       content_hash       = %s,
                       last_updated       = NOW()
                 WHERE project_id = %s
            """, (
                row["project_name"],     row["sector"],
                row["status"],           row["project_stage"],
                row["consulting_signal"],
                row["early_signal_score"], row["score_reason"],
                row["firm_fit_score"],   row["firm_fit_reason"],
                row["adjusted_score"],
                row.get("start_estimate"), row.get("end_estimate"),
                row["content_hash"],
                row["project_id"],
            ))

        conn.commit()
    finally:
        cur.close()
    return res


def _get_pending_alerts(conn) -> List[dict]:
    """Fetch projects pending notification (notified=0, score >= threshold)."""
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT project_id, project_name, country, sector, project_stage,
                   adjusted_score, early_signal_score, firm_fit_score,
                   consulting_signal, project_url, approval_date,
                   start_estimate, end_estimate
              FROM world_bank_early_pipeline
             WHERE notified = 0 AND adjusted_score >= %s
             ORDER BY adjusted_score DESC
        """, (_MIN_ALERT_SCORE,))
        return cur.fetchall()
    finally:
        cur.close()


def _mark_notified(conn, project_ids: List[str]) -> None:
    if not project_ids:
        return
    placeholders = ", ".join(["%s"] * len(project_ids))
    cur = conn.cursor()
    try:
        cur.execute(
            f"UPDATE world_bank_early_pipeline "
            f"SET notified=1 WHERE project_id IN ({placeholders})",
            project_ids,
        )
        conn.commit()
    finally:
        cur.close()


def _get_all_for_excel(conn) -> List[dict]:
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT project_id, project_name, country, region, sector,
                   approval_date, status, project_stage,
                   consulting_signal, early_signal_score, last_signal_score,
                   firm_fit_score, firm_fit_reason, adjusted_score,
                   score_reason, start_estimate, end_estimate,
                   project_url, description, first_seen
              FROM world_bank_early_pipeline
             ORDER BY adjusted_score DESC, approval_date DESC
        """)
        return cur.fetchall()
    finally:
        cur.close()


# =============================================================================
# SECTION 12  —  Excel Output  (Tasks 5 & 6)
# =============================================================================

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Column definition: (name, width)
_EARLY_COLUMNS = [
    ("Project ID",            12),
    ("Project Name",          45),
    ("Stage",                 12),
    ("Country",               14),
    ("Sectors / Themes",      32),
    ("Approval Date",         14),
    ("Expected Tender Window", 22),
    ("Consulting Signals",    32),
    ("Signal Score",          12),
    ("Firm Fit",              10),
    ("Adjusted Score",        13),
    ("Action",                11),   # Task 4: BID SOON / TRACK / LOW
    ("Score Breakdown",       30),
    ("Project URL",           40),
    ("My Interest",           16),   # user input — gold header
    ("Next Action",           22),   # user input — gold header
]

_HDR_FILL     = PatternFill("solid", fgColor="1F3864")   # navy (matches WB style)
_USR_HDR_FILL = PatternFill("solid", fgColor="FFD966")   # gold — user input columns
_SCORE_HIGH   = PatternFill("solid", fgColor="C6EFCE")   # green  ≥70
_SCORE_MED    = PatternFill("solid", fgColor="FFEB9C")   # amber  40-69
_SCORE_LOW    = PatternFill("solid", fgColor="FFC7CE")   # red    <40
_ALT_FILL     = PatternFill("solid", fgColor="F5F8FF")
_STAGE_FILLS  = {
    "pipeline": PatternFill("solid", fgColor="BDD7EE"),  # light blue
    "approved": PatternFill("solid", fgColor="C6EFCE"),  # light green
    "active":   PatternFill("solid", fgColor="E2EFDA"),  # dark green tint
}
_HDR_FONT   = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
_USR_HDR_FNT = Font(name="Calibri", color="7B5E00", bold=True, size=11)
_BODY_FONT  = Font(name="Calibri", size=10)
_LINK_FONT  = Font(name="Calibri", size=10, color="1155CC", underline="single")
_BORD = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)

_USER_COLS = {"My Interest", "Next Action"}


def _fmt_score_delta(last: int, current: int) -> str:
    """Format firm_fit_score with explicit sign, e.g. '+15' or '-5'."""
    if current > 0:
        return f"+{current}"
    if current < 0:
        return str(current)
    return "0"


def _save_early_pipeline_excel(rows: List[dict]) -> None:
    if not rows:
        print("[wb_early] No rows to write — skipping Excel")
        return

    if Path(WB_EXCEL_PATH).exists():
        try:
            wb = load_workbook(WB_EXCEL_PATH)
        except Exception:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if "Early Pipeline" in wb.sheetnames:
        del wb["Early Pipeline"]
    ws = wb.create_sheet("Early Pipeline")

    col_names = [c[0] for c in _EARLY_COLUMNS]

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 38
    for ci, (col_name, width) in enumerate(_EARLY_COLUMNS, 1):
        cell = ws.cell(1, ci, col_name)
        if col_name in _USER_COLS:
            cell.font  = _USR_HDR_FNT
            cell.fill  = _USR_HDR_FILL
        else:
            cell.font  = _HDR_FONT
            cell.fill  = _HDR_FILL
        cell.border    = _BORD
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = ws.cell(2, 3)

    stage_ci   = col_names.index("Stage") + 1
    adj_ci     = col_names.index("Adjusted Score") + 1
    action_ci  = col_names.index("Action") + 1          # Task 4
    sig_ci     = col_names.index("Signal Score") + 1
    fit_ci     = col_names.index("Firm Fit") + 1
    url_ci     = col_names.index("Project URL") + 1
    user_cis   = {col_names.index(c) + 1 for c in _USER_COLS}

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 44
        alt          = _ALT_FILL if ri % 2 == 0 else None
        adj_score    = int(row.get("adjusted_score") or 0)
        base_score   = int(row.get("early_signal_score") or 0)
        fit_score    = int(row.get("firm_fit_score") or 0)
        stage        = str(row.get("project_stage") or "active").lower()

        # Build window display — prefer stored label, fall back to computing
        start_est = row.get("start_estimate")
        end_est   = row.get("end_estimate")
        if start_est and end_est:
            try:
                s = datetime.strptime(str(start_est)[:10], "%Y-%m-%d")
                e = datetime.strptime(str(end_est)[:10],   "%Y-%m-%d")
                window_str = f"{s.strftime('%b %Y')} – {e.strftime('%b %Y')}"
            except ValueError:
                window_str = "—"
        else:
            _, _, window_str = _procurement_window(
                str(row.get("approval_date") or "")[:10] or None
            )

        values = {
            "Project ID":             row.get("project_id", ""),
            "Project Name":           row.get("project_name", ""),
            "Stage":                  stage.capitalize(),
            "Country":                row.get("country", ""),
            "Sectors / Themes":       row.get("sector", ""),
            "Approval Date":          str(row.get("approval_date") or ""),
            "Expected Tender Window": window_str,
            "Consulting Signals":     row.get("consulting_signal", ""),
            "Signal Score":           base_score,
            "Firm Fit":               _fmt_score_delta(0, fit_score),
            "Adjusted Score":         adj_score,
            "Action":                 row.get("action_category") or _action_category(adj_score),
            "Score Breakdown":        row.get("score_reason", ""),
            "Project URL":            row.get("project_url", ""),
            "My Interest":            "",
            "Next Action":            "",
        }

        for ci, col_name in enumerate(col_names, 1):
            val  = values.get(col_name, "")
            cell = ws.cell(ri, ci, val)
            cell.border    = _BORD
            cell.alignment = Alignment(wrap_text=True, vertical="top",
                                       horizontal="left")

            if ci in user_cis:
                # User input columns — light gold background
                cell.font = _BODY_FONT
                cell.fill = PatternFill("solid", fgColor="FFFDE7")
            elif ci == url_ci and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font      = _LINK_FONT
                if alt: cell.fill = alt
            elif ci == stage_ci:
                cell.fill = _STAGE_FILLS.get(stage, _ALT_FILL)
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == adj_ci:
                cell.fill = (_SCORE_HIGH if adj_score >= 70
                             else _SCORE_MED if adj_score >= 40
                             else _SCORE_LOW)
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == action_ci:
                # Task 4: BID SOON = red-orange, TRACK = amber, LOW = grey
                _act = str(val)
                if _act == "BID SOON":
                    cell.fill = PatternFill("solid", fgColor="FF4C4C")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                elif _act == "TRACK":
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="7B4F00")
                else:
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    cell.font = Font(name="Calibri", size=10, color="888888")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == sig_ci:
                cell.fill = (_SCORE_HIGH if base_score >= 70
                             else _SCORE_MED if base_score >= 40
                             else _SCORE_LOW)
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif ci == fit_ci:
                # Green if positive, red if negative, neutral if zero
                fit_num = row.get("firm_fit_score", 0) or 0
                if fit_num > 0:
                    cell.fill = PatternFill("solid", fgColor="E2EFDA")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="375623")
                elif fit_num < 0:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
                    cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                else:
                    cell.font = _BODY_FONT
                    if alt: cell.fill = alt
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.font = _BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_EARLY_COLUMNS))}1"
    wb.save(WB_EXCEL_PATH)
    print(f"[wb_early] Excel saved → {WB_EXCEL_PATH}  "
          f"(sheet: 'Early Pipeline', {len(rows)} row(s))")


# =============================================================================
# SECTION 13  —  Alert Messages  (Task 5)
# =============================================================================

def _score_emoji(score: int) -> str:
    return "🔥" if score >= 80 else "⭐" if score >= 60 else "📌"


def _build_alert_message(
    new_high:     List[dict],
    upgraded:     List[Tuple[dict, ChangeResult]],
    stage_changed: List[Tuple[dict, ChangeResult]],
) -> str:
    """
    Build a single Telegram-ready message covering all three alert types.
    Returns empty string if nothing to report.
    """
    sections = []

    # ── New high-score projects ────────────────────────────────────────────────
    if new_high:
        lines = [
            "🌍 *World Bank — New Early Pipeline*",
            f"_{len(new_high)} new project(s) with high consulting potential_",
            "",
        ]
        for p in new_high[:6]:
            score  = p.get("adjusted_score", 0)
            window = ""
            se, ee = p.get("start_estimate"), p.get("end_estimate")
            if se and ee:
                try:
                    s = datetime.strptime(str(se)[:10], "%Y-%m-%d")
                    e = datetime.strptime(str(ee)[:10], "%Y-%m-%d")
                    window = f"  📆 Tender window: _{s.strftime('%b %Y')} – {e.strftime('%b %Y')}_\n"
                except ValueError:
                    pass
            lines += [
                f"{_score_emoji(score)} *{str(p['project_name'])[:75]}*",
                f"  🆔 {p['project_id']}  |  📍 {p.get('country','India')}  "
                f"|  🏷 {str(p.get('project_stage','')).capitalize()}",
                f"  🎯 Score: *{score}/100*  (base: {p.get('early_signal_score',0)}, "
                f"fit: {_fmt_score_delta(0, p.get('firm_fit_score',0))})",
                f"  💡 _{str(p.get('consulting_signal',''))[:80]}_",
                window.rstrip(),
                f"  🔗 [View Project]({p.get('project_url','')})",
                "",
            ]
        sections.append("\n".join(lines))

    # ── Stage changes ──────────────────────────────────────────────────────────
    if stage_changed:
        lines = [
            "🔄 *World Bank — Stage Upgrade*",
            f"_{len(stage_changed)} project(s) advanced to next stage_",
            "",
        ]
        for p, cr in stage_changed[:4]:
            score = p.get("adjusted_score", 0)
            lines += [
                f"⬆️ *{str(p['project_name'])[:70]}*",
                f"  🆔 {p['project_id']}  |  "
                f"{cr.old_stage.capitalize()} → *{cr.new_stage.capitalize()}*",
                f"  🎯 Score: {score}/100",
                f"  🔗 [View Project]({p.get('project_url','')})",
                "",
            ]
        sections.append("\n".join(lines))

    # ── Score upgrades ─────────────────────────────────────────────────────────
    if upgraded:
        lines = [
            "📈 *World Bank — Signal Upgrade*",
            f"_{len(upgraded)} project(s) significantly improved_",
            "",
        ]
        for p, cr in upgraded[:4]:
            lines += [
                f"⬆️ *{str(p['project_name'])[:70]}*",
                f"  🆔 {p['project_id']}  |  "
                f"Score: {cr.old_score} → *{cr.new_score}* (+{cr.score_delta})",
                f"  💡 _{str(p.get('consulting_signal',''))[:80]}_",
                f"  🔗 [View Project]({p.get('project_url','')})",
                "",
            ]
        sections.append("\n".join(lines))

    return "\n\n---\n\n".join(sections)


# =============================================================================
# SECTION 14  —  Debug Report  (Task 7)
# =============================================================================

@dataclass
class RunStats:
    fetched:      int = 0
    signaled:     int = 0
    new:          int = 0
    upgraded:     int = 0
    changed:      int = 0
    unchanged:    int = 0
    stage_chg:    int = 0
    # Task 5: drop counters (populated only in debug mode)
    drops_age:    int = 0
    drops_signal: int = 0
    drops_infra:  int = 0
    score_deltas: List[int] = field(default_factory=list)
    top_upgraded: List[Tuple[dict, ChangeResult]] = field(default_factory=list)


def _print_debug_report(stats: RunStats, top5: List[dict]) -> None:
    sep  = "─" * 62
    avg  = (sum(stats.score_deltas) / len(stats.score_deltas)
            if stats.score_deltas else 0)

    print(f"\n[wb_early] {sep}")
    print(f"[wb_early] DEBUG REPORT  —  World Bank Early Pipeline  (v2)")
    print(f"[wb_early] {sep}")
    print(f"[wb_early]   Projects fetched          : {stats.fetched}")
    print(f"[wb_early]   With consulting signals   : {stats.signaled}")
    print(f"[wb_early]   ─────────────────────────────────────────────")
    # Task 5: drop breakdown
    total_dropped = stats.drops_age + stats.drops_signal + stats.drops_infra
    print(f"[wb_early]   DROPPED  ({total_dropped} total):")
    print(f"[wb_early]     Age / expired window    : {stats.drops_age}")
    print(f"[wb_early]     Weak signal (< {_MIN_SIGNAL_SCORE} pts) : {stats.drops_signal}")
    print(f"[wb_early]     Infra (no consulting)   : {stats.drops_infra}")
    print(f"[wb_early]   ─────────────────────────────────────────────")
    print(f"[wb_early]   New this run              : {stats.new}")
    print(f"[wb_early]   Updated (score ↑≥{_UPGRADE_THRESHOLD})     : {stats.upgraded}")
    print(f"[wb_early]   Content changed           : {stats.changed}")
    print(f"[wb_early]   Stage changed             : {stats.stage_chg}")
    print(f"[wb_early]   Unchanged                 : {stats.unchanged}")
    print(f"[wb_early]   Avg score delta (updated) : {avg:+.1f}")
    print(f"[wb_early] {sep}")

    if top5:
        print(f"[wb_early]   TOP 5 — HIGHEST ADJUSTED SCORE:")
        for i, p in enumerate(top5[:5], 1):
            fit   = p.get("firm_fit_score", 0) or 0
            fit_s = f"+{fit}" if fit > 0 else str(fit)
            print(f"[wb_early]   {i}. [{p['adjusted_score']:>3}/100] "
                  f"{p['project_id']}  {str(p['project_name'])[:50]}")
            print(f"[wb_early]        Stage  : {p.get('project_stage','—').capitalize()}"
                  f"  |  Fit: {fit_s}"
                  f"  |  Window: {p.get('window_label','—')}")
            print(f"[wb_early]        Signals: {str(p.get('consulting_signal',''))[:70]}")

    if stats.top_upgraded:
        print(f"[wb_early] {sep}")
        print(f"[wb_early]   TOP UPGRADED PROJECTS (score increase):")
        for p, cr in sorted(stats.top_upgraded,
                             key=lambda x: x[1].score_delta, reverse=True)[:3]:
            print(f"[wb_early]   ↑  {p['project_id']}  "
                  f"{cr.old_score} → {cr.new_score}  (+{cr.score_delta})"
                  f"  {str(p['project_name'])[:50]}")

    print(f"[wb_early] {sep}\n")


# =============================================================================
# SECTION 15  —  Main run()
# =============================================================================

_module_debug: bool = False


def set_debug(flag: bool) -> None:
    global _module_debug
    _module_debug = flag


def run(debug: bool = False) -> Tuple[List[dict], List[dict]]:
    """
    Entry point: fetch → score → diff → store → alert → export.
    Returns (new_projects, all_rows) in standard pipeline contract.
    """
    debug = debug or _module_debug
    if debug:
        print("[wb_early] Debug mode ON")

    print("[wb_early] Fetching World Bank project data…")
    raw_projects = _fetch_projects(debug=debug)
    print(f"[wb_early] {len(raw_projects)} raw projects fetched")

    # ── Build enriched rows ───────────────────────────────────────────────────
    # Task 5: pass drops dict in debug mode to track per-filter drop counts
    drops: Optional[dict] = {} if debug else None
    early_rows: List[dict] = []
    for rec in raw_projects:
        row = _build_early_row(rec, drops=drops)
        if row:
            early_rows.append(row)
    print(f"[wb_early] {len(early_rows)} with consulting signals")

    # ── DB upsert with change detection ──────────────────────────────────────
    stats = RunStats(
        fetched=len(raw_projects),
        signaled=len(early_rows),
        drops_age=drops.get("age", 0) if drops else 0,
        drops_signal=drops.get("signal", 0) if drops else 0,
        drops_infra=drops.get("infra", 0) if drops else 0,
    )
    new_projects:   List[dict] = []
    alerted_new:    List[dict] = []
    alerted_upg:    List[Tuple[dict, ChangeResult]] = []
    alerted_stage:  List[Tuple[dict, ChangeResult]] = []

    try:
        conn = get_connection()

        for row in early_rows:
            cr = _upsert_early_project(conn, row)

            if cr.is_new:
                stats.new += 1
                new_projects.append(row)
                if cr.new_score >= _MIN_ALERT_SCORE:
                    alerted_new.append(row)
            else:
                if cr.is_upgraded:
                    stats.upgraded += 1
                    alerted_upg.append((row, cr))
                    stats.top_upgraded.append((row, cr))
                if cr.is_changed:
                    stats.changed += 1
                if cr.stage_changed:
                    stats.stage_chg += 1
                    alerted_stage.append((row, cr))
                if not cr.is_upgraded and not cr.is_changed and not cr.stage_changed:
                    stats.unchanged += 1
                if cr.old_score > 0:
                    stats.score_deltas.append(cr.score_delta)

        # Fetch pending alerts (score >= threshold, notified=0)
        pending = _get_pending_alerts(conn)
        all_db_rows = _get_all_for_excel(conn)
        conn.close()

    except Exception as exc:
        print(f"[wb_early] DB error: {exc}")
        pending     = []
        all_db_rows = early_rows

    # ── Send notifications ────────────────────────────────────────────────────
    msg = _build_alert_message(alerted_new, alerted_upg, alerted_stage)
    if msg and pending:
        try:
            from notifier import send_rich_alert
            send_rich_alert(msg)
            print(f"[wb_early] Alert sent: "
                  f"{len(alerted_new)} new, "
                  f"{len(alerted_upg)} upgraded, "
                  f"{len(alerted_stage)} stage-changed")
            try:
                conn2 = get_connection()
                _mark_notified(conn2, [p["project_id"] for p in pending])
                conn2.close()
            except Exception:
                pass
        except Exception as exc:
            print(f"[wb_early] Notification error (non-fatal): {exc}")

    # ── Excel ─────────────────────────────────────────────────────────────────
    # Attach window_label for rows coming from DB (they lack it)
    for r in all_db_rows:
        if "window_label" not in r:
            _, _, r["window_label"] = _procurement_window(
                str(r.get("approval_date") or "")[:10] or None
            )
    try:
        _save_early_pipeline_excel(all_db_rows)
    except Exception as exc:
        print(f"[wb_early] Excel error: {exc}")

    # ── Debug report ──────────────────────────────────────────────────────────
    if debug:
        top5 = sorted(early_rows, key=lambda r: r["adjusted_score"], reverse=True)
        _print_debug_report(stats, top5)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f"\n{'='*58}")
    print(f"  EARLY PIPELINE SUMMARY — World Bank (v2)")
    print(f"{'='*58}")
    print(f"  Projects fetched       : {stats.fetched}")
    print(f"  With consulting signal : {stats.signaled}")
    print(f"  New                    : {stats.new}")
    print(f"  Upgraded (score ↑≥{_UPGRADE_THRESHOLD})  : {stats.upgraded}")
    print(f"  Stage changes          : {stats.stage_chg}")
    print(f"  Unchanged              : {stats.unchanged}")
    if early_rows:
        top = max(early_rows, key=lambda r: r["adjusted_score"])
        print(f"  Highest adjusted score : {top['adjusted_score']}/100")
        print(f"    → {top['project_id']}  {str(top['project_name'])[:48]}")
        print(f"    Tender window: {top.get('window_label','—')}")
    print(f"{'='*58}\n")

    return new_projects, early_rows


# =============================================================================
# SECTION 16  —  CLI Entry Point
# =============================================================================

if __name__ == "__main__":
    import sys as _sys
    _debug = "--debug" in _sys.argv
    new, rows = run(debug=_debug)
    print(f"\nNew projects : {len(new)}")
    print(f"Total rows   : {len(rows)}")
