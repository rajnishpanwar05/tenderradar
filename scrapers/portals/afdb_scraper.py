# =============================================================================
# afdb_pipeline.py — African Development Bank (AfDB) Consultants Pipeline
#
# Site   : https://www.afdb.org/en/about-us/careers/current-vacancies/consultants
# Method : requests + BeautifulSoup (server-rendered Drupal HTML, table layout)
# Login  : None required
# CAPTCHA: None
#
# What this does:
#   1. Paginates through all consultant vacancy pages (?page=0, 1, 2, ...)
#   2. Parses the HTML table — columns: Title+Link, Type, Published, Closing
#   3. Visits each detail page for full description + country
#   4. Scores relevance against IDCG expertise
#   5. Deduplicates via MySQL — only NEW opportunities trigger notification
#   6. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
#
# AfDB covers: firm consultants, individual consultants, EOIs across Africa/globally
# Focus: International development, infrastructure, environment, social sectors
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://www.afdb.org"
LISTING_URL = (
    f"{BASE_URL}/en/about-us/careers/current-vacancies/consultants"
    "?order=field_consultant_type&sort=asc"
)
DELAY       = 1.0   # polite crawl delay
MAX_PAGES   = 20    # 20 items/page × 20 pages = 400 max

AFDB_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "AfDB_Consultants_Master.xlsx")

HEADERS = {
    "User-Agent":      ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/120.0.0.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Skip firm consultants only posting goods/logistics — keep everything consulting
SKIP_PATTERNS = [
    "supply of", "goods", "catering", "cleaning", "printing services",
    "transport", "vehicle", "it equipment", "hardware",
]

def _is_goods(text: str) -> bool:
    t = text.lower()
    return any(p in t for p in SKIP_PATTERNS)


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",           65),
    ("Consultant Type", 22),
    ("Country",         25),
    ("Published",       16),
    ("Closing Date",    18),
    ("Description",     65),
    ("Relevance",       40),
    ("Detail Link",     55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse listing table
# =============================================================================

def _parse_table(soup: BeautifulSoup) -> list:
    """Parse the AfDB consultant vacancy table rows."""
    entries = []

    # AfDB uses an article > table structure
    table = soup.find('table')
    if not table:
        # Fallback: look for any div/article with links to /consultants/
        for link in soup.find_all('a', href=re.compile(r'/consultants/')):
            href  = link.get('href', '')
            title = link.get_text(strip=True)
            if title and len(title) > 8:
                entries.append({
                    "Title":           title,
                    "Consultant Type": "",
                    "Published":       "",
                    "Closing Date":    "",
                    "detail_url":      href if href.startswith('http') else BASE_URL + href,
                })
        return entries

    rows = table.find_all('tr')
    for row in rows[1:]:   # skip header row
        cells = row.find_all(['td', 'th'])
        if len(cells) < 2:
            continue

        # Column 0: Title + Link
        title_cell = cells[0]
        link_tag   = title_cell.find('a', href=True)
        title      = title_cell.get_text(strip=True)
        href       = ""
        if link_tag:
            href  = link_tag.get('href', '')
            title = link_tag.get_text(strip=True) or title

        if not title or len(title) < 8:
            continue

        detail_url = href if href.startswith('http') else (BASE_URL + href if href else "")

        consultant_type = cells[1].get_text(strip=True) if len(cells) > 1 else ""
        published       = cells[2].get_text(strip=True) if len(cells) > 2 else ""
        closing         = cells[3].get_text(strip=True) if len(cells) > 3 else ""

        slug      = re.sub(r'[^a-zA-Z0-9]', '_', href.split('/')[-1] or title)[:80]
        tender_id = f"AFDB_{slug}"

        entries.append({
            "Title":           title,
            "Consultant Type": consultant_type,
            "Published":       published,
            "Closing Date":    closing,
            "detail_url":      detail_url,
            "tender_id":       tender_id,
        })

    return entries


# =============================================================================
# SECTION 2 — Fetch detail page
# =============================================================================

def _fetch_detail(session: requests.Session, url: str) -> dict:
    """
    Fetch detail page and extract description + country.
    On 403 (access restricted), return empty detail gracefully — listing-level
    data (title, consultant type, dates) is still preserved by the caller.
    """
    result = {"Description": "", "Country": ""}
    if not url:
        return result
    try:
        r = session.get(url, headers=HEADERS, timeout=20)
        if r.status_code == 403:
            print(f"[afdb]   Detail 403 — access restricted, using listing data only ({url[:70]})")
            return result
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')

        # ── Country: Strategy 1 — Drupal dl/dt/dd or table th/td structure ──────
        for dt in soup.find_all(['dt', 'th']):
            label = dt.get_text(strip=True).lower()
            if any(k in label for k in ('country', 'duty station', 'location',
                                        'posting location', 'place of work')):
                sibling = dt.find_next_sibling(['dd', 'td'])
                if sibling:
                    result["Country"] = sibling.get_text(strip=True)[:80]
                    break

        # ── Country: Strategy 2 — Drupal field-label / field--label CSS classes ─
        if not result["Country"]:
            for lbl in soup.find_all(class_=re.compile(
                    r'field-label|field--label|label|views-label', re.I)):
                label_text = lbl.get_text(strip=True).lower()
                if any(k in label_text for k in ('country', 'duty station',
                                                   'location', 'posting')):
                    # Drupal places content in adjacent field-items / field--item
                    parent = lbl.parent
                    item = (
                        parent.find(class_=re.compile(r'field-item|field--item', re.I))
                        if parent else None
                    )
                    if not item:
                        item = lbl.find_next_sibling()
                    if item:
                        result["Country"] = item.get_text(strip=True)[:80]
                        break

        # ── Country: Strategy 3 — regex on page text (terminators safe for
        #    space-joined text — no \n dependency) ────────────────────────────────
        if not result["Country"]:
            text = soup.get_text(' ', strip=True)
            for pattern in [
                r'(?:Duty\s+Station|Country|Location|Posting\s+Location)'
                r'[:\s]+([A-Za-z][A-Za-z\s,/\-]+?)(?:\s{2,}|;|\.\s)',
                r'(?:based\s+in|located\s+in)[:\s]+'
                r'([A-Za-z][A-Za-z\s,]+?)(?:\s{2,}|;|\.\s)',
            ]:
                m = re.search(pattern, text, re.I)
                if m:
                    country = m.group(1).strip()
                    if 3 <= len(country) <= 80:
                        result["Country"] = country
                        break

        # ── Description — try specific body-content selectors in priority order ───
        content = None
        for selector in [
            ('div', re.compile(r'\bfield-body\b|\bfield--body\b|\bcontent-body\b', re.I)),
            ('div', re.compile(r'\bbody\b', re.I)),
        ]:
            content = soup.find(selector[0], class_=selector[1])
            if content:
                break
        if not content:
            content = soup.find('article') or soup.find('main')
        if content:
            result["Description"] = content.get_text(' ', strip=True)[:1800]

    except requests.exceptions.HTTPError as e:
        code = e.response.status_code if e.response is not None else "?"
        print(f"[afdb]   Detail HTTP {code} error — using listing data ({url[:60]})")
    except Exception as e:
        print(f"[afdb]   Detail fetch error ({url[:60]}): {e}")
    return result


# =============================================================================
# SECTION 3 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AfDB Consultants"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")    + 1
    link_idx = col_names.index("Detail Link")  + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 50
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(AFDB_EXCEL_PATH)
    print(f"[afdb] Excel saved: {AFDB_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 3b — Selenium driver (Cloudflare bypass)
# =============================================================================

def _build_driver():
    """Headless Chrome — bypasses Cloudflare Bot Management on AfDB."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument(
        "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


# =============================================================================
# SECTION 4 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the African Development Bank Consultants pipeline.
    Returns (new_tenders, all_rows).
    """
    print("\n" + "=" * 65)
    print("[afdb] African Development Bank Consultants Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(AFDB_EXCEL_PATH):
        try:
            os.remove(AFDB_EXCEL_PATH)
            print("[afdb] Cleared old Excel")
        except Exception:
            pass

    session      = requests.Session()
    session.headers.update(HEADERS)
    all_entries  = []
    seen_ids     = set()
    driver       = None
    cookies_done = False

    try:
        driver = _build_driver()

        # ── Paginate listing via Selenium (bypasses Cloudflare 403) ───────────
        for page_num in range(MAX_PAGES):
            url = f"{LISTING_URL}&page={page_num}"
            print(f"[afdb]   Page {page_num}...", end=" ", flush=True)
            try:
                driver.get(url)
                try:
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located(
                            (By.CSS_SELECTOR, "table, .views-row, article, .view-content")
                        )
                    )
                except Exception:
                    pass
                time.sleep(1.5)

                soup = BeautifulSoup(driver.page_source, "html.parser")

                # Transfer Cloudflare cookies to requests.Session for detail pages
                if not cookies_done:
                    for ck in driver.get_cookies():
                        session.cookies.set(ck["name"], ck["value"])
                    cookies_done = True

                entries = _parse_table(soup)

                if not entries:
                    print("empty — done.")
                    break

                new_on_page = [e for e in entries if e.get("tender_id", "") not in seen_ids]
                print(f"{len(new_on_page)} new listings (total: {len(entries)})")

                for e in new_on_page:
                    seen_ids.add(e.get("tender_id", ""))
                all_entries.extend(new_on_page)

                if len(new_on_page) == 0 and page_num > 0:
                    print("[afdb]   No new entries — last page reached.")
                    break

                time.sleep(DELAY)

            except Exception as e:
                print(f"ERROR: {e}")
                break

    except Exception as outer_e:
        print(f"[afdb] Driver setup error: {outer_e}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    print(f"[afdb] Total listings found: {len(all_entries)}")

    # ── Process each entry ─────────────────────────────────────────────────────
    for i, entry in enumerate(all_entries, 1):
        title = entry["Title"]

        if _is_goods(title):
            print(f"[afdb]   [{i:>3}/{len(all_entries)}] SKIP (goods): {title[:60]}")
            continue

        print(f"[afdb]   [{i:>3}/{len(all_entries)}] {title[:65]}")

        detail    = _fetch_detail(session, entry["detail_url"])
        relevance = score_relevance(title, detail.get("Description", ""))

        row = {
            "Title":           title,
            "Consultant Type": entry["Consultant Type"],
            "Country":         detail.get("Country", ""),
            "Published":       entry["Published"],
            "Closing Date":    entry["Closing Date"],
            "Description":     detail.get("Description", "")[:1500],
            "Relevance":       relevance,
            "Detail Link":     entry["detail_url"],
        }
        all_rows.append(row)

        tid = entry.get("tender_id", f"AFDB_{re.sub(r'[^a-zA-Z0-9]', '_', title)[:80]}")
        if check_if_new(tid):
            mark_as_seen(tid, title=title, source_site="AfDB", url=entry["detail_url"])
            new_tenders.append({
                "title":    title,
                "deadline": entry["Closing Date"],
                "value":    entry["Consultant Type"] or "See AfDB",
                "url":      entry["detail_url"],
            })
            print(f"           → NEW | Relevance: {relevance or '—'}")
        else:
            print(f"           → seen | Relevance: {relevance or '—'}")

        time.sleep(DELAY)

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[afdb] Done — {len(all_rows)} listings, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
