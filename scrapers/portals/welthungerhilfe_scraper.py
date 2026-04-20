# =============================================================================
# welthungerhilfe_scraper.py — Deutsche Welthungerhilfe Tenders
#
# Site   : https://www.welthungerhilfe.org/tenders
# Method : requests + BeautifulSoup — ZERO Selenium needed
#          HTML is server-rendered; div.tender__list__item per tender
#
# Fix (v2): replaced Selenium with direct requests.get().
#   Root cause of "timeout" bug: site was initialising Chrome for a page that
#   returns clean static HTML with no JS rendering requirement. Page loads in
#   <1 second with plain requests.
#
# Structure:
#   div.tender__list__item
#     ├─ div.tender__list__item__date     → Published date
#     ├─ div.tender__list__item__title a  → Title + URL
#     ├─ div.tender__list__item__client   → Contracting Authority
#     └─ div.tender__list__item__deadline → Response Deadline
# =============================================================================

from __future__ import annotations

import os
import re

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://www.welthungerhilfe.org"
LISTING_URL = f"{BASE_URL}/tenders"

WHH_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "Welthungerhilfe_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "en-GB,en;q=0.9",
    "Referer":         BASE_URL,
}

# Phrases indicating site has no active tenders
_NO_TENDER_PHRASES = [
    "no tenders available", "currently no tenders",
    "there are currently no tenders", "no active tenders",
    "no open tenders", "please try again later",
    "there are no open tenders", "we currently have no open tenders",
]

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",          65),
    ("Authority",      35),
    ("Country",        22),
    ("Published",      16),
    ("Deadline",       18),
    ("Relevance",      40),
    ("Detail Link",    55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Scraper
# =============================================================================

def _clean_date(raw: str) -> str:
    """Strip soft-hyphens and whitespace from WHH date strings."""
    return re.sub(r"[\u00ad\s]+", " ", raw).strip()


def _parse_items(html: str) -> list[dict]:
    """Parse div.tender__list__item elements from WHH tenders page."""
    soup  = BeautifulSoup(html, "html.parser")
    items = soup.find_all("div", class_="tender__list__item")
    results = []

    for item in items:
        # Title + URL
        title_div = item.find("div", class_="tender__list__item__title")
        if not title_div:
            continue
        link = title_div.find("a")
        if not link:
            continue
        title = link.get_text(strip=True)
        url   = link.get("href", "").strip()
        if not url.startswith("http"):
            url = BASE_URL + url

        # Published date
        date_div = item.find("div", class_="tender__list__item__date")
        published = ""
        if date_div:
            raw = date_div.get_text(strip=True)
            # Remove "Date of Publication:" prefix
            published = _clean_date(
                re.sub(r"Date of Publication\s*:\s*", "", raw, flags=re.I)
            )

        # Contracting authority
        client_div = item.find("div", class_="tender__list__item__client")
        authority = ""
        if client_div:
            raw = client_div.get_text(strip=True)
            authority = re.sub(r"Contracting Authority\s*:\s*", "", raw, flags=re.I).strip()

        # Deadline
        deadline_div = item.find("div", class_="tender__list__item__deadline")
        deadline = ""
        if deadline_div:
            raw = deadline_div.get_text(strip=True)
            # "Response Deadline (CET): ­12.04.2026 11:04"
            deadline = _clean_date(
                re.sub(r"Response Deadline\s*\([^)]*\)\s*:\s*", "", raw, flags=re.I)
            )
            # Keep only the date part (drop time)
            deadline = deadline.split(" ")[0] if " " in deadline else deadline

        if not title:
            continue

        results.append({
            "title":     title,
            "url":       url,
            "published": published,
            "deadline":  deadline,
            "authority": authority,
        })

    return results


def _infer_country(authority: str) -> str:
    """Extract country from authority string like 'Welthungerhilfe Afghanistan'."""
    if not authority:
        return ""
    # Strip 'Welthungerhilfe' / 'Deutsche Welthungerhilfe e.V.' prefixes
    country = re.sub(r"Deutsche?\s+Welthungerhilfe\s*(e\.?\s*V\.?)?\s*", "",
                     authority, flags=re.I).strip()
    return country if len(country) > 1 else ""


# =============================================================================
# SECTION 2 — Excel writer
# =============================================================================

def _save_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WHH Tenders"
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    col_names = [c[0] for c in MASTER_COLUMNS]
    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    rel_idx  = col_names.index("Relevance")   + 1
    link_idx = col_names.index("Detail Link") + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 40
        alt = ALT_FILL if ri % 2 == 0 else None

        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10,
                                 color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt:
                        cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt:
                    cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(WHH_EXCEL_PATH)
    print(f"[whh] Excel saved: {WHH_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 3 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Scrape Welthungerhilfe tenders page and return (new_tenders, all_rows).
    Uses plain requests — no Selenium required.
    """
    print("\n[whh] Welthungerhilfe tenders starting (requests-only)...")

    new_tenders: list[dict] = []
    all_rows:    list[dict] = []

    try:
        r = requests.get(LISTING_URL, headers=HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as exc:
        print(f"[whh] Page fetch failed: {exc}")
        return [], []

    # Check for "no tenders" message
    page_text = r.text.lower()
    if any(p in page_text for p in _NO_TENDER_PHRASES) and "tender__list__item" not in page_text:
        print("[whh] No active tenders on site right now.")
        return [], []

    items = _parse_items(r.text)
    print(f"[whh] Found {len(items)} tenders on listing page")

    if not items:
        print("[whh] WARNING: 0 items parsed — page structure may have changed")
        return [], []

    for notice in items:
        title     = notice["title"]
        url       = notice["url"]
        published = notice["published"]
        deadline  = notice["deadline"]
        authority = notice["authority"]
        country   = _infer_country(authority)

        # Stable ID from URL (eu-supply.com PID or last segment)
        pid_match = re.search(r"PID=(\d+)", url)
        slug      = pid_match.group(1) if pid_match else re.sub(r"[^a-zA-Z0-9]", "_", url[-40:])
        tender_id = f"WHH_{slug}"

        relevance = score_relevance(title, authority)

        row = {
            "Title":       title,
            "Authority":   authority,
            "Country":     country,
            "Published":   published,
            "Deadline":    deadline,
            "Relevance":   relevance,
            "Detail Link": url,
        }
        all_rows.append(row)

        if check_if_new(tender_id):
            mark_as_seen(tender_id, title=title, source_site="Welthungerhilfe", url=url)
            new_tenders.append({
                "title":    title,
                "deadline": deadline,
                "value":    authority or "Welthungerhilfe",
                "url":      url,
            })
            print(f"[whh]   NEW: {title[:65]}")
        else:
            print(f"[whh]   seen: {title[:65]}")

    if all_rows:
        _save_excel(all_rows)
    else:
        print("[whh] No rows to save")

    print(f"\n[whh] Done — {len(all_rows)} tenders, {len(new_tenders)} NEW")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nResult: {len(rows)} total, {len(new)} new")
