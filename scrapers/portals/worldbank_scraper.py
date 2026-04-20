# =============================================================================
# worldbank_scraper.py — World Bank Consulting Procurement Notices
#
# Source: search.worldbank.org/api/v2/procnotices (Procurement Notices API)
# Scope : Consulting Services (procurement_group=CS) — EOIs, RFPs, GPN
#         No PDF downloads. No PDF parsing. Pure JSON API.
#
# Flow:
#   1. Fetch recent CS notices from last 30 days (rolling window)
#   2. Score relevance on project_name + bid_description
#   3. Keep score > 0, deduplicate against DB
#   4. Save Excel + return (new, all_rows)
# =============================================================================

import os
import re
import time
from datetime import datetime, timedelta, timezone
from typing import Any

import requests

from config.config import WB_EXCEL_PATH
from core.base_scraper import IntelligentBaseScraper
from core.quality_engine import TenderResult, make_tender_result
from intelligence.keywords import score_relevance


# =============================================================================
# CONFIG
# =============================================================================

_API_URL   = "https://search.worldbank.org/api/v2/procnotices"
_NOTICE_URL = "https://projects.worldbank.org/en/projects-operations/procurement-detail/{id}"

# Only pull consulting services notices
_PROC_GROUP = "CS"

# Notice types to capture (skip Contract Awards — those are already closed)
_NOTICE_TYPES = [
    "Request for Expression of Interest",
    "General Procurement Notice",
    "Specific Procurement Notice",
]

# Rolling lookback window in days
_LOOKBACK_DAYS = 30

# Max pages per notice type (100 rows/page → 500 notices max per type)
_MAX_PAGES  = 5
_PAGE_SIZE  = 100

# Fields we want from the API
_FL = (
    "id,notice_type,noticedate,submission_deadline_date,"
    "project_ctry_name,project_id,project_name,"
    "bid_reference_no,bid_description,"
    "procurement_method_name,contact_organization"
)

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept": "application/json",
}

SCRAPER_META = {
    "flag":        "wb",
    "label":       "World Bank",
    "group":       "api",
    "timeout":     180,
    "max_retries": 2,
    "auto":        True,
}


# =============================================================================
# SCRAPER CLASS
# =============================================================================

class WorldBankScraper(IntelligentBaseScraper):
    """
    World Bank consulting procurement notices — no PDFs, pure API.

    Fetches EOIs and GPNs for consulting services from the WB procurement
    notices API, scores relevance, and returns matched opportunities.
    """

    SOURCE_NAME = "World Bank"
    SOURCE_URL  = "https://projects.worldbank.org/en/projects-operations/procurement"
    EXCEL_PATH  = WB_EXCEL_PATH

    # No strict schema validation needed — API is stable JSON
    EXPECTED_SCHEMA_FIELDS = []

    # =========================================================================
    # STEP 1 — fetch_data
    # =========================================================================

    def fetch_data(self) -> dict:
        """Fetch all recent CS procurement notices from the WB API."""
        since = (datetime.now(timezone.utc) - timedelta(days=_LOOKBACK_DAYS)
                 ).strftime("%Y-%m-%d")

        all_notices = []
        for notice_type in _NOTICE_TYPES:
            notices = self._fetch_notice_type(notice_type, since)
            all_notices.extend(notices)
            self._log.info(
                "[wb] %s — %d notice(s) fetched", notice_type, len(notices)
            )
            time.sleep(0.5)

        self._log.info("[wb] Total raw notices fetched: %d", len(all_notices))
        return {"notices": all_notices}

    def _fetch_notice_type(self, notice_type: str, since: str) -> list:
        """Paginate through notices of a given type since `since` date."""
        results = []
        for page in range(_MAX_PAGES):
            offset = page * _PAGE_SIZE
            params = {
                "format":         "json",
                "rows":           _PAGE_SIZE,
                "os":             offset,
                "procurement_group": _PROC_GROUP,
                "notice_type":    notice_type,
                "strdate":        since,
                "fl":             _FL,
            }
            try:
                resp = requests.get(
                    _API_URL, params=params, headers=_HEADERS, timeout=30
                )
                if resp.status_code != 200:
                    self._log.warning(
                        "[wb] HTTP %d for %s (page %d)",
                        resp.status_code, notice_type, page + 1,
                    )
                    break
                data    = resp.json()
                batch   = data.get("procnotices", [])
                total   = int(data.get("total", 0))
                if not batch:
                    break
                results.extend(batch)
                # Stop if we've pulled everything
                if offset + len(batch) >= total:
                    break
                time.sleep(0.3)
            except Exception as exc:
                self._log.warning("[wb] API error (page %d): %s", page + 1, exc)
                break

        return results

    # =========================================================================
    # STEP 2 — validate_schema (lightweight — API is stable)
    # =========================================================================

    def validate_schema(self, raw_data: dict) -> bool:
        notices = raw_data.get("notices", [])
        if not notices:
            return True
        sample     = notices[:10]
        violations = sum(1 for n in sample if not n.get("id") or not n.get("project_name"))
        return (violations / len(sample)) <= 0.20

    # =========================================================================
    # STEP 3 — extract_rows
    # =========================================================================

    def extract_rows(self, raw_data: dict) -> list[dict]:
        """
        Score each notice for relevance. Return all scored rows
        (filtering to score > 0 happens downstream).
        """
        notices = raw_data.get("notices", [])
        rows    = []

        for n in notices:
            nid         = n.get("id", "")
            project     = (n.get("project_name") or "").strip()
            description = (n.get("bid_description") or "").strip()
            country     = (n.get("project_ctry_name") or "").strip()
            org         = (n.get("contact_organization") or "").strip()
            method      = (n.get("procurement_method_name") or "").strip()
            notice_type = (n.get("notice_type") or "").strip()
            ref_no      = (n.get("bid_reference_no") or "").strip()

            # Build URL
            url = _NOTICE_URL.format(id=nid) if nid else self.SOURCE_URL

            # Parse deadline
            deadline = _parse_date(n.get("submission_deadline_date") or "")
            published = _parse_wb_date(n.get("noticedate") or "")

            # Title = description if available, else project name
            title = description[:200] if description else project[:200]
            if not title:
                continue

            # Score relevance against project + description
            score_text = f"{project} {description}"
            relevance  = score_relevance(score_text, method)

            rows.append({
                "id":          nid,
                "title":       title,
                "project":     project,
                "description": description,
                "country":     country,
                "organization": org or "World Bank",
                "method":      method,
                "notice_type": notice_type,
                "ref_no":      ref_no,
                "url":         url,
                "deadline":    deadline,
                "published":   published,
                "relevance":   relevance,      # string: matched sector labels
                "rel_score":   1 if relevance else 0,  # numeric proxy
                # These get filled by quality_engine in base class
                "quality_score":    0,
                "consulting_type":  "",
                "sector":           "",
            })

        self._log.info("[wb] %d rows after extraction (before relevance filter)", len(rows))
        return rows

    # =========================================================================
    # STEP 4 — to_standard_format
    # =========================================================================

    def to_standard_format(self, row: dict) -> TenderResult:
        return make_tender_result(
            title           = row.get("title", ""),
            url             = row.get("url", self.SOURCE_URL),
            deadline        = row.get("deadline", ""),
            organization    = row.get("organization", "World Bank"),
            sector          = row.get("sector", ""),
            consulting_type = row.get("consulting_type", ""),
            quality_score   = int(row.get("quality_score") or row.get("relevance") or 0),
            source          = self.SOURCE_NAME,
        )

    # =========================================================================
    # STEP 5 — get_tender_id
    # =========================================================================

    def get_tender_id(self, row: dict) -> str:
        nid    = str(row.get("id", "")).strip()
        ref_no = re.sub(r"\s+", "_", str(row.get("ref_no", "")).strip())[:40]
        if nid:
            return f"WB_{nid}"
        if ref_no:
            return f"WB_{ref_no}"
        return f"WB_{hash(row.get('title',''))}"

    # =========================================================================
    # STEP 6 — on_run_end (Excel output)
    # =========================================================================

    def on_run_end(self, all_rows: list[dict]) -> None:
        if not all_rows:
            return
        try:
            _save_wb_excel(all_rows)
        except Exception as exc:
            self._log.warning("[wb] Excel save error: %s", exc)


# =============================================================================
# HELPERS
# =============================================================================

def _parse_date(raw: str) -> str:
    """Parse ISO datetime '2026-04-03T00:00:00Z' → '03 Apr 2026'."""
    if not raw:
        return ""
    try:
        dt = datetime.strptime(raw[:10], "%Y-%m-%d")
        return dt.strftime("%d %b %Y")
    except ValueError:
        return raw[:10]


def _parse_wb_date(raw: str) -> str:
    """Parse WB notice date '20-Mar-2026' → '20 Mar 2026'."""
    if not raw:
        return ""
    try:
        dt = datetime.strptime(raw.strip(), "%d-%b-%Y")
        return dt.strftime("%d %b %Y")
    except ValueError:
        return raw[:12]


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

_COLUMNS = [
    ("Title / Assignment",     55),
    ("Project",                35),
    ("Country",                18),
    ("Organization",           30),
    ("Notice Type",            28),
    ("Method",                 25),
    ("Ref No",                 22),
    ("Published",              14),
    ("Deadline",               14),
    ("Relevance",              12),
    ("Quality Score",          14),
    ("Sector",                 18),
    ("Link",                   40),
]

_HDR  = PatternFill("solid", fgColor="1F3864")
_ALT  = PatternFill("solid", fgColor="F5F8FF")
_REL  = PatternFill("solid", fgColor="E2EFDA")
_QHI  = PatternFill("solid", fgColor="C6EFCE")
_QMID = PatternFill("solid", fgColor="FFEB9C")
_QLOW = PatternFill("solid", fgColor="FFC7CE")
_WFNT = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
_BDY  = Font(name="Calibri", size=10)
_RELF = Font(name="Calibri", size=10, color="375623", bold=True)
_BORD = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


def _save_wb_excel(rows: list[dict]) -> None:
    # Sort by quality_score desc, then relevance desc
    rows = sorted(rows, key=lambda r: (
        -int(r.get("quality_score") or 0),
        -int(r.get("rel_score") or 0),
    ))

    if os.path.exists(WB_EXCEL_PATH):
        try:
            os.remove(WB_EXCEL_PATH)
        except Exception as _e:
            import logging as _logging
            _logging.getLogger("tenderradar.wb_scraper").warning(
                "[wb_scraper] Could not remove old Excel file %s: %s", WB_EXCEL_PATH, _e
            )

    wb = Workbook(); ws = wb.active
    ws.title = "WB Consulting Notices"

    col_names = [c[0] for c in _COLUMNS]
    ws.row_dimensions[1].height = 36

    for ci, (name, width) in enumerate(_COLUMNS, 1):
        c = ws.cell(1, ci, name)
        c.font = _WFNT; c.fill = _HDR; c.border = _BORD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = ws.cell(2, 1)
    rel_col = col_names.index("Relevance") + 1
    qs_col  = col_names.index("Quality Score") + 1
    lnk_col = col_names.index("Link") + 1

    field_map = {
        "Title / Assignment": "title",
        "Project":            "project",
        "Country":            "country",
        "Organization":       "organization",
        "Notice Type":        "notice_type",
        "Method":             "method",
        "Ref No":             "ref_no",
        "Published":          "published",
        "Deadline":           "deadline",
        "Relevance":          "relevance",
        "Quality Score":      "quality_score",
        "Sector":             "sector",
        "Link":               "url",
    }

    for ri, row in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 40
        qs  = int(row.get("quality_score") or 0)
        alt = _ALT if ri % 2 == 0 else None

        for ci, col_name in enumerate(col_names, 1):
            key = field_map.get(col_name, col_name.lower())
            val = row.get(key, "")
            cell = ws.cell(ri, ci, val)
            cell.border = _BORD
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if ci == lnk_col and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
                if alt: cell.fill = alt
            elif ci == rel_col and val:
                cell.fill = _REL; cell.font = _RELF
            elif ci == qs_col:
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.fill = _QHI if qs >= 70 else (_QMID if qs >= 40 else _QLOW)
            else:
                cell.font = _BDY
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"
    wb.save(WB_EXCEL_PATH)
    print(f"[wb] Excel saved: {WB_EXCEL_PATH} ({len(rows)} rows)")


# =============================================================================
# MODULE-LEVEL run() SHIM — required by registry
# =============================================================================

_module_debug: bool = False


def set_debug(flag: bool) -> None:
    global _module_debug
    _module_debug = flag


def run() -> tuple[list, list]:
    return WorldBankScraper().run(debug=_module_debug)


if __name__ == "__main__":
    import sys
    debug = "--debug" in sys.argv
    new, rows = run()
    print(f"\nNew tenders : {len(new)}")
    print(f"Total rows  : {len(rows)}")
    if rows:
        print("\nTop 5 by quality score:")
        for r in sorted(rows, key=lambda x: -int(x.get("quality_score") or 0))[:5]:
            print(f"  [{r.get('quality_score',0):>3}] {r.get('title','')[:80]}")
            print(f"       Country: {r.get('country','')}  Deadline: {r.get('deadline','')}")
