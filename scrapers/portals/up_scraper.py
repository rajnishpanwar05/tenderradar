# =============================================================================
# up_scraper.py — Uttar Pradesh eTenders (etender.up.nic.in) Pipeline
#
# Site: https://etender.up.nic.in/nicgep/app
# Approach: requests.Session + "Tenders by Organisation" listing page
#   → 132 org links accessible WITHOUT CAPTCHA within the same session.
#   → Each org page returns full tender list with title, dates, tender ID.
#   → Identical NIC GePNIC (Tapestry/Dojo) framework as mahatenders.gov.in
#     and eprocure.gov.in — same table format, same sp= session tokens.
# No Selenium, no login required.
#
# Key orgs: Directorate of Local Bodies (~4641), Food & Civil Supplies (914),
#   Jal Nigam, UPMSCL, Housing & Development Board, UPSRTC, PWD,
#   Yamuna Expressway IDA, and 120+ more.
# =============================================================================

import os
import re
import time
import random

SCRAPER_META = {
    "flag":        "upetender",
    "label":       "UP eTenders",
    "group":       "requests",
    "timeout":     900,
    "max_retries": 1,
    "auto":        True,
}

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import LOG_FILE, PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance
from scrapers.portals.nic_detail_utils import (
    enrich_nic_row_with_detail,
    seed_nic_listing_metadata,
)

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL     = "https://etender.up.nic.in/nicgep/app"
ORG_LIST_URL = f"{BASE_URL}?page=FrontEndTendersByOrganisation&service=page"

EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "UP_Tenders_Master.xlsx")

REQUEST_DELAY     = 0.8   # seconds between requests (polite scraping)
MAX_ORGS          = 200   # safety cap (132 orgs currently on the portal)
MAX_PAGES_PER_ORG = 10    # page cap per org (most have 1–2 pages)
RETRY_ATTEMPTS    = 2
MAX_DETAIL_ENRICH_PER_ORG = 60  # raise ceiling so more tenders carry detail/doc links for deep extraction

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Tender ID",       28),
    ("Title",           65),
    ("Organisation",    40),
    ("Published Date",  20),
    ("Closing Date",    20),
    ("Opening Date",    20),
    ("Relevance",       42),
    ("URL",             55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Parsing helpers
# =============================================================================

def _extract_brackets(text: str):
    """
    NIC GePNIC framework stores the title cell as:
      [Title Text] [Ref No][Tender ID]
    Returns (title, ref_no, tender_id).
    """
    bracket_groups = re.findall(r"\[([^\[\]]+)\]", text)
    clean_title = re.sub(r"\[[^\[\]]+\]", "", text).strip()
    clean_title = re.sub(r"\s+", " ", clean_title).strip()

    ref_no    = ""
    tender_id = ""
    if len(bracket_groups) >= 3:
        ref_no    = bracket_groups[-2].strip()
        tender_id = bracket_groups[-1].strip()
        if not clean_title or len(clean_title) < 5:
            clean_title = bracket_groups[0].strip()
    elif len(bracket_groups) == 2:
        ref_no    = bracket_groups[0].strip()
        tender_id = bracket_groups[1].strip()
        if not clean_title or len(clean_title) < 5:
            clean_title = bracket_groups[0].strip()
    elif len(bracket_groups) == 1:
        tender_id = bracket_groups[0].strip()

    return clean_title, ref_no, tender_id


def _parse_org_tender_table(html: str, org_name: str):
    """
    Parse the tender table on a UP eTender org page.
    NIC GePNIC format columns (same across all state instances):
      0: S.No  1: e-Published Date  2: Closing Date  3: Opening Date
      4: Title and Ref.No./Tender ID  5: Organisation Chain
    Returns list of tender dicts.
    """
    soup    = BeautifulSoup(html, "html.parser")
    tenders = []

    EXACT_HEADERS = ["S.No", "e-Published Date", "Closing Date", "Opening Date",
                     "Title and Ref.No./Tender ID", "Organisation Chain"]

    target_table = None
    for row in soup.find_all("tr"):
        cells = row.find_all(["th", "td"])
        if [c.get_text(strip=True) for c in cells] == EXACT_HEADERS:
            target_table = row.find_parent("table")
            break

    if target_table is None:
        return tenders

    rows = target_table.find_all("tr")
    header_row = None
    for i, row in enumerate(rows):
        cells = row.find_all(["th", "td"])
        if [c.get_text(strip=True) for c in cells] == EXACT_HEADERS:
            header_row = i
            break

    if header_row is None:
        return tenders

    for row in rows[header_row + 1:]:
        cells = row.find_all(["td", "th"])
        if len(cells) < 5:
            continue

        raw_title_cell = cells[4].get_text(" ", strip=True)
        if not raw_title_cell or len(raw_title_cell) < 5:
            continue

        title, ref_no, tender_id = _extract_brackets(raw_title_cell)

        if not title or len(title) < 5:
            continue

        # Detail URL from link in title cell
        detail_url = BASE_URL
        link = cells[4].find("a", href=True)
        if link:
            href = link["href"]
            if href.startswith("http"):
                detail_url = href
            elif href.startswith("/"):
                detail_url = "https://etender.up.nic.in" + href
            else:
                detail_url = "https://etender.up.nic.in/nicgep/" + href.lstrip("/")

        org_chain = cells[5].get_text(" ", strip=True) if len(cells) > 5 else org_name
        if not org_chain:
            org_chain = org_name

        tenders.append({
            "Tender ID":      tender_id or ref_no or f"up/{title[:40]}",
            "Title":          title,
            "Organisation":   org_chain,
            "Published Date": cells[1].get_text(strip=True),
            "Closing Date":   cells[2].get_text(strip=True),
            "Opening Date":   cells[3].get_text(strip=True),
            "URL":            detail_url,
        })

    return tenders


def _get_next_page_url(html: str, current_url: str):
    """Look for 'Next' pagination link. Returns next URL or None."""
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True).lower()
        if text in ("next", "next »", ">", ">>", "next page"):
            href = a["href"]
            if href.startswith("http"):
                return href
            if href.startswith("?"):
                return BASE_URL + href
            if href.startswith("/"):
                return "https://etender.up.nic.in" + href
    return None


# =============================================================================
# Session / fetching helpers
# =============================================================================

def _make_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def _get(session, url, retries=RETRY_ATTEMPTS):
    for attempt in range(retries):
        try:
            r = session.get(url, timeout=30)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                print(f"[up]   GET failed ({url[:80]}): {e}", flush=True)
                return None
    return None


def _fetch_org_links(session):
    """
    Fetch the 'Tenders by Organisation' page and extract all org (name, url) tuples.
    UP portal org table: S.No | Organisation Name | Tender Count
    The sp= link is in the Tender Count cell.
    """
    print("[up] Fetching Tenders-by-Organisation listing...", flush=True)
    r = _get(session, ORG_LIST_URL)
    if not r:
        print("[up] Failed to load org listing page.", flush=True)
        return []

    soup = BeautifulSoup(r.text, "html.parser")
    orgs = []

    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        header_cells = rows[0].find_all(["th", "td"])
        if [c.get_text(strip=True) for c in header_cells] != ["S.No", "Organisation Name", "Tender Count"]:
            continue

        for row in rows[1:]:
            cells = row.find_all(["td", "th"])
            if len(cells) < 3:
                continue
            org_name = cells[1].get_text(strip=True)
            if not org_name:
                continue
            link = cells[2].find("a", href=True)
            if not link:
                continue
            href = link["href"]
            if href.startswith("http"):
                url = href
            elif href.startswith("/"):
                url = "https://etender.up.nic.in" + href
            else:
                url = "https://etender.up.nic.in/" + href.lstrip("/")
            orgs.append((org_name, url))
        break

    print(f"[up] Found {len(orgs)} organisation links.", flush=True)
    return orgs[:MAX_ORGS]


# =============================================================================
# Excel output
# =============================================================================

def _save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "UP Tenders"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    relevance_idx   = col_names.index("Relevance") + 1
    link_idx        = col_names.index("URL") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill:
                        cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill:
                    cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(EXCEL_PATH)
    print(f"[up] Excel saved: {EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[up] UP eTenders (etender.up.nic.in) Pipeline", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []

    if os.path.exists(EXCEL_PATH):
        try:
            os.remove(EXCEL_PATH)
        except Exception:
            pass

    session = _make_session()

    # Warm up session (captures JSESSIONID cookie)
    print("[up] Warming up session...", flush=True)
    _get(session, BASE_URL)
    time.sleep(1)

    # Fetch org listing
    org_links = _fetch_org_links(session)
    if not org_links:
        print("[up] No org links found — aborting.", flush=True)
        return new_tenders, all_rows

    # Visit each org and collect tenders
    total_scraped = 0
    for org_idx, (org_name, org_url) in enumerate(org_links, 1):
        print(f"[up]  [{org_idx}/{len(org_links)}] {org_name[:60]}", flush=True)

        page_url = org_url
        page_num = 1
        org_tenders = []
        org_detail_enriched = 0

        while page_url and page_num <= MAX_PAGES_PER_ORG:
            r = _get(session, page_url)
            if not r:
                break

            page_tenders = _parse_org_tender_table(r.text, org_name)
            if page_tenders:
                print(f"[up]    page {page_num}: {len(page_tenders)} tenders", flush=True)
                org_tenders.extend(page_tenders)
            else:
                break

            next_url = _get_next_page_url(r.text, page_url)
            if next_url and next_url != page_url:
                page_url = next_url
                page_num += 1
                time.sleep(REQUEST_DELAY)
            else:
                break

        for t in org_tenders:
            seed_nic_listing_metadata(t, "upetender")
            combined = " ".join(filter(None, [
                t["Title"],
                t.get("Organisation", ""),
                t.get("Description", ""),
            ]))
            t["Relevance"] = score_relevance(combined)
            tender_id  = t["Tender ID"]
            detail_url = t.get("URL", BASE_URL)
            is_new = check_if_new(tender_id)

            if (t.get("Relevance") or is_new) and org_detail_enriched < MAX_DETAIL_ENRICH_PER_ORG:
                enrich_nic_row_with_detail(session, t)
                org_detail_enriched += 1

            all_rows.append(t)
            total_scraped += 1

            if is_new:
                mark_as_seen(tender_id, title=t["Title"], source_site="UP eTender", url=detail_url)
                if t.get("Relevance"):
                    new_tenders.append({
                        "tender_id": tender_id,
                        "title":    t["Title"],
                        "description": t.get("Description", ""),
                        "organization": t.get("Organisation", ""),
                        "deadline": t.get("Closing Date", ""),
                        "value":    t.get("Value", ""),
                        "url":      detail_url,
                        "source_portal": "upetender",
                    })
                    print(f"[up]     → NEW + RELEVANT | {t['Relevance'][:60]}", flush=True)
                else:
                    print(f"[up]     → NEW (saved to Excel only)", flush=True)

        time.sleep(REQUEST_DELAY + random.uniform(0, 0.4))

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(
        f"\n[up] Done — {len(all_rows)} tenders scraped across {len(org_links)} orgs, "
        f"{len(new_tenders)} NEW+RELEVANT, {relevant} relevant total",
        flush=True,
    )
    return new_tenders, all_rows
