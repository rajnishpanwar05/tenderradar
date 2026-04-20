# =============================================================================
# jtds_pipeline.py — Jharkhand Tribal Development Society (JTDS) Tenders
#
# Site   : http://jtdsjharkhand.com/tender/
# Method : requests + BeautifulSoup (static WordPress HTML)
# Login  : None required
# CAPTCHA: None
#
# Page structure:
#   WordPress site — each tender is a region > heading + PDF download link
#   No structured metadata (dates/deadlines embedded in PDF, not HTML)
#   All tenders on a single scrollable page — no pagination
#
# What this does:
#   1. Fetches the tenders listing page
#   2. Extracts each tender: title + PDF link
#   3. Scores relevance against IDCG expertise keywords
#   4. Deduplicates via MySQL — only NEW tenders trigger notification
#   5. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
#
# JTDS focus areas: tribal welfare, livelihoods, skill development,
# forest rights, social protection — relevant to IDCG's social sector work
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "http://jtdsjharkhand.com"
LISTING_URL = f"{BASE_URL}/tender/"
DELAY       = 0.5

JTDS_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "JTDS_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# Skip pure goods/equipment tenders
SKIP_PATTERNS = [
    "supply of", "purchase of", "procurement of vehicle", "rate contract",
    "annual maintenance", "amc", "printing of", "furniture",
]

def _is_goods(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in SKIP_PATTERNS)


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",       65),
    ("Relevance",   40),
    ("PDF / Link",  55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse tenders
# =============================================================================

def _parse_tenders(soup: BeautifulSoup) -> list:
    """
    JTDS WordPress structure: heading + PDF link per tender.
    No structured date metadata in HTML.
    """
    entries    = []
    seen_hrefs = set()

    # Strategy 1: headings paired with PDF links
    for heading in soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6']):
        title = heading.get_text(strip=True)
        if not title or len(title) < 8:
            continue

        # Find the nearest PDF link (sibling or parent → child)
        pdf_link = ""
        search_el = heading

        # Look for link in heading itself
        link_in_heading = heading.find('a', href=True)
        if link_in_heading:
            href = link_in_heading.get('href', '')
            if href:
                pdf_link = href if href.startswith('http') else BASE_URL + href

        # Look in next siblings
        if not pdf_link:
            for sibling in heading.find_next_siblings(['a', 'p', 'div', 'li'], limit=4):
                link = sibling if sibling.name == 'a' else sibling.find('a', href=True)
                if link:
                    href = link.get('href', '') if hasattr(link, 'get') else ''
                    if href and (href.endswith('.pdf') or 'download' in href.lower() or
                                 'wp-content' in href.lower()):
                        pdf_link = href if href.startswith('http') else BASE_URL + href
                        break

        # JS-style download links: onclick="..." or data-href/data-url attributes
        if not pdf_link:
            search_nodes = [heading] + list(heading.find_next_siblings(limit=5))
            for node in search_nodes:
                candidates = [node] if node.name == 'a' else (
                    node.find_all('a') if hasattr(node, 'find_all') else []
                )
                for el in candidates:
                    if not hasattr(el, 'get'):
                        continue
                    # Check data-* attributes
                    for attr in ('data-href', 'data-url', 'data-link', 'data-file'):
                        dval = el.get(attr, '')
                        if dval and ('.pdf' in dval.lower() or 'upload' in dval.lower()):
                            pdf_link = dval if dval.startswith('http') else BASE_URL + dval
                            break
                    # Check onclick="..." for embedded file paths
                    if not pdf_link:
                        onclick = el.get('onclick', '')
                        if onclick:
                            m = re.search(r"['\"]([^'\"]*\.pdf[^'\"]*)['\"]", onclick, re.I)
                            if not m:
                                m = re.search(r"['\"]([^'\"]*upload[^'\"]*)['\"]", onclick, re.I)
                            if m:
                                href = m.group(1)
                                pdf_link = href if href.startswith('http') else BASE_URL + href
                    if pdf_link:
                        break
                if pdf_link:
                    break

        if pdf_link in seen_hrefs:
            continue
        if pdf_link:
            seen_hrefs.add(pdf_link)

        slug      = re.sub(r'[^a-zA-Z0-9]', '_',
                           (pdf_link.split('/')[-1] if pdf_link else title[:50]))[:80]
        tender_id = f"JTDS_{slug}"

        entries.append({
            "Title":      title,
            "PDF / Link": pdf_link or LISTING_URL,  # fallback to listing page
            "url":        pdf_link or LISTING_URL,
            "tender_id":  tender_id,
        })

    # Strategy 2: fallback — any PDF link with a descriptive anchor text
    if not entries:
        for link in soup.find_all('a', href=True):
            href  = link.get('href', '')
            title = link.get_text(strip=True)
            if not title or len(title) < 8:
                continue
            if not (href.endswith('.pdf') or 'upload' in href.lower()):
                continue
            if href in seen_hrefs:
                continue
            seen_hrefs.add(href)
            pdf_url   = href if href.startswith('http') else BASE_URL + href
            slug      = re.sub(r'[^a-zA-Z0-9]', '_', href.split('/')[-1])[:80]
            tender_id = f"JTDS_{slug}"
            entries.append({
                "Title":      title,
                "PDF / Link": pdf_url,
                "tender_id":  tender_id,
            })

    return entries


# =============================================================================
# SECTION 2 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "JTDS Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")   + 1
    link_idx = col_names.index("PDF / Link")  + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(JTDS_EXCEL_PATH)
    print(f"[jtds] Excel saved: {JTDS_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 3 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the JTDS Tenders pipeline.
    Returns (new_tenders, all_rows).
    """
    print("\n" + "=" * 65)
    print("[jtds] Jharkhand Tribal Development Society Tenders Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(JTDS_EXCEL_PATH):
        try:
            os.remove(JTDS_EXCEL_PATH)
            print("[jtds] Cleared old Excel")
        except Exception:
            pass

    try:
        r = requests.get(LISTING_URL, headers=HEADERS, timeout=25, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
    except Exception as e:
        print(f"[jtds] ERROR fetching page: {e}")
        return [], []

    entries = _parse_tenders(soup)
    print(f"[jtds] Found {len(entries)} tenders")

    if not entries:
        print("[jtds] No tenders found — page structure may have changed.")
        return [], []

    for i, entry in enumerate(entries, 1):
        title = entry["Title"]

        if _is_goods(title):
            print(f"[jtds]   [{i:>2}/{len(entries)}] SKIP (goods): {title[:60]}")
            continue

        relevance = score_relevance(title, "")

        row = {
            "Title":      title,
            "Relevance":  relevance,
            "PDF / Link": entry["PDF / Link"],
        }
        all_rows.append(row)

        tid = entry["tender_id"]
        if check_if_new(tid):
            mark_as_seen(tid, title=title, source_site="JTDS",
                         url=entry["PDF / Link"] or LISTING_URL)
            new_tenders.append({
                "title":    title,
                "deadline": "See PDF",
                "value":    "JTDS Jharkhand",
                "url":      entry["PDF / Link"] or LISTING_URL,
            })
            print(f"[jtds]   → NEW: {title[:70]} | {relevance or '—'}")
        else:
            print(f"[jtds]   → seen: {title[:70]}")

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[jtds] Done — {len(all_rows)} tenders, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
