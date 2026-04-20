# =============================================================================
# nic_states_pipeline.py — Unified NIC eProcurement State Portals Pipeline
#
# Portals covered (all run the same NIC eProcurement platform):
#   UP          : https://etender.up.nic.in/nicgep/app
#   Punjab      : https://eproc.punjab.gov.in/nicgep/app
#   TN (TNEARD) : https://tntenders.gov.in/nicgep/app
#   Tripura     : https://tripuratenders.gov.in/nicgep/app
#   Odisha      : https://tendersodisha.gov.in/nicgep/app
#   Nagaland    : https://nagalandtenders.gov.in/nicgep/app
#   Maharashtra : https://mahatenders.gov.in/nicgep/app
#   Uttarakhand : https://www.uktenders.gov.in/nicgep/app
#   Jharkhand   : https://jharkhandtenders.gov.in/nicgep/app
#   Himachal Pradesh: https://hptenders.gov.in/nicgep/app
#   Karnataka   : https://eproc.karnataka.gov.in/eprocportal/pages/index.jsp  (different UI)
#
# Method : Selenium VISIBLE Chrome (CAPTCHA must be solved manually)
# Login  : None required (public tenders listing)
# CAPTCHA: Present on listing page — one manual solve per portal per run
#
# ⚠️  EXCLUDED FROM DEFAULT RUN — requires manual CAPTCHA interaction
#     Run explicitly: python3 main.py --nic
#
# Each portal is scraped sequentially. For each:
#   1. Opens the "Latest Active Tenders" page in a visible Chrome window
#   2. Waits for user to solve CAPTCHA (120s timeout)
#   3. Scrapes all paginated tender results
#   4. Filters for Consultancy / Services category
#   5. Scores relevance and deduplicates via MySQL
# =============================================================================

import os, re, time, base64
import requests

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

# ── NIC State Portal Registry ──────────────────────────────────────────────────
# Each entry: (label, base_url)
NIC_PORTALS = [
    ("UP eProcurement",           "https://etender.up.nic.in/nicgep/app"),
    ("Punjab eProcurement",       "https://eproc.punjab.gov.in/nicgep/app"),
    ("TN TNEARD",                 "https://tntenders.gov.in/nicgep/app"),
    ("Tripura Tenders",           "https://tripuratenders.gov.in/nicgep/app"),
    ("Odisha eProcurement",       "https://tendersodisha.gov.in/nicgep/app"),
    ("Nagaland Tenders",          "https://nagalandtenders.gov.in/nicgep/app"),
    ("Maharashtra Tenders",       "https://mahatenders.gov.in/nicgep/app"),
    ("Uttarakhand Tenders",       "https://www.uktenders.gov.in/nicgep/app"),
    ("Jharkhand Tenders",         "https://jharkhandtenders.gov.in/nicgep/app"),
    ("Himachal Pradesh Tenders",  "https://hptenders.gov.in/nicgep/app"),
    ("Kerala eProcurement",       "https://etenders.kerala.gov.in/nicgep/app"),
    ("MP Tenders",                "https://mptenders.gov.in/nicgep/app"),
]
# Karnataka uses a different UI — handled separately at the end
KARNATAKA_URL = "https://eproc.karnataka.gov.in/eprocportal/pages/index.jsp"

NIC_EXCEL_PATH      = os.path.join(PORTAL_EXCELS_DIR, "NIC_States_Tenders_Master.xlsx")
CAPTCHA_WAIT        = 120   # seconds
ACTION_DELAY        = 2.0
MAX_PAGES_PER_PORTAL = 50
CAPTCHA_API_KEY = os.environ.get("CAPTCHA_API_KEY") or os.environ.get("TWOCAPTCHA_API_KEY", "")
NIC_AUTO_CAPTCHA = str(os.environ.get("NIC_AUTO_CAPTCHA", "")).strip().lower() in {"1", "true", "yes", "on"}
CAPTCHA_IN_URL = "https://2captcha.com/in.php"
CAPTCHA_RES_URL = "https://2captcha.com/res.php"
CAPTCHA_POLL_INTERVAL = 5
CAPTCHA_POLL_MAX = 20

# Consulting-relevant categories on NIC portals
CONSULT_CATEGORIES = [
    "consultancy", "consultant", "service", "services", "advisory",
    "survey", "study", "research", "evaluation", "assessment",
    "training", "capacity", "audit", "technical assistance",
]

def _is_consulting_category(cat: str) -> bool:
    c = cat.lower()
    return any(kw in c for kw in CONSULT_CATEGORIES)

def _is_goods_title(title: str) -> bool:
    t = title.lower()
    goods = ["supply of", "purchase of", "procurement of", "rate contract",
             "annual maintenance", "amc for", "printing of"]
    return any(g in t for g in goods)


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Portal",         22),
    ("Tender ID",      22),
    ("Title",          65),
    ("Organisation",   35),
    ("Category",       20),
    ("Tender Type",    14),
    ("Closing Date",   18),
    ("Published",      16),
    ("Relevance",      40),
    ("Detail Link",    55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Selenium driver (VISIBLE — CAPTCHA required)
# =============================================================================

def _make_driver():
    opts = Options()
    # NOT headless — user must solve CAPTCHA
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1280,900")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


# =============================================================================
# SECTION 2 — CAPTCHA wait helper
# =============================================================================

NIC_RESULT_KEYWORDS = [
    "tender id", "tender title", "closing date", "organisation name",
    "published date", "no tender found", "no record", "bid submission",
    "tender document", "closing on",
]

def _wait_for_results(driver, portal_label: str, timeout: int = CAPTCHA_WAIT) -> bool:
    """Wait until NIC portal renders the tender table (post-CAPTCHA)."""
    print(f"\n[nic] ⏳ Waiting for CAPTCHA solve on {portal_label}...", flush=True)
    print(f"[nic]    → Chrome window is open — solve CAPTCHA and click Search/Submit", flush=True)
    print(f"[nic]    → Timeout: {timeout}s", flush=True)

    start = time.time()
    while time.time() - start < timeout:
        html = driver.page_source.lower()
        if any(kw in html for kw in NIC_RESULT_KEYWORDS):
            print(f"[nic] ✅ Results loaded for {portal_label}!", flush=True)
            return True
        time.sleep(1.5)

    print(f"[nic] ❌ Timed out on {portal_label} — skipping.", flush=True)
    return False


def _two_captcha_submit_image(image_b64: str) -> str:
    if not CAPTCHA_API_KEY or not image_b64:
        return ""
    try:
        resp = requests.post(
            CAPTCHA_IN_URL,
            data={
                "key": CAPTCHA_API_KEY,
                "method": "base64",
                "body": image_b64,
                "json": 1,
            },
            timeout=60,
        )
        data = resp.json() if resp.ok else {}
        if not data or data.get("status") != 1:
            print(f"[nic] auto-captcha submit failed: {data}", flush=True)
            return ""
        return str(data.get("request") or "")
    except Exception as exc:
        print(f"[nic] auto-captcha submit error: {exc}", flush=True)
        return ""


def _two_captcha_poll(captcha_id: str) -> str:
    if not captcha_id:
        return ""
    for _ in range(CAPTCHA_POLL_MAX):
        try:
            time.sleep(CAPTCHA_POLL_INTERVAL)
            resp = requests.get(
                CAPTCHA_RES_URL,
                params={"key": CAPTCHA_API_KEY, "action": "get", "id": captcha_id, "json": 1},
                timeout=30,
            )
            data = resp.json() if resp.ok else {}
            if not data:
                continue
            if data.get("status") == 1:
                return str(data.get("request") or "")
            if data.get("request") not in ("CAPTCHA_NOT_READY", "CAPCHA_NOT_READY"):
                print(f"[nic] auto-captcha poll failed: {data}", flush=True)
                return ""
        except Exception:
            continue
    return ""


def _try_auto_solve_captcha(driver, portal_label: str) -> bool:
    if not NIC_AUTO_CAPTCHA:
        return False
    if not CAPTCHA_API_KEY:
        print(f"[nic] auto-captcha enabled but CAPTCHA_API_KEY missing ({portal_label})", flush=True)
        return False
    try:
        images = driver.find_elements(By.CSS_SELECTOR, "img[src*='captcha'], img[id*='captcha'], img[class*='captcha']")
        if not images:
            return False
        img = next((i for i in images if i.is_displayed()), images[0])
        png = img.screenshot_as_png
        if not png:
            return False
        captcha_id = _two_captcha_submit_image(base64.b64encode(png).decode("ascii"))
        answer = _two_captcha_poll(captcha_id)
        if not answer:
            return False

        inputs = driver.find_elements(By.CSS_SELECTOR, "input[id*='captcha'], input[name*='captcha'], input[placeholder*='captcha']")
        if not inputs:
            inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
        target = next((i for i in inputs if i.is_displayed() and i.is_enabled()), None)
        if not target:
            return False
        target.clear()
        target.send_keys(answer)

        clicked = False
        for xp in [
            "//input[@type='submit']",
            "//button[@type='submit']",
            "//button[contains(translate(normalize-space(.),'SEARCHSUBMIT','searchsubmit'),'search')]",
            "//input[contains(translate(@value,'SEARCHSUBMIT','searchsubmit'),'search')]",
            "//button[contains(translate(normalize-space(.),'SEARCHSUBMIT','searchsubmit'),'submit')]",
        ]:
            try:
                btn = driver.find_element(By.XPATH, xp)
                if btn.is_displayed() and btn.is_enabled():
                    driver.execute_script("arguments[0].click();", btn)
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            target.submit()

        for _ in range(18):
            html = driver.page_source.lower()
            if any(kw in html for kw in NIC_RESULT_KEYWORDS):
                print(f"[nic] ✅ auto-captcha solved for {portal_label}", flush=True)
                return True
            time.sleep(1.5)
    except Exception as exc:
        print(f"[nic] auto-captcha error on {portal_label}: {exc}", flush=True)
    return False


# =============================================================================
# SECTION 3 — Parse NIC tender table (shared across all NIC portals)
# =============================================================================

def _parse_nic_table(html: str, portal_base_url: str) -> list:
    """Parse the standard NIC eProcurement tender listing table."""
    soup    = BeautifulSoup(html, 'html.parser')
    results = []

    for table in soup.find_all('table'):
        rows = table.find_all('tr')
        if len(rows) < 2:
            continue

        header_row = rows[0]
        headers    = [th.get_text(strip=True).lower() for th in header_row.find_all(['th', 'td'])]
        header_str = " ".join(headers)

        if not any(kw in header_str for kw in
                   ["tender id", "tender title", "closing date", "organisation", "bid submission"]):
            continue

        # Map header names to column indices
        col = {}
        for i, h in enumerate(headers):
            if "tender id"     in h:                   col.setdefault("tender_id",  i)
            if "tender title"  in h or "title" in h:   col.setdefault("title",      i)
            if "organisation"  in h or "dept"  in h:   col.setdefault("org",        i)
            if "category"      in h:                   col.setdefault("category",   i)
            if "type"          in h and "tender" in h: col.setdefault("type",       i)
            if "closing"       in h or "end date" in h:col.setdefault("closing",    i)
            if "published"     in h or "posted"  in h: col.setdefault("published",  i)
            if "ref"           in h:                   col.setdefault("ref",        i)

        def gcell(row_cells, key):
            idx = col.get(key)
            if idx is not None and idx < len(row_cells):
                return row_cells[idx].get_text(" ", strip=True)
            return ""

        for row in rows[1:]:
            cells   = row.find_all(['td', 'th'])
            if len(cells) < 3:
                continue
            if len(row.get_text(strip=True)) < 5:
                continue

            title    = gcell(cells, "title")
            if not title:
                title = max(cells, key=lambda c: len(c.get_text())).get_text(strip=True)
            if not title or len(title) < 5:
                continue

            category = gcell(cells, "category")

            # Skip non-consulting categories
            if category and not _is_consulting_category(category):
                continue

            # Skip goods titles even when category passes
            if _is_goods_title(title):
                continue

            # Detail link
            detail_link = ""
            for a in row.find_all('a', href=True):
                href = a['href']
                if any(kw in href.lower() for kw in ['viewtender', 'view', 'detail', 'tender']):
                    detail_link = href if href.startswith('http') else portal_base_url + href
                    break

            entry = {
                "Tender ID":    gcell(cells, "tender_id"),
                "Title":        title,
                "Organisation": gcell(cells, "org"),
                "Category":     category,
                "Tender Type":  gcell(cells, "type"),
                "Closing Date": gcell(cells, "closing"),
                "Published":    gcell(cells, "published"),
                "Detail Link":  detail_link,
            }
            results.append(entry)

        if results:
            return results

    return results


# =============================================================================
# SECTION 4 — Paginate NIC portal
# =============================================================================

def _get_next_page(driver) -> bool:
    """Click Next page button. Returns True if clicked successfully."""
    next_xpaths = [
        "//a[normalize-space()='Next']",
        "//a[contains(text(),'Next')]",
        "//a[contains(text(),'>')]",
        "//input[@value='Next']",
        "//li[contains(@class,'next')]/a",
        "//span[contains(@class,'next')]/a",
    ]
    for xpath in next_xpaths:
        try:
            el = driver.find_element(By.XPATH, xpath)
            if el.is_displayed() and el.is_enabled():
                driver.execute_script("arguments[0].click();", el)
                time.sleep(ACTION_DELAY)
                return True
        except Exception:
            continue
    return False


def _scrape_portal(driver, portal_label: str, portal_url: str) -> list:
    """Scrape a single NIC portal. Returns list of raw tender dicts."""
    active_url = f"{portal_url}?page=FrontEndLatestActiveTenders&service=page"
    print(f"\n[nic] Opening portal: {portal_label}", flush=True)
    print(f"[nic]   URL: {active_url}", flush=True)

    try:
        driver.get(active_url)
        time.sleep(2)
    except Exception as e:
        print(f"[nic]   ERROR loading {portal_label}: {e}", flush=True)
        return []

    # Check if CAPTCHA needed
    html_lower = driver.page_source.lower()
    needs_captcha = any(kw in html_lower for kw in ['captcha', 'enter captcha', 'type captcha'])

    if needs_captcha:
        solved = _try_auto_solve_captcha(driver, portal_label)
        if not solved and not _wait_for_results(driver, portal_label):
            return []
    else:
        # No CAPTCHA — results should already be loading
        time.sleep(ACTION_DELAY)

    results  = []
    page_num = 1

    while page_num <= MAX_PAGES_PER_PORTAL:
        print(f"[nic]   Page {page_num}...", end=" ", flush=True)
        html = driver.page_source

        if any(kw in html.lower() for kw in ['no tender found', 'no record', '0 tender']):
            print("no tenders found.")
            break

        entries = _parse_nic_table(html, portal_url)
        print(f"{len(entries)} consulting tenders")
        results.extend(entries)

        if not _get_next_page(driver):
            print(f"[nic]   Last page for {portal_label}.")
            break
        page_num += 1

    return results


# =============================================================================
# SECTION 5 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "NIC State Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")    + 1
    link_idx = col_names.index("Detail Link")  + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(NIC_EXCEL_PATH)
    print(f"[nic] Excel saved: {NIC_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 6 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the unified NIC State Portals pipeline.
    ⚠️  Requires manual CAPTCHA solve in visible Chrome window.
    Returns (new_tenders, all_rows).
    """
    print("\n" + "=" * 65, flush=True)
    print("[nic] NIC State Portals Pipeline starting...", flush=True)
    print("[nic] ⚠️  CAPTCHA handling: auto via 2captcha (if NIC_AUTO_CAPTCHA=1), else manual", flush=True)
    print("[nic]   Portals:", ", ".join(p[0] for p in NIC_PORTALS), flush=True)
    print("=" * 65, flush=True)

    new_tenders, all_rows = [], []

    if os.path.exists(NIC_EXCEL_PATH):
        try:
            os.remove(NIC_EXCEL_PATH)
            print("[nic] Cleared old Excel")
        except Exception:
            pass

    driver = None
    try:
        driver = _make_driver()
        all_raw = []

        for portal_label, portal_url in NIC_PORTALS:
            raw = _scrape_portal(driver, portal_label, portal_url)
            for entry in raw:
                entry["Portal"] = portal_label
            all_raw.extend(raw)
            print(f"[nic]   {portal_label}: {len(raw)} tenders extracted")
            time.sleep(2)

        print(f"\n[nic] Total raw tenders across all portals: {len(all_raw)}")

        # Deduplicate by Tender ID
        seen_portal_ids = set()
        unique = []
        for entry in all_raw:
            pid   = f"{entry.get('Portal','')}_{entry.get('Tender ID','') or entry.get('Title','')[:40]}"
            if pid not in seen_portal_ids:
                seen_portal_ids.add(pid)
                unique.append(entry)

        print(f"[nic] After dedup: {len(unique)} unique tenders")

        # Score relevance + DB dedup
        for t in unique:
            combined  = (t.get("Title", "") + " " + t.get("Organisation", "") +
                         " " + t.get("Category", ""))
            relevance = score_relevance(combined)
            t["Relevance"] = relevance
            all_rows.append(t)

            portal_prefix = re.sub(r'[^A-Z0-9]', '', t.get("Portal", "NIC").upper())[:6]
            tid_raw   = t.get("Tender ID", "").strip()
            tender_id = (f"NIC_{portal_prefix}_{tid_raw}"
                         if tid_raw
                         else f"NIC_{portal_prefix}_{t['Title'][:50]}")
            tender_id = re.sub(r'\s+', '_', tender_id)[:200]

            detail_url = t.get("Detail Link", "")

            if check_if_new(tender_id):
                mark_as_seen(tender_id, title=t["Title"], source_site=f"NIC-{t.get('Portal','')}",
                             url=detail_url or "")
                new_tenders.append({
                    "title":    t["Title"],
                    "deadline": t.get("Closing Date", ""),
                    "value":    t.get("Category", ""),
                    "url":      detail_url or "",
                })
                print(f"[nic]   → NEW: {t['Title'][:70]} | {relevance[:40] or '—'}", flush=True)

    except Exception as e:
        print(f"[nic] FATAL error: {e}")
    finally:
        if driver:
            try:
                driver.quit()
                print("[nic] Browser closed.")
            except Exception:
                pass

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(f"\n[nic] Done — {len(all_rows)} tenders, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
