from __future__ import annotations

# =============================================================================
# ec_scraper.py — European Commission Funding & Tenders Portal  v2
#
# Source  : https://ec.europa.eu/info/funding-tenders/opportunities/portal/
#           screen/opportunities/calls-for-tenders
#
# API     : https://api.tech.ec.europa.eu/search-api/prod/rest/search
#           Officially documented at:
#           https://ec.europa.eu/info/funding-tenders/opportunities/portal/
#           screen/support/apis
#
#           IMPORTANT: This API uses multipart/form-data (NOT JSON).
#           Parameters are passed as form fields.  No auth required.
#           Rate limit: not documented; ~1 req/sec is safe.
#
# Scope:
#   This scraper targets EC INSTITUTIONAL contracts — calls for tenders and
#   proposals issued by European Commission DGs, executive agencies, and
#   EU bodies (e.g. DEVCO/DG INTPA, EuropeAid, EU Delegations, ECHO).
#   These cover:
#     • Technical assistance and capacity building in developing countries
#     • Evaluation and monitoring studies
#     • Development cooperation programmes
#
#   NOTE: Member-state public procurement is already covered by the TED-EU
#   scraper (api.ted.europa.eu/v3) — no duplication here.
#
# =============================================================================

import json
import os
import time
from datetime import datetime, timedelta
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from config.config import EC_API_KEY
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Registry metadata (override static defaults) ─────────────────────────────
SCRAPER_META = {
    "flag":        "ec",
    "label":       "European Commission (EC)",
    "group":       "api",
    "timeout":     600,
    "max_retries": 1,
    "auto":        True,
}

# ── Constants ─────────────────────────────────────────────────────────────────
EC_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "EC_Tenders_Master.xlsx")
CACHE_PATH    = os.path.join(os.path.dirname(EC_EXCEL_PATH), ".ec_cache.json")

SOURCE_NAME   = "EC"

# Confirmed endpoint (from API research + network inspection):
_API_URL  = "https://api.tech.ec.europa.eu/search-api/prod/rest/search"
# Portal base for building direct links:
_PORTAL   = "https://ec.europa.eu/info/funding-tenders/opportunities/portal"

PAGE_SIZE = 50     # EC portal default is 50
MAX_PAGES = 6      # 6 × 50 = 300 results per query

# IDCG-targeted search queries — EC calls for technical assistance, evaluation,
# development cooperation  (DG INTPA / EuropeAid / DEVCO focus)
_QUERIES = [
    "technical assistance evaluation monitoring",
    "capacity building assessment developing countries",
    "impact evaluation development cooperation",
    "monitoring evaluation south asia",
    "monitoring evaluation africa",
    "independent evaluation governance",
    "third party monitoring humanitarian",
    "social assessment environment climate",
]

_MAX_RETRIES     = 2
_RETRY_BASE_WAIT = 2

# ── Excel styles ──────────────────────────────────────────────────────────────
COLUMNS = [
    ("Identifier",   16),
    ("Title",        60),
    ("Programme",    22),
    ("Type",         20),
    ("Deadline",     13),
    ("Budget (€)",   16),
    ("Countries",    20),
    ("Relevance",    38),
    ("Link",         55),
]
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_ALT_FILL  = PatternFill("solid", fgColor="F5F8FF")
_REL_FILL  = PatternFill("solid", fgColor="E2EFDA")
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_LINK_FONT = Font(name="Calibri", size=10, color="1155CC", underline="single")
_HIGH_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),  right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin",  color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


def _save_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "EC Tenders"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    col_names = [c[0] for c in COLUMNS]
    for ci, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    rel_idx  = col_names.index("Relevance") + 1
    link_idx = col_names.index("Link")      + 1

    for ri, row_data in enumerate(rows, 2):
        alt = _ALT_FILL if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 40
        vals = [
            row_data.get("identifier", ""),
            row_data.get("title", ""),
            row_data.get("programme", ""),
            row_data.get("call_type", ""),
            row_data.get("deadline", ""),
            row_data.get("budget", ""),
            row_data.get("countries", ""),
            row_data.get("relevance", ""),
            row_data.get("link", ""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = _LINK_FONT
            elif ci == rel_idx:
                if val:
                    cell.fill = _REL_FILL
                    cell.font = _HIGH_FONT
                else:
                    cell.font = _BODY_FONT
                    if alt:
                        cell.fill = alt
            else:
                cell.font = _BODY_FONT
                if alt:
                    cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(EC_EXCEL_PATH)
    print(f"[ec] Excel saved: {EC_EXCEL_PATH}  ({len(rows)} rows)")


# ── Cache helpers ─────────────────────────────────────────────────────────────

def _load_cache() -> list[dict]:
    try:
        if os.path.exists(CACHE_PATH):
            with open(CACHE_PATH, encoding="utf-8") as f:
                data = json.load(f)
            age_h = (time.time() - data.get("saved_at", 0)) / 3600
            if age_h < 12:
                print(f"[ec] Using cache ({age_h:.1f}h old, {len(data['calls'])} items)")
                return data["calls"]
    except Exception:
        pass
    return []


def _save_cache(calls: list[dict]) -> None:
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump({"saved_at": time.time(), "calls": calls}, f)
    except Exception:
        pass


# ── API helpers ───────────────────────────────────────────────────────────────

_SESSION = requests.Session()
_SESSION.headers.update({
    "Accept":          "application/json",
    "Accept-Language": "en",
    "User-Agent":      "TenderMonitor/2.0 (+https://idcg.co.in; monitoring@idcg.co.in)",
    "Referer":         f"{_PORTAL}/screen/opportunities/calls-for-tenders",
    "Origin":          _PORTAL,
    "Content-Type":    "application/json",
})
# EC Funding & Tenders uses shared frontend keys (no user-generated keys).
# Always prefer shared keys, ignore custom EC_API_KEY values.
_EC_SHARED_KEYS = ["SEDIA", "SEDIA-NEWS"]
_ACTIVE_EC_KEY = _EC_SHARED_KEYS[0]
_SESSION.headers["apikey"] = _ACTIVE_EC_KEY


def _fetch_page(query: str, page: int) -> Optional[dict]:
    """
    Fetch one page from the EC Funding & Tenders search API.

    API confirmed to use multipart/form-data (NOT JSON body).
    Parameters from the portal URL:
      isExactMatch, order, pageNumber, pageSize, sortBy, statusFilter, query
    """
    # The API accepts parameters as URL query string (GET) or multipart form (POST).
    # We use GET with query params — simpler and confirmed to work.
    params = {
        "text":         query,
        "isExactMatch": "false",
        "order":        "DESC",
        "pageNumber":   str(page),
        "pageSize":     str(PAGE_SIZE),
        "sortBy":       "startDate",
        "statusFilter": "OPEN",
        "programme":    "allProgrammes",
    }

    # Browser-style payload (JSON body)
    query_payload = {
        "bool": {
            "must": [
                {"term": {"type": "TENDER"}},
                {"term": {"status": "OPEN"}},
            ]
        }
    }
    languages_payload = ["en"]
    sort_payload = [{"startDate": {"order": "desc"}}]
    display_fields = [
        "id", "title", "status", "type", "deadlineDate", "publicationDate",
        "url", "programme", "identifier", "summary", "keywords",
    ]

    wait = _RETRY_BASE_WAIT
    active_key = _ACTIVE_EC_KEY
    tried_simple = False
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            # POST with request parameters + JSON body (browser behavior)
            if not tried_simple:
                payload = {
                    "query": query_payload,
                    "languages": languages_payload,
                    "sort": sort_payload,
                    "displayFields": display_fields,
                    "text": params["text"],
                    "isExactMatch": params["isExactMatch"],
                    "order": params["order"],
                    "pageNumber": params["pageNumber"],
                    "pageSize": params["pageSize"],
                    "sortBy": params["sortBy"],
                    "statusFilter": params["statusFilter"],
                    "programme": params["programme"],
                }
                r = _SESSION.post(
                    _API_URL,
                    params={**params, "apiKey": active_key},
                    json=payload,
                    timeout=30,
                )
            else:
                # Minimal fallback: JSON body only (no params)
                payload = {
                    "query": query_payload,
                    "languages": languages_payload,
                    "sort": sort_payload,
                    "displayFields": display_fields,
                    "text": params["text"],
                    "pageNumber": params["pageNumber"],
                    "pageSize": params["pageSize"],
                }
                r = _SESSION.post(
                    _API_URL,
                    params={"apiKey": active_key},
                    json=payload,
                    timeout=30,
                )

            if r.status_code == 200:
                try:
                    return r.json()
                except ValueError:
                    return None

            if r.status_code == 400 and "api key" in (r.text or "").lower():
                # Swap to shared fallback key and retry once
                if active_key != _EC_SHARED_KEYS[-1]:
                    active_key = _EC_SHARED_KEYS[-1]
                    _SESSION.headers["apikey"] = active_key
                    print("[ec]   apiKey rejected — retrying with SEDIA-NEWS")
                    time.sleep(1)
                    continue
                print("[ec]   EC API rejected apiKey — continuing with retry policy")

            if r.status_code == 429:
                retry_after = int(r.headers.get("Retry-After", wait))
                print(f"[ec]   Rate-limited — waiting {retry_after}s")
                time.sleep(retry_after)
                wait *= 2
                continue

            if r.status_code >= 500:
                if not tried_simple:
                    print("[ec]   Server error — retrying with minimal params-only request")
                    tried_simple = True
                    time.sleep(1)
                    continue
                print(f"[ec]   Server error {r.status_code} — retrying in {wait}s")
                time.sleep(wait)
                wait *= 2
                continue

            if r.status_code == 404:
                print(f"[ec]   404 — endpoint may have changed. Check: {_API_URL}")
                return None

            body_snip = (r.text or "")[:400].replace("\\n", " ")
            print(f"[ec]   HTTP {r.status_code} on page {page} body='{body_snip}'")
            return None

        except requests.Timeout:
            print(f"[ec]   Timeout (attempt {attempt}/{_MAX_RETRIES})")
            time.sleep(wait)
            wait *= 2
        except requests.ConnectionError as exc:
            print(f"[ec]   Connection error: {exc}")
            time.sleep(wait)
            wait *= 2
        except Exception as exc:
            print(f"[ec]   Unexpected error: {exc}")
            return None

    return None


def _extract_call(item: dict) -> Optional[dict]:
    """
    Parse a single result item from the EC API response.
    Handles multiple response structures (the EC API response varies).
    """
    if not isinstance(item, dict):
        return None
    # The EC API wraps fields in 'metadata' and 'content' sub-objects
    # Some fields may be arrays of one element, some may be plain strings
    metadata = item.get("metadata") or {}
    content  = item.get("content")  or {}
    if not isinstance(metadata, dict):
        metadata = {}
    if not isinstance(content, dict):
        content = {}

    def _first(val):
        """Extract first element if list, otherwise return as-is."""
        if isinstance(val, list):
            return val[0] if val else ""
        return val or ""

    # Title (required)
    title = str(_first(
        metadata.get("title") or content.get("title") or item.get("title") or ""
    )).strip()
    if not title or len(title) < 5:
        return None

    identifier = str(_first(
        metadata.get("identifier") or metadata.get("callIdentifier")
        or content.get("identifier") or item.get("identifier") or ""
    )).strip()

    # Call type
    call_type_raw = str(_first(
        metadata.get("callType") or metadata.get("type")
        or content.get("type") or item.get("type") or ""
    ))
    call_type = call_type_raw.replace("_", " ").title()

    # Deadline — skip already-closed calls (> 7 days past)
    deadline_raw = str(_first(
        metadata.get("deadlineDate") or metadata.get("closingDate")
        or content.get("deadlineDate") or ""
    ))
    deadline = deadline_raw[:10] if deadline_raw else ""
    if deadline:
        try:
            if datetime.strptime(deadline, "%Y-%m-%d") < datetime.now() - timedelta(days=7):
                return None
        except ValueError:
            pass

    # Programme / funding source
    programme = str(_first(
        metadata.get("programmeName") or metadata.get("fundingScheme")
        or metadata.get("programme") or ""
    ))[:80]

    # Budget
    budget = str(_first(
        metadata.get("budgetOverall") or metadata.get("budgetTopicAction")
        or metadata.get("budget") or ""
    ))[:30]

    # Countries / locations
    locations_raw = metadata.get("locations") or metadata.get("countries") or []
    if isinstance(locations_raw, list):
        countries = ", ".join(str(c) for c in locations_raw[:5])
    else:
        countries = str(locations_raw)[:80]

    # Build direct link to the call
    if identifier:
        link = (
            f"{_PORTAL}/screen/opportunities/competitive-calls-grants/{identifier}"
        )
    else:
        link = (
            metadata.get("callUrl") or metadata.get("url")
            or content.get("url") or item.get("url") or ""
        )

    # Summary / description (when present)
    summary = str(_first(
        metadata.get("summary") or metadata.get("shortSummary")
        or metadata.get("description") or content.get("summary")
        or content.get("description") or item.get("summary")
        or item.get("description") or ""
    ))

    return {
        "identifier": identifier,
        "title":      title,
        "call_type":  call_type,
        "programme":  programme,
        "deadline":   deadline,
        "budget":     budget,
        "countries":  countries,
        "link":       str(link),
        "url":        str(link),
        "description": summary[:4000],
    }


def _fetch_all_calls() -> list[dict]:
    """Run all search queries, paginate, deduplicate by identifier."""
    seen_ids: set[str] = set()
    all_calls: list[dict] = []

    for qi, query in enumerate(_QUERIES, 1):
        print(f"[ec]   Query {qi}/{len(_QUERIES)}: '{query}'...")
        query_count = 0

        for page in range(1, MAX_PAGES + 1):
            data = _fetch_page(query, page)
            if data is None:
                break
            if not isinstance(data, (dict, list)):
                print(f"[ec]   Unexpected response type: {type(data)}")
                break

            # Response structure varies — try several known key names
            if isinstance(data, list):
                results = data
                total = len(results)
            else:
                results = (
                    data.get("results")
                    or data.get("hits")
                    or data.get("data")
                    or data.get("items")
                    or []
                )
                total = int(
                    data.get("totalResults")
                    or data.get("total")
                    or data.get("totalCount")
                    or 0
                )

            if not results:
                if page == 1:
                    print(f"[ec]   Query returned no results (response keys: {list(data.keys())[:5]})")
                break

            for item in results:
                call = _extract_call(item)
                if call is None:
                    continue
                uid = call.get("identifier") or call.get("title", "")[:60]
                if uid in seen_ids:
                    continue
                seen_ids.add(uid)
                all_calls.append(call)
                query_count += 1

            fetched = (page - 1) * PAGE_SIZE + len(results)
            if fetched >= total:
                break

            time.sleep(1.0)   # EC rate limit: ~1 req/sec is safe

        print(f"[ec]   → {query_count} new calls")
        time.sleep(2.0)

    return all_calls


# ── Main ──────────────────────────────────────────────────────────────────────

def run() -> tuple:
    """
    Run the EC Funding & Tenders pipeline.
    Returns (new_tenders: list, all_rows: list).
    """
    print("\n[ec] European Commission Funding & Tenders Scraper v2")
    print("[ec] Targeting: EC-issued contracts (DG INTPA, EuropeAid, EU agencies)")
    print("[ec] Note: Member-state procurement is handled by the TED-EU scraper")
    # Shared keys are used; no user-generated EC key required.

    raw_calls = _fetch_all_calls()

    if not raw_calls:
        print("[ec] Live API returned nothing — trying cache...")
        raw_calls = _load_cache()
        if not raw_calls:
            print("[ec] No cache — skipping.")
            return [], []

    if raw_calls:
        _save_cache(raw_calls)

    print(f"[ec] Total unique calls: {len(raw_calls)}")

    new_tenders: list[dict] = []
    all_rows:    list[dict] = []

    for call in raw_calls:
        title = (call.get("title") or "").strip()
        if not title:
            continue

        relevance = score_relevance(title, "")
        row = {**call, "relevance": relevance}
        all_rows.append(row)

        ident = call.get("identifier") or ""
        tid   = f"EC_{ident}" if ident else f"EC_{hash(title) & 0xFFFFFF}"
        if check_if_new(tid):
            mark_as_seen(tid, title, SOURCE_NAME, call.get("link", ""))
            new_tenders.append({
                "title":    title,
                "deadline": call.get("deadline", ""),
                "value":    call.get("budget", ""),
                "url":      call.get("link", ""),
            })

    relevant = sum(1 for r in all_rows if r.get("relevance"))
    print(f"[ec] {len(all_rows)} rows ({relevant} relevant, {len(new_tenders)} NEW)")

    if all_rows:
        _save_excel(all_rows)
    else:
        print("[ec] No rows to save.")

    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nResult: {len(rows)} total, {len(new)} new")
