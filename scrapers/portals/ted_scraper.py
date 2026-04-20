# =============================================================================
# ted_pipeline.py — TED (Tenders Electronic Daily) EU Pipeline
#
# Site   : https://ted.europa.eu/en/search/result
# Method : requests + BeautifulSoup (React/MUI table, server-side rendered HTML)
# Login  : None required
# CAPTCHA: None
#
# What this does:
#   1. Runs targeted keyword searches on TED for India/South-Asia development tenders
#   2. Filters for consulting/advisory/technical-assistance type tenders only
#   3. Paginates through results (?page=N, 50 per page)
#   4. Extracts: title, country, deadline, notice number, link
#   5. Scores relevance against IDCG expertise keywords
#   6. Deduplicates via MySQL — only NEW notices trigger notification
#   7. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
#
# TED = EU's Official Journal Supplement — covers:
#   • EU-funded development contracts in India/South Asia (EuropeAid, INTPA, FPI)
#   • Technical assistance, M&E, research, capacity building tenders
#   • 500K+ active notices total — we search with tight keyword filters
#
# Search strategy:
#   Run multiple targeted queries, deduplicate by notice number:
#   Query 1: "India" + "technical assistance"
#   Query 2: "India" + "consulting" OR "consultancy"
#   Query 3: "South Asia" + "development"
#   Query 4: India + CPV 73000000/79000000 (R&D / Business services)
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import urlencode

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://ted.europa.eu"
SEARCH_URL  = f"{BASE_URL}/en/search/result"
# TED relaunched as a React SPA (~2023). The HTML endpoint returns a JS shell.
# Use the official TED Open Data REST API v3 instead for reliable JSON responses.
TED_API_URL = "https://api.ted.europa.eu/v3/notices/search"

TED_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "TED_EU_Tenders_Master.xlsx")

DELAY       = 1.5    # polite crawl
MAX_PAGES   = 10     # 50/page × 10 = 500 max per query
MAX_RESULTS = 200    # total across all queries before cutting off (TED is huge)

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer":         BASE_URL,
}

# TED API v3 expert query syntax: FIELD~VALUE (~ = contains, AND/OR supported)
# Field codes: TI=title, ND=notice number, CY=country, PD=publish date, DT=deadline
# Note: free-text queries are NOT supported — must use field operators.
SEARCH_QUERIES = [
    # India-specific
    "TI~India AND TI~consulting",
    "TI~India AND TI~evaluation",
    "TI~India AND TI~advisory",
    "TI~India AND TI~monitoring",
    "TI~India AND TI~assessment",
    "TI~India AND TI~capacity",
    "TI~India AND TI~research",
    # South Asia
    "TI~\"South Asia\" AND TI~development",
    "TI~Bangladesh AND TI~consulting",
    "TI~Nepal AND TI~consulting",
    # Africa
    "TI~Africa AND TI~evaluation",
    "TI~Africa AND TI~consulting",
    "TI~Africa AND TI~\"technical assistance\"",
    # Global development consulting
    "TI~\"technical assistance\" AND TI~development",
    "TI~\"impact evaluation\" AND TI~development",
    "TI~\"monitoring and evaluation\"",
    "TI~\"capacity building\" AND TI~development",
    "TI~\"baseline survey\"",
    "TI~\"third party monitoring\"",
]

# Skip goods/infrastructure/construction tenders
SKIP_CPV_KEYWORDS = [
    "works", "construction", "supply", "installation",
    "maintenance", "road", "bridge", "building",
]

def _is_goods(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in [
        "supply of", "works contract", "construction of", "building of",
        "installation of", "procurement of goods", "maintenance of",
    ])


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Notice No.",   18),
    ("Title",        65),
    ("Country",      18),
    ("Published",    14),
    ("Deadline",     18),
    ("Relevance",    40),
    ("Detail Link",  55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse TED search results table
# =============================================================================

def _parse_results(soup: BeautifulSoup) -> list:
    """
    Parse TED's MUI table structure.
    Columns: [checkbox] [Notice No.] [Description] [Country] [Published] [Deadline]
    """
    entries = []

    # Try standard table first
    table = soup.find('table')
    if table:
        rows = table.find_all('tr')
        for row in rows[1:]:   # skip header
            cells = row.find_all(['td', 'th'])
            if len(cells) < 4:
                continue

            # Notice number — cell with link pattern like "174374-2026"
            notice_no  = ""
            detail_url = ""
            for cell in cells:
                link = cell.find('a', href=True)
                if link:
                    txt = link.get_text(strip=True)
                    if re.match(r'\d{5,}-\d{4}', txt):
                        notice_no  = txt
                        href       = link.get('href', '')
                        detail_url = href if href.startswith('http') else BASE_URL + href
                        break

            if not notice_no:
                continue

            # Description — longest text cell
            description = ""
            for cell in cells:
                txt = cell.get_text(' ', strip=True)
                if len(txt) > len(description) and not re.match(r'\d{5,}-\d{4}', txt):
                    description = txt

            # Country, dates — detect by pattern
            country   = ""
            published = ""
            deadline  = ""
            for cell in cells:
                txt = cell.get_text(strip=True)
                # Date pattern DD/MM/YYYY
                if re.match(r'\d{2}/\d{2}/\d{4}', txt):
                    if not published:
                        published = txt
                    elif not deadline:
                        deadline = txt
                # Country — 2-letter code or short country name
                elif re.match(r'^[A-Z]{2,3}$', txt) and not country:
                    country = txt
                elif 2 < len(txt) < 40 and txt.replace(' ', '').isalpha() and not country:
                    country = txt

            entries.append({
                "Notice No.":  notice_no,
                "Title":       description[:300].strip(),
                "Country":     country,
                "Published":   published,
                "Deadline":    deadline,
                "Detail Link": detail_url,
                "tender_id":   f"TED_{notice_no}",
            })
        return entries

    # Fallback: look for notice links anywhere in the page
    for link in soup.find_all('a', href=re.compile(r'/en/notice/')):
        href = link.get('href', '')
        txt  = link.get_text(strip=True)
        if not txt:
            # Try parent text
            parent_txt = link.parent.get_text(' ', strip=True) if link.parent else ""
            txt = parent_txt[:200]

        notice_m  = re.search(r'(\d{5,}-\d{4})', href + txt)
        notice_no = notice_m.group(1) if notice_m else re.sub(r'[^a-zA-Z0-9]', '_', href.split('/')[-1])[:40]
        if not notice_no:
            continue

        detail_url = href if href.startswith('http') else BASE_URL + href
        entries.append({
            "Notice No.":  notice_no,
            "Title":       txt[:300].strip(),
            "Country":     "",
            "Published":   "",
            "Deadline":    "",
            "Detail Link": detail_url,
            "tender_id":   f"TED_{notice_no}",
        })

    return entries


# =============================================================================
# SECTION 2a — TED API v3 search (preferred: JSON, no browser needed)
# =============================================================================

def _search_api(session: requests.Session, query: str, seen_ids: set) -> list:
    """
    Query the TED Open Data API v3 (JSON).
    IMPORTANT: API v3 requires POST with JSON body + expert query syntax.
    Expert syntax: TI~India (title contains India), AND/OR supported.
    GET requests return 405. Free-text queries return 400 syntax error.
    """
    results = []
    api_headers = {**HEADERS, "Accept": "application/json",
                   "Content-Type": "application/json"}

    for page in range(1, MAX_PAGES + 1):
        body = {
            "query":  query,          # expert query syntax e.g. "TI~India AND TI~consulting"
            "scope":  "ACTIVE",
            "limit":  50,
            "page":   page,
            "fields": ["ND", "TI", "CY", "PD", "DT"],
        }
        try:
            r = session.post(TED_API_URL, json=body,
                             headers=api_headers, timeout=30)
            r.raise_for_status()
            data    = r.json()
            notices = data.get("notices") or []
        except Exception as e:
            print(f"[ted]   API page {page} error: {e}")
            break

        if not notices:
            break

        print(f"[ted]   API page {page}: {len(notices)} notices")

        for n in notices:
            nd = (n.get("ND") or n.get("nd") or "").strip()
            if not nd:
                continue

            tender_id = f"TED_{nd}"
            if tender_id in seen_ids:
                continue
            seen_ids.add(tender_id)

            # Title is a dict keyed by language code, or plain string
            ti = n.get("TI") or n.get("ti") or {}
            if isinstance(ti, dict):
                title = (ti.get("ENG") or ti.get("FRA") or
                         (list(ti.values())[0] if ti else ""))
            else:
                title = str(ti)
            if isinstance(title, list):
                title = title[0] if title else ""
            title = str(title).strip()

            cy = n.get("CY") or n.get("cy") or ""
            if isinstance(cy, list):
                cy = ", ".join(cy)

            # Dates arrive as YYYYMMDD strings — reformat to YYYY-MM-DD
            def _fmt(d):
                d = str(d or "").strip()
                return f"{d[:4]}-{d[4:6]}-{d[6:8]}" if len(d) == 8 and d.isdigit() else d

            published  = _fmt(n.get("PD") or n.get("pd") or "")
            deadline   = _fmt(n.get("DT") or n.get("dt") or "")
            detail_url = (n.get("uri") or
                          f"https://ted.europa.eu/en/notice/-/detail/{nd}")

            results.append({
                "Notice No.":  nd,
                "Title":       title[:300],
                "Country":     str(cy),
                "Published":   published,
                "Deadline":    deadline,
                "Detail Link": detail_url,
                "tender_id":   tender_id,
            })

        time.sleep(DELAY)

        # Stop when we've consumed all results
        total = int(data.get("total") or 0)
        if not notices or page * 50 >= total:
            break

    return results


# =============================================================================
# SECTION 2b — Run one search query (API → HTML fallback)
# =============================================================================

def _search(session: requests.Session, query: str, seen_ids: set) -> list:
    """Run a single TED search query and paginate. Returns new entries."""
    results = []
    print(f"[ted]   Query: '{query}'")

    # ── Preferred: TED Open Data API v3 (JSON) ─────────────────────────────────
    api_results = _search_api(session, query, seen_ids)
    if api_results:
        print(f"[ted]   API returned {len(api_results)} entries")
        return api_results

    # ── Fallback: legacy HTML scraping (works only if server-renders results) ──
    print(f"[ted]   API returned 0 — trying HTML fallback (may be empty on SPA)...")
    for page_num in range(MAX_PAGES):
        params = {
            "scope":       "ACTIVE",
            "q":           query,
            "page":        page_num,
        }
        url = f"{SEARCH_URL}?{urlencode(params)}"

        try:
            r = session.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            soup    = BeautifulSoup(r.text, 'html.parser')
            entries = _parse_results(soup)

            if not entries:
                print(f"[ted]     Page {page_num}: empty — done")
                break

            new_on_page = [e for e in entries if e["tender_id"] not in seen_ids]
            print(f"[ted]     Page {page_num}: {len(new_on_page)} new / {len(entries)} total")

            for e in new_on_page:
                seen_ids.add(e["tender_id"])
            results.extend(new_on_page)

            if len(new_on_page) == 0 and page_num > 0:
                break

            time.sleep(DELAY)

        except Exception as e:
            print(f"[ted]     Page {page_num} ERROR: {e}")
            break

    return results


# =============================================================================
# SECTION 3 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TED EU Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")   + 1
    link_idx = col_names.index("Detail Link") + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(TED_EXCEL_PATH)
    print(f"[ted] Excel saved: {TED_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 4 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the TED EU Tenders pipeline.
    Returns (new_tenders, all_rows).
    """
    print("\n" + "=" * 65)
    print("[ted] TED (Tenders Electronic Daily) EU Pipeline starting...")
    print(f"[ted] Running {len(SEARCH_QUERIES)} targeted queries for India/South Asia")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(TED_EXCEL_PATH):
        try:
            os.remove(TED_EXCEL_PATH)
            print("[ted] Cleared old Excel")
        except Exception:
            pass

    session  = requests.Session()
    session.headers.update(HEADERS)
    all_entries = []
    seen_ids    = set()

    # ── Run each search query ──────────────────────────────────────────────────
    for query in SEARCH_QUERIES:
        results = _search(session, query, seen_ids)
        all_entries.extend(results)
        print(f"[ted]   → {len(results)} new results for this query")
        time.sleep(DELAY * 2)

        # Safety cap — TED is massive
        if len(all_entries) >= MAX_RESULTS:
            print(f"[ted]   Reached MAX_RESULTS={MAX_RESULTS} cap — stopping early.")
            break

    print(f"\n[ted] Total unique entries across all queries: {len(all_entries)}")

    # ── Score relevance + DB dedup ─────────────────────────────────────────────
    for i, entry in enumerate(all_entries, 1):
        title = entry["Title"]

        if _is_goods(title):
            continue

        relevance = score_relevance(title, "")

        row = {
            "Notice No.":  entry["Notice No."],
            "Title":       title,
            "Country":     entry["Country"],
            "Published":   entry["Published"],
            "Deadline":    entry["Deadline"],
            "Relevance":   relevance,
            "Detail Link": entry["Detail Link"],
        }
        all_rows.append(row)

        tid = entry["tender_id"]
        if check_if_new(tid):
            mark_as_seen(tid, title=title[:200], source_site="TED-EU",
                         url=entry["Detail Link"])
            new_tenders.append({
                "title":    title[:120],
                "deadline": entry["Deadline"],
                "value":    "See TED",
                "url":      entry["Detail Link"],
            })
            print(f"[ted]   → NEW [{entry['Notice No.']}]: {title[:70]} | {relevance or '—'}")

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[ted] Done — {len(all_rows)} notices, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
