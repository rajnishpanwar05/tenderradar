# =============================================================================
# giz_pipeline.py — GIZ Tender Pipeline
#
# Source: https://ausschreibungen.giz.de  (Official GIZ procurement platform)
#         English version: https://ausschreibungen.giz.de/?lang=en
#
# Strategy:
#   1. Loads the GIZ procurement marketplace (Selenium headless Chrome)
#   2. Paginates through all pages of the tender table (#PROJECT_RESULT)
#   3. Filters rows mentioning India / indien / south asia
#   4. Scores relevance with shared keywords.py
#   5. Deduplicates via DB and saves to GIZ_India_Tenders_Master.xlsx
#
# No login required. No CAPTCHA.
# To submit bids: qn_quotation@giz.de
# =============================================================================

import os
import re
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Registry metadata (overrides static defaults) ────────────────────────────
SCRAPER_META = {
    "flag":        "giz",
    "label":       "GIZ India",
    "group":       "selenium",
    "timeout":     600,   # Selenium + pagination can be slow
    "max_retries": 1,
    "auto":        True,
}

# ── Config ─────────────────────────────────────────────────────────────────────
GIZ_BASE_URL   = "https://ausschreibungen.giz.de"
GIZ_EN_URL     = "https://ausschreibungen.giz.de/?lang=en"
GIZ_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "GIZ_India_Tenders_Master.xlsx")

# India / South Asia filter terms (German + English)
# Expanded: includes states, major cities, Indian orgs, and South Asian countries
INDIA_TERMS = [
    # Core
    "india", "indien", "indian",
    # South Asia broad
    "south asia", "südasien", "south asian", "saarc",
    # Cities
    "new delhi", "delhi", "mumbai", "bangalore", "bengaluru", "hyderabad",
    "chennai", "kolkata", "pune", "ahmedabad", "jaipur", "lucknow",
    "bhopal", "bhubaneswar", "guwahati", "chandigarh", "patna", "ranchi",
    # States / regions
    "rajasthan", "gujarat", "maharashtra", "karnataka", "kerala", "tamil",
    "andhra", "telangana", "odisha", "jharkhand", "assam", "meghalaya",
    "uttarakhand", "himachal", "punjab", "haryana", "uttar pradesh",
    "madhya pradesh", "chhattisgarh", "bihar", "west bengal",
    # Indian orgs / programmes
    "niti aayog", "ministry of", "government of india", "goi",
    "world bank india", "undp india", "sidbi", "nabard", "nhpc",
    # Neighbouring South Asia (GIZ SA programme covers these too)
    "nepal", "nepal ", "bangladesh", "sri lanka", "bhutan", "maldives",
    "pakistan",  # GIZ SA projects sometimes span the region
]

PAGE_LOAD_WAIT = 20    # seconds for page to render
ACTION_DELAY   = 2.0   # between page clicks (slightly faster)
MAX_PAGES      = 25    # cap; GIZ has ~20 active pages

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",           70),
    ("Ref No",          18),
    ("Type",            12),
    ("Deadline",        20),
    ("Published",       18),
    ("Organisation",    40),
    ("Relevance",       40),
    ("Detail Link",     55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Browser helpers
# =============================================================================

def _create_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)


def _wait_for_table(driver, timeout=PAGE_LOAD_WAIT):
    """Wait for the procurement table to appear on the page."""
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR,
                 "#PROJECT_RESULT, table.resultTable, table#resultTable, "
                 ".listTable, table[summary], #mainBox table")
            )
        )
        return True
    except Exception:
        return False


def _get_next_page(driver):
    """Click the next page arrow. Returns True if navigated."""
    for xpath in [
        "//a[contains(@href,'selectedTablePagePROJECT_RESULT')]"
        "[not(contains(@class,'disabled'))][last()]",
        "//a[@title='nächste Seite' or @title='next page' or @title='Next']",
        "//a[contains(text(),'›') or contains(text(),'>') or contains(text(),'Next')]",
        "//img[contains(@src,'next') or contains(@alt,'next')]/..",
        "//a[contains(@class,'nextPage') or contains(@class,'next')]",
    ]:
        try:
            btn = driver.find_element(By.XPATH, xpath)
            if btn.is_displayed() and btn.is_enabled():
                href = btn.get_attribute("href") or ""
                # Don't click if it's the last/current page link without next
                driver.execute_script("arguments[0].click();", btn)
                time.sleep(ACTION_DELAY)
                return True
        except Exception:
            continue
    return False


# =============================================================================
# Parsing
# =============================================================================

def _is_india_related(text: str) -> bool:
    t = text.lower()
    return any(term in t for term in INDIA_TERMS)


def _parse_giz_table(html: str):
    """
    Parse the GIZ ausschreibungen.giz.de tender table.
    Table ID: PROJECT_RESULT (or similar)
    Columns: Published | Deadline | Title | Type | Organisation | Action
    """
    soup    = BeautifulSoup(html, "html.parser")
    tenders = []

    # Try to find the main results table
    table = (
        soup.find("table", id="PROJECT_RESULT") or
        soup.find("table", id=re.compile(r"result|tender|project", re.I)) or
        soup.find("table", class_=re.compile(r"result|list|tender", re.I))
    )
    if table is None:
        # Fallback: try any table with recognisable column headers
        for t in soup.find_all("table"):
            headers_text = t.get_text(" ", strip=True).lower()
            if any(kw in headers_text for kw in
                   ["bezeichnung", "title", "deadline", "frist", "organisation", "typ"]):
                table = t
                break

    if table is None:
        print("[giz]   WARNING: Could not locate the GIZ tender table.", flush=True)
        return tenders

    rows = table.find_all("tr")
    if not rows:
        return tenders

    # Detect column positions from header row
    header_row  = rows[0]
    header_cells = [th.get_text(strip=True).lower()
                    for th in header_row.find_all(["th", "td"])]
    col = {}
    for i, h in enumerate(header_cells):
        # GIZ EN uses "short description"; GIZ DE uses "bezeichnung" / "betreff"
        if any(kw in h for kw in ["bezeichnung", "title", "betreff", "subject",
                                   "short description", "description", "kurzbeschreibung",
                                   "short", "tender name", "project"]):
            col.setdefault("title", i)
        if any(kw in h for kw in ["frist", "deadline", "closing", "abgabe",
                                   "submission", "end date", "closing date"]):
            col.setdefault("deadline", i)
        if any(kw in h for kw in ["veröffentlicht", "published", "date", "datum",
                                   "publication", "start date"]):
            col.setdefault("published", i)
        if any(kw in h for kw in ["typ", "type", "art", "category"]):
            col.setdefault("type", i)
        if any(kw in h for kw in ["stelle", "organisation", "auftraggeber", "contracting",
                                   "authority", "client", "buyer"]):
            col.setdefault("org", i)

    def cell_text(cells, key):
        idx = col.get(key)
        if idx is not None and idx < len(cells):
            return cells[idx].get_text(" ", strip=True)
        return ""

    print(f"[giz]   Table found: {len(rows)-1} rows, headers={header_cells}", flush=True)

    for row in rows[1:]:
        cells = row.find_all(["td", "th"])
        if len(cells) < 2:
            continue

        row_text = row.get_text(" ", strip=True)
        if len(row_text) < 5:
            continue

        # Get title (prefer detected column, fallback to longest cell)
        title = cell_text(cells, "title")
        if not title:
            title = max(cells, key=lambda c: len(c.get_text())).get_text(strip=True)

        if not title or len(title) < 3:
            continue

        # Extract detail link
        detail_link = ""
        for a in row.find_all("a", href=True):
            href = a["href"]
            if any(kw in href.lower() for kw in
                   ["projectforwarding", "detail", "project", "tender", "ausschreibung"]):
                detail_link = (href if href.startswith("http")
                               else GIZ_BASE_URL + "/" + href.lstrip("/"))
                break
        if not detail_link:
            # Take any link in the row
            a_tag = row.find("a", href=True)
            if a_tag:
                href = a_tag["href"]
                detail_link = href if href.startswith("http") else GIZ_BASE_URL + "/" + href.lstrip("/")

        # Extract ref number from title/row
        ref_m = re.search(r"\b([0-9]{6,10})\b", row_text)
        ref_no = ref_m.group(1) if ref_m else ""

        type_m = re.search(r"\b(RFQ|RFP|EOI|CFP|ITB|VgV|UVgO|TNW)\b", row_text, re.I)
        proc_type = type_m.group(1).upper() if type_m else ""

        tenders.append({
            "Title":        title,
            "Ref No":       ref_no,
            "Type":         proc_type,
            "Deadline":     cell_text(cells, "deadline"),
            "Published":    cell_text(cells, "published"),
            "Organisation": cell_text(cells, "org"),
            "Detail Link":  detail_link,
            "_row_text":    row_text,   # used for India filter, removed before saving
        })

    return tenders


# =============================================================================
# Excel writer
# =============================================================================

def _save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "GIZ Tenders"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes  = ws.cell(row=2, column=1)
    relevance_idx    = col_names.index("Relevance") + 1
    link_idx         = col_names.index("Detail Link") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill: cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill: cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(GIZ_EXCEL_PATH)
    print(f"[giz] Excel saved: {GIZ_EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[giz] GIZ Tender Pipeline starting...", flush=True)
    print(f"[giz] Source: {GIZ_BASE_URL}", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []

    # Clear old Excel
    if os.path.exists(GIZ_EXCEL_PATH):
        try:
            os.remove(GIZ_EXCEL_PATH)
        except Exception:
            pass

    driver = None
    try:
        driver = _create_driver()
    except Exception as exc:
        print(f"[giz] Browser init failed: {exc}", flush=True)
        print("[giz] Selenium unavailable — skipping this run without marking a failure.", flush=True)
        return new_tenders, all_rows
    raw    = []
    try:
        print("[giz] Loading GIZ procurement platform...", flush=True)
        driver.get(GIZ_EN_URL)
        time.sleep(3)

        # Try English language switcher if available
        try:
            en_link = driver.find_element(
                By.XPATH, "//a[contains(@href,'lang=en') or contains(text(),'EN') "
                          "or contains(text(),'English')]"
            )
            driver.execute_script("arguments[0].click();", en_link)
            time.sleep(2)
        except Exception:
            pass

        if not _wait_for_table(driver):
            print("[giz] WARNING: Timed out waiting for GIZ table.", flush=True)
            print(f"[giz]   Page title: {driver.title}", flush=True)

        page_num = 1
        seen_page_signatures: set = set()
        while page_num <= MAX_PAGES:
            print(f"[giz]   Page {page_num}...", flush=True)
            page_tenders = _parse_giz_table(driver.page_source)
            print(f"[giz]   {len(page_tenders)} tenders on page {page_num}", flush=True)

            # ── Loop detection: stop if this page is identical to a previous one ──
            # Use first 3 titles as a page fingerprint
            sig = "|".join(t.get("Title","")[:30] for t in page_tenders[:3])
            if sig and sig in seen_page_signatures:
                print(f"[giz]   Loop detected on page {page_num} — all tenders collected.", flush=True)
                break
            if sig:
                seen_page_signatures.add(sig)

            raw.extend(page_tenders)

            if not _get_next_page(driver):
                print("[giz] No more pages.", flush=True)
                break
            page_num += 1
            _wait_for_table(driver, timeout=10)

    finally:
        if driver:
            driver.quit()
            print("[giz] Browser closed.", flush=True)

    if not raw:
        print("[giz] No tenders extracted from GIZ platform.", flush=True)
        return new_tenders, all_rows

    # ── Dedup within this run (same ref_no or same title can appear on multiple pages) ──
    seen_keys: set = set()
    deduped: list = []
    for t in raw:
        ref   = t.get("Ref No", "").strip()
        title = t.get("Title", "").strip()
        # Skip rows where title looks like an org name (no digits, very short, or GIZ boilerplate)
        if not title or len(title) < 8:
            continue
        # Skip if title is clearly an org header row (all words capitalized, no numbers, >4 words of org jargon)
        if title.lower().startswith("deutsche gesellschaft") or title.lower() == "giz":
            continue
        key = ref if ref else title[:80]
        if key in seen_keys:
            continue
        seen_keys.add(key)
        deduped.append(t)

    print(f"[giz] After dedup: {len(raw)} raw → {len(deduped)} unique", flush=True)

    # ── Non-consulting junk filter ─────────────────────────────────────────────
    # GIZ portal has ~50 active tenders at any time. Most are consulting/advisory.
    # Remove obvious hardware / logistics / supply items; keep everything else.
    # Intelligence layer will score India/IDCG relevance per tender.
    _GIZ_JUNK = [
        "minibusse", "minibus", "mini bus", "bus für", "fahrzeuge", "vehicle",
        "hardware", "it-hardware", "laptop", "computer", "workstation",
        "generator", "generatoren", "sim ", "esim", "mobilfunk",
        "jahresabo", "subscription", "furniture", "möbel",
        "printing", "druckerzeugnisse", "stationery",
        "catering", "cleaning", "reinigung",
    ]

    def _is_junk(title: str) -> bool:
        t_lower = title.lower()
        return any(j in t_lower for j in _GIZ_JUNK)

    consulting_tenders = [t for t in deduped if not _is_junk(t.get("Title", ""))]
    junk_count = len(deduped) - len(consulting_tenders)

    # Separate India-specific vs Global consulting
    india_specific  = [t for t in consulting_tenders if _is_india_related(t["Title"] + " " + t["_row_text"])]
    global_advisory = [t for t in consulting_tenders if t not in india_specific]

    print(f"[giz] {len(deduped)} unique | {junk_count} junk removed | "
          f"{len(india_specific)} India-specific | {len(global_advisory)} global advisory", flush=True)

    # Strategy: India-specific always included; global advisory included only if
    # it has a consulting/advisory signal (GIZ portal is consulting-only so nearly all pass)
    _ADVISORY_SIGNALS = ["advisory", "consultancy", "consulting", "technical assistance",
                         "capacity", "policy", "assessment", "evaluation", "research",
                         "study", "expertise", "support", "strengthening", "development",
                         "training", "analysis", "monitoring", "audit", "framework"]

    def _has_advisory_signal(title: str) -> bool:
        t_lower = title.lower()
        return any(s in t_lower for s in _ADVISORY_SIGNALS)

    global_consulting = [t for t in global_advisory if _has_advisory_signal(t.get("Title", ""))]
    india_tenders = india_specific + global_consulting

    print(f"[giz] Final pool: {len(india_specific)} India + {len(global_consulting)} global consulting "
          f"= {len(india_tenders)} total for scoring", flush=True)

    if not india_tenders:
        print("[giz] No tenders after filter — using all consulting tenders.", flush=True)
        india_tenders = consulting_tenders

    # ── Score relevance + DB deduplication ────────────────────────────────────
    for t in india_tenders:
        row_text  = t.pop("_row_text", "")
        combined  = t.get("Title", "") + " " + t.get("Organisation", "") + " " + row_text
        relevance = score_relevance(combined)
        t["Relevance"] = relevance
        all_rows.append(t)

        ref_no    = t.get("Ref No", "").strip()
        tender_id = f"GIZ/{ref_no}" if ref_no else f"GIZ/{t['Title'][:60]}"
        detail_url = t.get("Detail Link", GIZ_BASE_URL)

        if check_if_new(tender_id):
            mark_as_seen(tender_id, title=t["Title"], source_site="GIZ", url=detail_url)
            # ── Intelligence-layer compatible format ──────────────────────────
            # Use title as description fallback — GIZ titles are descriptive
            # (e.g. "81319336-Renewable Energy Policy Advisory Services...")
            description = (
                t.get("Title", "") + ". " +
                t.get("Organisation", "") + ". " +
                t.get("Type", "") + " procurement. "
                "GIZ South Asia / India programme."
            )
            new_tenders.append({
                "title":       t["Title"],
                "description": description,
                "deadline":    t.get("Deadline", ""),
                "value":       "",
                "url":         detail_url,
                "source":      "GIZ",
                "org":         t.get("Organisation", ""),
                "type":        t.get("Type", ""),
                "ref_no":      ref_no,
                # Intelligence layer field aliases
                "Description": description,
                "Deadline":    t.get("Deadline", ""),
            })
            print(f"[giz]   → NEW: {t['Title'][:70]} | {relevance[:40] or '—'}", flush=True)

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(f"\n[giz] Done — {len(all_rows)} total, {len(new_tenders)} NEW, {relevant} relevant", flush=True)
    return new_tenders, all_rows
