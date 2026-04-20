# =============================================================================
# phfi_pipeline.py — Public Health Foundation of India (PHFI) Tenders Pipeline
#
# Site   : https://phfi.org/tenders/
# Method : requests + BeautifulSoup (server-rendered Drupal HTML)
# Login  : None required
# CAPTCHA: None
#
# Note   : PHFI posts tenders intermittently — page may show "no tenders"
#          when nothing is active. The scraper handles this gracefully.
#
# Status : LOW_FREQUENCY portal — returning 0 rows is NORMAL, not a bug.
#          Do NOT raise alerts on empty runs. No code fix required.
#
# What this does:
#   1. Fetches the tenders listing page
#   2. Finds all tender cards / list items with title and links
#   3. Visits each detail page for deadline + description
#   4. Scores relevance against IDCG expertise
#   5. Deduplicates via MySQL — only NEW tenders trigger notification
#   6. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
#
# PHFI = Public Health Foundation of India
# Focus: public health, epidemiology, nutrition, health policy, research
# Strong overlap with IDCG's health & social sectors expertise
# =============================================================================

import os, re, time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://phfi.org"
LISTING_URL = f"{BASE_URL}/tenders/"
DELAY       = 0.8

PHFI_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "PHFI_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer":         BASE_URL,
}

# Skip goods/equipment
SKIP_PATTERNS = [
    "supply of", "purchase of", "procurement of equipment", "rate contract",
    "annual maintenance", "amc", "printing of", "hardware",
]

def _is_goods(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in SKIP_PATTERNS)


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",        65),
    ("Deadline",     22),
    ("Description",  65),
    ("Relevance",    40),
    ("Detail Link",  55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Parse listing page
# =============================================================================

def _parse_listings(soup: BeautifulSoup) -> list:
    """
    PHFI uses Drupal — tenders appear as nodes/articles.
    Try multiple selectors to be robust to theme changes.
    """
    entries    = []
    seen_hrefs = set()

    # Check for "no tenders" message first
    page_text = soup.get_text().lower()
    if any(msg in page_text for msg in
           ["no tenders available", "no tender", "check back later", "nothing found"]):
        print("[phfi]   Page says: no tenders currently available.")
        return entries

    # Strategy 1: article elements
    cards = soup.find_all('article')

    # Strategy 2: view rows (Drupal views)
    if not cards:
        cards = soup.find_all('div', class_=re.compile(r'views-row|tender|node|item', re.I))

    # Strategy 3: any link pointing to /tenders/ sub-pages
    if not cards:
        tender_links = soup.find_all('a', href=re.compile(r'/tenders?/', re.I))
        for lnk in tender_links:
            href = lnk.get('href', '')
            if href and href not in seen_hrefs and href != LISTING_URL:
                cards.append(lnk.parent)

    for card in cards:
        # Find title + link
        title_tag = card.find(['h2', 'h3', 'h4', 'h5'])
        link_tag  = card.find('a', href=True)

        title = ""
        href  = ""
        if title_tag:
            title = title_tag.get_text(strip=True)
        if link_tag:
            href = link_tag.get('href', '')
            if not title:
                title = link_tag.get_text(strip=True)

        if not title or len(title) < 8:
            continue

        detail_url = href if href.startswith('http') else (BASE_URL + href if href else "")
        if not detail_url or detail_url in seen_hrefs:
            continue
        # Skip the listing page itself
        if detail_url.rstrip('/') == LISTING_URL.rstrip('/'):
            continue
        seen_hrefs.add(detail_url)

        # Deadline — look for date patterns in card text
        deadline  = ""
        card_text = card.get_text('\n', strip=True)
        for line in card_text.split('\n'):
            if re.search(r'deadline|last date|closing|submit', line, re.I):
                date_m = re.search(
                    r'(\d{1,2}[/-]\d{1,2}[/-]\d{4}|\d{1,2}\s+\w+\s+\d{4}|\w+\s+\d{1,2},?\s+\d{4})',
                    line
                )
                if date_m:
                    deadline = date_m.group(1)
                    break

        slug      = re.sub(r'[^a-zA-Z0-9]', '_', href.split('/')[-1] or title[:50])[:80]
        tender_id = f"PHFI_{slug}"

        entries.append({
            "Title":      title,
            "Deadline":   deadline,
            "detail_url": detail_url,
            "tender_id":  tender_id,
        })

    return entries


# =============================================================================
# SECTION 2 — Fetch detail page
# =============================================================================

def _fetch_detail(session: requests.Session, url: str) -> dict:
    result = {"Description": "", "Deadline": ""}
    if not url:
        return result
    try:
        r = session.get(url, headers=HEADERS, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')

        # Main content
        content = (
            soup.find('div', class_=re.compile(r'field-item|body|content|article', re.I)) or
            soup.find('article') or
            soup.find('main')
        )
        if content:
            result["Description"] = content.get_text(' ', strip=True)[:1800]

        # Better deadline extraction from detail page
        text = soup.get_text('\n', strip=True)
        for line in text.split('\n'):
            if re.search(r'deadline|last date|closing', line, re.I):
                date_m = re.search(
                    r'(\d{1,2}[/-]\d{1,2}[/-]\d{4}|\d{1,2}\s+\w+\s+\d{4}|\w+\s+\d{1,2},?\s+\d{4})',
                    line
                )
                if date_m:
                    result["Deadline"] = date_m.group(1)
                    break

    except Exception as e:
        print(f"[phfi]   Detail fetch error ({url[:60]}): {e}")
    return result


# =============================================================================
# SECTION 3 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PHFI Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")   + 1
    link_idx = col_names.index("Detail Link") + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 50
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(PHFI_EXCEL_PATH)
    print(f"[phfi] Excel saved: {PHFI_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# SECTION 4 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the PHFI Tenders pipeline.
    Returns (new_tenders, all_rows).
    Note: PHFI posts tenders intermittently. Returns empty lists when none active.
    """
    print("\n" + "=" * 65)
    print("[phfi] PHFI (Public Health Foundation of India) Tenders Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(PHFI_EXCEL_PATH):
        try:
            os.remove(PHFI_EXCEL_PATH)
            print("[phfi] Cleared old Excel")
        except Exception:
            pass

    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Fetch listing ──────────────────────────────────────────────────────────
    try:
        r = session.get(LISTING_URL, timeout=25)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, 'html.parser')
    except Exception as e:
        print(f"[phfi] ERROR fetching listing: {e}")
        return [], []

    entries = _parse_listings(soup)
    print(f"[phfi] Tenders found on listing: {len(entries)}")

    if not entries:
        print("[phfi] No active tenders at this time — will check again next run.")
        return [], []

    # ── Process each entry ─────────────────────────────────────────────────────
    for i, entry in enumerate(entries, 1):
        title = entry["Title"]

        if _is_goods(title):
            print(f"[phfi]   [{i:>2}/{len(entries)}] SKIP (goods): {title[:60]}")
            continue

        print(f"[phfi]   [{i:>2}/{len(entries)}] {title[:65]}")

        detail    = _fetch_detail(session, entry["detail_url"])
        deadline  = detail.get("Deadline") or entry["Deadline"]
        relevance = score_relevance(title, detail.get("Description", ""))

        row = {
            "Title":       title,
            "Deadline":    deadline,
            "Description": detail.get("Description", "")[:1500],
            "Relevance":   relevance,
            "Detail Link": entry["detail_url"],
        }
        all_rows.append(row)

        tid = entry["tender_id"]
        if check_if_new(tid):
            mark_as_seen(tid, title=title, source_site="PHFI", url=entry["detail_url"])
            new_tenders.append({
                "title":    title,
                "deadline": deadline,
                "value":    "See PHFI",
                "url":      entry["detail_url"],
            })
            print(f"           → NEW | Relevance: {relevance or '—'}")
        else:
            print(f"           → seen | Relevance: {relevance or '—'}")

        time.sleep(DELAY)

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[phfi] Done — {len(all_rows)} tenders, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
