# =============================================================================
# karnataka_scraper.py — Karnataka KPPP Portal (kppp.karnataka.gov.in)
#
# Site   : https://kppp.karnataka.gov.in
# Method : Direct REST API — pure requests + JSON (NO Selenium, NO BeautifulSoup)
#
# Background:
#   Karnataka ran a legacy JBoss Seam + JSF portal (eproc.karnataka.gov.in).
#   In June 2024, all procurement migrated to the new KPPP (Karnataka Public
#   Procurement Portal) — an Angular SPA backed by Spring Boot microservices.
#   The old portal is now empty ("No Tenders Found" on every search).
#
# API Discovery:
#   Reverse-engineered from KPPP's Angular main.js bundle.
#   The Angular app calls internal REST APIs at:
#     https://kppp.karnataka.gov.in/supplier-registration-service/v1/api/
#   Endpoints return JSON (not HTML). No CAPTCHA, no auth required for search.
#
# Endpoints:
#   POST /portal-service/search-eproc-tenders            → GOODS
#   POST /portal-service/works/search-eproc-tenders      → WORKS
#   POST /portal-service/services/search-eproc-tenders   → SERVICES
#   Query params: page=N&size=50&order-by-tender-publish=true
#   X-Total-Count response header gives total count for pagination.
#
# Total tenders: ~6,000+ (GOODS: ~730, WORKS: ~3,800, SERVICES: ~1,500)
# =============================================================================

import os
import re
import time
import math
from urllib.parse import urljoin

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_API    = "https://kppp.karnataka.gov.in/supplier-registration-service/v1/api"
PORTAL_URL  = "https://kppp.karnataka.gov.in/portal/searchTender/live"

EXCEL_PATH  = os.path.join(PORTAL_EXCELS_DIR, "Karnataka_Tenders_Master.xlsx")

PAGE_SIZE      = 50     # records per API page
REQUEST_DELAY  = 0.5    # seconds between pagination requests
TIMEOUT        = 25     # per-request timeout
RETRY_ATTEMPTS = 2
MAX_DOC_ENRICH_PER_CATEGORY = 350

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept":          "application/json, text/plain, */*",
    "Content-Type":    "application/json",
    "Referer":         PORTAL_URL,
    "Origin":          "https://kppp.karnataka.gov.in",
    "Accept-Language": "en-US,en;q=0.9",
}

# Categories to scrape: (label, endpoint_suffix, request_body_category)
CATEGORIES = [
    ("WORKS",    "/portal-service/works/search-eproc-tenders",    "WORKS"),
    ("SERVICES", "/portal-service/services/search-eproc-tenders", "SERVICES"),
    ("GOODS",    "/portal-service/search-eproc-tenders",          "GOODS"),
]

_DOC_EXT_RE = re.compile(r"\.(pdf|docx?|xlsx?|xlsm|zip|rar|csv)(?:$|[?#])", re.I)
_URL_KEY_HINTS = ("url", "href", "download", "link", "path", "document")
_NAME_KEY_HINTS = ("name", "filename", "file_name", "documentname", "docname", "title")

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Tender ID",       28),
    ("Title",           65),
    ("Organisation",    40),
    ("Category",        14),
    ("Est. Value (₹)",  18),
    ("Published Date",  20),
    ("Closing Date",    20),
    ("Relevance",       42),
    ("URL",             55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# HTTP helpers
# =============================================================================

def _make_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def _post_page(session, endpoint_suffix, category, page_num):
    """
    POST one page of tender results for the given category.
    Returns (records: list, total_count: int) or ([], 0) on failure.
    """
    url = (f"{BASE_API}{endpoint_suffix}"
           f"?page={page_num}&size={PAGE_SIZE}&order-by-tender-publish=true")
    payload = {
        "category":   category,
        "status":     "PUBLISHED",
        "tenderType": "OPEN",
    }

    for attempt in range(RETRY_ATTEMPTS):
        try:
            r = session.post(url, json=payload, timeout=TIMEOUT)
            if r.status_code == 200:
                total = int(r.headers.get("X-Total-Count", 0) or 0)
                return r.json(), total
            else:
                print(f"[karnataka] {category} page {page_num}: HTTP {r.status_code}", flush=True)
                return [], 0
        except Exception as e:
            if attempt < RETRY_ATTEMPTS - 1:
                time.sleep(2)
            else:
                print(f"[karnataka] {category} page {page_num}: {e}", flush=True)
    return [], 0


# =============================================================================
# Data extraction
# =============================================================================

def _record_to_row(rec, category_label):
    """
    Convert one API JSON record to our standard row dict.
    """
    ecv = rec.get("ecv") or 0
    ecv_str = f"₹{ecv:,.0f}" if ecv else ""

    pub_date = rec.get("publishedDate", "") or ""
    clo_date = rec.get("tenderClosureDate", "") or ""
    # Dates come as "DD-MM-YYYY HH:MM:SS" — trim time for display
    if " " in pub_date:
        pub_date = pub_date.split(" ")[0]
    if " " in clo_date:
        clo_date = clo_date.split(" ")[0]

    dept  = rec.get("deptName", "") or ""
    loc   = rec.get("locationName", "") or ""
    org   = f"{dept} — {loc}" if loc and loc != dept else dept
    title = rec.get("title", "") or rec.get("description", "")
    tender_id = rec.get("tenderNumber", "") or f"KA/{rec.get('id','')}"
    doc_urls, doc_names = _extract_documents_from_record(rec)
    summary_parts = [part for part in [
        rec.get("description", "") or "",
        f"Category: {category_label}" if category_label else "",
        f"Estimated value: {ecv_str}" if ecv_str else "",
        f"Department: {dept}" if dept else "",
        f"Location: {loc}" if loc else "",
        f"Published: {pub_date}" if pub_date else "",
        f"Closing: {clo_date}" if clo_date else "",
        f"Tender ID: {tender_id}" if tender_id else "",
        f"NIT ID: {rec.get('nitId', '')}" if rec.get("nitId") else "",
        f"Documents: {', '.join(doc_names[:8])}" if doc_names else "",
        f"Document links: {' | '.join(doc_urls[:10])}" if doc_urls else "",
    ] if part]

    row = {
        "Tender ID":       tender_id,
        "tender_id":       tender_id,
        "Title":           title,
        "Organisation":    org,
        "organization":    org,
        "Category":        category_label,
        "Est. Value (₹)":  ecv_str,
        "Value":           ecv_str,
        "Published Date":  pub_date,
        "Closing Date":    clo_date,
        "Description":     " | ".join(summary_parts)[:10000],
        "URL":             PORTAL_URL,
        "source_portal":   "karnataka",
        "_nit_id":         rec.get("nitId", ""),
        "_raw_id":         rec.get("id", ""),
    }
    if doc_names:
        row["Document Names"] = ", ".join(doc_names[:15])
    if doc_urls:
        row["Document URLs"] = " | ".join(doc_urls[:20])
    return row


def _extract_documents_from_record(rec):
    """
    Recursively extract likely document links/names from KPPP record payload.
    This keeps deep enrichment grounded even when docs are nested in API JSON.
    """
    urls = []
    names = []
    seen_urls = set()
    seen_names = set()

    def _maybe_add_url(val):
        if not isinstance(val, str):
            return
        raw = val.strip()
        if not raw:
            return
        low = raw.lower()
        if not (_DOC_EXT_RE.search(low) or any(h in low for h in ("download", "tenderdoc", "document", "attachment"))):
            return
        full = raw if low.startswith("http") else urljoin("https://kppp.karnataka.gov.in", raw)
        if full in seen_urls:
            return
        seen_urls.add(full)
        urls.append(full)

    def _maybe_add_name(val):
        if not isinstance(val, str):
            return
        raw = re.sub(r"\s+", " ", val).strip()
        if not raw:
            return
        if not _DOC_EXT_RE.search(raw.lower()):
            return
        key = raw.lower()
        if key in seen_names:
            return
        seen_names.add(key)
        names.append(raw)

    def _walk(node):
        if isinstance(node, dict):
            for k, v in node.items():
                k_low = str(k or "").lower()
                if isinstance(v, str):
                    if any(h in k_low for h in _URL_KEY_HINTS):
                        _maybe_add_url(v)
                    if any(h in k_low for h in _NAME_KEY_HINTS):
                        _maybe_add_name(v)
                    _maybe_add_url(v)
                    _maybe_add_name(v)
                else:
                    _walk(v)
        elif isinstance(node, list):
            for item in node:
                _walk(item)

    pre_urls = rec.get("_document_urls") if isinstance(rec, dict) else None
    pre_names = rec.get("_document_names") if isinstance(rec, dict) else None
    if isinstance(pre_urls, list):
        for u in pre_urls:
            _maybe_add_url(u)
    if isinstance(pre_names, list):
        for n in pre_names:
            _maybe_add_name(n)

    _walk(rec or {})
    return urls, names


def _extract_list_payload(payload):
    """Best-effort extraction of list payloads from varied API response shapes."""
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for k in ("data", "content", "documents", "items", "records", "response"):
            v = payload.get(k)
            if isinstance(v, list):
                return v
            if isinstance(v, dict):
                for kk in ("data", "content", "documents", "items", "records"):
                    vv = v.get(kk)
                    if isinstance(vv, list):
                        return vv
    return []


def _fetch_tender_documents(session, category_label, raw_id):
    """
    Fetch Karnataka tender document metadata from portal-service APIs.
    Returns (doc_urls, doc_names).
    """
    cat = str(category_label or "").strip().upper()
    doc_cfg = {
        "WORKS": ("get-works-tender-files", "works-tender-file"),
        "SERVICES": ("get-services-tender-files", "services-tender-file"),
        "GOODS": ("get-goods-tender-files", "goods-tender-file"),
    }.get(cat)
    if not doc_cfg or not raw_id:
        return [], []

    list_ep, dl_ep = doc_cfg
    list_url = f"{BASE_API}/portal-service/{raw_id}/{list_ep}"
    doc_urls = []
    doc_names = []
    seen_u = set()
    seen_n = set()

    try:
        r = session.get(list_url, timeout=15)
        if r.status_code != 200:
            return [], []
        docs = _extract_list_payload(r.json())
        for d in docs:
            if not isinstance(d, dict):
                continue
            file_id = (
                d.get("id") or d.get("fileId") or d.get("file_id")
                or d.get("documentId") or d.get("docId")
            )
            name = (
                d.get("fileName") or d.get("filename") or d.get("documentName")
                or d.get("name") or d.get("title") or d.get("docName") or ""
            )
            href = d.get("url") or d.get("href") or d.get("downloadUrl") or d.get("downloadURL") or ""

            if isinstance(name, str):
                n = re.sub(r"\s+", " ", name).strip()
                if n and n.lower() not in seen_n:
                    seen_n.add(n.lower())
                    doc_names.append(n)

            if isinstance(href, str) and href.strip():
                u = href.strip()
                u = u if u.lower().startswith("http") else urljoin("https://kppp.karnataka.gov.in", u)
                if u not in seen_u:
                    seen_u.add(u)
                    doc_urls.append(u)

            if file_id:
                u = f"{BASE_API}/portal-service/{raw_id}/{dl_ep}/{file_id}/download-file"
                if u not in seen_u:
                    seen_u.add(u)
                    doc_urls.append(u)

    except Exception:
        return doc_urls, doc_names

    return doc_urls, doc_names


# =============================================================================
# Excel output
# =============================================================================

def _save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Karnataka Tenders"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    relevance_idx = col_names.index("Relevance") + 1
    link_idx      = col_names.index("URL") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill:
                        cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill:
                    cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(EXCEL_PATH)
    print(f"[karnataka] Excel saved: {EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[karnataka] Karnataka KPPP (kppp.karnataka.gov.in) Pipeline", flush=True)
    print("[karnataka] Method: REST API — no Selenium, no HTML parsing", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []
    seen_ids    = set()

    if os.path.exists(EXCEL_PATH):
        try:
            os.remove(EXCEL_PATH)
        except Exception:
            pass

    # ── Pre-flight: check portal is reachable ─────────────────────────────────
    print("[karnataka] Pre-flight check...", flush=True)
    try:
        probe = requests.head(
            "https://kppp.karnataka.gov.in",
            timeout=8,
            headers={"User-Agent": HEADERS["User-Agent"]},
        )
        if probe.status_code >= 500:
            print(f"[karnataka] SKIPPED — portal returned HTTP {probe.status_code}", flush=True)
            return new_tenders, all_rows
    except Exception as e:
        print(f"[karnataka] SKIPPED — portal unreachable: {e}", flush=True)
        return new_tenders, all_rows

    session = _make_session()

    # ── Scrape each category ──────────────────────────────────────────────────
    for cat_label, endpoint, cat_code in CATEGORIES:
        print(f"\n[karnataka] ── {cat_label} ──────────────────────────────", flush=True)
        docs_enriched = 0

        # Page 0 to get total count
        records, total = _post_page(session, endpoint, cat_code, page_num=0)
        if not records:
            print(f"[karnataka] {cat_label}: no records returned — skipping.", flush=True)
            continue

        total_pages = math.ceil(total / PAGE_SIZE) if total else 1
        print(f"[karnataka] {cat_label}: {total} tenders across {total_pages} pages", flush=True)

        # Process page 0 results first
        page_data = records

        for page_num in range(total_pages):
            if page_num > 0:
                time.sleep(REQUEST_DELAY)
                page_data, _ = _post_page(session, endpoint, cat_code, page_num)
                if not page_data:
                    print(f"[karnataka] {cat_label} page {page_num}: empty — stopping.", flush=True)
                    break

            print(f"[karnataka]   page {page_num + 1}/{total_pages}: {len(page_data)} records", flush=True)

            for rec in page_data:
                if docs_enriched < MAX_DOC_ENRICH_PER_CATEGORY:
                    raw_id = rec.get("id")
                    doc_urls, doc_names = _fetch_tender_documents(session, cat_label, raw_id)
                    if doc_urls:
                        rec["_document_urls"] = doc_urls
                    if doc_names:
                        rec["_document_names"] = doc_names
                    if doc_urls or doc_names:
                        docs_enriched += 1
                row = _record_to_row(rec, cat_label)
                raw_tid = row["Tender ID"]
                db_key = f"KA_{re.sub(r'[^A-Za-z0-9]', '_', raw_tid)}"
                row["Tender ID"] = db_key
                row["tender_id"] = db_key
                tid = row["Tender ID"]

                if tid in seen_ids:
                    continue
                seen_ids.add(tid)

                combined      = row["Title"] + " " + row["Organisation"]
                row["Relevance"] = score_relevance(combined)

                # Drop internal keys before saving to Excel
                excel_row = {k: v for k, v in row.items() if not k.startswith("_")}
                all_rows.append(excel_row)

                if check_if_new(db_key):
                    mark_as_seen(db_key, title=row["Title"][:255],
                                 source_site="Karnataka", url=PORTAL_URL)
                    if row.get("Relevance"):
                        new_tenders.append({
                            "tender_id": db_key,
                            "title":    row["Title"],
                            "description": row.get("Description", ""),
                            "organization": row.get("Organisation", ""),
                            "deadline": row.get("Closing Date", ""),
                            "value":    row.get("Value", ""),
                            "url":      PORTAL_URL,
                            "source_portal": "karnataka",
                        })
                        print(f"[karnataka]   → NEW + RELEVANT | {row['Relevance'][:55]}", flush=True)
                    else:
                        print(f"[karnataka]   → NEW (Excel only)", flush=True)

        cat_relevant = sum(1 for r in all_rows
                          if r.get("Category") == cat_label and r.get("Relevance"))
        print(f"[karnataka] {cat_label}: {cat_relevant} relevant | docs enriched: {docs_enriched}", flush=True)

    # ── Save Excel ─────────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(
        f"\n[karnataka] Done — {len(all_rows)} tenders scraped, "
        f"{len(new_tenders)} NEW+RELEVANT, {relevant} relevant total",
        flush=True,
    )
    return new_tenders, all_rows
