# =============================================================================
# dtvp_pipeline.py — Deutsches Vergabeportal (DTVP) — German Procurement Portal
#
# Source   : https://www.dtvp.de/Center/common/project/search.do
# English  : https://en.dtvp.de
# Method   : Selenium headless Chrome (results loaded via JavaScript/AJAX)
#
# Why included:
#   DTVP lists European procurement including international development
#   contracts placed by German agencies (GIZ, KfW, BMZ) and EU institutions.
#   The English-accessible backend can be searched for India/development terms.
#
# Fix (2026-03-A): Switched from requests POST to Selenium (AJAX issue).
# Fix (2026-03-B): Replaced manual form-fill with ?searchString= URL param.
# Fix (2026-03-C): German date parsing; broadened search queries; type field.
#
#   How DTVP works:
#   1. Page initialises TenderSearch("/Center/api/v2/project/search", form,
#      submitFormOnLoad=true, …)
#   2. On DOM-ready the page reads getUrlParameter('searchString') and sets
#      [name='searchText'] BEFORE TenderSearch auto-submits.
#   3. TenderSearch POSTs to /Center/api/v2/project/search with a session JWT
#      (injected from the browser session — cannot be replicated via requests).
#   4. On success Handlebars renders results into #listTemplate.
#
#   Scraper strategy: load the search URL with &searchString=<query> appended,
#   then wait for #listTemplate to be populated by the auto-submit.  No manual
#   form interaction needed; the JWT is handled by the browser session.
#
#   Column order after Handlebars render:
#   td[0] = publishingDate   (Published — <abbr title="DD.MM.YYYY">)
#   td[1] = relevantDate     (Deadline  — <abbr title="DD.MM.YYYY"> or German phrase)
#   td[2] = title            (Kurzbezeichnung — no link)
#   td[3] = type             (contractingRule + publicationType)
#   td[4] = organisationName (Vergabeplattform / Veröffentlicher)
#   td[5] = Aktion           (links.ENTER_PROJECTROOM + earmark buttons)
#                             earmark <a> carries data-earmarked-id="{{projectId}}"
#
# Note: Public tender listings only (no login required for search results).
# =============================================================================

import hashlib
import os
import re
import time
from urllib.parse import quote

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import sys
sys.path.insert(0, os.path.expanduser("~/tender_system"))

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

# ── Constants ─────────────────────────────────────────────────────────────────
DTVP_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "DTVP_Tenders_Master.xlsx")
SEARCH_URL      = ("https://www.dtvp.de/Center/common/project/"
                   "search.do?method=showExtendedSearch&fromExternal=true")
BASE_URL        = "https://www.dtvp.de"

# ── Search queries ─────────────────────────────────────────────────────────────
# Targeting international development / consulting tenders from German agencies
# (GIZ, KfW, BMZ) and EU institutions that publish on DTVP.
# Kept to English since DTVP search works across English-language tender titles.
SEARCH_QUERIES = [
    # India-specific
    "India consulting",
    "India technical assistance",
    "India advisory",
    "India monitoring evaluation",
    "GIZ India",
    "KfW India",
    "BMZ India",
    # Broader development agency terms
    "South Asia consulting",
    "Asia development advisory",
    "capacity building consulting",
    "development evaluation",
    "technical assistance consulting",
    # German agency catch-all
    "GIZ consulting",
    "KfW consulting",
]

# ── German date parsing ────────────────────────────────────────────────────────
_MONTH_MAP = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}

# German phrases that mean "no deadline / not applicable"
_NO_DEADLINE_DE = [
    "nicht vorhanden",   # not available
    "keine frist",       # no deadline
    "entfällt",          # not applicable
    "ohne frist",        # without deadline
    "kein datum",        # no date
    "n.a.",
    "n/a",
    "keine angabe",      # no information
    "auf anfrage",       # on request
]


def _parse_german_date(raw: str) -> str:
    """
    Convert DTVP date strings to clean English format.

    Input variants seen on DTVP:
      • "31.03.2026"                                          → "31 Mar 2026"
      • "31.03.2026 um 12:00 Uhr"                            → "31 Mar 2026  12:00"
      • "31.03.2026 12:00"                                   → "31 Mar 2026  12:00"
      • "Die Frist ist bei diesem Verfahren nicht vorhanden
         (z.B. EU-Vorinformation)"                           → "No deadline set"
      • ""  / "—"  / whitespace                              → "—"
    """
    if not raw:
        return "—"

    s = raw.strip()
    if not s or s in ("—", "-", "N/A", "n/a"):
        return "—"

    s_lower = s.lower()

    # German "no deadline" phrases
    if any(phrase in s_lower for phrase in _NO_DEADLINE_DE):
        return "No deadline set"

    # Extract DD.MM.YYYY and optional HH:MM time
    m = re.search(r'(\d{1,2})\.(\d{2})\.(\d{4})(?:.*?(\d{2}:\d{2}))?', s)
    if m:
        day   = int(m.group(1))
        month = m.group(2)
        year  = m.group(3)
        time_ = m.group(4)          # may be None
        mon   = _MONTH_MAP.get(month, month)
        if time_:
            return f"{day} {mon} {year}  {time_}"
        return f"{day} {mon} {year}"

    # Couldn't parse — return as-is so we at least have something
    return s


# ── Excel styles ───────────────────────────────────────────────────────────────
DARK_BLUE  = PatternFill("solid", fgColor="1F3864")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
WHITE      = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT  = Font(name="Calibri", size=10)
THIN       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Ref / ID",          22),
    ("Title",             60),
    ("Contracting Auth.", 35),
    ("Type",              28),
    ("Deadline",          18),
    ("Published",         14),
    ("Link",              50),
    ("Relevance",         30),
]


def _init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "DTVP Tenders"
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        c = ws.cell(1, col_idx, col_name)
        c.fill      = DARK_BLUE
        c.font      = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = col_width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLUMNS)).column_letter}1"
    return wb, ws


def _save_excel(rows: list):
    if os.path.exists(DTVP_EXCEL_PATH):
        wb = load_workbook(DTVP_EXCEL_PATH)
        ws = wb.active
    else:
        wb, ws = _init_excel()

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.add(str(row[0]))

    for r in rows:
        key = r.get("ref_id", "")
        if key in existing:
            continue
        row_idx = ws.max_row + 1
        fill    = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
        ws.append([
            r["ref_id"], r["title"], r["authority"], r.get("type", ""),
            r["deadline"], r["published"], r["link"], r["relevance"],
        ])
        for col_i in range(1, len(COLUMNS) + 1):
            c = ws.cell(row_idx, col_i)
            c.font      = CELL_FONT
            c.fill      = fill
            c.border    = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        existing.add(key)

    wb.save(DTVP_EXCEL_PATH)


def _parse_results(html: str) -> list:
    """
    Parse DTVP results from Selenium-rendered HTML.

    After Handlebars renders the 'entry-template' into #listTemplate the table
    columns are:
      td[0] publishingDate   — <abbr title="DD.MM.YYYY">DD.MM.</abbr>
      td[1] relevantDate     — deadline, same abbr pattern (or long German phrase)
      td[2] title            — Kurzbezeichnung (plain text, no link)
      td[3] type             — contractingRule + publicationType
      td[4] organisationName — Vergabeplattform / Veröffentlicher
      td[5] Aktion           — project-room link (target=_blank) + earmark buttons
                               earmark <a> carries data-earmarked-id="{{projectId}}"
    """
    soup  = BeautifulSoup(html, "html.parser")
    rows  = []

    # Results live inside #listTemplate > table (rendered by Handlebars)
    table = soup.select_one("#listTemplate table")
    if not table:
        lt = soup.select_one("#listTemplate")
        if lt:
            msg = lt.get_text(strip=True)
            if msg:
                print(f"  DTVP: #listTemplate text: {msg[:120]}")
        return rows

    for tr in table.select("tbody tr"):
        tds = tr.find_all("td")
        if len(tds) < 3:
            continue

        # Published — prefer the long date from <abbr title="…">
        pub_abbr   = tds[0].find("abbr")
        pub_raw    = (pub_abbr.get("title") or pub_abbr.get_text(strip=True)) \
                     if pub_abbr else tds[0].get_text(strip=True)
        published  = _parse_german_date(pub_raw)

        # Deadline — same pattern; abbr title may hold full German phrase
        dl_abbr    = tds[1].find("abbr")
        dl_raw     = (dl_abbr.get("title") or dl_abbr.get_text(strip=True)) \
                     if dl_abbr else tds[1].get_text(strip=True)
        deadline   = _parse_german_date(dl_raw)

        # Title
        title = tds[2].get_text(strip=True)
        if not title or len(title) < 4:
            continue

        # Contract type (td[3])
        tender_type = tds[3].get_text(separator=" ", strip=True) if len(tds) > 3 else ""

        # Authority / publisher (td[4])
        authority = tds[4].get_text(strip=True) if len(tds) > 4 else ""

        # Link — project-room link preferred; else build from projectId earmark
        link = ""
        if len(tds) > 5:
            pr_link = tds[5].find("a", target="_blank")
            if pr_link:
                href = pr_link.get("href", "")
                link = href if href.startswith("http") else BASE_URL + href

            if not link:
                ear = tds[5].find("a", attrs={"data-earmarked-id": True})
                if ear:
                    pid = ear.get("data-earmarked-id", "").strip()
                    if pid:
                        link = (f"{BASE_URL}/Center/public/project/"
                                f"{pid}/publications.html")

        rows.append({
            "ref_id"   : "",   # not shown in DTVP search results — built later
            "title"    : title,
            "authority": authority,
            "type"     : tender_type,
            "deadline" : deadline,
            "published": published,
            "link"     : link or SEARCH_URL,
        })

    return rows


def _make_driver():
    """Headless Chrome driver for JavaScript-rendered DTVP results."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


def _search_with_selenium(driver, query: str) -> list:
    """
    Load DTVP with ?searchString=<query> so TenderSearch auto-submits our query.

    How this works:
      1. The page's $(function(){}) handler reads getUrlParameter('searchString')
         and sets [name='searchText'] BEFORE TenderSearch is constructed.
      2. TenderSearch is constructed with submitFormOnLoad=true, so it immediately
         calls serializeForm() → reads searchText → POST to /Center/api/v2/project/search
         with the browser's session JWT (X-JWT header).
      3. On success Handlebars renders the results into #listTemplate.

    We just need to wait for #listTemplate to be non-empty, then parse.
    No manual form interaction — the JWT token is handled automatically.
    """
    try:
        url = f"{SEARCH_URL}&searchString={quote(query)}"
        driver.get(url)

        # Wait for Handlebars to populate #listTemplate
        # (fires after AJAX response from /Center/api/v2/project/search)
        wait = WebDriverWait(driver, 30)
        try:
            wait.until(lambda d: len(d.find_elements(
                By.CSS_SELECTOR, "#listTemplate *")) > 0)
        except TimeoutException:
            print(f"  DTVP: timeout waiting for #listTemplate for '{query}'")

        time.sleep(1.0)   # let Handlebars finish full render
        html    = driver.page_source
        results = _parse_results(html)
        print(f"  DTVP: '{query}' → {len(results)} rows")
        return results

    except Exception as e:
        print(f"  DTVP: Selenium error for '{query}': {e}")
        return []


def _make_ref_id(title: str, authority: str) -> str:
    """
    Generate a stable, unique ref_id from title + authority.

    DTVP search results don't expose a project ID in the HTML.
    Using title[:40] alone was fragile (collisions on similar titles).
    This adds an 8-char hash of the full title+authority to guarantee uniqueness.
    """
    slug  = re.sub(r"[^A-Za-z0-9]", "_", title)[:36].strip("_")
    h8    = hashlib.md5(f"{title}|{authority}".encode()).hexdigest()[:8]
    return f"{slug}_{h8}"


def run():
    print("\n" + "=" * 65)
    print("[dtvp] DTVP Germany Pipeline starting (Selenium mode)...")
    print("=" * 65)

    seen_ids    = set()
    all_rows    = []
    new_tenders = []
    driver      = None

    try:
        driver = _make_driver()

        for query in SEARCH_QUERIES:
            print(f"  DTVP querying: '{query}' ...")
            parsed = _search_with_selenium(driver, query)

            for r in parsed:
                ref_id = _make_ref_id(r["title"], r["authority"])
                tid    = f"DTVP_{ref_id}"

                if tid in seen_ids:
                    continue
                seen_ids.add(tid)

                relevance = score_relevance(r["title"])
                if not relevance and not title_is_relevant(r["title"]):
                    continue

                r["relevance"] = relevance
                r["ref_id"]    = ref_id
                r["tender_id"] = tid
                all_rows.append(r)

                if check_if_new(tid):
                    mark_as_seen(tid, r["title"], "DTVP", r.get("link") or SEARCH_URL)
                    new_tenders.append(r)

            time.sleep(1.5)

    except Exception as e:
        print(f"[dtvp] FATAL: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    _save_excel(all_rows)
    print(f"DTVP done — {len(all_rows)} relevant, {len(new_tenders)} NEW")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, all_ = run()
    print(f"\n✅  DTVP: {len(new)} new / {len(all_)} total")
