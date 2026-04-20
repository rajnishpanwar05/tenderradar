# =============================================================================
# scrapers/portals/devbusiness_scraper.py — UN Development Business scraper
#
# ⚠️  PORTAL PERMANENTLY SHUT DOWN — 31 March 2025
#
# Site   : https://devbusiness.un.org  (shows shutdown notice only)
# Status : UN Development Business (UNDB) ceased all operations on 31-Mar-2025
#          after 46 years of service. The website now only shows a shutdown page.
#
# History:
#   UNDB aggregated procurement from World Bank, UNDP, UNOPS, GIZ, ADB, AfDB,
#   KfW, JICA, and 40+ donors. It is now permanently closed.
#
# Replacement portals (listed on the shutdown page):
#   - World Bank:    https://projects.worldbank.org/en/projects-operations/procurement
#   - ADB:           https://www.adb.org/projects/tenders
#   - IDB:           https://www.iadb.org/en/projects
#   - AfDB:          https://www.afdb.org/en/projects-and-operations/procurement
#   - UN Global:     https://ungm.org/
#
#   All of the above are already integrated in TenderRadar.
#
# This scraper returns [], [] immediately (no-op) and logs a clear warning.
# Set auto=False to suppress it from the registry run list.
#
# Run:
#   python3 main.py --devbusiness
# =============================================================================

from __future__ import annotations

import os
import re
import time
import json
import logging
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import DEVBUSINESS_USER, DEVBUSINESS_PASS, PORTAL_EXCELS_DIR
from database.db  import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# SCRAPER_META — required for auto-discovery by core/registry.py
SCRAPER_META = {
    "flag":        "devbusiness",
    "label":       "Dev Business UN (SHUT DOWN)",
    "group":       "requests",
    "timeout":     10,
    "max_retries": 0,
    "auto":        False,   # excluded — portal permanently closed 31-Mar-2025
}

logger = logging.getLogger("tenderradar.devbusiness")

# =============================================================================
# Configuration
# =============================================================================

BASE_URL     = "https://devbusiness.un.org"
LOGIN_URL    = f"{BASE_URL}/user/login"
OPPS_URL     = f"{BASE_URL}/opportunities"
OPPS_JSON    = f"{BASE_URL}/api/v1/opportunities"

EXCEL_PATH   = os.path.join(PORTAL_EXCELS_DIR, "DevBusiness_UN_Tenders_Master.xlsx")
PAGE_SIZE    = 20
MAX_PAGES    = 25
DELAY        = 1.2
DETAIL_DELAY = 0.8

HEADERS = {
    "User-Agent":      ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/120.0.0.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept":          "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer":         BASE_URL,
}

_SKIP_TYPES = {
    "goods", "supply", "equipment", "works", "construction", "hardware",
    "vehicle", "furniture", "printing", "ict", "it services", "software license",
    "annual maintenance", "rate contract", "empanelment",
}

_CONSULT_TYPES = {
    "consulting", "advisory", "consultancy", "technical assistance",
    "evaluation", "research", "capacity building", "training",
    "assessment", "study", "survey", "review",
}

# =============================================================================
# Excel styling
# =============================================================================

_MASTER_COLUMNS = [
    ("Title",        70),
    ("Organization", 35),
    ("Country",      22),
    ("Deadline",     18),
    ("Value (USD)",  18),
    ("Notice Type",  22),
    ("Description",  70),
    ("Relevance",    45),
    ("URL",          60),
]
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_ALT_FILL    = PatternFill("solid", fgColor="F5F8FF")
_REL_FILL    = PatternFill("solid", fgColor="E2EFDA")
_WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
_BODY_FONT   = Font(name="Calibri", size=10)
_REL_FONT    = Font(name="Calibri", size=10, color="375623", bold=True)
_NOREL_FONT  = Font(name="Calibri", size=10, color="999999")
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin",  color="D0D7E3"), bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Selenium WAF bypass + Drupal login
# =============================================================================

def _make_driver():
    """Headless Chrome driver — executes AWS WAF JS challenge transparently."""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except Exception:
        service = Service()   # fall back to system chromedriver

    opts = Options()
    opts.add_argument("--headless")       # use classic headless (stable on macOS)
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    # Point to installed Chrome on macOS if present
    _chrome_mac = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    if os.path.exists(_chrome_mac):
        opts.binary_location = _chrome_mac
    driver = webdriver.Chrome(service=service, options=opts)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def _selenium_login(driver) -> bool:
    """
    Navigate to Dev Business login page (clears WAF challenge automatically)
    and submit Drupal credentials.  Returns True on success.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    user = DEVBUSINESS_USER.strip()
    pw   = DEVBUSINESS_PASS.strip()

    if not user or not pw:
        logger.warning("[devbusiness] No credentials — set DEVBUSINESS_USER/PASS in .env")
        return False

    try:
        logger.info("[devbusiness] Selenium: loading login page (clearing WAF)…")
        driver.get(LOGIN_URL)

        # Wait for WAF challenge to resolve and actual login form to appear
        wait = WebDriverWait(driver, 45)
        try:
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "form#user-login-form, form#user-login, input[name='name']")
            ))
        except Exception:
            # WAF may still be showing interstitial — wait a bit more
            time.sleep(8)
            driver.refresh()
            wait.until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, "input[name='name'], input[name='mail']")
            ))

        # Fill in credentials
        try:
            username_field = driver.find_element(By.CSS_SELECTOR, "input[name='name']")
        except Exception:
            username_field = driver.find_element(By.CSS_SELECTOR, "input[name='mail']")
        username_field.clear()
        username_field.send_keys(user)

        password_field = driver.find_element(By.CSS_SELECTOR, "input[name='pass']")
        password_field.clear()
        password_field.send_keys(pw)

        # Submit the form
        try:
            submit = driver.find_element(
                By.CSS_SELECTOR, "input[type='submit'], button[type='submit']"
            )
            submit.click()
        except Exception:
            password_field.submit()

        # Wait for redirect away from login page
        time.sleep(4)
        wait.until(lambda d: "/user/login" not in d.current_url or
                   d.find_elements(By.CSS_SELECTOR, ".messages--error"))

        if "/user/login" in driver.current_url:
            # Check for error message
            try:
                err = driver.find_element(By.CSS_SELECTOR, ".messages--error")
                logger.error("[devbusiness] Login error: %s", err.text[:200])
            except Exception:
                logger.error("[devbusiness] Login failed — still on login page")
            return False

        logger.info("[devbusiness] Selenium login successful → %s", driver.current_url)
        return True

    except Exception as exc:
        logger.error("[devbusiness] Selenium login exception: %s", exc)
        return False


def _transfer_cookies(driver, session: requests.Session) -> None:
    """Copy all Selenium cookies (including WAF tokens) into a requests session."""
    for cookie in driver.get_cookies():
        session.cookies.set(
            cookie["name"],
            cookie["value"],
            domain=cookie.get("domain", "devbusiness.un.org"),
        )
    session.headers.update(HEADERS)
    logger.info("[devbusiness] Transferred %d cookies to requests session",
                len(driver.get_cookies()))


# =============================================================================
# Listing page scraper (requests-based, after cookie transfer)
# =============================================================================

def _try_json_api(session: requests.Session, page: int) -> Optional[list]:
    """Attempt JSON API — returns list or None."""
    try:
        params = {"page": page, "limit": PAGE_SIZE, "_format": "json"}
        r = session.get(OPPS_JSON, params=params, timeout=30)
        if r.status_code != 200:
            return None
        if "json" not in r.headers.get("content-type", ""):
            return None
        data = r.json()
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for k in ("rows", "data", "results", "items"):
                if isinstance(data.get(k), list):
                    return data[k]
        return None
    except Exception:
        return None


_FIELD_LABELS_RE = (
    "Organization|Agency|Client|Donor|Country|Countries|Location|"
    "Deadline|Closing|Due Date|Value|Contract Value|Notice Type|Type|Category"
)


def _parse_html_listing(html: str) -> list:
    """Parse Dev Business HTML listing page — handles Drupal Views layout."""
    soup    = BeautifulSoup(html, "html.parser")
    entries = []

    cards = soup.find_all("div", class_=re.compile(r"views-row"))
    if not cards:
        cards = soup.find_all("article")
    if not cards:
        cards = soup.find_all(
            lambda tag: tag.name in ("li", "tr") and
            tag.find("a", href=re.compile(r"/opportunity|/contract|/tender"))
        )

    seen_urls = set()

    for card in cards:
        link = (
            card.find("a", class_=re.compile(r"opportunity|title|contract", re.I)) or
            (card.find("h3") and card.find("h3").find("a")) or
            (card.find("h2") and card.find("h2").find("a")) or
            card.find("a", href=re.compile(r"/opportunity|/contract|/tender"))
        )
        if not link:
            continue
        title = link.get_text(strip=True)
        if not title or len(title) < 6:
            continue
        href = link.get("href", "").strip()
        if not href:
            continue
        url = (BASE_URL + href) if not href.startswith("http") else href
        if url in seen_urls:
            continue
        seen_urls.add(url)

        card_text = card.get_text(" ", strip=True)

        def _extract_field(*labels: str) -> str:
            for lbl in labels:
                m = re.search(
                    rf"(?:{re.escape(lbl)})\s*[:\-]\s*(.{{3,120}}?)(?=\s*(?:{_FIELD_LABELS_RE})|$)",
                    card_text, re.I,
                )
                if m:
                    return m.group(1).strip()[:120]
            return ""

        entries.append({
            "title":    title,
            "org":      _extract_field("Organization", "Agency", "Client", "Donor"),
            "country":  _extract_field("Country", "Countries", "Location"),
            "deadline": _extract_field("Deadline", "Closing Date", "Submission Deadline"),
            "value":    _extract_field("Value", "Contract Value", "Estimated Value"),
            "ntype":    _extract_field("Notice Type", "Type", "Contract Type"),
            "url":      url,
        })

    return entries


def _fetch_listing_page(session: requests.Session, page: int) -> tuple:
    """Fetch one listing page. Returns (entries: list, has_next: bool)."""
    json_rows = _try_json_api(session, page)
    if json_rows is not None:
        entries = []
        for row in json_rows:
            title = (
                row.get("title") or row.get("field_title") or
                row.get("field_opportunity_title", "")
            ).strip()
            if not title:
                continue
            path = row.get("path") or row.get("url") or row.get("field_url") or ""
            url  = (BASE_URL + path) if path and not path.startswith("http") else path
            entries.append({
                "title":    title,
                "org":      str(row.get("field_agency") or row.get("field_organization") or ""),
                "country":  str(row.get("field_country") or row.get("field_location") or ""),
                "deadline": str(row.get("field_deadline") or row.get("field_closing_date") or ""),
                "value":    str(row.get("field_value") or row.get("field_contract_value") or ""),
                "ntype":    str(row.get("field_notice_type") or row.get("field_type") or ""),
                "url":      url,
            })
        return entries, len(json_rows) >= PAGE_SIZE

    # HTML fallback
    try:
        r = session.get(OPPS_URL, params={"page": page}, timeout=40)
        r.raise_for_status()
    except Exception as exc:
        logger.warning("[devbusiness] listing page %d failed: %s", page, exc)
        return [], False

    entries = _parse_html_listing(r.text)
    soup    = BeautifulSoup(r.text, "html.parser")
    has_next = bool(
        soup.find("a", class_=re.compile(r"pager.*next|next.*pager", re.I)) or
        soup.find("a", rel="next") or
        soup.find("li", class_="next")
    )
    if not has_next and len(entries) >= PAGE_SIZE:
        has_next = True
    return entries, has_next


# =============================================================================
# Detail page fetcher
# =============================================================================

def _fetch_detail(session: requests.Session, url: str) -> str:
    """Fetch detail page and return main text (up to 2500 chars)."""
    try:
        r = session.get(url, timeout=30)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup.find_all(["nav", "header", "footer", "aside", "script", "style"]):
            tag.decompose()
        for sel in [
            "div.field--type-text-long", "div.field--name-body",
            "article.opportunity", "div#block-system-main",
            "main", "div#content", "div.content-region",
        ]:
            el = soup.select_one(sel)
            if el:
                text = el.get_text(" ", strip=True)
                if len(text) > 100:
                    return text[:2500]
        return soup.get_text(" ", strip=True)[:2500]
    except Exception as exc:
        logger.debug("[devbusiness] detail fetch failed %s: %s", url[:80], exc)
        return ""


# =============================================================================
# Skip filter
# =============================================================================

def _is_goods_tender(title: str, ntype: str) -> bool:
    combined = (title + " " + ntype).lower()
    return any(s in combined for s in _SKIP_TYPES)


# =============================================================================
# Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dev Business Opportunities"

    for col_idx, (col_name, col_width) in enumerate(_MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = _WHITE_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, 2):
        fill = _REL_FILL if row.get("Relevance") else (_ALT_FILL if row_idx % 2 == 0 else None)
        font = _REL_FONT if row.get("Relevance") else _BODY_FONT

        values = [
            row.get("Title",        ""),
            row.get("Organization", ""),
            row.get("Country",      ""),
            row.get("Deadline",     ""),
            row.get("Value",        ""),
            row.get("NoticeType",   ""),
            row.get("Description",  ""),
            row.get("Relevance",    ""),
            row.get("URL",          ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value or ""))
            cell.font      = font
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (3, 8)))
            cell.border    = _THIN_BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 36

    try:
        os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
        wb.save(EXCEL_PATH)
        logger.info("[devbusiness] Excel saved (%d rows)", len(rows))
    except Exception as exc:
        logger.warning("[devbusiness] Excel save failed: %s", exc)


# =============================================================================
# Main run() — public entry point
# =============================================================================

def run() -> tuple:
    """
    Dev Business UN — PERMANENTLY SHUT DOWN on 31 March 2025.

    Returns [], [] immediately. The replacement portals (World Bank, ADB,
    AfDB, UNGM) are all already integrated in TenderRadar.
    """
    print(
        "[devbusiness] ⚠️  UN Development Business (UNDB) permanently closed 31-Mar-2025.\n"
        "              Replacement portals already covered: World Bank, ADB, AfDB, UNGM.\n"
        "              Skipping.",
        flush=True,
    )
    return [], []

    # ── Dead code below — kept for reference ─────────────────────────────────
    if not DEVBUSINESS_USER.strip() or not DEVBUSINESS_PASS.strip():
        print("[devbusiness] ⚠️  No credentials — set DEVBUSINESS_USER/DEVBUSINESS_PASS in .env")
        return [], []

    driver  = None
    session = requests.Session()

    try:
        driver = _make_driver()
        logged_in = _selenium_login(driver)

        if not logged_in:
            print("[devbusiness] Login failed — aborting.", flush=True)
            return [], []

        # Transfer cookies (WAF token + Drupal session) to requests
        _transfer_cookies(driver, session)

    except Exception as exc:
        logger.error("[devbusiness] Selenium setup failed: %s", exc)
        print(f"[devbusiness] Selenium error: {exc}", flush=True)
        return [], []
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    # ── Paginate with requests ────────────────────────────────────────────────
    new_tenders:    list = []
    all_rows:       list = []
    total_seen      = 0
    total_new       = 0
    total_skipped   = 0

    for page in range(MAX_PAGES):
        print(f"[devbusiness] Fetching page {page + 1}/{MAX_PAGES}…", flush=True)
        entries, has_next = _fetch_listing_page(session, page)

        if not entries:
            print(f"[devbusiness] No entries on page {page + 1} — stopping.", flush=True)
            break

        for entry in entries:
            title = entry.get("title", "").strip()
            url   = entry.get("url",   "").strip()
            if not title or not url:
                continue

            total_seen += 1
            ntype = entry.get("ntype", "")

            if _is_goods_tender(title, ntype):
                total_skipped += 1
                continue

            description = _fetch_detail(session, url)
            time.sleep(DETAIL_DELAY)

            relevance = score_relevance(title, description)

            row = {
                "Title":        title,
                "Organization": entry.get("org",      ""),
                "Country":      entry.get("country",  ""),
                "Deadline":     entry.get("deadline", ""),
                "Value":        entry.get("value",    ""),
                "NoticeType":   ntype,
                "Description":  description[:1500],
                "Relevance":    relevance,
                "URL":          url,
            }
            all_rows.append(row)

            path_part = url.replace(BASE_URL, "").strip("/").replace("/", "_")
            tender_id = (
                f"devbusiness::{path_part[:120]}"
                if path_part else
                f"devbusiness::{title[:80]}"
            )

            if check_if_new(tender_id):
                mark_as_seen(tender_id, title=title, source_site="Dev Business UN", url=url)
                rel_count = len([r for r in relevance.split(",") if r.strip()]) if relevance else 0
                if rel_count >= 1:
                    new_tenders.append({
                        "tender_id":      tender_id,
                        "title":          title,
                        "organization":   entry.get("org",     ""),
                        "country":        entry.get("country", ""),
                        "deadline":       entry.get("deadline",""),
                        "url":            url,
                        "source_site":    "Dev Business UN",
                        "description":    description[:1000],
                        "relevance_score":rel_count * 10,
                    })
                    total_new += 1
                    print(f"[devbusiness]   NEW + ALERT | {title[:55]} | {relevance[:50]}", flush=True)
                else:
                    print(f"[devbusiness]   NEW (no alert — no keyword match)", flush=True)
            else:
                print(f"[devbusiness]   seen: {title[:60]}", flush=True)

        time.sleep(DELAY)
        if not has_next:
            print("[devbusiness] Last page — stopping.", flush=True)
            break

    if all_rows:
        _save_excel(all_rows)

    print(
        f"\n[devbusiness] Done — {total_seen} scanned, "
        f"{total_skipped} skipped (goods/supply), "
        f"{len(all_rows)} processed, {total_new} NEW + relevant",
        flush=True,
    )
    return new_tenders, all_rows
