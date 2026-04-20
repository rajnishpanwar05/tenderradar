# =============================================================================
# ilo_scraper.py — ILO (International Labour Organization) Procurement Notices
#
# Source  : https://www.ilo.org/webcommon/php/ungm-rss.php
#           (ILO's own HTML feed of their notices posted to UNGM)
# Method  : requests + BeautifulSoup (server-rendered HTML list)
# Login   : None required
# CAPTCHA : None
#
# Why include alongside UNGM:
#   • The UNGM scraper fetches ALL UN agency notices (often overwhelmed/slow)
#   • This feed is ILO-SPECIFIC — only ILO consultancy, advisory, and TA notices
#   • ILO is a major IDCG client sector (labor, skills, capacity building)
#   • Notices appear here ~24h before UNGM's main search index
#
# Content captured (134 items typically):
#   • ILO consultancy assignments (individual/firm)
#   • Training facilitation & curriculum development
#   • Research & evaluation contracts
#   • Technical advisory services
#   • Capacity building for labor/employment agencies
#
# Run with: python3 main.py --ilo
# Included in default run automatically.
# =============================================================================

import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

# Max concurrent UNGM detail fetches — reduces ~134 sequential fetches to ~17
# parallel batches, cutting worst-case time from ~670s to ~80s.
_DETAIL_MAX_WORKERS = 8

# ── Scraper metadata (auto-discovered by core/registry.py) ────────────────────
SCRAPER_META = {
    "flag":        "ilo",
    "label":       "ILO Procurement",
    "group":       "requests",
    "timeout":     600,
    "max_retries": 1,
    "auto":        True,
}

# ── Config ─────────────────────────────────────────────────────────────────────
RSS_URL      = "https://www.ilo.org/webcommon/php/ungm-rss.php"
AWARDS_URL   = "https://www.ilo.org/webcommon/php/ungm-rss-award.php"
UNGM_BASE    = "https://www.ungm.org"
DELAY        = 0.5    # seconds between requests
ILO_EXCEL    = os.path.join(PORTAL_EXCELS_DIR, "ILO_Procurement_Master.xlsx")

HEADERS = {
    "User-Agent":      "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         "https://www.ilo.org/about-ilo/procurement-ilo",
}

# Skip pure goods/logistics/facilities
_SKIP_PATTERNS = [
    "vehicle repair", "hotel services", "catering", "printing", "cleaning",
    "security services", "maintenance", "supply of", "purchase of",
    "request for quotation", "rfq", "maquetacion", "maquetación",
]


# ── Excel styles ───────────────────────────────────────────────────────────────
COLUMNS = [
    ("Title",        65),
    ("Deadline",     18),
    ("Posted",       16),
    ("Country",      22),
    ("Type",         18),
    ("Description",  65),
    ("Relevance",    38),
    ("Link",         55),
]

HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F0F4FA")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
REL_FONT       = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="9E9E9E")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ILO Procurement"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in COLUMNS]

    for ci, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance") + 1
    link_idx = col_names.index("Link") + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = REL_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt:
                        cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt:
                    cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(ILO_EXCEL)
    print(f"[ilo] Excel saved: {ILO_EXCEL}  ({len(rows)} rows)", flush=True)


def _is_skip(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in _SKIP_PATTERNS)


def _fetch_ungm_detail(session: requests.Session, ungm_url: str) -> dict:
    """
    Fetch an UNGM notice detail page for additional metadata.
    Returns dict with keys: description, country, opp_type.
    Falls back gracefully on any error.
    """
    result = {"description": "", "country": "", "opp_type": ""}
    if not ungm_url or not ungm_url.startswith("http"):
        return result
    try:
        r = session.get(ungm_url, timeout=20)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        # Remove nav/footer noise
        for tag in soup.find_all(["nav", "header", "footer", "script", "style"]):
            tag.decompose()

        # Country
        country_label = soup.find(string=re.compile(r"Country|Location", re.I))
        if country_label and country_label.parent:
            nxt = country_label.parent.find_next_sibling()
            if nxt:
                result["country"] = nxt.get_text(strip=True)

        # Type
        type_label = soup.find(string=re.compile(r"Procurement method|Notice type|Type", re.I))
        if type_label and type_label.parent:
            nxt = type_label.parent.find_next_sibling()
            if nxt:
                result["opp_type"] = nxt.get_text(strip=True)

        # Description — main content area
        main = soup.find("div", class_=re.compile(r"content|description|main|notice", re.I))
        if main:
            result["description"] = main.get_text(" ", strip=True)[:1800]
        else:
            result["description"] = soup.get_text(" ", strip=True)[:1500]

    except Exception as e:
        print(f"[ilo]   detail fetch error ({ungm_url[:60]}): {e}", flush=True)
    return result


def _parse_rss_items(html: str) -> list:
    """
    Parse the ILO UNGM RSS HTML page.
    Each item is an <li> with:
      - <h5><a href="ungm_url">Title</a></h5>
      - <p class="item-info">metadata</p>
      - <p>Deadline: DD Month YYYY</p>
    """
    soup  = BeautifulSoup(html, "html.parser")
    items = soup.find_all("li")
    results = []

    for li in items:
        a_tag = li.find("a", href=True)
        if not a_tag:
            continue

        title = a_tag.get_text(strip=True)
        if not title or len(title) < 5:
            continue

        url = a_tag["href"]
        if not url.startswith("http"):
            url = UNGM_BASE + url

        # Extract deadline from the <p>Deadline: ...</p> text
        deadline = ""
        for p in li.find_all("p"):
            text = p.get_text(strip=True)
            if text.startswith("Deadline:"):
                deadline = text.replace("Deadline:", "").strip()
                break

        # Extract reference number from URL (ungm.org/Public/notice/NNNN)
        notice_id = ""
        m = re.search(r"/notice/(\d+)", url)
        if m:
            notice_id = m.group(1)

        results.append({
            "title":     title,
            "url":       url,
            "deadline":  deadline,
            "notice_id": notice_id,
        })

    return results


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[ilo] ILO Procurement Notices Pipeline starting...", flush=True)
    print("[ilo] Source: ILO UNGM RSS feed (134 ILO-specific notices)", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []

    if os.path.exists(ILO_EXCEL):
        try:
            os.remove(ILO_EXCEL)
        except Exception:
            pass

    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Fetch RSS HTML ─────────────────────────────────────────────────────────
    for attempt in (1, 2, 3):
        try:
            r = session.get(RSS_URL, timeout=90)
            r.raise_for_status()
            raw_items = _parse_rss_items(r.text)
            break
        except Exception as e:
            if attempt == 3:
                print(f"[ilo] ERROR fetching RSS feed: {e}", flush=True)
                return new_tenders, all_rows
            print("[ilo] Retry fetching RSS after error…", flush=True)
            time.sleep(2)

    print(f"[ilo] RSS parsed: {len(raw_items)} notices found", flush=True)

    # ── Pre-filter: skip goods/logistics by title alone (no detail fetch needed) ─
    candidates = []
    seen_ids: set = set()
    for item in raw_items:
        title     = item["title"]
        notice_id = item["notice_id"]
        if _is_skip(title):
            continue
        key = notice_id or title[:80]
        if key in seen_ids:
            continue
        seen_ids.add(key)
        candidates.append(item)

    print(
        f"[ilo] After goods pre-filter: {len(candidates)}/{len(raw_items)} notices to fetch details for",
        flush=True,
    )

    # ── Fetch detail pages concurrently ───────────────────────────────────────
    total = len(candidates)
    details: list = [None] * total
    with ThreadPoolExecutor(max_workers=_DETAIL_MAX_WORKERS) as pool:
        future_to_idx = {
            pool.submit(_fetch_ungm_detail, session, item["url"]): idx
            for idx, item in enumerate(candidates)
        }
        done = 0
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            done += 1
            try:
                details[idx] = future.result() or {}
            except Exception:
                details[idx] = {}
            print(
                f"[ilo]   [{done:>3}/{total}] {candidates[idx]['title'][:60]}",
                flush=True,
            )

    # ── Score + dedup (sequential — no I/O) ──────────────────────────────────
    for idx, item in enumerate(candidates):
        title     = item["title"]
        url       = item["url"]
        notice_id = item["notice_id"]
        deadline  = item["deadline"]
        detail    = details[idx] or {}

        description = detail.get("description", "")
        country     = detail.get("country", "")
        opp_type    = detail.get("opp_type", "")

        relevance = score_relevance(title, description)

        # Only keep if relevant OR title has IDCG consulting keywords
        if not relevance and not title_is_relevant(title):
            continue

        row = {
            "Title":         title,
            "Deadline":      deadline,
            "Posted":        "",
            "Country":       country,
            "Type":          opp_type,
            "Description":   description[:1500] if description else "",
            "Relevance":     relevance,
            "Link":          url,
            # Explicit portal slug so materialize_normalized_batch classifies
            # these as 'ilo' rather than falling back to 'ungm' via URL inference
            "source_portal": "ilo",
            "tender_id":     f"ILO/{notice_id}" if notice_id
                             else f"ILO/{''.join(c if c.isalnum() else '_' for c in title[:60])}",
        }
        all_rows.append(row)

        tender_id = (
            f"ILO/{notice_id}" if notice_id
            else f"ILO/{re.sub(r'[^a-zA-Z0-9]', '_', title[:60])}"
        )

        if check_if_new(tender_id):
            mark_as_seen(tender_id, title=title, source_site="ILO", url=url)
            new_tenders.append({
                "title":    title,
                "deadline": deadline,
                "value":    opp_type or "ILO Consultancy",
                "url":      url,
            })
            print(f"[ilo]   → NEW | {relevance[:60] if relevance else '—'}", flush=True)
        else:
            print(f"[ilo]   → seen | {relevance[:50] if relevance else '—'}", flush=True)

    # ── Save Excel ─────────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)

    print(
        f"\n[ilo] Done — {len(all_rows)} relevant notices, {len(new_tenders)} NEW",
        flush=True,
    )
    return new_tenders, all_rows


if __name__ == "__main__":
    new, all_ = run()
    print(f"\n✅ ILO: {len(new)} new / {len(all_)} total")
