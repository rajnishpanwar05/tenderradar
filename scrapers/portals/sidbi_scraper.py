# =============================================================================
# sidbi_pipeline.py — SIDBI (Small Industries Development Bank of India) Tenders
#
# Site   : https://sidbi.in/en/tenders
# Method : Primary — JSON API (DataTables AJAX endpoint, no Selenium needed)
#          Fallback — Selenium headless Chrome
# Login  : None required
# CAPTCHA: None on the JSON API path
#
# Key discovery (2026-04): The DataTables AJAX endpoint returns all tenders
# as JSON when the request includes:
#   X-Requested-With: XMLHttpRequest
#   Referer: https://www.sidbi.in/en/tenders
# This completely bypasses Selenium and any CAPTCHA.
#
# What this does:
#   1. Fetches JSON API: JSONtender_front.php?show=frontend  (primary, fast)
#   2. Falls back to Selenium if JSON API fails
#   3. Scores relevance against IDCG expertise
#   4. Deduplicates via MySQL — only NEW tenders trigger notification
#   5. Saves formatted Excel + returns (new_tenders, all_rows) for main.py
# =============================================================================

import os, re, time
import requests

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://sidbi.in"
LISTING_URL = f"{BASE_URL}/en/tenders"

SIDBI_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "SIDBI_Tenders_Master.xlsx")

PAGE_LOAD_WAIT  = 20
ACTION_DELAY    = 2.0
MAX_PAGES       = 20   # 50/page × 20 = 1000 max

# Goods/supply patterns to skip
SKIP_TITLE_PATTERNS = [
    "supply of", "purchase of", "procurement of vehicle", "procurement of furniture",
    "procurement of equipment", "annual maintenance", "amc for", "rate contract",
    "printing of", "empanelment of vendor",
]

def _is_goods_tender(title: str) -> bool:
    t = title.lower()
    return any(p in t for p in SKIP_TITLE_PATTERNS)


# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Title",          65),
    ("Tender Date",    16),
    ("Last Date",      16),
    ("Remark",         30),
    ("Description",    65),
    ("Relevance",      40),
    ("Detail Link",    55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
HIGH_FONT      = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"), right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 1 — Selenium driver setup
# =============================================================================

def _make_driver(headless=True):
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


# =============================================================================
# SECTION 2 — Parse tender table rows
# =============================================================================

def _parse_table(html: str, base_url: str = BASE_URL) -> list:
    """Extract rows from the SIDBI tender DataTable."""
    soup    = BeautifulSoup(html, 'html.parser')
    entries = []

    # Find DataTable — SIDBI uses id="tenderLIST" (confirmed live 2026-04-02)
    table = soup.find('table', id=re.compile(r'tenderLIST|tender|DataTable|myTable', re.I))
    if not table:
        # Fallback: any table with tender-like headers
        for t in soup.find_all('table'):
            headers_text = t.get_text()[:300].lower()
            if any(kw in headers_text for kw in ['title', 'tender date', 'last date', 'submission']):
                table = t
                break

    if not table:
        return entries

    rows = table.find_all('tr')
    if not rows:
        return entries

    # Identify header row and column positions
    header_row = rows[0]
    headers = [th.get_text(strip=True).lower() for th in header_row.find_all(['th', 'td'])]

    col = {}
    for i, h in enumerate(headers):
        if 'title' in h or 'tender' in h:       col.setdefault('title',       i)
        if 'tender date' in h or 'issue' in h:  col.setdefault('tender_date', i)
        if 'last date' in h or 'deadline' in h: col.setdefault('last_date',   i)
        if 'submission' in h:                   col.setdefault('last_date',   i)
        if 'remark' in h or 'status' in h:      col.setdefault('remark',      i)

    for row in rows[1:]:
        cells = row.find_all(['td', 'th'])
        if len(cells) < 2:
            continue
        row_text = row.get_text(strip=True)
        if len(row_text) < 5:
            continue

        def gcell(key):
            idx = col.get(key)
            return cells[idx].get_text(' ', strip=True) if idx is not None and idx < len(cells) else ""

        title = gcell('title')
        if not title or len(title) < 5:
            # Use longest cell as title
            title = max(cells, key=lambda c: len(c.get_text())).get_text(strip=True)

        # Extract detail link
        detail_link = ""
        title_cell  = cells[col.get('title', 0)] if col.get('title', 0) < len(cells) else cells[0]
        link_tag    = title_cell.find('a', href=True) or row.find('a', href=True)
        if link_tag:
            href = link_tag.get('href', '').strip()
            detail_link = href if href.startswith('http') else base_url + href

        if not title or len(title) < 5:
            continue

        tender_date = gcell('tender_date')
        last_date   = gcell('last_date')
        remark      = gcell('remark')

        safe_id = re.sub(r'[^a-zA-Z0-9]', '_', (detail_link or title)[:80])
        tender_id = f"SIDBI_{safe_id}"

        entries.append({
            "Title":       title,
            "Tender Date": tender_date,
            "Last Date":   last_date,
            "Remark":      remark,
            "detail_url":  detail_link,
            "tender_id":   tender_id,
        })

    return entries


def _fetch_detail(driver, url: str) -> str:
    """Fetch tender detail page and return description text."""
    if not url or not url.startswith('http'):
        return ""
    try:
        driver.get(url)
        WebDriverWait(driver, PAGE_LOAD_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, 'body'))
        )
        time.sleep(1.5)
        soup    = BeautifulSoup(driver.page_source, 'html.parser')
        content = (
            soup.find('div', class_=re.compile(r'content|article|body|tender', re.I)) or
            soup.find('main') or
            soup.find('article')
        )
        if content:
            return content.get_text(' ', strip=True)[:1800]
        return soup.get_text(' ', strip=True)[:1500]
    except Exception as e:
        print(f"[sidbi]   Detail fetch error: {e}")
        return ""


# =============================================================================
# SECTION 3 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SIDBI Tenders"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = WHITE_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    rel_idx  = col_names.index("Relevance")    + 1
    link_idx = col_names.index("Detail Link")  + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None
        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif ci == rel_idx:
                if val:
                    cell.fill = RELEVANCE_FILL; cell.font = HIGH_FONT
                else:
                    cell.font = NO_REL_FONT
                    if alt: cell.fill = alt
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(SIDBI_EXCEL_PATH)
    print(f"[sidbi] Excel saved: {SIDBI_EXCEL_PATH}  ({len(rows)} rows)")


# ── JSON API constants ─────────────────────────────────────────────────────────
# Discovered 2026-04: DataTables AJAX endpoint returns all tenders as JSON.
# Requires XHR headers; returns "Request Ignored" without them.
_JSON_API_URL = "https://www.sidbi.in/head/engine/json/JSONtender_front.php?show=frontend"
_XHR_HEADERS  = {
    "User-Agent":        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept":            "application/json, text/javascript, */*; q=0.01",
    "Accept-Language":   "en-US,en;q=0.9",
    "X-Requested-With":  "XMLHttpRequest",
    "Referer":           "https://www.sidbi.in/en/tenders",
    "Origin":            "https://www.sidbi.in",
}


# =============================================================================
# SECTION 3b — requests-first helper (JSON API — no Selenium, no CAPTCHA)
# =============================================================================

def _try_requests_first() -> list:
    """
    Primary fetch path: call the DataTables AJAX JSON endpoint with XHR headers.
    Returns list of parsed entries, or [] on failure (triggers Selenium fallback).

    JSON fields per tender:
      tender_title, tender_title_eng, tender_date, tender_last_date,
      tender_remarks, tender_url, status
    """
    try:
        sess = requests.Session()
        # Load main page first to get session cookies (PHPSESSID)
        sess.get(LISTING_URL, headers=_XHR_HEADERS, timeout=15)

        r = sess.get(_JSON_API_URL, headers=_XHR_HEADERS, timeout=20)
        r.raise_for_status()

        # Server returns "Request Ignored" (plain text) when XHR headers missing
        if not r.text.strip().startswith("{"):
            print(f"[sidbi] JSON API returned non-JSON: {r.text[:60]!r}")
            return []

        data  = r.json()
        items = data.get("data", [])
        if not items:
            print("[sidbi] JSON API: no items in response")
            return []

        print(f"[sidbi] JSON API: {len(items)} tender(s) received")
        entries = []
        for item in items:
            title = (item.get("tender_title_eng") or item.get("tender_title") or "").strip()
            if not title or len(title) < 5:
                continue

            tender_url = (item.get("tender_url") or "").strip()
            detail_link = (
                tender_url if tender_url.startswith("http")
                else (BASE_URL + "/" + tender_url.lstrip("/") if tender_url else LISTING_URL)
            )

            safe_id   = re.sub(r"[^a-zA-Z0-9]", "_", (detail_link or title)[:80])
            tender_id = f"SIDBI_{safe_id}"

            entries.append({
                "Title":       title,
                "Tender Date": (item.get("tender_date") or "").strip(),
                "Last Date":   (item.get("tender_last_date") or "").strip(),
                "Remark":      (item.get("tender_remarks") or "").strip(),
                "detail_url":  detail_link,
                "tender_id":   tender_id,
            })
        return entries

    except Exception as e:
        print(f"[sidbi] JSON API failed: {e}")
        return []


# =============================================================================
# SECTION 4 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the SIDBI Tenders pipeline.
    Returns (new_tenders, all_rows).
    """
    print("\n" + "=" * 65)
    print("[sidbi] SIDBI Tenders Pipeline starting...")
    print("=" * 65)

    new_tenders, all_rows = [], []

    if os.path.exists(SIDBI_EXCEL_PATH):
        try:
            os.remove(SIDBI_EXCEL_PATH)
            print("[sidbi] Cleared old Excel")
        except Exception:
            pass

    # ── Try plain requests before launching Selenium (bypasses Selenium CAPTCHA) ─
    print("[sidbi] Attempting requests-first (no browser)...")
    rf_entries = _try_requests_first()
    if rf_entries:
        print(f"[sidbi] Requests succeeded ({len(rf_entries)} rows) — Selenium not needed.")
        for entry in rf_entries:
            title = entry["Title"]
            if _is_goods_tender(title):
                continue
            relevance = score_relevance(title, "")
            row = {
                "Title":       title,
                "Tender Date": entry["Tender Date"],
                "Last Date":   entry["Last Date"],
                "Remark":      entry["Remark"],
                "Description": "",          # no detail page in requests-fast path
                "Relevance":   relevance,
                "Detail Link": entry["detail_url"] or LISTING_URL,
            }
            all_rows.append(row)
            tid = entry["tender_id"]
            if check_if_new(tid):
                mark_as_seen(tid, title=title, source_site="SIDBI",
                             url=entry["detail_url"] or LISTING_URL)
                new_tenders.append({
                    "title":    title,
                    "deadline": entry["Last Date"],
                    "value":    "See SIDBI",
                    "url":      entry["detail_url"] or LISTING_URL,
                })
        if all_rows:
            _save_excel(all_rows)
        relevant = sum(1 for r in all_rows if r["Relevance"])
        print(f"\n[sidbi] Done (requests) — {len(all_rows)} listings, "
              f"{len(new_tenders)} NEW, {relevant} relevant")
        return new_tenders, all_rows

    print("[sidbi] Requests returned 0 rows — launching Selenium browser...")
    driver = None
    try:
        driver = _make_driver(headless=True)
        driver.get(LISTING_URL)

        # Wait for DataTables JS to render the table (SIDBI uses id="tenderLIST")
        try:
            WebDriverWait(driver, PAGE_LOAD_WAIT).until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    '#tenderLIST tbody tr, table[id*="tender"] tbody tr, table tr td'
                ))
            )
        except TimeoutException:
            print("[sidbi] Timeout waiting for table — page may need login or CAPTCHA.")
            return [], []

        time.sleep(ACTION_DELAY)

        # Check for CAPTCHA (Selenium path)
        page_src_lower = driver.page_source.lower()
        if any(kw in page_src_lower for kw in ['captcha', 'recaptcha', 'verify you are human',
                                                 'are you a robot', 'hcaptcha', 'cf-challenge']):
            print("[sidbi] WARNING: CAPTCHA detected in Selenium — skipping SIDBI this run.")
            print("[sidbi] TIP: Run manually with headless=False to solve CAPTCHA once.")
            return [], []

        # Try to show 50 rows per page if dropdown available
        try:
            length_select = driver.find_element(By.CSS_SELECTOR, 'select[name*="length"], select[name*="DataTable"]')
            Select(length_select).select_by_value('50')
            time.sleep(ACTION_DELAY)
        except Exception:
            pass

        all_entries = []
        seen_titles = set()

        for page_num in range(1, MAX_PAGES + 1):
            print(f"[sidbi]   Page {page_num}...", end=" ", flush=True)
            html    = driver.page_source
            entries = _parse_table(html)
            new_on_page = [e for e in entries if e["tender_id"] not in seen_titles]
            print(f"{len(new_on_page)} tenders")

            for e in new_on_page:
                seen_titles.add(e["tender_id"])
            all_entries.extend(new_on_page)

            if not new_on_page and page_num > 1:
                print("[sidbi]   No new rows — done.")
                break

            # Click Next button (DataTables pagination)
            try:
                next_btn = driver.find_element(
                    By.CSS_SELECTOR,
                    '#DataTables_Table_0_next, .next, [id$="_next"], a[data-dt-idx]'
                )
                if 'disabled' in (next_btn.get_attribute('class') or ''):
                    print("[sidbi]   Last page reached.")
                    break
                driver.execute_script("arguments[0].click();", next_btn)
                time.sleep(ACTION_DELAY)
            except NoSuchElementException:
                print(f"[sidbi]   No pagination after page {page_num}.")
                break

        print(f"[sidbi] Total listings: {len(all_entries)}")

        # Process each entry
        for i, entry in enumerate(all_entries, 1):
            title = entry["Title"]

            if _is_goods_tender(title):
                print(f"[sidbi]   [{i:>3}/{len(all_entries)}] SKIP (goods): {title[:60]}")
                continue

            print(f"[sidbi]   [{i:>3}/{len(all_entries)}] {title[:65]}...")

            desc      = _fetch_detail(driver, entry["detail_url"]) if entry["detail_url"] else ""
            relevance = score_relevance(title, desc)

            row = {
                "Title":       title,
                "Tender Date": entry["Tender Date"],
                "Last Date":   entry["Last Date"],
                "Remark":      entry["Remark"],
                "Description": desc[:1500] if desc else "",
                "Relevance":   relevance,
                "Detail Link": entry["detail_url"] or LISTING_URL,
            }
            all_rows.append(row)

            tid = entry["tender_id"]
            if check_if_new(tid):
                mark_as_seen(tid, title=title, source_site="SIDBI",
                             url=entry["detail_url"] or LISTING_URL)
                new_tenders.append({
                    "title":    title,
                    "deadline": entry["Last Date"],
                    "value":    "See SIDBI",
                    "url":      entry["detail_url"] or LISTING_URL,
                })
                print(f"           → NEW | Relevance: {relevance or '—'}")
            else:
                print(f"           → seen | Relevance: {relevance or '—'}")

            time.sleep(1.0)

    except Exception as e:
        print(f"[sidbi] FATAL error: {e}")
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r["Relevance"])
    print(f"\n[sidbi] Done — {len(all_rows)} listings, {len(new_tenders)} NEW, {relevant} relevant")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nNew: {len(new)}")
