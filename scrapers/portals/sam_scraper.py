from __future__ import annotations

# =============================================================================
# sam_scraper.py — SAM.gov (System for Award Management) v3
#
# Source  : https://api.sam.gov/opportunities/v2/search  (free public REST API)
# Method  : GET → JSON  (no Selenium)
# Auth    : Free API key — generate at sam.gov → Profile → API Keys
#           IDCG accounts: safa@idcg.co.in / ajit@idcg.co.in
#           Add to config/.env:  SAM_API_KEY=<your_key>
#
# Robustness features (v3):
#   • Rate-limit handling  : 429 → exponential backoff (up to 4 retries)
#   • Key-expiry detection : 401/403 → clear warning + graceful exit
#   • Result caching       : saves last good JSON → falls back if API fails
#   • Endpoint probing     : tests key once before full run
#   • Targeted queries     : IDCG-specific India/South Asia consulting focus
#   • Dedup across queries : single seen_notice_ids set prevents duplicates
# =============================================================================

import json
import os
import re
import time
from datetime import datetime, timedelta
from typing import Optional

# Auto-registry metadata — overrides static registry timeout
SCRAPER_META = {
    "flag":    "sam",
    "label":   "SAM.gov",
    "group":   "api",
    "timeout": 600,        # SAM API is slow; 5 queries × 2 pages × ~25s = ~250s worst-case
    "auto":    True,
}

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR, SAM_API_KEY
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Constants ─────────────────────────────────────────────────────────────────
SAM_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "SAM_Opportunities_Master.xlsx")
CACHE_PATH     = os.path.join(os.path.dirname(SAM_EXCEL_PATH), ".sam_cache.json")

API_URLS   = [
    "https://api.sam.gov/opportunities/v2/search",
    "https://api-alpha.sam.gov/opportunities/v2/search",
]
LIMIT       = 100     # max per API call (SAM.gov hard cap)
MAX_PAGES   = 2       # 2 × 100 = 200 results per query (API is slow; keep under timeout)
POSTED_DAYS = 30      # look back 30 days (45 was adding lag with no benefit)

# IDCG-targeted search queries — phrase-quoted so SAM does AND matching, not OR.
# SAM.gov API v2 treats q="phrase" as phrase search when URL-encoded as %22phrase%22.
# Fewer, sharper queries beat 10 broad ones that return 27k+ noise records.
SEARCH_QUERIES = [
    "USAID India",
    "India technical assistance evaluation",
    "India consulting advisory",
    "South Asia development consulting",
    "India monitoring assessment",
]

_PROCUREMENT_TYPES = ["p", "r", "s", "o", "k", "i"]

# Opportunity types to keep (SAM.gov type codes)
_KEEP_TYPES = {
    "o",    # Solicitation
    "p",    # Pre-Solicitation
    "r",    # Sources Sought
    "s",    # Special Notice
    "k",    # Combined Synopsis/Solicitation
    "i",    # Intent to Bundle Requirements
}

HEADERS = {
    "Accept":     "application/json",
    "User-Agent": "TenderMonitor/3.0",
}

# Retry config
_MAX_RETRIES     = 4
_RETRY_BASE_WAIT = 5   # seconds; doubles each attempt (5, 10, 20, 40)


# ── Excel styles ───────────────────────────────────────────────────────────────
COLUMNS = [
    ("Sol. Number",  22),
    ("Title",        58),
    ("Agency",       35),
    ("Type",         14),
    ("Posted",       13),
    ("Deadline",     13),
    ("NAICS",        10),
    ("Relevance",    35),
    ("Link",         52),
]
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_ALT_FILL  = PatternFill("solid", fgColor="F5F8FF")
_REL_FILL  = PatternFill("solid", fgColor="E2EFDA")
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=10)
_LINK_FONT = Font(name="Calibri", size=10, color="1155CC", underline="single")
_HIGH_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),  right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin",  color="D0D7E3"),  bottom=Side(style="thin", color="D0D7E3"),
)


def _save_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SAM Opportunities"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    col_names = [c[0] for c in COLUMNS]
    for ci, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    rel_idx  = col_names.index("Relevance") + 1
    link_idx = col_names.index("Link")      + 1

    for ri, row_data in enumerate(rows, 2):
        alt = _ALT_FILL if ri % 2 == 0 else None
        ws.row_dimensions[ri].height = 40

        vals = [
            row_data.get("sol_num",   ""),
            row_data.get("title",     ""),
            row_data.get("agency",    ""),
            row_data.get("opp_type",  ""),
            row_data.get("posted",    ""),
            row_data.get("deadline",  ""),
            row_data.get("naics",     ""),
            row_data.get("relevance", ""),
            row_data.get("link",      ""),
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = _LINK_FONT
            elif ci == rel_idx:
                if val:
                    cell.fill = _REL_FILL
                    cell.font = _HIGH_FONT
                else:
                    cell.font = _BODY_FONT
                    if alt:
                        cell.fill = alt
            else:
                cell.font = _BODY_FONT
                if alt:
                    cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    wb.save(SAM_EXCEL_PATH)
    print(f"[sam] Excel saved: {SAM_EXCEL_PATH}  ({len(rows)} rows)")


# ── Cache helpers ──────────────────────────────────────────────────────────────

def _load_cache() -> list[dict]:
    """Load cached raw opportunities from last successful API run."""
    try:
        if os.path.exists(CACHE_PATH):
            with open(CACHE_PATH, encoding="utf-8") as f:
                data = json.load(f)
            age_hours = (time.time() - data.get("saved_at", 0)) / 3600
            if age_hours < 12:
                print(f"[sam] Using cached data ({age_hours:.1f}h old, {len(data['opps'])} opps)")
                return data["opps"]
            print(f"[sam] Cache too old ({age_hours:.1f}h) — will fetch fresh")
    except Exception:
        pass
    return []


def _save_cache(opps: list[dict]) -> None:
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as f:
            json.dump({"saved_at": time.time(), "opps": opps}, f)
    except Exception:
        pass


# ── API helpers ────────────────────────────────────────────────────────────────

def _sam_get(params: dict) -> tuple[Optional[requests.Response], str]:
    """
    Try SAM endpoints in order and return (response, base_url).
    Falls through on 404 to accommodate endpoint migrations.
    """
    for base in API_URLS:
        try:
            r = requests.get(base, params=params, headers=HEADERS, timeout=25)
            if r.status_code == 404:
                continue
            return r, base
        except Exception:
            continue
    return None, ""


def _probe_key(api_key: str, max_wait: int = 120) -> tuple[bool, str]:
    """
    Test the API key with a minimal request.
    Retries once on 429 (rate limit) after waiting Retry-After seconds.
    Returns (ok: bool, message: str).
    """
    today = datetime.utcnow()
    date_to   = today.strftime("%m/%d/%Y")
    date_from = (today - timedelta(days=7)).strftime("%m/%d/%Y")
    params = {
        "api_key":    api_key,
        "q":          "consulting",
        "postedFrom": date_from,
        "postedTo":   date_to,
        "limit":      1,
        "status":     "active",
    }
    try:
        for attempt in range(2):
            r, base = _sam_get(params)
            if r is None:
                return False, "SAM API endpoint unreachable (404/connection)"
            if r.status_code == 200:
                try:
                    total = r.json().get("totalRecords", 0)
                    print(f"[sam] Probe OK — API key valid, {total} total records in range")
                except Exception:
                    pass
                return True, "OK"
            if r.status_code in (401, 403):
                body = (r.text or "")[:200]
                return False, f"Invalid/expired key (HTTP {r.status_code}) — {body}"
            if r.status_code == 429:
                if attempt == 0:
                    wait = min(int(r.headers.get("Retry-After", 60)), max_wait)
                    print(f"[sam] Rate limited on probe — waiting {wait}s before retry…")
                    time.sleep(wait)
                    continue
                return False, "Rate limited — daily quota likely exhausted; will retry next run"
            body = (r.text or "")[:200]
            return False, f"Unexpected HTTP {r.status_code} — {body}"
    except Exception as exc:
        return False, f"Network error: {exc}"
    return False, "Probe failed after retries"


def _fetch_page(
    api_key: str,
    query: str,
    offset: int,
    date_from: str,
    date_to: str,
) -> Optional[dict]:
    """
    Fetch one page of results with exponential-backoff retry.
    Returns parsed JSON dict or None on permanent failure.
    """
    params = {
        "api_key":    api_key,
        "q":          query,           # SAM v2 uses 'q' for keyword search, not 'title'
        "limit":      LIMIT,
        "offset":     offset,
        "postedFrom": date_from,
        "postedTo":   date_to,
        "ptype":      ",".join(_PROCUREMENT_TYPES),  # comma-separated, not list
        "status":     "active",
    }
    wait = _RETRY_BASE_WAIT
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            r, base = _sam_get(params)
            if r is None:
                print("[sam]   Endpoint unreachable (404/connection) — aborting query")
                return None

            if r.status_code == 200:
                return r.json()

            if r.status_code == 429:
                # Rate limited — back off and retry
                retry_after = int(r.headers.get("Retry-After", wait))
                print(f"[sam]   Rate limited (429) — waiting {retry_after}s (attempt {attempt}/{_MAX_RETRIES})")
                time.sleep(retry_after)
                wait *= 2
                continue

            if r.status_code in (401, 403):
                print(f"[sam]   Auth error ({r.status_code}) — API key invalid or expired")
                return None  # Permanent — stop retrying

            if r.status_code >= 500:
                print(f"[sam]   Server error ({r.status_code}) — retrying in {wait}s")
                time.sleep(wait)
                wait *= 2
                continue

            body_snip = r.text[:500].replace("\n", " ") if hasattr(r, "text") else ""
            print(f"[sam]   HTTP {r.status_code} for query '{query}' offset {offset} body='{body_snip}'")
            return None

        except requests.Timeout:
            print(f"[sam]   Timeout (attempt {attempt}/{_MAX_RETRIES}) — retrying in {wait}s")
            time.sleep(wait)
            wait *= 2
        except requests.ConnectionError as exc:
            print(f"[sam]   Connection error: {exc} — retrying in {wait}s")
            time.sleep(wait)
            wait *= 2
        except Exception as exc:
            print(f"[sam]   Unexpected error: {exc}")
            return None

    print(f"[sam]   Gave up after {_MAX_RETRIES} attempts")
    return None


def _fetch_all_opportunities(api_key: str, queries: list[str] | None = None) -> list[dict]:
    """
    Run all queries, paginate each, deduplicate by noticeId.
    Returns list of raw opportunity dicts.
    """
    queries = queries or SEARCH_QUERIES
    today     = datetime.utcnow()
    date_to   = today.strftime("%m/%d/%Y")
    date_from = (today - timedelta(days=POSTED_DAYS)).strftime("%m/%d/%Y")

    seen_ids: set[str] = set()
    all_opps: list[dict] = []

    for qi, query in enumerate(queries, 1):
        print(f"[sam]   Query {qi}/{len(queries)}: '{query}'...")
        query_count = 0

        for page in range(MAX_PAGES):
            offset = page * LIMIT
            data   = _fetch_page(api_key, query, offset, date_from, date_to)

            if data is None:
                break   # auth error or permanent failure — stop this query

            opps  = data.get("opportunitiesData") or []
            total = int(data.get("totalRecords") or 0)

            for opp in opps:
                nid = (opp.get("noticeId") or "").strip()
                if nid and nid not in seen_ids:
                    seen_ids.add(nid)
                    all_opps.append(opp)
                    query_count += 1

            fetched_so_far = offset + len(opps)
            if not opps or fetched_so_far >= total:
                break

            time.sleep(0.5)   # polite pause between pages

        print(f"[sam]   → {query_count} new opportunities")
        time.sleep(1.0)   # pause between queries

    # If nothing came back, run a final catch-all query
    if not all_opps:
        print("[sam]   No results from focused queries — running catch-all '*'...")
        data = _fetch_page(api_key, "consulting", 0, date_from, date_to)
        if data:
            for opp in (data.get("opportunitiesData") or []):
                nid = (opp.get("noticeId") or "").strip()
                if nid and nid not in seen_ids:
                    seen_ids.add(nid)
                    all_opps.append(opp)
            print(f"[sam]   Catch-all added {len(all_opps)} opportunities")
    return all_opps


# ── Row builder ───────────────────────────────────────────────────────────────

def _build_row(opp: dict) -> Optional[dict]:
    """
    Convert a raw SAM.gov opportunity dict to our row format.
    Returns None if the opportunity should be skipped.
    """
    notice_id = (opp.get("noticeId") or "").strip()
    title     = (opp.get("title")    or "").strip()
    if not notice_id or not title:
        return None

    # Type filter — skip awards, modifications of no interest
    opp_type_code = (opp.get("type") or "").lower().strip()
    if opp_type_code and opp_type_code not in _KEEP_TYPES and opp_type_code.startswith("a"):
        return None   # "a" = Award Notice — not relevant

    description = (opp.get("description") or "").strip()
    relevance   = score_relevance(title, description)

    # Require at least a minimal consulting/advisory signal
    title_lower = title.lower()
    has_signal  = relevance or any(k in title_lower for k in (
        "consult", "advisory", "technical assistance", "evaluation",
        "assess", "capacity", "monitoring", "research", "survey",
        "social", "environment", "governance", "development",
    ))
    if not has_signal:
        return None

    sol_num  = (opp.get("solicitationNumber") or notice_id).strip()
    agency   = (
        opp.get("fullParentPathName")
        or opp.get("organizationHierarchy")
        or opp.get("organizationName")
        or ""
    ).strip()
    naics    = (opp.get("naicsCode") or "").strip()

    # Deadline and posted date — strip to YYYY-MM-DD
    deadline = (opp.get("responseDeadLine") or "")[:10]
    posted   = (opp.get("postedDate")       or "")[:10]

    link = (
        opp.get("uiLink")
        or f"https://sam.gov/opp/{notice_id}/view"
    ).strip()

    # Preserve any explicit resource/attachment URLs exposed by SAM payload.
    discovered_urls: list[str] = []
    seen_urls = set()
    for key in ("resourceLinks", "attachments", "relatedUrls", "additionalInfoLink", "archiveUrl", "link"):
        value = opp.get(key)
        if isinstance(value, str):
            candidate = value.strip()
            if candidate.startswith("http") and candidate not in seen_urls:
                seen_urls.add(candidate)
                discovered_urls.append(candidate)
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, str):
                    candidate = item.strip()
                elif isinstance(item, dict):
                    candidate = str(
                        item.get("url")
                        or item.get("link")
                        or item.get("href")
                        or item.get("resourceLink")
                        or ""
                    ).strip()
                else:
                    candidate = ""
                if candidate.startswith("http") and candidate not in seen_urls:
                    seen_urls.add(candidate)
                    discovered_urls.append(candidate)
        elif isinstance(value, dict):
            candidate = str(
                value.get("url")
                or value.get("link")
                or value.get("href")
                or value.get("resourceLink")
                or ""
            ).strip()
            if candidate.startswith("http") and candidate not in seen_urls:
                seen_urls.add(candidate)
                discovered_urls.append(candidate)

    rich_description = description
    if discovered_urls:
        rich_description = (description + "\n\nDocument/Resource URLs:\n" + "\n".join(discovered_urls)).strip()

    return {
        "tender_id": f"SAM_{notice_id}",
        "sol_num":   sol_num,
        "title":     title,
        "agency":    agency,
        "opp_type":  opp_type_code.upper(),
        "posted":    posted,
        "deadline":  deadline,
        "naics":     naics,
        "link":      link,
        "relevance": relevance,
        "description": rich_description[:8000],
    }


# ── Main ───────────────────────────────────────────────────────────────────────

def run() -> tuple:
    """
    Run the SAM.gov opportunities pipeline.
    Returns (new_tenders: list, all_rows: list).
    """
    print("\n[sam] SAM.gov Opportunities Pipeline (v3)")

    # ── 1. Check API key ───────────────────────────────────────────────────────
    api_key = SAM_API_KEY if (
        SAM_API_KEY and SAM_API_KEY.strip() not in ("", "YOUR_SAM_API_KEY")
    ) else ""

    raw_opps: list[dict] = []

    if not api_key:
        print("[sam] ⚠  No API key configured.")
        print("[sam]    To fix: log in at sam.gov with safa@idcg.co.in")
        print("[sam]    → Profile → API Keys → Generate Public Key")
        print("[sam]    → Add to config/.env:  SAM_API_KEY=<your_key>")
        print("[sam]    Attempting to use cached data...")
        raw_opps = _load_cache()
        if not raw_opps:
            print("[sam]    No cache available — skipping this run.")
            return [], []
    else:
        # ── 2. Probe key ──────────────────────────────────────────────────────
        ok, msg = _probe_key(api_key)
        if not ok:
            print(f"[sam] ⚠  API key check failed: {msg}")
            print("[sam]    → Skipping this run (not marking unstable). Regenerate key at sam.gov → Profile → API Keys")
            return [], []

        # ── 3. Fetch fresh data ───────────────────────────────────────────
        print(f"[sam] API key OK — fetching opportunities (last {POSTED_DAYS} days)...")
        raw_opps = _fetch_all_opportunities(api_key)
        if not raw_opps:
            print("[sam]    API returned 0 rows — skipping this run (not unstable).")
            return [], []

        print(f"[sam] Total unique opportunities fetched: {len(raw_opps)}")
        if raw_opps:
            _save_cache(raw_opps)

    # ── 4. Build rows + dedup ──────────────────────────────────────────────────
    new_tenders: list[dict] = []
    all_rows:    list[dict] = []

    for opp in raw_opps:
        row = _build_row(opp)
        if row is None:
            continue

        tid = row.pop("tender_id")
        all_rows.append(row)

        if check_if_new(tid):
            mark_as_seen(tid, row["title"], "SAM.gov", row["link"])
            new_tenders.append({
                "title":    row["title"],
                "deadline": row["deadline"],
                "value":    row["agency"] or "US Federal",
                "url":      row["link"],
            })

    relevant = sum(1 for r in all_rows if r.get("relevance"))
    print(f"[sam] {len(all_rows)} rows passed filter ({relevant} relevant, {len(new_tenders)} NEW)")

    if all_rows:
        _save_excel(all_rows)
    else:
        print("[sam] No rows to save — filters may be too strict or no results for these queries")

    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nResult: {len(rows)} total, {len(new)} new")
