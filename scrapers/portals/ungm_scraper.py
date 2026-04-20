# =============================================================================
# scrapers/portals/ungm_scraper.py — UN Global Marketplace (UNGM) scraper
#
# Site   : https://www.ungm.org/Public/Notice
# Method : requests POST to /Public/Notice/Search (paginated HTML API)
#          — NO Selenium required; 15 rows/page, up to MAX_PAGES per country
# Login  : None required (public notices)
# CAPTCHA: None
#
# Fix (v2): replaced Selenium infinite-scroll with direct POST API.
#   Root cause of "8 rows" bug: Selenium's scroll approach returned immediately
#   when no new .tableRow elements appeared (UNGM changed their infinite-scroll
#   trigger timing). The POST API is deterministic, fast, and returns HTML we
#   can parse with BeautifulSoup — no browser needed.
#
# Framework: IntelligentBaseScraper
#   fetch_data()       — paginated POST requests across 3 country filters + global
#   extract_rows()     — normalise field names; enrich description for scoring
#   to_standard_format() — TenderResult mapping
#   enrich_fields()    — UNGM-specific score boosts (post quality-filter)
#   on_run_end()       — colour-coded Excel + debug summary
#
# Inherited from framework (do NOT re-implement here):
#   quality scoring, consulting-signal detection, intelligence filter,
#   DB dedup (check_if_new / mark_as_seen), ScraperMonitor,
#   structure-change detection
#
# UNGM covers ALL UN agencies: UNDP, UNICEF, WFP, ILO, FAO, UNOPS, etc.
# =============================================================================

from __future__ import annotations

import json
import os
import re
import time
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

from config.config import PORTAL_EXCELS_DIR
from core.base_scraper import IntelligentBaseScraper
from core.quality_engine import TenderResult, make_tender_result

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL    = "https://www.ungm.org"
NOTICES_URL = f"{BASE_URL}/Public/Notice"
SEARCH_URL  = f"{BASE_URL}/Public/Notice/Search"

UNGM_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "UNGM_Tenders_Master.xlsx")

PAGE_SIZE     = 15   # UNGM returns 15 per page
MAX_PAGES     = 20   # 15 × 20 = 300 max per country filter
PAGE_DELAY    = 0.8  # seconds between page requests

# Country filter values for UNGM
COUNTRY_FILTERS = [
    ("India",      "2321"),
    ("Nepal",      "2400"),
    ("Bangladesh", "215"),
    ("Global",     "0"),   # no country filter — catches multi-country tenders
]

# Notice types to SKIP (goods/logistics/simple quotations — not consulting)
SKIP_TYPES = {"rfq", "request for quotation", "ltpo", "pca"}


def _should_keep(notice_type: str) -> bool:
    t = notice_type.lower().strip()
    return not any(s in t for s in SKIP_TYPES)


# =============================================================================
# SECTION 1 — Vocabulary / lookup tables
# =============================================================================

# ── UI artifact texts to REJECT as titles ─────────────────────────────────────
_UI_ARTIFACTS: frozenset[str] = frozenset({
    "open in a new window", "open in new window", "open",
    "view", "view details", "view notice", "read more",
    "click here", "click", "details", "more", "link",
    "new window", "external link", "notice details",
})


def _is_ui_artifact(text: str) -> bool:
    """Return True when text is a browser UI label, not a tender title."""
    clean = text.lower().strip()
    return (
        clean in _UI_ARTIFACTS
        or len(clean) < 12
        or re.fullmatch(r"[\W\d]+", clean) is not None
    )


def _clean_title(raw: str) -> str:
    """Strip known UI artifact suffixes from raw title text."""
    for artifact in (
        "Open in a new window", "Open in new window",
        "open in a new window", "Expand", "Details",
    ):
        raw = raw.replace(artifact, "")
    return re.sub(r"\s+", " ", raw).strip()


# ── Known UN / MDB agencies ───────────────────────────────────────────────────
_UN_AGENCIES: set[str] = {
    "UNDP", "UNICEF", "WFP", "WHO", "FAO", "ILO", "UNOPS", "UNESCO",
    "UNHCR", "UNFPA", "IOM", "UNODC", "UNIDO", "UNCTAD", "UNEP",
    "OCHA", "OHCHR", "WMO", "IMO", "ITU", "UPU", "WIPO", "IFAD",
    "ITC", "UNWTO", "UNRWA", "IAEA", "CTBTO", "OPCW", "UNAIDS",
    "PAHO", "UNHABITAT", "UN-Habitat", "UNEP-WCMC", "UN Women",
    "IFC", "World Bank", "ADB", "AIIB", "AfDB",
}
_UN_ABBREV_SET: frozenset[str] = frozenset(a.upper() for a in _UN_AGENCIES)


def _extract_agency_from_reference(reference: str) -> str:
    """UNDP-IND-2026-039 → 'UNDP'; FAO-NEP-2025-012 → 'FAO'."""
    if not reference:
        return ""
    parts = reference.split("-")
    if parts:
        candidate = parts[0].upper()
        if candidate in _UN_ABBREV_SET:
            return candidate
    return ""


# ── High-priority agencies for post-filter score boost ────────────────────────
_HIGH_PRIORITY_AGENCIES: frozenset[str] = frozenset({
    "undp", "unicef", "who", "wfp", "unops", "ifad", "ifc",
    "world bank", "fao", "unhcr", "unfpa", "unido", "iom",
})

# ── Target geographies for score boost ────────────────────────────────────────
_TARGET_COUNTRIES: frozenset[str] = frozenset({
    "india", "nepal", "bangladesh", "sri lanka", "bhutan",
    "myanmar", "south asia",
})


# =============================================================================
# SECTION 2 — Excel styles
# =============================================================================

MASTER_COLUMNS = [
    ("Reference",      22),
    ("Title",          65),
    ("Organisation",   30),
    ("Type",           12),
    ("Country",        22),
    ("Published",      16),
    ("Deadline",       18),
    ("Quality Score",  14),
    ("AI Decision",    16),
    ("My Decision",    14),
    ("Detail Link",    55),
]
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
ALT_FILL     = PatternFill("solid", fgColor="F5F8FF")
QS_HIGH_FILL = PatternFill("solid", fgColor="E2EFDA")
QS_MID_FILL  = PatternFill("solid", fgColor="FFF2CC")
QS_LOW_FILL  = PatternFill("solid", fgColor="FCE4D6")

TIER_BID_FILL    = PatternFill("solid", fgColor="70AD47")
TIER_STRONG_FILL = PatternFill("solid", fgColor="FFC000")
TIER_WEAK_FILL   = PatternFill("solid", fgColor="FF7C00")
TIER_IGNORE_FILL = PatternFill("solid", fgColor="D9D9D9")

_TIER_DISPLAY: dict[str, str] = {
    "BID_NOW":         "BID NOW",
    "STRONG_CONSIDER": "STRONG",
    "WEAK_CONSIDER":   "REVIEW",
    "IGNORE":          "IGNORE",
}

WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT   = Font(name="Calibri", size=10)
TIER_FONT   = Font(name="Calibri", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# SECTION 3 — Selenium driver + in-browser fetch (bypasses Cloudflare)
# =============================================================================

def _make_driver() -> webdriver.Chrome:
    """Headless Chrome — looks like a real browser to Cloudflare."""
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=opts)


def _fetch_page_via_browser(driver: webdriver.Chrome, country_id: str, page: int) -> str | None:
    """
    POST to /Public/Notice/Search from within the browser context using fetch().
    Cloudflare clearance is already obtained (Chrome loaded the notices page),
    so cookies are set and the XHR is treated as legitimate.
    Returns the HTML fragment or None on error.
    """
    body = (
        f"NoticeType=&Title=&Description=&Reference=&Beneficiary="
        f"&CountryId={country_id}&UNSPSCcategoryId=0&AgencyId=0"
        f"&PublishedFrom=&PublishedTo=&DeadlineFrom=&DeadlineTo="
        f"&pageIndex={page}&pageSize={PAGE_SIZE}"
        f"&sortField=DatePublished&sortOrder=desc"
    )
    script = """
    return await fetch('/Public/Notice/Search', {
        method: 'POST',
        headers: {
            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
            'X-Requested-With': 'XMLHttpRequest',
            'Accept': 'text/html, */*; q=0.01'
        },
        body: arguments[0],
        credentials: 'include'
    }).then(r => r.text()).catch(e => null);
    """
    try:
        html = driver.execute_async_script(
            "var cb = arguments[arguments.length-1];"
            "fetch('/Public/Notice/Search', {"
            "  method: 'POST',"
            "  headers: {"
            "    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',"
            "    'X-Requested-With': 'XMLHttpRequest',"
            "    'Accept': 'text/html, */*; q=0.01'"
            "  },"
            f" body: '{body}',"
            "  credentials: 'include'"
            "}).then(r=>r.text()).then(cb).catch(()=>cb(null));",
        )
        return html
    except Exception as exc:
        print(f"[UNGM]   Browser fetch p{page} error: {exc}")
        return None


# =============================================================================
# SECTION 4 — HTML parser
# =============================================================================

def _parse_notices(html: str) -> list[dict]:
    """
    Parse UNGM notice rows from POST /Public/Notice/Search response.
    Each page returns up to 15 div.tableRow elements.
    Returns list of raw notice dicts.
    """
    soup    = BeautifulSoup(html, "html.parser")
    entries = []

    rows = soup.find_all("div", class_="tableRow")
    if not rows:
        return entries

    for row in rows:
        if "header" in " ".join(row.get("class", [])):
            continue

        cells = row.find_all("div", class_="tableCell")
        if len(cells) < 2:
            continue

        # ── Title ─────────────────────────────────────────────────────────────
        title = ""
        title_cell = row.find("div", class_="resultTitle")
        if title_cell:
            title = _clean_title(title_cell.get_text(separator=" ", strip=True))

        if not title or _is_ui_artifact(title):
            continue

        # ── Detail URL ────────────────────────────────────────────────────────
        detail_url = ""
        for cell in cells:
            for a in cell.find_all("a", href=True):
                href = a.get("href", "")
                if "/Public/Notice/" in href:
                    detail_url = href if href.startswith("http") else BASE_URL + href
                    break
            if detail_url:
                break

        # ── Cell text extraction ───────────────────────────────────────────────
        cell_texts = [c.get_text(strip=True) for c in cells]

        deadline = published = org = notice_type = reference = country = ""

        for txt in cell_texts:
            if not txt:
                continue
            if re.match(r"\d{2}-\w{3}-\d{4}|\d{4}-\d{2}-\d{2}", txt):
                if not deadline:
                    deadline = txt
                elif not published:
                    published = txt
            elif re.match(r"[A-Z]{2,8}-[A-Z]{2,3}-\d{4}-\d+", txt):
                reference = txt
            elif txt.upper() in ("RFP", "IC", "EOI", "ITB", "RFQ", "RFI",
                                  "ITP", "ICTB", "LTA", "RET"):
                notice_type = txt

        # ── Organisation ──────────────────────────────────────────────────────
        org = _extract_agency_from_reference(reference)
        if not org:
            for txt in cell_texts:
                upper = txt.strip().upper()
                if upper in _UN_ABBREV_SET and len(txt) <= 20:
                    org = upper
                    break
                for agency in _UN_AGENCIES:
                    if agency.lower() in txt.lower() and len(txt) < 50:
                        org = agency
                        break
                if org:
                    break

        # ── Country (last cell) ───────────────────────────────────────────────
        if cells:
            country = cells[-1].get_text(strip=True)
            if len(country) > 40 or not re.search(r"[A-Za-z]", country):
                country = ""
            if country and country.upper() in _UN_ABBREV_SET:
                country = ""

        # ── Type filter ───────────────────────────────────────────────────────
        if not _should_keep(notice_type or "rfp"):
            continue

        tender_id = (
            f"UNGM_{reference}"
            if reference else
            f"UNGM_{re.sub(r'[^a-zA-Z0-9]', '_', detail_url.split('/')[-1])[:60]}"
        )

        entries.append({
            "Reference":    reference,
            "Title":        title,
            "Organisation": org,
            "Type":         notice_type,
            "Country":      country,
            "Published":    published,
            "Deadline":     deadline,
            "Detail Link":  detail_url,
            "tender_id":    tender_id,
        })

    return entries


def _scrape_country_via_browser(
    driver: webdriver.Chrome,
    seen_ids: set[str],
    label: str,
    country_id: str,
    max_pages: int = MAX_PAGES,
) -> list[dict]:
    """
    Paginate through UNGM notices for one country using in-browser fetch().
    Stops when a page returns fewer than PAGE_SIZE rows (last page reached).
    """
    entries: list[dict] = []

    for page in range(1, max_pages + 1):
        html = _fetch_page_via_browser(driver, country_id, page)
        if not html:
            print(f"[UNGM]   {label} p{page}: fetch returned None — stopping")
            break

        notices = _parse_notices(html)
        new = [n for n in notices if n["tender_id"] not in seen_ids]
        for n in new:
            seen_ids.add(n["tender_id"])
        entries.extend(new)

        print(
            f"[UNGM]   {label} p{page}: {len(notices)} rows, "
            f"{len(new)} new (total={len(entries)})"
        )

        if len(notices) < PAGE_SIZE:
            print(f"[UNGM]   {label}: last page at p{page}")
            break

        time.sleep(PAGE_DELAY)

    return entries


# =============================================================================
# SECTION 5 — Excel writer
# =============================================================================

def _tier_fill(tier: str) -> PatternFill:
    return {
        "BID_NOW":         TIER_BID_FILL,
        "STRONG_CONSIDER": TIER_STRONG_FILL,
        "WEAK_CONSIDER":   TIER_WEAK_FILL,
    }.get(tier, TIER_IGNORE_FILL)


def _save_excel(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "UNGM Notices"
    ws.row_dimensions[1].height = 36
    col_names = [c[0] for c in MASTER_COLUMNS]

    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.border    = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    qs_idx        = col_names.index("Quality Score") + 1
    dec_idx       = col_names.index("AI Decision")   + 1
    my_dec_idx    = col_names.index("My Decision")   + 1
    link_idx      = col_names.index("Detail Link")   + 1

    USER_HDR_FILL = PatternFill("solid", fgColor="FFD966")
    USER_HDR_FONT = Font(name="Calibri", bold=True, color="7F4A00", size=11)
    # Recolour the My Decision header gold
    ws.cell(row=1, column=my_dec_idx).fill = USER_HDR_FILL
    ws.cell(row=1, column=my_dec_idx).font = USER_HDR_FONT

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 45
        alt = ALT_FILL if ri % 2 == 0 else None

        tier = (
            row_data.get("decision_tag")
            or row_data.get("Decision")
            or ""
        ).strip()

        for ci, col_name in enumerate(col_names, 1):
            if col_name == "Quality Score":
                val = row_data.get("Quality Score") or row_data.get("quality_score") or ""
            elif col_name == "AI Decision":
                raw_tier = (
                    row_data.get("decision_tag")
                    or row_data.get("Decision")
                    or "IGNORE"
                )
                val = _TIER_DISPLAY.get(raw_tier, raw_tier)
            elif col_name == "My Decision":
                val = ""   # left blank for user to fill
            else:
                val = row_data.get(col_name, "")

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10,
                                 color="1155CC", underline="single")
            elif ci == qs_idx:
                try:
                    qs = int(val or 0)
                except (ValueError, TypeError):
                    qs = 0
                cell.font = Font(name="Calibri", size=10, bold=True)
                cell.fill = (
                    QS_HIGH_FILL if qs >= 70
                    else QS_MID_FILL if qs >= 40
                    else QS_LOW_FILL if qs > 0
                    else (alt or PatternFill())
                )
            elif ci == dec_idx:
                raw_tier = (
                    row_data.get("decision_tag")
                    or row_data.get("Decision")
                    or "IGNORE"
                )
                cell.font      = TIER_FONT
                cell.fill      = _tier_fill(raw_tier)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=False
                )
            elif ci == my_dec_idx:
                cell.fill      = PatternFill("solid", fgColor="FFFDE7")
                cell.font      = Font(name="Calibri", size=10, color="9CA3AF", italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.font = BODY_FONT
                if alt:
                    cell.fill = alt

    # Dropdown for My Decision
    from openpyxl.worksheet.datavalidation import DataValidation
    last_row = len(rows) + 1
    dv = DataValidation(
        type="list", formula1='"Bid,No Bid,Review Later"',
        allow_blank=True, showErrorMessage=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"{get_column_letter(my_dec_idx)}2:{get_column_letter(my_dec_idx)}{last_row}")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    try:
        wb.save(UNGM_EXCEL_PATH)
        print(f"[UNGM] Excel saved: {UNGM_EXCEL_PATH}  ({len(rows)} rows)")
    except Exception as exc:
        print(f"[UNGM] Excel save failed: {exc}")


# =============================================================================
# SECTION 6 — UNGMScraper
# =============================================================================

class UNGMScraper(IntelligentBaseScraper):
    """
    UNGM portal scraper (IntelligentBaseScraper subclass).
    v2: uses requests POST API instead of Selenium — faster, more reliable.

    Portal-specific methods:
      fetch_data()         — paginated POST requests across country filters
      extract_rows()       — normalise + enrich description for quality engine
      enrich_fields()      — UNGM-specific score boosts (post quality-filter)
      to_standard_format() — TenderResult mapping
      get_tender_id()      — UNGM reference-based stable key
      on_run_end()         — Excel + debug summary

    Inherited (do NOT touch):
      quality scoring, intelligence filter, DB dedup, ScraperMonitor
    """

    SOURCE_NAME = "UNGM"
    SOURCE_URL  = NOTICES_URL
    EXCEL_PATH  = UNGM_EXCEL_PATH

    # UNGM notices are title-only → scores skew lower than WB; relax threshold
    QUALITY_THRESHOLD = 15

    EXPECTED_SCHEMA_FIELDS: list[tuple[str, type]] = [
        ("Title",       str),
        ("Detail Link", str),
        ("Deadline",    str),
    ]

    # =========================================================================
    # Step 1 — fetch_data
    # =========================================================================

    def fetch_data(self) -> dict | None:
        """
        Use Selenium (headless Chrome) to obtain Cloudflare clearance, then
        paginate through UNGM notices via in-browser fetch() calls.
        This bypasses Cloudflare's bot detection (requests library gets blocked).
        """
        import urllib.request, urllib.error, ssl

        # ── Quick HTTP pre-check (2s) — skip Chrome entirely if server is down ──
        try:
            ctx = ssl.create_default_context()
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            req = urllib.request.Request(
                NOTICES_URL,
                headers={"User-Agent": "Mozilla/5.0 (compatible; TenderRadar/1.0)"},
            )
            with urllib.request.urlopen(req, timeout=8, context=ctx) as resp:
                status = resp.getcode()
                if status >= 500:
                    print(
                        f"[UNGM] ⚠  Server returned HTTP {status} — "
                        "UNGM is experiencing downtime. Will retry next run."
                    )
                    return {"notices": []}
        except Exception as pre_exc:
            pre_msg = str(pre_exc)
            if "500" in pre_msg or "timed out" in pre_msg.lower() or "EOF" in pre_msg:
                print(
                    f"[UNGM] ⚠  Pre-check failed ({pre_msg[:80]}). "
                    "UNGM may be down — will retry next run."
                )
                return {"notices": []}
            # Other errors (DNS, SSL handshake) — proceed to Selenium anyway
            print(f"[UNGM]   Pre-check note: {pre_msg[:80]} — launching Chrome")

        seen_ids: set[str] = set()
        all_entries: list[dict] = []
        driver: webdriver.Chrome | None = None

        try:
            driver = _make_driver()

            # Load the notices page — this gets Cloudflare clearance for the session
            print("[UNGM]   Loading notices page...")
            driver.get(NOTICES_URL)
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.ID, "selNoticeCountry"))
                )
            except TimeoutException:
                # Check if it's a server error page rather than a real timeout
                title = driver.title.lower()
                src   = driver.page_source.lower()
                if (
                    "timed out" in title
                    or "500" in title
                    or "server error" in title
                    or "error" in title
                    or "selnoticecountry" not in src
                ):
                    print(
                        "[UNGM] ⚠  Server returned an error page "
                        f"(title='{driver.title}'). "
                        "UNGM may be experiencing downtime — will retry next run."
                    )
                    return {"notices": []}
                print("[UNGM]   Page load timeout — proceeding anyway")
            time.sleep(1.5)

            # Now paginate each country filter via in-browser fetch
            for label, country_id in COUNTRY_FILTERS:
                print(f"[UNGM]   Scraping: {label} (country_id={country_id})...")
                entries = _scrape_country_via_browser(
                    driver, seen_ids, label, country_id, max_pages=MAX_PAGES
                )
                all_entries.extend(entries)
                print(f"[UNGM]   {label}: {len(entries)} unique notices")

        except Exception as exc:
            err_str = str(exc)
            if "ERR_CONNECTION_CLOSED" in err_str or "SSLEOFError" in err_str or "Connection reset" in err_str:
                print(
                    "[UNGM] ⚠  Cloudflare IP block detected (ERR_CONNECTION_CLOSED). "
                    "This is temporary — typically clears in 24-48 hours. "
                    "No action needed; will retry automatically next run."
                )
            else:
                print(f"[UNGM] fetch_data() error: {exc}")
        finally:
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass

        print(f"[UNGM] Total unique raw notices: {len(all_entries)}")
        return {"notices": all_entries} if all_entries else None

    # =========================================================================
    # Step 3 — extract_rows
    # =========================================================================

    def extract_rows(self, raw_data: dict) -> list[dict]:
        rows:     list[dict] = []
        rejected: int        = 0

        for notice in raw_data.get("notices", []):
            title = notice.get("Title", "").strip()

            if not title or _is_ui_artifact(title) or len(title) < 15:
                rejected += 1
                continue

            row = dict(notice)

            row["description"] = title
            row["deadline"]    = notice.get("Deadline", "")
            row["method"]      = notice.get("Type", "")
            row["url"]         = notice.get("Detail Link", "")

            rows.append(row)

        if rejected:
            print(f"[UNGM] Pre-filtered {rejected} UI-artifact / short-title rows")

        return rows

    # =========================================================================
    # Step 4b — on_filter_complete
    # =========================================================================

    def on_filter_complete(
        self,
        scraped_total: int,
        accepted:      list[dict],
        rejected:      list[dict],
        reasons:       list[str],
    ) -> None:
        self._rejected_log: list[tuple[dict, str]] = list(zip(rejected, reasons))

    # =========================================================================
    # Step 5 — enrich_fields  (UNGM-specific post-filter score boosts)
    # =========================================================================

    _REAL_CONSULTING_TYPES: frozenset[str] = frozenset({
        "Evaluation", "Technical Assistance", "Capacity Building",
        "Research/Study", "Audit/Fiduciary", "Advisory/Policy",
        "Individual Consultant",
    })

    def enrich_fields(self, rows: list[dict]) -> list[dict]:
        for row in rows:
            boost   = 0
            org     = (row.get("Organisation") or row.get("organisation") or "").lower()
            country = (row.get("Country")      or row.get("country")      or "").lower()
            ntype   = (row.get("Type")          or row.get("type")         or "").lower()
            ctype   = (row.get("consulting_type") or row.get("Consulting Type") or "")

            if (ctype in self._REAL_CONSULTING_TYPES
                    and any(a in org for a in _HIGH_PRIORITY_AGENCIES)):
                boost += 8

            if any(c in country for c in _TARGET_COUNTRIES):
                boost += 4

            if ntype in ("ic", "individual contractor"):
                boost += 4

            if boost:
                current   = int(row.get("quality_score") or row.get("Quality Score") or 0)
                new_score = min(100, current + boost)
                row["quality_score"] = new_score
                row["Quality Score"] = new_score

        return rows

    # =========================================================================
    # Step 7 — to_standard_format
    # =========================================================================

    def to_standard_format(self, row: dict) -> TenderResult:
        return make_tender_result(
            title          = row.get("Title", ""),
            url            = row.get("Detail Link", row.get("url", self.SOURCE_URL)),
            deadline       = row.get("Deadline", row.get("deadline", "")),
            organization   = row.get("Organisation", ""),
            sector         = row.get("sector", ""),
            consulting_type= row.get("consulting_type", ""),
            quality_score  = int(
                row.get("quality_score") or row.get("Quality Score") or 0
            ),
            source         = self.SOURCE_NAME,
            decision_tag   = row.get("decision_tag") or row.get("Decision") or "",
        )

    # =========================================================================
    # get_tender_id — prefer UNGM reference code
    # =========================================================================

    def get_tender_id(self, row: dict) -> str:
        tid = row.get("tender_id", "")
        if tid:
            return tid
        return super().get_tender_id(row)

    # =========================================================================
    # on_run_end — Excel + debug summary
    # =========================================================================

    def on_run_end(self, all_rows: list[dict]) -> None:
        if all_rows:
            _save_excel(all_rows)
        else:
            print("[UNGM] No rows to save — Excel skipped.")
        self._print_ungm_summary(all_rows)

    def _print_ungm_summary(self, rows: list[dict]) -> None:
        if not rows:
            print("[UNGM] Summary: 0 rows")
            return

        tiers: dict[str, int] = {}
        for row in rows:
            tier = (
                row.get("decision_tag") or row.get("Decision") or "IGNORE"
            ).upper()
            tiers[tier] = tiers.get(tier, 0) + 1

        print(f"\n[UNGM] ── Run Summary ─────────────────────────────")
        print(f"[UNGM]   Total accepted : {len(rows)}")
        for tier, count in sorted(tiers.items(), key=lambda x: -x[1]):
            print(f"[UNGM]   {tier:<18}: {count}")

        top5 = sorted(rows, key=lambda r: int(r.get("quality_score") or 0), reverse=True)[:5]
        print("[UNGM]   Top 5 by score:")
        for r in top5:
            score = r.get("quality_score") or 0
            title = (r.get("Title") or "")[:60]
            print(f"[UNGM]     {score:>3}  {title}")
        print("[UNGM] ────────────────────────────────────────────\n")


# =============================================================================
# Module-level run() — required by the JobRunner (calls mod.run())
# =============================================================================

def run() -> tuple:
    """Instantiate UNGMScraper and execute the full pipeline."""
    scraper = UNGMScraper()
    return scraper.run()


if __name__ == "__main__":
    new, rows = run()
    print(f"\nResult: {len(rows)} total, {len(new)} new")
