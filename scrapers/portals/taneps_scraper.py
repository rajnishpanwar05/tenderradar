# =============================================================================
# taneps_scraper.py — Tanzania Procurement Portal (NeST / PPRA)
#
# Primary  : https://nest.go.tz/           (National e-Procurement System)
# Authority: https://www.ppra.go.tz/       (Tanzania PPRA — regulatory site)
# Legacy   : https://www.taneps.go.tz/     (old TANEPS — unreachable 2026-03)
#
# Technology: Angular SPA + GraphQL backend (microservices: nest-uaa, nest-app)
#
# Status (2026-04):
#   NeST (nest.go.tz) is Tanzania's official national e-procurement system.
#   It requires:
#     a) Selenium to render the Angular SPA (all routes return same HTML shell)
#     b) A registered TENDERER account to query tenders via GraphQL
#
#   The PPRA regulatory site (www.ppra.go.tz) has no active tender tables —
#   it links to NeST for all procurement activity.
#
#   Until a NeST account is registered, the pipeline returns [], [] gracefully.
#   Once credentials are added as NEST_USER / NEST_PASS in .env, this scraper
#   will authenticate via Selenium and scrape the tender-lookup page.
#
# Run:
#   python3 main.py --taneps
# =============================================================================

import os
import re
import time
import logging

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

import sys
sys.path.insert(0, os.path.expanduser("~/tender_system"))

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance, title_is_relevant

logger = logging.getLogger("tenderradar.taneps")

# SCRAPER_META — required for auto-discovery by core/registry.py
SCRAPER_META = {
    "flag":        "taneps",
    "label":       "Tanzania NeST (PPRA)",
    "group":       "selenium",    # Angular SPA — requires Selenium
    "timeout":     600,
    "max_retries": 1,
    "auto":        True,
}

# ── Constants ─────────────────────────────────────────────────────────────────
TANEPS_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "TANEPS_Tenders_Master.xlsx")

NEST_BASE_URL  = "https://nest.go.tz"
NEST_LOGIN_URL = f"{NEST_BASE_URL}/login"
NEST_TENDER_URL = f"{NEST_BASE_URL}/tender-lookup"

# GraphQL endpoint (authenticated queries only)
NEST_GRAPHQL_URL = f"{NEST_BASE_URL}/graphql"

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/120.0.0.0 Safari/537.36"),
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "en-US,en;q=0.9",
}
TIMEOUT   = 40
MAX_PAGES = 15

# Credentials (optional — scraper degrades gracefully if not set)
NEST_USER = os.environ.get("NEST_USER", "").strip()
NEST_PASS = os.environ.get("NEST_PASS", "").strip()

# ── Excel styles ───────────────────────────────────────────────────────────────
DARK_BLUE  = PatternFill("solid", fgColor="1F3864")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
WHITE      = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT  = Font(name="Calibri", size=10)
THIN       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

COLUMNS = [
    ("Ref No.",    20),
    ("Title",      60),
    ("Entity",     35),
    ("Category",   18),
    ("Deadline",   16),
    ("Link",       50),
    ("Relevance",  30),
]


def _init_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "TANEPS Tenders"
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        c = ws.cell(1, col_idx, col_name)
        c.fill      = DARK_BLUE
        c.font      = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = THIN
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = col_width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(COLUMNS)).column_letter}1"
    return wb, ws


def _save_excel(rows: list):
    if os.path.exists(TANEPS_EXCEL_PATH):
        wb = load_workbook(TANEPS_EXCEL_PATH)
        ws = wb.active
    else:
        wb, ws = _init_excel()

    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing.add(str(row[0]))

    for r in rows:
        key = r.get("ref_no", "")
        if key in existing:
            continue
        row_idx = ws.max_row + 1
        fill    = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
        ws.append([
            r["ref_no"], r["title"], r["entity"], r["category"],
            r["deadline"], r["link"], r["relevance"],
        ])
        for col_i in range(1, len(COLUMNS) + 1):
            c = ws.cell(row_idx, col_i)
            c.font      = CELL_FONT
            c.fill      = fill
            c.border    = THIN
            c.alignment = Alignment(wrap_text=True, vertical="top")
        existing.add(key)

    wb.save(TANEPS_EXCEL_PATH)


# =============================================================================
# Selenium-based scraper for NeST tender-lookup page
# =============================================================================

def _make_driver():
    """Headless Chrome driver for the Angular NeST SPA."""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    try:
        from webdriver_manager.chrome import ChromeDriverManager
        service = Service(ChromeDriverManager().install())
    except Exception:
        service = Service()

    opts = Options()
    opts.add_argument("--headless")       # classic headless — stable on macOS
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1400,900")
    opts.add_argument("--disable-gpu")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    _chrome_mac = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    if os.path.exists(_chrome_mac):
        opts.binary_location = _chrome_mac
    return webdriver.Chrome(service=service, options=opts)


def _selenium_login(driver) -> bool:
    """
    Log into NeST with TENDERER credentials.
    Returns True on success.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    if not NEST_USER or not NEST_PASS:
        return False

    try:
        driver.get(NEST_LOGIN_URL)
        wait = WebDriverWait(driver, 30)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='text'], input[formcontrolname]")))
        time.sleep(2)

        # Angular reactive form fields
        user_field = None
        for sel in ["input[formcontrolname='username']", "input[type='text']", "input[name='username']"]:
            try:
                user_field = driver.find_element(By.CSS_SELECTOR, sel)
                break
            except Exception:
                continue

        pass_field = None
        for sel in ["input[formcontrolname='password']", "input[type='password']"]:
            try:
                pass_field = driver.find_element(By.CSS_SELECTOR, sel)
                break
            except Exception:
                continue

        if not user_field or not pass_field:
            logger.warning("[taneps] Could not find login fields on NeST")
            return False

        user_field.send_keys(NEST_USER)
        pass_field.send_keys(NEST_PASS)

        # Submit
        for sel in ["button[type='submit']", "input[type='submit']"]:
            try:
                driver.find_element(By.CSS_SELECTOR, sel).click()
                break
            except Exception:
                continue

        time.sleep(5)
        return "/login" not in driver.current_url

    except Exception as exc:
        logger.error("[taneps] NeST login failed: %s", exc)
        return False


def _parse_tender_cards(html: str, base_url: str) -> list:
    """
    Parse tender cards from the Angular-rendered tender-lookup page.
    Angular renders mat-card or similar components with tender data.
    """
    soup = BeautifulSoup(html, "html.parser")
    rows = []

    # Try Material Design card elements
    cards = (
        soup.find_all("mat-card") or
        soup.find_all("div", class_=re.compile(r"card|tender-item|procurement-item")) or
        soup.find_all("tr", class_=re.compile(r"tender|procurement"))
    )

    for card in cards:
        text = card.get_text(" ", strip=True)
        if len(text) < 20:
            continue

        title = ""
        # Try heading tags first
        for tag in card.find_all(["h1", "h2", "h3", "h4", "strong", "b"]):
            t = tag.get_text(strip=True)
            if len(t) > 10 and len(t) < 300:
                title = t
                break

        if not title:
            # Try first long text block
            for el in card.find_all(["span", "p", "td"]):
                t = el.get_text(strip=True)
                if len(t) > 15:
                    title = t[:200]
                    break

        if not title:
            continue

        link_tag = card.find("a", href=True)
        href = link_tag["href"] if link_tag else ""
        full_link = href if href.startswith("http") else (base_url + href if href else "")

        rows.append({
            "ref_no":   "",
            "title":    title,
            "entity":   "",
            "category": "",
            "deadline": "",
            "link":     full_link,
        })

    # Fallback: table-based layout
    if not rows:
        for table in soup.find_all("table"):
            trs = table.find_all("tr")
            for tr in trs[1:]:
                tds = tr.find_all("td")
                if len(tds) < 2:
                    continue
                texts = [td.get_text(strip=True) for td in tds]
                title = texts[1] if len(texts) > 1 else texts[0]
                if not title or len(title) < 5:
                    continue
                a_tag = tr.find("a", href=True)
                href  = a_tag["href"] if a_tag else ""
                rows.append({
                    "ref_no":   texts[0],
                    "title":    title,
                    "entity":   texts[2] if len(texts) > 2 else "",
                    "category": texts[3] if len(texts) > 3 else "",
                    "deadline": texts[4] if len(texts) > 4 else "",
                    "link":     href if href.startswith("http") else base_url + href,
                })

    return rows


def _scrape_with_selenium() -> list:
    """
    Use Selenium to render the Angular NeST SPA and extract tender data.
    Returns a list of raw tender dicts.
    """
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    driver = None
    all_rows = []

    try:
        driver = _make_driver()

        logged_in = _selenium_login(driver)
        if not logged_in:
            logger.warning("[taneps] Not logged in — NeST requires auth. "
                           "Set NEST_USER / NEST_PASS in .env")
            print("[taneps] NeST requires TENDERER account. Add NEST_USER/NEST_PASS to .env", flush=True)
            return []

        print(f"[taneps] Navigating to {NEST_TENDER_URL}…", flush=True)
        driver.get(NEST_TENDER_URL)

        wait = WebDriverWait(driver, 30)
        time.sleep(5)   # let Angular render

        for page_num in range(1, MAX_PAGES + 1):
            html = driver.page_source
            rows = _parse_tender_cards(html, NEST_BASE_URL)
            print(f"[taneps] Page {page_num}: {len(rows)} tender cards found", flush=True)
            all_rows.extend(rows)

            if not rows:
                break

            # Try to click next page button
            try:
                next_btn = driver.find_element(
                    By.CSS_SELECTOR,
                    "button[aria-label*='next' i], button.mat-paginator-navigation-next, "
                    "[class*='next']:not([disabled])"
                )
                if next_btn.is_enabled():
                    next_btn.click()
                    time.sleep(3)
                else:
                    break
            except Exception:
                break

    except Exception as exc:
        logger.error("[taneps] Selenium scrape failed: %s", exc)
        print(f"[taneps] Error: {exc}", flush=True)
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass

    return all_rows


# =============================================================================
# Main run()
# =============================================================================

def run():
    print("[taneps] Tanzania NeST (PPRA) scraper starting…", flush=True)

    if not NEST_USER or not NEST_PASS:
        print(
            "[taneps] NeST portal requires a registered TENDERER account.\n"
            "         Register free at: https://nest.go.tz/tenderer_registration\n"
            "         Then add to ~/tender_system/config/.env:\n"
            "           NEST_USER=your@email.com\n"
            "           NEST_PASS=your_password",
            flush=True,
        )
        return [], []

    raw_rows = _scrape_with_selenium()

    all_rows    = []
    new_tenders = []
    seen_ids    = set()

    for r in raw_rows:
        ref_no    = r.get("ref_no") or re.sub(r"[^A-Za-z0-9]", "_", r["title"])[:50]
        tender_id = f"TANEPS_{ref_no}"

        if tender_id in seen_ids:
            continue
        seen_ids.add(tender_id)

        relevance = score_relevance(r["title"])
        if not relevance and not title_is_relevant(r["title"]):
            continue

        r["relevance"] = relevance
        r["tender_id"] = tender_id
        all_rows.append(r)

        if check_if_new(tender_id):
            mark_as_seen(tender_id, r["title"], "TANEPS", r.get("link") or NEST_BASE_URL)
            new_tenders.append(r)

    _save_excel(all_rows)
    print(f"[taneps] Done — {len(all_rows)} relevant, {len(new_tenders)} NEW", flush=True)
    return new_tenders, all_rows


if __name__ == "__main__":
    new, all_ = run()
    print(f"\n✅  TANEPS: {len(new)} new / {len(all_)} total")
