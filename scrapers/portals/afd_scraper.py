# =============================================================================
# afd_scraper.py — AFD DGMarket Procurement Portal Pipeline
#
# Site   : https://afd.dgmarket.com  (AFD's official procurement portal)
# Method : requests + BeautifulSoup — pure HTML, ZERO JavaScript dependency
# Pages  : brandedNoticeList.do?page=N  (9 pages × 20 rows ≈ 170 notices)
# Detail : /tender/{id} — structured HTML with full eligibility + English text
#
# CRITICAL NOTE: afd.fr/calls-for-projects = grants for French NGOs (WRONG).
#                afd.dgmarket.com           = real consulting contracts (RIGHT).
#
# Pipeline:
#   1. List pages 1-9  → extract country, title, published, deadline, detail_url
#   2. Pre-filter      → drop HARD_REJECT (goods/construction/awards) at title level
#   3. Detail fetch    → parse notice_type, reference, buyer, eligibility, full text
#   4. Type filter     → skip "Avis d'attribution" (contract awards)
#   5. Score           → multilingual consulting keyword scoring (FR/EN/PT/ES)
#   6. Dedup + save    → MySQL + Excel
#
# Output: 30-45 consulting/advisory tenders from Africa, Asia, MENA per run
# =============================================================================

from __future__ import annotations

import os
import re
import time
import hashlib
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance

# ── Constants ──────────────────────────────────────────────────────────────────

BASE_URL     = "https://afd.dgmarket.com"
LIST_URL     = f"{BASE_URL}/tenders/brandedNoticeList.do"
DETAIL_URL   = f"{BASE_URL}/tender"
MAX_PAGES    = 9
# The tender list is always the 4th table on the page (index 3)
# Page 1 = base URL, pages 2+ = ?selPageNumber=N
_TENDER_TABLE_IDX = 3
DELAY        = 0.5   # seconds between requests — polite but fast
DETAIL_WORKERS = 5   # parallel detail-page fetches for NEW tenders

AFD_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "AFD_Tenders_Master.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
    "Referer":         BASE_URL,
}

# ── Multilingual consulting signals (FR / EN / PT / ES) ───────────────────────
# Keep a tender if ANY of these substrings appear in the title (case-insensitive)
_CONSULTING_SIGNALS: list[str] = [
    # French
    "consultant", "consultanc", "assistance technique", "formation profession",
    "évaluation", "evaluation", "étude ", "etude ", "accompagnement",
    "manifestation d'intérêt", "manifestation d'interet", "mission de suivi",
    "audit ", "conseil technique", "expertise", "appui technique",
    "recrutement d'un", "sélection d'un", "selection d'un",
    "prestation de service", "prestation intellectuelle",
    # English
    "consulting", "consultancy", "technical assistance", "advisory",
    "assessment", "capacity building", "expression of interest",
    "feasibility", "monitoring", "supervision of", "research study",
    # Portuguese
    "consultoria", "assistência técnica", "assessoria", "avaliação",
    # Spanish
    "consultoría", "asistencia técnica", "evaluación", "asesoría",
    "fiscalización",   # construction supervision — borderline but often needs consultants
]

# Hard reject at list-page level — these are clearly goods/works, skip detail fetch
_HARD_REJECT: list[str] = [
    "supply and delivery", "supply of ", "procurement of ",
    "fournitures", "acquisition de matériels", "acquisition et la livraison",
    "marché de fournitures", "achat de ",
    "travaux de construction", "travaux de réhabilitation",
    "réhabilitation et construction", "rénovation de ", "renovation de ",
    "installation of ", "installation et ", "pose de ",
    "geysers at", "insulation tester", "filter elements",
    "power box", "hard disc", "fuel oil", "intrants avicoles",
    "piézomètres", "piezometros", "heating pipe", "heating equipment",
    "contract award notice",  # already closed
]

# Notice types that are closed/irrelevant — skip after detail fetch
_SKIP_NOTICE_TYPES: list[str] = [
    "avis d'attribution",
    "contract award",
    "avis de marché passé",
    "avis général",               # general notice — no deadline, usually framework
]

# Excel column definitions
MASTER_COLUMNS = [
    ("Title",            62),
    ("Notice Type",      26),
    ("Country",          20),
    ("Reference",        30),
    ("Published",        14),
    ("Deadline",         18),
    ("Financing Agency", 35),
    ("Buyer",            40),
    ("Description",      70),
    ("Relevance",        42),
    ("Detail Link",      55),
]

# ── Excel styles ───────────────────────────────────────────────────────────────
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_ALT_FILL  = PatternFill("solid", fgColor="F5F8FF")
_REL_FILL  = PatternFill("solid", fgColor="E2EFDA")
_HDR_FONT  = Font(name="Arial", color="FFFFFF", bold=True, size=10)
_BODY_FONT = Font(name="Arial", size=9)
_HIGH_FONT = Font(name="Arial", size=9, color="375623", bold=True)
_GRAY_FONT = Font(name="Arial", size=9, color="999999")
_LINK_FONT = Font(name="Arial", size=9, color="1155CC", underline="single")
_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Section 1 — List page scraper
# =============================================================================

def _should_reject_title(title: str) -> bool:
    """Return True if the title clearly signals goods/works — skip detail fetch."""
    t = title.lower()
    return any(sig in t for sig in _HARD_REJECT)


def _could_be_consulting(title: str) -> bool:
    """Return True if title matches any consulting signal."""
    t = title.lower()
    return any(sig in t for sig in _CONSULTING_SIGNALS)


def _scrape_list_pages(session: requests.Session) -> list[dict]:
    """
    Scrape all list pages and return raw notice dicts with basic metadata.
    Pre-filter at title level to avoid unnecessary detail fetches.
    """
    notices = []
    seen_urls = set()

    for page_num in range(1, MAX_PAGES + 1):
        # Page 1 = base URL; pages 2+ use selPageNumber param
        url = LIST_URL if page_num == 1 else f"{LIST_URL}?selPageNumber={page_num}"
        print(f"[afd]   List page {page_num}/{MAX_PAGES} → {url}")
        try:
            r = session.get(url, timeout=20)
            r.raise_for_status()
        except Exception as e:
            print(f"[afd]   List page {page_num} error: {e}")
            time.sleep(DELAY * 2)
            continue

        soup = BeautifulSoup(r.text, "html.parser")

        # The tender list is the 4th table (index 3) — tables[0-2] are nav/search/pagination
        all_tables = soup.find_all("table")
        if len(all_tables) <= _TENDER_TABLE_IDX:
            print(f"[afd]   Tender table not found on page {page_num} — stopping")
            break
        table = all_tables[_TENDER_TABLE_IDX]

        rows = table.find_all("tr")
        page_count = 0

        for row in rows:
            cells = row.find_all("td")
            if len(cells) < 3:
                continue   # skip header row (has th/links but no td)

            country = cells[0].get_text(strip=True)

            # Title + link are in second cell
            title_cell = cells[1]
            link_tag   = title_cell.find("a")
            if not link_tag:
                continue

            title      = link_tag.get_text(strip=True)
            href       = link_tag.get("href", "")
            detail_url = (href if href.startswith("http")
                          else BASE_URL + href)

            if detail_url in seen_urls:
                continue
            seen_urls.add(detail_url)

            published = cells[2].get_text(strip=True) if len(cells) > 2 else ""
            deadline  = cells[3].get_text(strip=True) if len(cells) > 3 else ""

            # Pre-filter: hard reject at title level
            if _should_reject_title(title):
                print(f"[afd]     REJECT (title): {title[:60]}")
                continue

            notices.append({
                "title":      title,
                "country":    country,
                "published":  published,
                "deadline":   deadline,
                "detail_url": detail_url,
            })
            page_count += 1

        print(f"[afd]   Page {page_num}: {page_count} notices kept after title filter")

        # Check if there's a next page
        page_text = soup.get_text()
        total_match = re.search(r"1-\d+ de (\d+)", page_text)
        if total_match:
            total = int(total_match.group(1))
            if page_num * 20 >= total:
                print(f"[afd]   Reached end ({total} total notices)")
                break

        time.sleep(DELAY)

    return notices


# =============================================================================
# Section 2 — Detail page parser
# =============================================================================

def _parse_detail_page(soup: BeautifulSoup, url: str) -> dict:
    """
    Parse an AFD DGMarket detail page.
    Returns a dict with structured fields, or {} if invalid/award notice.
    """
    # Notice type (below title)
    notice_type = ""
    type_tags = soup.find_all(["p", "div", "span", "h3", "h4"])
    for tag in type_tags:
        text = tag.get_text(strip=True)
        if any(kw in text.lower() for kw in [
            "appel à manifestation", "appel d'offres", "avis d'attribution",
            "expression of interest", "request for proposal", "avis général",
            "contract award", "avis de marché"
        ]):
            notice_type = text
            break

    # Skip contract awards
    if any(skip in notice_type.lower() for skip in _SKIP_NOTICE_TYPES):
        return {}

    # Parse structured info table (key: value rows)
    info = {}
    # DGMarket uses a definition-list style or table for metadata
    for row in soup.find_all("tr"):
        cells = row.find_all("td")
        if len(cells) >= 2:
            key   = cells[0].get_text(strip=True).rstrip(":").strip()
            value = cells[1].get_text(" ", strip=True)
            if key:
                info[key] = value

    # Also try label/value pattern via text scanning
    full_text = soup.get_text(" ", strip=True)

    def _extract_field(patterns: list[str]) -> str:
        for pat in patterns:
            m = re.search(pat, full_text, re.I)
            if m:
                return m.group(1).strip()[:200]
        return ""

    reference = (
        info.get("Numéro de l'avis/du contrat", "")
        or info.get("Notice/Contract Number", "")
        or info.get("Référence", "")
        or _extract_field([
            r"Num[eé]ro de l.avis[^:]*:\s*([A-Z0-9/_\-\.]+)",
            r"Reference[^:]*:\s*([A-Z0-9/_\-\.]+)",
        ])
    )

    financing_agency = (
        info.get("Agence de financement", "")
        or info.get("Financing Agency", "")
        or "Agence Française de Développement"
    )

    buyer = (
        info.get("Acheteur", "")
        or info.get("Buyer", "")
        or info.get("Purchaser", "")
        or ""
    )

    # Eligibility / description — the richest text on the page
    desc_el = (
        soup.find("td", string=re.compile(r"Eligibilit", re.I))
        or soup.find(string=re.compile(r"Eligibilit", re.I))
    )
    description = ""
    if desc_el:
        parent = desc_el.parent if hasattr(desc_el, "parent") else None
        if parent:
            # Get the sibling td (the value cell)
            sibling = parent.find_next_sibling("td")
            if sibling:
                description = sibling.get_text(" ", strip=True)[:2000]

    # Fallback: grab the main content area
    if not description:
        main = soup.find("div", class_=re.compile(r"content|main|body", re.I))
        if main:
            description = main.get_text(" ", strip=True)[:2000]
        else:
            description = full_text[:2000]

    # Try to extract English version from "Texte original"
    eng_match = re.search(
        r"ENGLISH VERSION.*?(?:\n|\r\n?)(.*?)(?:VERSION FRANÇAISE|$)",
        full_text, re.I | re.S
    )
    if eng_match:
        eng_text = eng_match.group(1).strip()[:1500]
        if len(eng_text) > 100:
            description = eng_text   # prefer English over French

    return {
        "notice_type":      notice_type,
        "reference":        reference,
        "financing_agency": financing_agency,
        "buyer":            buyer,
        "description":      description,
    }


# =============================================================================
# Section 3 — Excel writer
# =============================================================================

def _save_excel(rows: list) -> None:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "AFD Tenders"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    col_names = [c[0] for c in MASTER_COLUMNS]
    for ci, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = col_width

    rel_idx  = col_names.index("Relevance")    + 1
    link_idx = col_names.index("Detail Link")  + 1

    for ri, row_data in enumerate(rows, 2):
        ws.row_dimensions[ri].height = 40
        alt = _ALT_FILL if ri % 2 == 0 else None

        for ci, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if ci == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = _LINK_FONT
            elif ci == rel_idx:
                if val:
                    cell.fill = _REL_FILL
                    cell.font = _HIGH_FONT
                else:
                    cell.font = _GRAY_FONT
                    if alt:
                        cell.fill = alt
            else:
                cell.font = _BODY_FONT
                if alt:
                    cell.fill = alt

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(AFD_EXCEL_PATH)
    print(f"[afd] Excel saved: {AFD_EXCEL_PATH}  ({len(rows)} rows)")


# =============================================================================
# Section 4 — Detail helper (for parallel fetching)
# =============================================================================

def _fetch_detail(session: requests.Session, notice: dict) -> dict | None:
    """
    Fetch and parse a single AFD detail page.
    Returns enriched row dict (with _tender_id key), or None if skippable.
    Called in parallel for new tenders only.
    """
    title      = notice["title"]
    detail_url = notice["detail_url"]
    try:
        r = session.get(detail_url, timeout=25)
        r.raise_for_status()
        soup   = BeautifulSoup(r.text, "html.parser")
        detail = _parse_detail_page(soup, detail_url)
    except Exception as e:
        print(f"[afd]   Detail fetch error ({title[:40]}): {e}")
        return None

    if not detail:
        return None   # contract award / general notice

    combined_text = title + " " + detail.get("description", "")
    if not _could_be_consulting(combined_text):
        return None   # no consulting signal in full text

    relevance = score_relevance(
        title,
        detail.get("description", "") + " " + detail.get("buyer", "")
    )

    return {
        "_tender_id":       notice["_tender_id"],
        "Title":            title,
        "Notice Type":      detail.get("notice_type", ""),
        "Country":          notice.get("country", ""),
        "Reference":        detail.get("reference", ""),
        "Published":        notice.get("published", ""),
        "Deadline":         notice.get("deadline", ""),
        "Financing Agency": detail.get("financing_agency", ""),
        "Buyer":            detail.get("buyer", ""),
        "Description":      detail.get("description", "")[:1500],
        "Relevance":        relevance,
        "Detail Link":      detail_url,
    }


# =============================================================================
# Section 5 — Main run()
# =============================================================================

def run() -> tuple:
    """
    Run the AFD DGMarket procurement pipeline.
    Returns (new_tenders: list, all_rows: list).

    Speed optimisations (v2):
      1. Build tender_id from URL slug at list-page stage — no detail fetch needed.
      2. Skip detail fetch entirely for already-seen tenders (add slim row instead).
         On steady-state runs most tenders are seen → 80%+ of fetches skipped.
      3. Parallel detail fetch (DETAIL_WORKERS) for new tenders only.
      4. DELAY reduced from 1.2 s → 0.5 s.
    Result: ~228 s → ~25-40 s on repeat runs; ~60-80 s on first run.
    """
    print("\n" + "=" * 65)
    print("[afd] AFD DGMarket Procurement Pipeline starting...")
    print(f"[afd] Source: {BASE_URL}")
    print("=" * 65)

    new_tenders: list = []
    all_rows:    list = []

    session = requests.Session()
    session.headers.update(HEADERS)

    # ── Phase 1: Scrape all list pages ────────────────────────────────────────
    print(f"\n[afd] Phase 1: Scraping list pages (max {MAX_PAGES} pages)...")
    notices = _scrape_list_pages(session)
    print(f"[afd] Phase 1 done: {len(notices)} notices after title pre-filter")

    # ── Phase 1b: Tag each notice with its tender_id (slug from URL) ──────────
    run_seen: set[str] = set()
    unique_notices: list[dict] = []
    for notice in notices:
        slug = notice["detail_url"].rstrip("/").split("/")[-1]
        tid  = f"AFD_{slug}"
        if tid not in run_seen:
            run_seen.add(tid)
            notice["_tender_id"] = tid
            unique_notices.append(notice)
    notices = unique_notices

    # ── Phase 2: Split NEW vs already seen ────────────────────────────────────
    new_notices:  list[dict] = []
    seen_notices: list[dict] = []
    for notice in notices:
        if check_if_new(notice["_tender_id"]):
            new_notices.append(notice)
        else:
            seen_notices.append(notice)

    print(
        f"[afd] Phase 2: {len(new_notices)} NEW (fetch detail), "
        f"{len(seen_notices)} seen (skip detail)"
    )

    # Seen tenders — slim row, no network request needed
    for notice in seen_notices:
        relevance = score_relevance(notice["title"], "")
        all_rows.append({
            "Title":            notice["title"],
            "Notice Type":      "",
            "Country":          notice.get("country", ""),
            "Reference":        "",
            "Published":        notice.get("published", ""),
            "Deadline":         notice.get("deadline", ""),
            "Financing Agency": "Agence Française de Développement",
            "Buyer":            "",
            "Description":      "",
            "Relevance":        relevance,
            "Detail Link":      notice["detail_url"],
        })

    # ── Phase 3: Parallel detail fetch for NEW tenders ────────────────────────
    if new_notices:
        print(f"[afd] Phase 3: Fetching {len(new_notices)} new detail pages "
              f"({DETAIL_WORKERS} parallel workers)...")
        with ThreadPoolExecutor(max_workers=DETAIL_WORKERS) as pool:
            futures = {
                pool.submit(_fetch_detail, session, notice): notice
                for notice in new_notices
            }
            for future in as_completed(futures):
                try:
                    row = future.result()
                except Exception as exc:
                    print(f"[afd]   Worker error: {exc}")
                    row = None

                if row is None:
                    continue

                tid = row.pop("_tender_id")
                all_rows.append(row)
                mark_as_seen(tid, title=row["Title"],
                             source_site="AFD", url=row["Detail Link"])
                new_tenders.append({
                    "title":    row["Title"],
                    "deadline": row["Deadline"],
                    "value":    row["Buyer"] or "AFD-funded project",
                    "url":      row["Detail Link"],
                })
                print(f"[afd]   → NEW: {row['Title'][:65]}")

    # ── Phase 4: Save ─────────────────────────────────────────────────────────
    if all_rows:
        _save_excel(all_rows)
    else:
        print("[afd] WARNING: No rows to save — check connectivity or filters")

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(f"\n[afd] Done — {len(all_rows)} tenders "
          f"({len(new_tenders)} NEW, {relevant} relevant to IDCG)")
    return new_tenders, all_rows


if __name__ == "__main__":
    new, rows = run()
    print(f"\nResult: {len(rows)} total, {len(new)} new")
