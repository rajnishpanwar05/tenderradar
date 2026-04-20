# =============================================================================
# cg_scraper.py — Central Government eProcure (eprocure.gov.in) Pipeline
#
# Site: https://eprocure.gov.in/eprocure/app
# Approach: requests.Session + "Tenders by Organisation" listing page
#   → 241 org links accessible WITHOUT CAPTCHA within same session.
#   → Each org page returns full tender list with title, dates, tender ID.
# No Selenium, no login required.
# =============================================================================

import os
import re
import time
import random
import base64
import json
from urllib.parse import urljoin
from typing import Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config.config import LOG_FILE, PORTAL_EXCELS_DIR
from database.db import check_if_new, mark_as_seen
from intelligence.keywords import score_relevance
from scrapers.portals.nic_detail_utils import (
    enrich_nic_row_with_detail,
    seed_nic_listing_metadata,
)

# ── Registry metadata (overrides static defaults) ────────────────────────────
SCRAPER_META = {
    "flag":        "cg",
    "label":       "CG eProcurement",
    "group":       "requests",
    "timeout":     900,   # CG can take 7–12 minutes depending on org count
    "max_retries": 1,
    "auto":        True,
}

# ── Config ─────────────────────────────────────────────────────────────────────
BASE_URL     = "https://eprocure.gov.in/eprocure/app"
ORG_LIST_URL = f"{BASE_URL}?page=FrontEndTendersByOrganisation&service=page"

CG_EXCEL_PATH = os.path.join(PORTAL_EXCELS_DIR, "CG_Central_eProcure_Tenders.xlsx")

REQUEST_DELAY   = 0.8   # seconds between requests (polite scraping)
MAX_ORGS        = 250   # safety cap on number of orgs to visit
MAX_PAGES_PER_ORG = 10  # page cap per org (most have 1-2 pages)
RETRY_ATTEMPTS  = 2
MAX_DETAIL_ENRICH_PER_ORG = 60  # raise ceiling so more tenders carry detail/doc links for deep extraction
CAPTCHA_API_KEY = os.environ.get("CAPTCHA_API_KEY", "")
CAPTCHA_TIMEOUT  = 90
CAPTCHA_IN_URL = "https://2captcha.com/in.php"
CAPTCHA_RES_URL = "https://2captcha.com/res.php"
CAPTCHA_POLL_INTERVAL = 5
CAPTCHA_POLL_MAX = 20

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

# ── Excel styles ───────────────────────────────────────────────────────────────
MASTER_COLUMNS = [
    ("Tender ID",       28),
    ("Title",           65),
    ("Organisation",    40),
    ("Published Date",  20),
    ("Closing Date",    20),
    ("Opening Date",    20),
    ("Relevance",       42),
    ("URL",             55),
]
HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
ALT_FILL       = PatternFill("solid", fgColor="F5F8FF")
RELEVANCE_FILL = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT     = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BODY_FONT      = Font(name="Calibri", size=10)
RELEVANCE_FONT = Font(name="Calibri", size=10, color="375623", bold=True)
NO_REL_FONT    = Font(name="Calibri", size=10, color="999999")
THIN_BORDER    = Border(
    left=Side(style="thin", color="D0D7E3"),
    right=Side(style="thin", color="D0D7E3"),
    top=Side(style="thin", color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)


# =============================================================================
# Parsing helpers
# =============================================================================

def _extract_brackets(text: str):
    """
    eprocure.gov.in stores tender cell as:
      [Title Text] [Ref No][Tender ID]
    or
      [Title Text][Tender ID]

    The ref no may contain parentheses inside square brackets, e.g.:
      [XX-191/SO(DO)/Pharma/25-26/ME]

    We match only top-level SQUARE brackets (greedy, allowing any content
    including nested parens) to avoid splitting on internal parens.

    Returns (title, ref_no, tender_id).
    """
    # Match square bracket groups — allow any content including parens inside
    bracket_groups = re.findall(r"\[([^\[\]]+)\]", text)
    # Remove all [..] groups from the string to get clean title
    clean_title = re.sub(r"\[[^\[\]]+\]", "", text).strip()
    clean_title = re.sub(r"\s+", " ", clean_title).strip()

    ref_no    = ""
    tender_id = ""
    if len(bracket_groups) >= 3:
        # [Title][Ref No][Tender ID]
        ref_no    = bracket_groups[-2].strip()
        tender_id = bracket_groups[-1].strip()
        # title is the first bracket group (since clean_title would be empty)
        if not clean_title or len(clean_title) < 5:
            clean_title = bracket_groups[0].strip()
    elif len(bracket_groups) == 2:
        ref_no    = bracket_groups[0].strip()
        tender_id = bracket_groups[1].strip()
        if not clean_title or len(clean_title) < 5:
            clean_title = bracket_groups[0].strip()
    elif len(bracket_groups) == 1:
        tender_id = bracket_groups[0].strip()

    return clean_title, ref_no, tender_id


def _parse_org_tender_table(html: str, org_name: str):
    """
    Parse the tender table on an org page.
    Table columns (fixed positions):
      0: S.No  1: e-Published Date  2: Closing Date  3: Opening Date
      4: Title and Ref.No./Tender ID  5: Organisation Chain
    Title cell format: [Title Text] [Ref No][Tender ID]
    Returns list of dicts with keys matching MASTER_COLUMNS.
    """
    soup    = BeautifulSoup(html, "html.parser")
    tenders = []

    # Find the exact header row
    EXACT_HEADERS = ["S.No", "e-Published Date", "Closing Date", "Opening Date",
                     "Title and Ref.No./Tender ID", "Organisation Chain"]

    target_table = None
    for row in soup.find_all("tr"):
        cells = row.find_all(["th", "td"])
        if [c.get_text(strip=True) for c in cells] == EXACT_HEADERS:
            target_table = row.find_parent("table")
            break

    if target_table is None:
        return tenders

    rows = target_table.find_all("tr")
    header_row = None
    for i, row in enumerate(rows):
        cells = row.find_all(["th", "td"])
        if [c.get_text(strip=True) for c in cells] == EXACT_HEADERS:
            header_row = i
            break

    if header_row is None:
        return tenders

    for row in rows[header_row + 1:]:
        cells = row.find_all(["td", "th"])
        if len(cells) < 5:
            continue

        raw_title_cell = cells[4].get_text(" ", strip=True)
        if not raw_title_cell or len(raw_title_cell) < 5:
            continue

        title, ref_no, tender_id = _extract_brackets(raw_title_cell)

        # If title is empty after stripping brackets, skip
        if not title or len(title) < 5:
            continue

        # Detail URL from link in title cell
        detail_url = BASE_URL
        link = cells[4].find("a", href=True)
        if link:
            href = link["href"]
            if href.startswith("http"):
                detail_url = href
            elif href.startswith("/"):
                detail_url = "https://eprocure.gov.in" + href
            else:
                detail_url = "https://eprocure.gov.in/eprocure/" + href.lstrip("/")

        org_chain = cells[5].get_text(" ", strip=True) if len(cells) > 5 else org_name
        if not org_chain:
            org_chain = org_name

        tenders.append({
            "Tender ID":      tender_id or ref_no or f"eprocure/{title[:40]}",
            "Title":          title,
            "Organisation":   org_chain,
            "Published Date": cells[1].get_text(strip=True),
            "Closing Date":   cells[2].get_text(strip=True),
            "Opening Date":   cells[3].get_text(strip=True),
            "URL":            detail_url,
        })

    return tenders


def _get_next_page_url(html: str, current_url: str):
    """
    Look for a 'Next' pagination link on the page.
    Returns the next URL string or None.
    """
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        text = a.get_text(strip=True).lower()
        if text in ("next", "next »", ">", ">>", "next page"):
            href = a["href"]
            if href.startswith("http"):
                return href
            if href.startswith("?"):
                return BASE_URL + href
            if href.startswith("/"):
                return "https://eprocure.gov.in" + href
    return None


# =============================================================================
# Session / fetching helpers
# =============================================================================

def _make_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def _two_captcha_submit(payload: dict) -> str:
    try:
        resp = requests.post(CAPTCHA_IN_URL, data=payload, timeout=60)
        data = resp.json() if resp.ok else {}
        if not data or data.get("status") != 1:
            print(f"[cg] 2captcha submit failed: {data}")
            return ""
        return str(data.get("request") or "")
    except Exception as exc:
        print(f"[cg] 2captcha submit error: {exc}")
        return ""


def _two_captcha_poll(captcha_id: str) -> str:
    if not captcha_id:
        return ""
    for _ in range(CAPTCHA_POLL_MAX):
        try:
            time.sleep(CAPTCHA_POLL_INTERVAL)
            resp = requests.get(
                CAPTCHA_RES_URL,
                params={"key": CAPTCHA_API_KEY, "action": "get", "id": captcha_id, "json": 1},
                timeout=30,
            )
            data = resp.json() if resp.ok else {}
            if not data:
                continue
            if data.get("status") == 1:
                return str(data.get("request") or "")
            if data.get("request") not in ("CAPCHA_NOT_READY", "CAPTCHA_NOT_READY"):
                print(f"[cg] 2captcha poll failed: {data}")
                return ""
        except Exception:
            continue
    print("[cg] 2captcha poll timeout.")
    return ""


def _looks_like_captcha(html: str) -> bool:
    if not html:
        return False
    lower = html.lower()
    if "g-recaptcha" in lower or "hcaptcha" in lower:
        return True
    if "captcha" in lower:
        return True
    return False


def _detect_sitekey(html: str) -> tuple[str, str]:
    """Return (provider, sitekey) where provider in {'recaptcha','hcaptcha','funcaptcha'}."""
    if not html:
        return "", ""
    soup = BeautifulSoup(html, "html.parser")
    # reCAPTCHA / hCaptcha use data-sitekey
    node = soup.find(attrs={"data-sitekey": True})
    if node:
        sitekey = str(node.get("data-sitekey") or "").strip()
        if "hcaptcha" in html.lower():
            return "hcaptcha", sitekey
        return "recaptcha", sitekey
    # FunCaptcha / Arkose labs
    node = soup.find(attrs={"data-pkey": True})
    if node:
        return "funcaptcha", str(node.get("data-pkey") or "").strip()
    return "", ""


def _solve_image_captcha(session: requests.Session, html: str, page_url: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    img = None
    for cand in soup.find_all("img"):
        src = str(cand.get("src") or "")
        if "captcha" in src.lower():
            img = cand
            break
    if not img:
        return ""
    img_url = urljoin(page_url, img.get("src"))
    try:
        resp = session.get(img_url, timeout=30)
        resp.raise_for_status()
        b64 = base64.b64encode(resp.content).decode("ascii")
    except Exception:
        return ""

    captcha_id = _two_captcha_submit({
        "key": CAPTCHA_API_KEY,
        "method": "base64",
        "body": b64,
        "json": 1,
    })
    return _two_captcha_poll(captcha_id)


def _submit_captcha_form(session: requests.Session, html: str, page_url: str, captcha_value: str, provider: str) -> Optional[str]:
    if not captcha_value:
        return None
    soup = BeautifulSoup(html, "html.parser")
    form = soup.find("form")
    if not form:
        return None
    action = form.get("action") or page_url
    action_url = urljoin(page_url, action)

    payload = {}
    for inp in form.find_all("input"):
        name = inp.get("name")
        if not name:
            continue
        payload[name] = inp.get("value") or ""

    # Try to inject the captcha answer into an existing field
    injected = False
    for key in list(payload.keys()):
        if "captcha" in key.lower():
            payload[key] = captcha_value
            injected = True
            break

    # reCAPTCHA / hCaptcha uses a textarea named g-recaptcha-response/h-captcha-response
    if not injected and provider in {"recaptcha", "hcaptcha"}:
        payload["g-recaptcha-response"] = captcha_value
        payload["h-captcha-response"] = captcha_value
        injected = True

    if not injected:
        return None

    try:
        resp = session.post(action_url, data=payload, timeout=30)
        if resp.ok and resp.text:
            return resp.text
    except Exception:
        return None
    return None


def _solve_captcha_if_present(session: requests.Session, page_url: str, html: str) -> str:
    """
    Best-effort CAPTCHA solve flow. Returns updated HTML if solved, else original.
    """
    if not CAPTCHA_API_KEY:
        return html
    if not _looks_like_captcha(html):
        return html

    provider, sitekey = _detect_sitekey(html)
    captcha_value = ""

    if provider in {"recaptcha", "hcaptcha"} and sitekey:
        captcha_id = _two_captcha_submit({
            "key": CAPTCHA_API_KEY,
            "method": "userrecaptcha" if provider == "recaptcha" else "hcaptcha",
            "googlekey": sitekey,
            "pageurl": page_url,
            "json": 1,
        })
        captcha_value = _two_captcha_poll(captcha_id)
    elif provider == "funcaptcha" and sitekey:
        captcha_id = _two_captcha_submit({
            "key": CAPTCHA_API_KEY,
            "method": "funcaptcha",
            "publickey": sitekey,
            "pageurl": page_url,
            "json": 1,
        })
        captcha_value = _two_captcha_poll(captcha_id)
    else:
        captcha_value = _solve_image_captcha(session, html, page_url)

    if not captcha_value:
        return html

    updated = _submit_captcha_form(session, html, page_url, captcha_value, provider)
    return updated or html


def _get(session, url, retries=RETRY_ATTEMPTS):
    for attempt in range(retries):
        try:
            r = session.get(url, timeout=30)
            r.raise_for_status()
            return r
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
            else:
                print(f"[cg]   GET failed ({url[:80]}): {e}", flush=True)
                return None
    return None


def _fetch_org_links(session):
    """
    Fetch the 'Tenders by Organisation' page and extract all org links.
    The page has a table with headers: S.No | Organisation Name | Tender Count
    The link (sp= URL) is on the Tender Count cell; org name is in cell[1].
    Returns list of (org_name, sp_url) tuples.
    """
    print("[cg] Fetching Tenders-by-Organisation listing...", flush=True)
    r = _get(session, ORG_LIST_URL)
    if not r:
        print("[cg] Failed to load org listing page.", flush=True)
        return []

    html = _solve_captcha_if_present(session, ORG_LIST_URL, r.text)
    soup = BeautifulSoup(html, "html.parser")
    orgs = []

    # Find the table whose first row is exactly: S.No | Organisation Name | Tender Count
    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        if not rows:
            continue
        header_cells = rows[0].find_all(["th", "td"])
        if [c.get_text(strip=True) for c in header_cells] != ["S.No", "Organisation Name", "Tender Count"]:
            continue

        # Found the right table — parse data rows
        for row in rows[1:]:
            cells = row.find_all(["td", "th"])
            if len(cells) < 3:
                continue
            org_name = cells[1].get_text(strip=True)
            if not org_name:
                continue
            # The link is inside the Tender Count cell (cell[2])
            link = cells[2].find("a", href=True)
            if not link:
                continue
            href = link["href"]
            if href.startswith("http"):
                url = href
            elif href.startswith("/"):
                url = "https://eprocure.gov.in" + href
            else:
                url = "https://eprocure.gov.in" + "/" + href.lstrip("/")
            orgs.append((org_name, url))
        break

    print(f"[cg] Found {len(orgs)} organisation links.", flush=True)
    return orgs[:MAX_ORGS]


# =============================================================================
# Excel output
# =============================================================================

def _save_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Central eProcure Tenders"
    col_names = [c[0] for c in MASTER_COLUMNS]
    ws.row_dimensions[1].height = 36

    for col_idx, (col_name, col_width) in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = WHITE_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.freeze_panes = ws.cell(row=2, column=1)
    relevance_idx   = col_names.index("Relevance") + 1
    link_idx        = col_names.index("URL") + 1

    for row_idx, row_data in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 50
        alt_fill = ALT_FILL if row_idx % 2 == 0 else None

        for col_idx, col_name in enumerate(col_names, 1):
            val  = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

            if col_idx == link_idx and val and str(val).startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="1155CC", underline="single")
            elif col_idx == relevance_idx:
                if val:
                    cell.fill = RELEVANCE_FILL
                    cell.font = RELEVANCE_FONT
                else:
                    if alt_fill:
                        cell.fill = alt_fill
                    cell.font = NO_REL_FONT
            else:
                cell.font = BODY_FONT
                if alt_fill:
                    cell.fill = alt_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(len(MASTER_COLUMNS))}1"
    wb.save(CG_EXCEL_PATH)
    print(f"[cg] Excel saved: {CG_EXCEL_PATH}  ({len(rows)} rows)", flush=True)


# =============================================================================
# run() — pipeline entry point
# =============================================================================

def run():
    print("\n" + "=" * 65, flush=True)
    print("[cg] Central Government eProcure (eprocure.gov.in) Pipeline", flush=True)
    print("=" * 65, flush=True)

    new_tenders = []
    all_rows    = []

    # Clear old Excel
    if os.path.exists(CG_EXCEL_PATH):
        try:
            os.remove(CG_EXCEL_PATH)
        except Exception:
            pass

    session = _make_session()

    # ── Step 1: Warm up session by hitting the home page ──────────────────────
    print("[cg] Warming up session...", flush=True)
    _get(session, "https://eprocure.gov.in/eprocure/app")
    time.sleep(1)

    # ── Step 2: Fetch org listing ─────────────────────────────────────────────
    org_links = _fetch_org_links(session)
    if not org_links:
        print("[cg] No org links found — aborting.", flush=True)
        return new_tenders, all_rows

    # ── Step 3: Visit each org and collect tenders ────────────────────────────
    total_scraped = 0
    for org_idx, (org_name, org_url) in enumerate(org_links, 1):
        print(f"[cg]  [{org_idx}/{len(org_links)}] {org_name[:60]}", flush=True)

        page_url = org_url
        page_num = 1
        org_tenders = []
        org_detail_enriched = 0

        while page_url and page_num <= MAX_PAGES_PER_ORG:
            r = _get(session, page_url)
            if not r:
                break

            page_tenders = _parse_org_tender_table(r.text, org_name)
            if page_tenders:
                print(f"[cg]    page {page_num}: {len(page_tenders)} tenders", flush=True)
                org_tenders.extend(page_tenders)
            else:
                # No tenders on this page → stop paginating this org
                break

            next_url = _get_next_page_url(r.text, page_url)
            if next_url and next_url != page_url:
                page_url = next_url
                page_num += 1
                time.sleep(REQUEST_DELAY)
            else:
                break

        for t in org_tenders:
            seed_nic_listing_metadata(t, "cg")
            combined = " ".join(filter(None, [
                t["Title"],
                t.get("Organisation", ""),
                t.get("Description", ""),
            ]))
            t["Relevance"] = score_relevance(combined)
            tender_id  = t["Tender ID"]
            detail_url = t.get("URL", BASE_URL)
            is_new = check_if_new(tender_id)

            if (t.get("Relevance") or is_new) and org_detail_enriched < MAX_DETAIL_ENRICH_PER_ORG:
                enrich_nic_row_with_detail(session, t)
                org_detail_enriched += 1

            all_rows.append(t)
            total_scraped += 1

            if is_new:
                mark_as_seen(tender_id, title=t["Title"], source_site="CG", url=detail_url)
                # Only notify for keyword-relevant tenders (avoid flooding)
                if t.get("Relevance"):
                    new_tenders.append({
                        "tender_id": tender_id,
                        "title":    t["Title"],
                        "description": t.get("Description", ""),
                        "organization": t.get("Organisation", ""),
                        "deadline": t.get("Closing Date", ""),
                        "value":    t.get("Value", ""),
                        "url":      detail_url,
                        "source_portal": "cg",
                    })
                    print(f"[cg]     → NEW + RELEVANT | {t['Relevance'][:60]}", flush=True)
                else:
                    print(f"[cg]     → NEW (saved to Excel only)", flush=True)
            # else: already seen, no print to keep logs clean

        time.sleep(REQUEST_DELAY + random.uniform(0, 0.4))

    if all_rows:
        _save_excel(all_rows)

    relevant = sum(1 for r in all_rows if r.get("Relevance"))
    print(
        f"\n[cg] Done — {len(all_rows)} tenders scraped across {len(org_links)} orgs, "
        f"{len(new_tenders)} NEW+RELEVANT, {relevant} relevant total",
        flush=True,
    )
    return new_tenders, all_rows
